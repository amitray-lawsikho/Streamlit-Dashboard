import streamlit as st
import pandas as pd
import os
import pytz
import io
from datetime import datetime, date, time, timedelta
from google.cloud import bigquery
from google.oauth2 import service_account
import warnings
warnings.filterwarnings(
    "ignore",
    message="Please replace `st.components.v1.html` with `st.iframe`"
)
# ReportLab imports (used by both dashboards)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Flowable
from reportlab.lib.enums import TA_CENTER

# Openpyxl imports (used by Revenue dashboard)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client
import re

# --- GLOBAL CONFIG & CREDENTIALS ---
st.set_page_config(
    page_title="Lead Metrics — LawSikho",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"    
)

def get_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=credentials, project=info["project_id"])
    else:
        SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
        if os.path.exists(SERVICE_ACCOUNT_FILE):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
            return bigquery.Client()
        return None

@st.cache_resource
def get_cached_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=credentials, project=info["project_id"])
    else:
        SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
        if os.path.exists(SERVICE_ACCOUNT_FILE):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
            return bigquery.Client()
        return None

client = get_cached_bq_client()

# --- LOGIN MODULE ---

def _apply_role_filters():
    ri = st.session_state.get('auth_role_info', {'role': 'admin'})
    role    = ri.get('role', 'admin')
    teams   = ri.get('teams')      # list | None
    callers = ri.get('callers')    # list | None
    cname   = ri.get('caller_name')

    st.session_state['rf_role']        = role
    st.session_state['rf_teams']       = teams   or []
    st.session_state['rf_callers']     = callers or []
    st.session_state['rf_caller_name'] = cname   or ''
    # Roles that see team/vertical multiselects in the sidebar
    st.session_state['rf_full_filters']= role == 'admin'

# ── Hardcoded access tiers ─────────────────────────────────────
ADMIN_EMAILS = {
    "admin@lawsikho.in",
    "yash@lawsikho.in",
    "akanksha.patil@lawsikho.in",
    "devki.dt@lawsikho.in",
    "mansi.gupta@lawsikho.in",
    "shivamtejpal@lawsikho.in",
    "pratik.s@lawsikho.in",
    "suraj@lawsikho.in",
    "parul.nagar@lawsikho.in",
    "amitray@lawsikho.in",
    "rinku@lawsikho.in",
    "karunakarareddy@lawsikho.in"
    "priyansh.s@lawsikho.in"
}

VERTICAL_HEAD_TEAMS = {
    # email               : [list of Team Names they can see]
    "uzair@lawsikho.in"       : ["19th Jan US acc/women ai Closure Batch","CD Closures - Inayat","Contract Drafting","Corporate law - Anas",
                                 "Corporate law - Jyoti","Elite","Law firm trainees - Anas","US Acc Closure Specialist - Inayat","US Accounting - Inayat",
                                 "US accounting closures - Sana","US accounting trainees","Women ai Trainee - Umme","Women ai/CD closure"],

    "shivya.p@lawsikho.in"    : ["ID - 2","ID - 4","ID - 8","ID - 9"],
    
    "mayur@lawsikho.in"       : ["Changemakers"],
    "deepansi@lawsikho.in"    : ["US Accounting","US accounting - Closures"],
    "darshan.c@lawsikho.in"   : ["ID Closure"],
    "anmol.g@lawsikho.in"     : ["DSV- Aditya","DSV- Shivam","ID Closure - Anmol","Women ai"],
    "abhipsa@lawsikho.in"     : ["CD - Community","CD - Community Manager","Criminal - Community","Criminal - Community Manager","ID - Community","ID - Community Manager"],
}

AUTH_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf"
    "/pub?gid=0&single=true&output=csv"
)

# Sheet column names — update if your sheet differs
_COL_NAME    = "Caller Name"
_COL_EMAIL   = "Email id"
_COL_DESIG   = "Academic Counselor/TL/ATL"
_COL_TEAM    = "Team Name"
_COL_TRAINER = "Sales Leader"
_TL_VALS     = {"TL", "ATL", "AD", "TEAM LEAD", "TEAM LEADER"}


@st.cache_resource
def get_supabase():
    # Try top-level keys first, then nested under a section
    url = (
        st.secrets.get("SUPABASE_URL")
        or st.secrets.get("supabase", {}).get("url")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_URL")
        or ""
    )
    key = (
        st.secrets.get("SUPABASE_KEY")
        or st.secrets.get("supabase", {}).get("key")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_KEY")
        or ""
    )

    if not url or not key:
        st.error(
            "⚠️ Supabase credentials missing from secrets. "
            "Add SUPABASE_URL and SUPABASE_KEY to your secrets.toml at the TOP LEVEL "
            "(before any [section] header)."
        )
        st.stop()

    return create_client(url, key)

supa = get_supabase()


@st.cache_data(ttl=300, show_spinner=False)
def load_auth_sheet() -> pd.DataFrame:
    df = pd.read_csv(AUTH_SHEET_URL)
    df.columns = df.columns.str.strip()
    if _COL_EMAIL in df.columns:
        df['_email_norm'] = df[_COL_EMAIL].astype(str).str.strip().str.lower()
    return df


def _extract_trainer_email(cell: str) -> str | None:
    """Parses 'Name (email@domain.com)' → 'email@domain.com'"""
    m = re.search(r'\(([^)\s]+@[^)\s]+)\)', str(cell))
    return m.group(1).strip().lower() if m else None


def determine_role(email: str, df: pd.DataFrame) -> dict | None:
   
    el = email.strip().lower()

    # 1 ── Admin
    if el in {e.lower() for e in ADMIN_EMAILS}:
        return {'role': 'admin', 'teams': None, 'callers': None,
                'caller_name': None, 'display_name': email}

    # 2 ── Vertical Head (hardcoded)
    for vh_mail, vh_teams in VERTICAL_HEAD_TEAMS.items():
        if vh_mail.lower() == el:
            return {'role': 'vertical_head', 'teams': vh_teams, 'callers': None,
                    'caller_name': None, 'display_name': email}

    if '_email_norm' not in df.columns:
        return None

    # 3 ── Trainer  (Sales Leader column)
    if _COL_TRAINER in df.columns:
        df2 = df.copy()
        df2['_tr_email'] = df2[_COL_TRAINER].apply(_extract_trainer_email)
        trainer_rows = df2[df2['_tr_email'] == el]
        if not trainer_rows.empty:
            teams   = trainer_rows[_COL_TEAM].dropna().unique().tolist() if _COL_TEAM in trainer_rows.columns else []
            callers = trainer_rows[_COL_NAME].dropna().unique().tolist() if _COL_NAME in trainer_rows.columns else []
        
            # ── Extract trainer's OWN name from "Name (email)" format in Sales Leader cell ──
            _sl_cell   = str(trainer_rows[_COL_TRAINER].iloc[0])
            _name_match = re.match(r'^([^(]+)\s*\(', _sl_cell)
            _tr_name   = _name_match.group(1).strip() if _name_match else None
        
            # ── Fallback: look up trainer's own row by email ──
            if not _tr_name:
                _own_rows = df[df['_email_norm'] == el]
                _tr_name  = _own_rows.iloc[0][_COL_NAME] if not _own_rows.empty and _COL_NAME in _own_rows.columns else email
        
            return {'role': 'trainer', 'teams': teams, 'callers': callers,
                    'caller_name': None, 'display_name': _tr_name}

    user_rows = df[df['_email_norm'] == el]
    if user_rows.empty:
        return None  # email not in sheet → not authorised

    # 4 ── TL / AD / ATL
    if _COL_DESIG in user_rows.columns:
        tl_rows = user_rows[
            user_rows[_COL_DESIG].astype(str).str.strip().str.upper().isin(_TL_VALS)
        ]
        if not tl_rows.empty:
            team = tl_rows.iloc[0][_COL_TEAM] if _COL_TEAM in tl_rows.columns else None
            team_callers = (
                df[df[_COL_TEAM] == team][_COL_NAME].dropna().unique().tolist()
                if team else []
            )
            disp = tl_rows.iloc[0][_COL_NAME] if _COL_NAME in tl_rows.columns else email
            return {'role': 'tl', 'teams': [team] if team else [], 'callers': team_callers,
                    'caller_name': None, 'display_name': disp}

    # 5 ── Regular Caller
    caller_name = user_rows.iloc[0][_COL_NAME] if _COL_NAME in user_rows.columns else email
    return {'role': 'caller', 'teams': None, 'callers': None,
            'caller_name': caller_name, 'display_name': caller_name}


# ──────────────────────────────────────────────────────────────
# AUTH UI HELPERS
# ──────────────────────────────────────────────────────────────

def _complete_login(email: str, session):
    """Resolve role and store everything in session state."""
    df_auth   = load_auth_sheet()
    role_info = determine_role(email, df_auth)
    if role_info is None:
        st.error("⛔ Your email is not authorised to access this dashboard.")
        return
    st.session_state['password_correct'] = True
    st.session_state['current_user']     = email
    st.session_state['supabase_session'] = session
    st.session_state['auth_role_info']   = role_info
    st.rerun()


def _auth_sign_in_panel():
    """Email + Password login panel."""
    email = st.text_input("Email", key="si_email", placeholder="Your mail ID")
    pwd   = st.text_input("Password", type="password", key="si_pwd", placeholder="Your Password")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Sign In →", key="si_btn", width='stretch'):
            if not email or not pwd:
                st.error("Enter both email and password."); return
            try:
                resp = supa.auth.sign_in_with_password({"email": email, "password": pwd})
                _complete_login(resp.user.email, resp.session)
            except Exception as ex:
                st.error(f"Login failed: {ex}")
    with c2:
        if st.button("Forgot / Change Password", key="si_otp_switch", width='stretch'):
            st.session_state['auth_tab']           = "otp"
            st.session_state['otp_prefill_email']  = email
            st.session_state['otp_step']           = 1
            st.rerun()


def _auth_otp_panel():
    """OTP verification + password-set panel (first-time or reset)."""
    step = st.session_state.get('otp_step', 1)

    # ── Step 1: request OTP ──────────────────────────────────
    if step == 1:
        email = st.text_input(
            "Email", key="otp_email",
            value=st.session_state.get('otp_prefill_email', ''),
            placeholder="your@lawsikho.in"
        )
        if st.button("Send OTP →", key="otp_send_btn", width='stretch'):
            if not email:
                st.error("Enter your email."); return
            # Pre-check: is the email in the sheet?
            role_check = determine_role(email, load_auth_sheet())
            if role_check is None:
                st.error("⛔ This email is not authorised."); return
            try:
                supa.auth.sign_in_with_otp({
                    "email": email,
                    "options": {"should_create_user": True}
                })
                st.session_state['otp_step']          = 2
                st.session_state['otp_pending_email'] = email
                st.rerun()
            except Exception as ex:
                st.error(f"Could not send OTP: {ex}")

        if st.button("← Back to Sign In", key="otp_back1", width='stretch'):
            st.session_state['auth_tab'] = "signin"
            st.rerun()

    # ── Step 2: verify OTP + set password ───────────────────
    elif step == 2:
        pending_email = st.session_state.get('otp_pending_email', '')
        st.success(f"OTP sent to **{pending_email}** — check your inbox (also spam).")

        otp = st.text_input("OTP code from email", key="otp_code", max_chars=8, placeholder="OTP Received on Email")
        pw1  = st.text_input("Set new password",    type="password", key="otp_pw1")
        pw2  = st.text_input("Confirm password",    type="password", key="otp_pw2")

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Verify & Continue →", key="otp_verify_btn", width='stretch'):
                if len(otp) != 8:
                    st.error("Enter the 8-digit code."); return
                if pw1 != pw2:
                    st.error("Passwords don't match."); return
                if len(pw1) < 8:
                    st.error("Password must be ≥ 8 characters."); return
                try:
                    resp = supa.auth.verify_otp({
                        "email": pending_email,
                        "token": otp,
                        "type": "magiclink"
                    })
                    # Save the password (user is now logged in to Supabase)
                    supa.auth.update_user({"password": pw1})
                    st.session_state['otp_step'] = 1
                    _complete_login(resp.user.email, resp.session)
                except Exception as ex:
                    st.error(f"Verification failed — wrong code or it expired: {ex}")
        with c2:
            if st.button("← Back", key="otp_back2", width='stretch'):
                st.session_state['otp_step'] = 1
                st.rerun()

def show_homepage_with_login():
    # ── Full-page dark background + input styling ──
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    #MainMenu, header[data-testid="stHeader"] { display: none !important; }
    [data-testid="stToolbar"] { display: none !important; }
    [data-testid="stDecoration"] { display: none !important; }
    [data-testid="stStatusWidget"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"], .main { background: #0B1120 !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; margin-bottom: 0 !important; }
    [data-testid="stHorizontalBlock"] { gap: 0 !important; }

    div[data-testid="column"]:nth-child(2),
    div[data-testid="column"]:nth-child(2) > div,
    div[data-testid="column"]:nth-child(2) > div > div,
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"],
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #0f172a !important;
    }
    div[data-testid="column"]:nth-child(2) > div:first-child {
        background-color: #0f172a !important;
        border: 1px solid rgba(255,255,255,.12) !important;
        border-radius: 16px !important;
        padding: 1.6rem 1.8rem 1.8rem !important;
    }

    div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] {
        background-color: #1e293b !important;
        border-color: rgba(148,163,184,.3) !important;
        border-radius: 10px !important;
    }
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"]:focus-within,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"]:focus-within {
        border-color: #3B82F6 !important;
        box-shadow: 0 0 0 2px rgba(59,130,246,.2) !important;
    }

    div[data-testid="column"]:nth-child(2) [data-baseweb="input"] input,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] input,
    div[data-testid="column"]:nth-child(2) input[type="text"],
    div[data-testid="column"]:nth-child(2) input[type="password"],
    div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #3B82F6 !important;
        border: none !important;
    }
    div[data-testid="column"]:nth-child(2) input::placeholder {
        color: rgba(241,245,249,.32) !important;
        -webkit-text-fill-color: rgba(241,245,249,.32) !important;
    }

    div[data-testid="column"]:nth-child(2) input:-webkit-autofill,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:hover,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:focus,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:active {
        -webkit-box-shadow: 0 0 0px 1000px #1e293b inset !important;
        box-shadow: 0 0 0px 1000px #1e293b inset !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #3B82F6 !important;
    }

    div[data-testid="column"]:nth-child(2) label,
    div[data-testid="column"]:nth-child(2) label p {
        color: rgba(241,245,249,.55) !important;
        font-size: 0.8rem !important;
    }

    div[data-testid="column"]:nth-child(2) [data-baseweb="input"] button,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] button {
        background-color: transparent !important;
        color: rgba(241,245,249,.5) !important;
        border: none !important;
    }

    div[data-testid="column"]:nth-child(2) .stButton > button {
        width: 100% !important;
        background: linear-gradient(135deg, #3B82F6, #1D4ED8) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 11px !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        box-shadow: none !important;
    }
    div[data-testid="column"]:nth-child(2) .stButton > button:hover,
    div[data-testid="column"]:nth-child(2) .stButton > button:active,
    div[data-testid="column"]:nth-child(2) .stButton > button:focus {
        background: linear-gradient(135deg, #1D4ED8, #1E3A8A) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        outline: none !important;
        box-shadow: 0 4px 16px rgba(59,130,246,.35) !important;
        transform: translateY(-1px) !important;
    }

    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"],
    [data-theme="light"] div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
    }
    [data-theme="light"] div[data-testid="column"]:nth-child(2),
    [data-theme="light"] div[data-testid="column"]:nth-child(2) > div,
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"] {
        background-color: #0f172a !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── HERO HTML ──
    html_hero = """<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;800&family=Plus+Jakarta+Sans:wght@300;400;500&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet"/>
    <style>
    *{box-sizing:border-box;margin:0;padding:0;}
    html,body{font-family:'Plus Jakarta Sans',sans-serif;background:#0B1120;color:#E2E8F0;overflow-x:hidden;}
    body{background:
        radial-gradient(ellipse 80% 50% at 50% -10%,rgba(59,130,246,.12) 0%,transparent 60%),
        radial-gradient(ellipse 60% 40% at 90% 80%,rgba(59,130,246,.35) 0%,transparent 55%),
        #0B1120;}
    body::before{content:"";position:fixed;inset:0;
        background-image:linear-gradient(rgba(255,255,255,.022) 1px,transparent 1px),
            linear-gradient(90deg,rgba(255,255,255,.022) 1px,transparent 1px);
        background-size:48px 48px;pointer-events:none;}
    .hero{display:flex;flex-direction:column;align-items:center;text-align:center;padding:3rem 2rem 2.5rem;}
    .logos{display:flex;align-items:center;margin-bottom:1rem;}
    .lname{font-size:1.25rem;font-weight:700;color:#fff;padding:0 1.6rem;}
    .lsep{width:1px;height:48px;background:linear-gradient(180deg,transparent,#3B82F6E0,transparent);box-shadow:0 0 8px rgba(59,130,246,.35);}
    .tagline{font-family:'Fira Code',monospace;font-size:.75rem;color:rgba(255,255,255,.32);letter-spacing:1.5px;margin-bottom:2rem;}
    .eyebrow{display:inline-flex;align-items:center;gap:.45rem;font-family:'Fira Code',monospace;font-size:.65rem;
        letter-spacing:2.5px;text-transform:uppercase;color:#3B82F6;background:rgba(59,130,246,.2);
        border:1px solid rgba(59,130,246,.2);border-radius:100px;padding:.28rem .95rem;margin-bottom:1.2rem;}
    .dot{width:5px;height:5px;background:#3B82F6;border-radius:50%;box-shadow:0 0 5px #3B82F6;
        animation:p 2s ease-in-out infinite;}
    @keyframes p{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.45;transform:scale(1.4);}}
    .headline{font-family:'Playfair Display',serif;font-size:clamp(2rem,5vw,3.5rem);
        font-weight:800;line-height:1.09;color:#fff;letter-spacing:-1.5px;margin-bottom:.65rem;}
    .accent{background:linear-gradient(125deg,#3B82F6,#60A5FA 40%,#93C5FD);
        -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
    .sub{font-size:1rem;font-weight:300;color:rgba(255,255,255,.38);max-width:560px;}
    </style></head><body>
    <div class="hero">
      <div class="logos">
        <div class="lname">LawSikho</div>
        <div class="lsep"></div>
        <div class="lname">Skill Arbitrage</div>
      </div>
      <div class="tagline">India Learning &nbsp;📖&nbsp; India Earning</div>
      <div class="eyebrow"><span class="dot"></span>Lead Analytics Hub</div>
      <div class="headline">Lead Pipeline,<br><span class="accent">at a glance</span></div>
      <div class="sub">Real-time insights into lead assignments, breaches &amp; dial coverage</div>
    </div>
    </body></html>"""
    st.iframe(html_hero, height=420)

    # ── AUTH PANEL ──
    left, mid, right = st.columns([1, 1, 1])
    with mid:
        auth_tab = st.session_state.get('auth_tab', 'signin')

        st.markdown("""
        <div style="text-align:center;margin-bottom:.8rem;">
            <span style="font-family:'Playfair Display',serif;font-size:1.1rem;font-weight:600;color:#fff;">
                🔐 DASHBOARD ACCESS
            </span>
        </div>
        """, unsafe_allow_html=True)

        if auth_tab == 'signin':
            _auth_sign_in_panel()
        else:
            _auth_otp_panel()

        if st.session_state.get('password_correct'):
            st.rerun()

    # ── Simple footer ──
    st.markdown("""
    <div style='text-align:center;padding:2rem 1rem 1.2rem;border-top:1px solid rgba(255,255,255,.06);margin-top:2.5rem;'>
        <div style='font-family:"Fira Code",monospace;font-size:.64rem;color:rgba(255,255,255,.28);margin-bottom:.4rem;'>
            For Internal Use of Sales and Operations Team Only
            <span style='display:inline-block;width:3px;height:3px;background:#3B82F666;border-radius:50%;margin:0 .45rem;vertical-align:middle;'></span>
            All Rights Reserved
        </div>
        <div style='font-family:"Fira Code",monospace;font-size:.58rem;color:rgba(255,255,255,.14);'>
            Developed and Designed by Amit Ray
            <span style='display:inline-block;width:3px;height:3px;background:#3B82F666;border-radius:50%;margin:0 .45rem;vertical-align:middle;'></span>
            Reach out for Support and Queries
        </div>
    </div>
    """, unsafe_allow_html=True)

def run_leads_dashboard():
    st.markdown("""
    <style>
    [data-testid="stMainBlockContainer"] {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    .block-container { max-width: 100% !important; }
    </style>
    """, unsafe_allow_html=True)
 
    # ── CONSTANTS ──────────────────────────────────────────────────────────────
    _CSV_URL     = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf/pub?gid=0&single=true&output=csv"
    _LEADS_TABLE = "studious-apex-488820-c3.crm_dashboard.lsq_leads"
 
    _STAGE_MAP = {
        'FRESH'             : ['New Lead', 'Re-enquired Lead', 'Opportunity Created'],
        'DNP'               : ['Call Not Picking Up', 'Call Not Connected'],
        'CBL'               : ['Call Back Later'],
        'FLW-UP'            : ['Follow Up For Closure'],
        'COUNSELLED'        : ['Counselled lead'],
        'DISCOVERY'         : ['Discovery Call Done'],
        'ROADMAP'           : ['Roadmap Done'],
        'MBL'               : ['May buy later'],
        'ACTUALLY-ENROLLED' : ['Actually Enrolled'],
        'INVALID/NTINTRSTD' : ['Irrelevant lead', 'Not Interested', 'Invalid'],
        'BOOKING-RCVD'      : ['Booking fees received'],
        'LOAN-PNDG'         : ['Loan pending'],
        'COLL-DNE'          : ['Collections done'],
        'PRE-SALES'         : ['Pre-Sales Registrations'],
        'COURSE ENROLLED'   : ['Course Enrolled'],
    }
 
    _BREACHED_COL_MAP = {
        'CBL'       : ['Call Back Later'],
        'FLW-UP'    : ['Follow Up For Closure'],
        'COUNSELLED': ['Counselled lead'],
        'DISCOVERY' : ['Discovery Call Done'],
        'ROADMAP'   : ['Roadmap Done'],
    }
 
    _DIALLED_COL_MAP = {
        'CALL NOT PICKING UP': ['Call Not Picking Up'],
        'CALL NOT CONNECTED' : ['Call Not Connected'],
    }
 
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
 
:root {
    --ld-accent: #3B82F6; --ld-deep: #1D4ED8; --ld-deepest: #1e3a8a;
    --ld-light: #DBEAFE;
    --radius-sm: 8px; --radius-md: 12px; --radius-lg: 16px;
    --shadow-sm: 0 1px 3px rgba(0,0,0,.08);
    --shadow-md: 0 4px 16px rgba(0,0,0,.10);
    --shadow-lg: 0 8px 32px rgba(0,0,0,.14);
    --transition: all 0.22s cubic-bezier(.4,0,.2,1);
}
[data-testid="stAppViewContainer"]:not([class*="dark"]) {
    --bg-base:#EFF6FF;--bg-surface:#FFFFFF;--bg-muted:#DBEAFE;
    --border:rgba(59,130,246,.12);--text-primary:#111827;
    --text-muted:#6B7280;--metric-bg:#FFFFFF;
}
@media (prefers-color-scheme: dark) {
    :root{--bg-base:#0A0F1E;--bg-surface:#111827;--bg-elevated:#1E293B;
          --bg-muted:#1E3A5F;--border:rgba(59,130,246,.10);
          --text-primary:#EFF6FF;--text-muted:#93C5FD;--metric-bg:#1E293B;}
}
[data-theme="dark"]{
    --bg-base:#0A0F1E!important;--bg-surface:#111827!important;
    --bg-elevated:#1E293B!important;--bg-muted:#1E3A5F!important;
    --border:rgba(59,130,246,.10)!important;--text-primary:#EFF6FF!important;
    --text-muted:#93C5FD!important;--metric-bg:#1E293B!important;
}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif!important;}
footer{visibility:hidden;}
[data-testid="stStatusWidget"]{display:none!important;}
[data-testid="stSidebar"]{border-right:1px solid var(--border,rgba(59,130,246,.12));}
 
.ld-header{
    background:linear-gradient(135deg,#0f172a 0%,#1e3a8a 45%,#1e40af 100%);
    border-radius:var(--radius-lg);padding:1.5rem 2rem 1.2rem;
    margin-bottom:1.2rem;position:relative;overflow:hidden;box-shadow:var(--shadow-lg);
}
.ld-header::before{
    content:"";position:absolute;top:-40px;right:-40px;width:200px;height:200px;
    background:radial-gradient(circle,rgba(59,130,246,.25) 0%,transparent 70%);border-radius:50%;
}
.ld-title{font-size:1.65rem;font-weight:700;color:#FFF;letter-spacing:.5px;margin:0 0 .25rem;}
.ld-subtitle{font-size:.82rem;color:rgba(255,255,255,.6);margin:0;font-family:'DM Mono',monospace;}
.ld-badge{
    display:inline-flex;align-items:center;gap:5px;
    background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.18);
    border-radius:20px;padding:3px 10px;font-size:.73rem;
    color:rgba(255,255,255,.9);font-family:'DM Mono',monospace;
}
.ld-pulse{
    width:6px;height:6px;background:#60A5FA;border-radius:50%;
    display:inline-block;animation:pulse-ld 1.8s ease-in-out infinite;
}
@keyframes pulse-ld{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.5;transform:scale(1.4);}}
 
.ld-metric-card{
    background:var(--metric-bg,#fff);border:1px solid var(--border,rgba(59,130,246,.12));
    border-radius:var(--radius-md);padding:.9rem 1rem;transition:var(--transition);
    box-shadow:var(--shadow-sm);position:relative;overflow:hidden;text-align:center;
}
.ld-metric-card::before{
    content:"";position:absolute;top:0;left:0;width:100%;height:3px;
    background:linear-gradient(90deg,#3B82F6,#60A5FA);opacity:0;transition:opacity .2s;
}
.ld-metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md);}
.ld-metric-card:hover::before{opacity:1;}
.ld-metric-label{font-size:.68rem;font-weight:600;text-transform:uppercase;letter-spacing:.8px;
    color:var(--text-muted,#6B7280);margin:0 0 .3rem;}
.ld-metric-value{font-size:1.45rem;font-weight:700;color:var(--text-primary,#111827);
    line-height:1;font-family:'DM Mono',monospace;}
 
.ld-section-header{display:flex;align-items:center;gap:.6rem;margin:1.5rem 0 .8rem;}
.ld-section-line{flex:1;height:1px;background:linear-gradient(90deg,#3B82F6,transparent);opacity:.35;}
.ld-section-title{font-size:.78rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;
    color:#3B82F6;white-space:nowrap;text-align:center;}
 
.ld-insight-card{background:var(--metric-bg,#fff);border:1px solid var(--border,rgba(59,130,246,.12));
    border-radius:var(--radius-md);padding:1rem 1.1rem;margin-bottom:.6rem;
    box-shadow:var(--shadow-sm);transition:var(--transition);}
.ld-insight-card:hover{box-shadow:var(--shadow-md);}
.ld-insight-card.good{border-left:3px solid #3B82F6;}
.ld-insight-card.warn{border-left:3px solid #FBBF24;}
.ld-insight-card.bad{border-left:3px solid #F87171;}
.ld-insight-card.info{border-left:3px solid #60A5FA;}
 
div[data-testid="stDataFrame"] thead tr th{
    background:linear-gradient(135deg,#1e3a8a,#1d4ed8)!important;
    color:#fff!important;font-family:'DM Sans',sans-serif!important;
    font-size:.72rem!important;font-weight:700!important;letter-spacing:.6px;
    text-transform:uppercase;white-space:normal!important;word-wrap:break-word!important;
    text-align:center!important;vertical-align:middle!important;
    min-width:80px!important;padding:10px!important;
}
[data-testid="stSidebar"] .stButton>button{
    width:100%;font-family:'DM Sans',sans-serif!important;font-weight:600!important;
    font-size:.82rem!important;border-radius:var(--radius-sm);transition:var(--transition);
    background:linear-gradient(135deg,#1d4ed8,#1e3a8a)!important;
    color:#fff!important;border:none!important;
}
[data-testid="stSidebar"] .stButton>button:hover{opacity:.88!important;transform:translateY(-1px)!important;}
.stDownloadButton>button{
    background:linear-gradient(135deg,#1d4ed8,#1e3a8a)!important;color:#fff!important;
    border:none!important;border-radius:var(--radius-sm)!important;
    font-family:'DM Sans',sans-serif!important;font-weight:600!important;
    font-size:.82rem!important;width:100%!important;transition:var(--transition)!important;
}
hr{border-color:var(--border,rgba(59,130,246,.12))!important;margin:1.2rem 0!important;}
</style>
""", unsafe_allow_html=True)
 
    def _ld_section_header(label):
        st.markdown(f"""
        <div class="ld-section-header">
            <div class="ld-section-line"></div>
            <span class="ld-section-title">{label}</span>
            <div class="ld-section-line" style="background:linear-gradient(90deg,transparent,#3B82F6)"></div>
        </div>""", unsafe_allow_html=True)
 
    def _style_caller(row):
        if row.get('CALLER') == 'TOTAL':
            return ['font-weight:bold;background-color:#374151;color:#FFFFFF;'] * len(row)
        return [''] * len(row)
 
    def _style_team(row):
        if row.get('TEAM') == 'TOTAL':
            return ['font-weight:bold;background-color:#374151;color:#FFFFFF;'] * len(row)
        return [''] * len(row)
 
    def _merge_team_ld(df, df_meta):
        df_w = df.copy()
        df_w['merge_key'] = df_w['Owner'].astype(str).str.strip().str.lower()
        slim = df_meta[['merge_key','Caller Name','Team Name','Vertical']].drop_duplicates('merge_key')
        merged = pd.merge(df_w, slim, on='merge_key', how='left')
        merged['Owner']     = merged['Caller Name'].fillna(merged['Owner'])
        merged['Team Name'] = merged['Team Name'].fillna('Others')
        merged['Vertical']  = merged['Vertical'].fillna('Others')
        merged = merged[merged['Team Name'] != 'Others']
        return merged
 
    def _append_caller_total_ld(df):
        if df.empty: return df
        nc  = [c for c in df.columns if c not in ('CALLER','TEAM')]
        row = {'CALLER':'TOTAL','TEAM':'—'}
        for c in nc: row[c] = int(df[c].sum())
        return pd.concat([df, pd.DataFrame([row])], ignore_index=True)
 
    def _append_team_total_ld(df):
        if df.empty: return df
        nc  = [c for c in df.columns if c != 'TEAM']
        row = {'TEAM':'TOTAL'}
        for c in nc: row[c] = int(df[c].sum())
        return pd.concat([df, pd.DataFrame([row])], ignore_index=True)
 
    def _show_caller(df, msg="No data."):
        if df.empty: st.info(msg); return
        final = _append_caller_total_ld(df)
        h = min((len(final) + 1) * 35 + 20, 800)
        st.dataframe(final.style.apply(_style_caller, axis=1),
                     width='stretch', hide_index=True, height=h)
 
    def _show_team(df, msg="No data."):
        if df.empty: st.info(msg); return
        final = _append_team_total_ld(df)
        h = min((len(final) + 1) * 35 + 20, 600)
        st.dataframe(final.style.apply(_style_team, axis=1),
                     width='stretch', hide_index=True, height=h)
        
    def _build_leads_xlsx_bytes_ld(df_rows):
        EXPORT_COLS = [
            'AssignedOn', 'FirstName', 'LastName', 'Email', 'PhoneNumber',
            'Alternate_PhoneNumber', 'Owner', 'ContactStage', 'LastCalledDate',
            'Follow_up_date', 'Enquired_Course', 'Campaign_Name',
            'Phone_call_counter', 'Assigned_On_Call_Counter', 'Team', 'Vertical', 'AssignedBy'
        ]
        df = df_rows.copy()
        if 'Team' not in df.columns and 'Team Name' in df.columns:
            df['Team'] = df['Team Name']
        cols = [c for c in EXPORT_COLS if c in df.columns]
        df   = df[cols].reset_index(drop=True)

        HDR_FILL  = PatternFill("solid", start_color="1e3a8a", end_color="1e3a8a")
        HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        ALT_FILL  = PatternFill("solid", start_color="EFF6FF", end_color="EFF6FF")
        WHT_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        DATA_FONT = Font(name="Calibri", size=10)
        CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        BORDER    = Border(
            left=Side(style='thin', color='BFDBFE'), right=Side(style='thin', color='BFDBFE'),
            top=Side(style='thin',  color='BFDBFE'), bottom=Side(style='thin', color='BFDBFE'),
        )
        COL_WIDTHS = {
            'AssignedOn': 14, 'FirstName': 18, 'LastName': 18, 'Email': 32,
            'PhoneNumber': 14, 'Alternate_PhoneNumber': 18, 'Owner': 24,
            'ContactStage': 24, 'LastCalledDate': 14, 'Follow_up_date': 14,
            'Enquired_Course': 26, 'Campaign_Name': 22, 'Phone_call_counter': 12,
            'Assigned_On_Call_Counter': 14, 'Team': 24, 'Vertical': 18, 'AssignedBy': 20,
        }
        wb = Workbook(); ws = wb.active; ws.title = "Breached Leads"
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(1, c_idx, col.replace('_', ' ').upper())
            cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER
        ws.row_dimensions[1].height = 26
        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            fill = ALT_FILL if r_idx % 2 == 0 else WHT_FILL
            for c_idx, col in enumerate(cols, 1):
                val = row[col]
                try:
                    if pd.isna(val): val = ''
                except (TypeError, ValueError): pass
                if hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d')
                cell = ws.cell(r_idx, c_idx, val)
                cell.fill, cell.font, cell.alignment, cell.border = fill, DATA_FONT, LEFT, BORDER
        for c_idx, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(col, 16)
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()
 
    def _stage_counts(grp, col_map):
        row, total = {}, 0
        for col, stages in col_map.items():
            cnt = int(grp['ContactStage'].isin(stages).sum())
            row[col] = cnt; total += cnt
        row['TOTAL'] = total
        return row
 
    @st.cache_data(ttl=120, show_spinner=False)
    def _ld_get_metadata():
        df = pd.read_csv(_CSV_URL)
        df.columns = df.columns.str.strip()
        df['merge_key'] = df['Caller Name'].str.strip().str.lower()
        return sorted(df['Team Name'].dropna().unique()), sorted(df['Vertical'].dropna().unique()), df
 
    @st.cache_data(ttl=300, show_spinner=False)
    def _ld_last_update():
        try:
            r = client.query(
                f"SELECT MAX(updated_at_ampm) AS lu FROM `{_LEADS_TABLE}` WHERE updated_at_ampm IS NOT NULL"
            ).to_dataframe()
            return str(r['lu'].iloc[0]) if not r.empty and r['lu'].iloc[0] else "N/A"
        except Exception: return "N/A"
 
    @st.cache_data(ttl=600, show_spinner=False)
    def _ld_available_dates():
        try:
            r = client.query(
                f"SELECT MIN(AssignedOn) AS mn, MAX(AssignedOn) AS mx FROM `{_LEADS_TABLE}`"
            ).to_dataframe()
            if not r.empty and not pd.isna(r['mn'].iloc[0]):
                return r['mn'].iloc[0], r['mx'].iloc[0]
        except Exception: pass
        return date.today(), date.today()
 
    @st.cache_data(ttl=120, show_spinner=False)
    def _ld_fetch(start_date, end_date):
        q = f"""
            SELECT * FROM `{_LEADS_TABLE}`
            WHERE AssignedOn BETWEEN '{start_date}' AND '{end_date}'
        """
        df = client.query(q).to_dataframe()
        if not df.empty:
            df['Owner']        = df['Owner'].astype(str).str.strip()
            df['ContactStage'] = (
                df['ContactStage'].astype(str)
                .str.replace('\xa0', ' ', regex=False)  
                .str.replace('Â', '', regex=False) 
                .str.replace(r'\s+', ' ', regex=True)   
                .str.strip())
            df['ContactStage'] = df['ContactStage'].astype(str).str.strip()
            df['Follow_up_date'] = pd.to_datetime(df['Follow_up_date'], errors='coerce')
            df['LastCalledDate'] = pd.to_datetime(df['LastCalledDate'], errors='coerce')
            df['Assigned_On_Call_Counter'] = pd.to_numeric(
                df['Assigned_On_Call_Counter'], errors='coerce').fillna(0)
        return df
 
    def _proc_assigned_caller(df_m):
        rows = []
        for owner, g in df_m.groupby('Owner'):
            team = g['Team Name'].mode().iloc[0] if len(g) else 'Others'
            r = {'CALLER': owner, 'TEAM': team}
            r.update(_stage_counts(g, _STAGE_MAP))
            rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _proc_assigned_team(df_m):
        rows = []
        for team, g in df_m.groupby('Team Name'):
            r = {'TEAM': team}; r.update(_stage_counts(g, _STAGE_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _get_breached(df_m):
        now_ts = pd.Timestamp.now().normalize()
        cut    = now_ts - pd.Timedelta(days=3)
        all_s  = [s for v in _BREACHED_COL_MAP.values() for s in v]
        fup    = df_m['Follow_up_date'].isna() | (df_m['Follow_up_date'] < now_ts)
        lcd    = df_m['LastCalledDate'].notna() & (df_m['LastCalledDate'] < cut)
        return df_m[fup & lcd & df_m['ContactStage'].isin(all_s)].copy()
 
    def _proc_breached_caller(df_m):
        df_b = _get_breached(df_m); rows = []
        for owner, g in df_b.groupby('Owner'):
            team = g['Team Name'].mode().iloc[0] if len(g) else 'Others'
            r = {'CALLER': owner, 'TEAM': team}; r.update(_stage_counts(g, _BREACHED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _proc_breached_team(df_m):
        df_b = _get_breached(df_m); rows = []
        for team, g in df_b.groupby('Team Name'):
            r = {'TEAM': team}; r.update(_stage_counts(g, _BREACHED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _get_less_dialled(df_m):
        all_s = [s for v in _DIALLED_COL_MAP.values() for s in v]
        return df_m[(df_m['Assigned_On_Call_Counter'] < 11) & df_m['ContactStage'].isin(all_s)].copy()
 
    def _proc_ld_caller(df_m):
        df_ld = _get_less_dialled(df_m); rows = []
        for owner, g in df_ld.groupby('Owner'):
            team = g['Team Name'].mode().iloc[0] if len(g) else 'Others'
            r = {'CALLER': owner, 'TEAM': team}; r.update(_stage_counts(g, _DIALLED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _proc_ld_team(df_m):
        df_ld = _get_less_dialled(df_m); rows = []
        for team, g in df_ld.groupby('Team Name'):
            r = {'TEAM': team}; r.update(_stage_counts(g, _DIALLED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _compute_insights(df_ac, df_bc, df_ldc, df_at, df_bt, df_ldt):
        ins = []
        if not df_ac.empty:
            t = df_ac.iloc[0]
            ins.append({"type":"good","icon":"🏆",
                "title":f"Highest Assigned Leads — {t['CALLER']}",
                "body":(f"{t['CALLER']} holds the highest lead count with {t['TOTAL']} leads. "
                        f"Team: {t['TEAM']}. Fresh: {t.get('FRESH',0)} · "
                        f"Counselled: {t.get('COUNSELLED',0)} · Enrolled: {t.get('COURSE ENROLLED',0)}.")})
        if not df_at.empty:
            t = df_at.iloc[0]
            ins.append({"type":"good","icon":"🏅",
                "title":f"Highest Assigned Leads — Team: {t['TEAM']}",
                "body":(f"{t['TEAM']} received the most leads with {t['TOTAL']} assignments. "
                        f"Fresh: {t.get('FRESH',0)} · DNP: {t.get('DNP',0)} · "
                        f"Course Enrolled: {t.get('COURSE ENROLLED',0)}.")})
        if not df_bc.empty:
            t = df_bc.iloc[0]
            ins.append({"type":"bad","icon":"⚠️",
                "title":f"Highest Potential Breached — {t['CALLER']}",
                "body":(f"{t['CALLER']} has {t['TOTAL']} leads with overdue follow-up and no recent call. "
                        f"Team: {t['TEAM']}. CBL: {t.get('CBL',0)} · FLW-UP: {t.get('FLW-UP',0)} · "
                        f"Counselled: {t.get('COUNSELLED',0)}. Immediate action required.")})
        if not df_bt.empty:
            t = df_bt.iloc[0]
            ins.append({"type":"warn","icon":"🚨",
                "title":f"Highest Potential Breached — Team: {t['TEAM']}",
                "body":(f"{t['TEAM']} leads the breached list with {t['TOTAL']} overdue leads. "
                        f"CBL: {t.get('CBL',0)} · FLW-UP: {t.get('FLW-UP',0)} · "
                        f"Counselled: {t.get('COUNSELLED',0)}. Schedule recovery sessions.")})
        if not df_ldc.empty:
            t = df_ldc.iloc[0]
            ins.append({"type":"warn","icon":"📞",
                "title":f"Highest Less Dialled — {t['CALLER']}",
                "body":(f"{t['CALLER']} has {t['TOTAL']} DNP leads with <11 dial attempts. "
                        f"Team: {t['TEAM']}. Not Picking Up: {t.get('CALL NOT PICKING UP',0)} · "
                        f"Not Connected: {t.get('CALL NOT CONNECTED',0)}. Increase dial frequency.")})
        if not df_ldt.empty:
            t = df_ldt.iloc[0]
            ins.append({"type":"info","icon":"📈",
                "title":f"Highest Less Dialled — Team: {t['TEAM']}",
                "body":(f"{t['TEAM']} has {t['TOTAL']} less-dialled leads below 11 attempts. "
                        f"Maximise dial frequency before these leads degrade further.")})
        return ins[:6]
 
    @st.cache_data(show_spinner=False)
    def _ld_pdf_bytes() -> bytes:
        from reportlab.lib.pagesizes import A4 as _A4
        from reportlab.lib import colors as _colors
        from reportlab.lib.units import mm as _mm
        from reportlab.lib.styles import ParagraphStyle as _PS
        from reportlab.platypus import (SimpleDocTemplate as _SDT, Paragraph as _P,
                                         Spacer as _Sp, Table as _T, TableStyle as _TS,
                                         HRFlowable as _HR, Flowable as _F)
        from reportlab.lib.enums import TA_CENTER as _TAC
        import io as _io
 
        buf     = _io.BytesIO()
        BD      = _colors.HexColor("#1e3a8a"); BM=_colors.HexColor("#1d4ed8")
        BP      = _colors.HexColor("#DBEAFE"); BR=_colors.HexColor("#EFF6FF")
        GD      = _colors.HexColor("#374151"); GM=_colors.HexColor("#6B7280")
        W_      = _colors.white;               BK=_colors.HexColor("#111827")
        PW, PH  = _A4
 
        def sty(name,**kw):
            d=dict(fontName='Helvetica',fontSize=9,textColor=BK,spaceAfter=3,leading=14)
            d.update(kw); return _PS(name,**d)
 
        S={'b':sty('b_ld'),'l':sty('l_ld',fontName='Helvetica-Bold',fontSize=8,textColor=BD,spaceAfter=1),
           'f':sty('f_ld',fontName='Helvetica-Oblique',fontSize=8.5,textColor=BM,backColor=BP,leftIndent=8,rightIndent=8),
           'ft':sty('ft_ld',fontSize=7.5,textColor=GM,alignment=_TAC)}
 
        class Cover(_F):
            def __init__(self,w): _F.__init__(self); self.w=w; self.height=90
            def draw(self):
                c=self.canv
                c.setFillColor(BD); c.rect(0,52,self.w,38,fill=1,stroke=0)
                c.setFillColor(BM); c.rect(0,22,self.w,30,fill=1,stroke=0)
                c.setFillColor(_colors.HexColor("#0f172a")); c.rect(0,0,self.w,22,fill=1,stroke=0)
                c.setFillColor(W_); c.setFont("Helvetica-Bold",22)
                c.drawCentredString(self.w/2,66,"LEAD METRICS DASHBOARD")
                c.setFillColor(_colors.HexColor("#BFDBFE")); c.setFont("Helvetica-Bold",11)
                c.drawCentredString(self.w/2,34,"Logic & Metric Reference Guide")
                c.setFillColor(_colors.HexColor("#DBEAFE")); c.setFont("Helvetica",8.5)
                c.drawCentredString(self.w/2,8,
                    "LawSikho & Skill Arbitrage  \u00b7  Internal Use Only")
 
        class Ban(_F):
            def __init__(self,icon,title,color=None,w=None):
                _F.__init__(self); self.icon=icon; self.title=title
                self.color=color or BD; self.w=w or (PW-30*_mm); self.height=22
            def draw(self):
                c=self.canv; c.setFillColor(self.color)
                c.roundRect(0,0,self.w,self.height,4,fill=1,stroke=0)
                c.setFillColor(W_); c.setFont("Helvetica-Bold",11)
                c.drawString(10,6,f"{self.icon}  {self.title}")
 
        def bt(rows,cw=None):
            cw=cw or [44*_mm,116*_mm]
            data=[[_P(f"<b>{r[0]}</b>",S['l']),_P(r[1],S['b'])] for r in rows]
            t=_T(data,colWidths=cw,hAlign='LEFT')
            t.setStyle(_TS([('BACKGROUND',(0,0),(0,-1),BP),('VALIGN',(0,0),(-1,-1),'TOP'),
                ('GRID',(0,0),(-1,-1),0.3,_colors.HexColor("#BFDBFE")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[W_,BR]),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            return t
 
        def lt(rows):
            data=[[_P(f"<b>{r[0]}</b>",S['l']),_P(r[1],S['f'])] for r in rows]
            t=_T(data,colWidths=[52*_mm,108*_mm],hAlign='LEFT')
            t.setStyle(_TS([('VALIGN',(0,0),(-1,-1),'TOP'),
                ('GRID',(0,0),(-1,-1),0.3,_colors.HexColor("#93C5FD")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[W_,BR]),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            return t
 
        SP=_Sp; HR_=lambda:_HR(width="100%",thickness=0.6,color=_colors.HexColor("#93C5FD"),spaceAfter=6,spaceBefore=4)
        cw_=PW-30*_mm
 
        story=[
            SP(1,18*_mm),Cover(cw_),SP(1,10*_mm),
            _P("This document explains every metric, table, and column in the Lead Metrics Dashboard.",S['b']),
            SP(1,6*_mm),
            Ban("📋","SECTION 1 — TABS OVERVIEW"),SP(1,3*_mm),
            bt([("📊 Assigned Leads Report",
                 "Three callerwise tables: (1) Assigned Leads Distribution — all ContactStage buckets per caller; "
                 "(2) Potential Breached Leads — overdue follow-up + no recent call activity; "
                 "(3) Less Dialled Leads — DNP leads with <11 dial attempts."),
                ("🧠 Insights & Teamwise",
                 "Six auto-generated insights + three teamwise versions of all tables.")]),
            SP(1,6*_mm),
            Ban("📊","SECTION 2 — ASSIGNED LEADS DISTRIBUTION"),SP(1,3*_mm),
            _P("One row per caller. Sorted by TOTAL descending. Date filter = AssignedOn field.",S['b']),SP(1,2*_mm),
            lt([
                ("FRESH","New Lead + Re-enquired Lead + Opportunity Created"),
                ("DNP","Call Not Picking Up + Call Not Connected"),
                ("CBL","Call Back Later"),("FLW-UP","Follow Up For Closure"),
                ("COUNSELLED","Counselled lead"),("DISCOVERY","Discovery Call Done"),
                ("ROADMAP","Roadmap Done"),("MBL","May buy later"),
                ("ACTUALLY-ENROLLED","Actually Enrolled"),
                ("INVALID/NTINTRSTD","Irrelevant lead + Not Interested + Invalid"),
                ("BOOKING-RCVD","Booking fees received"),("LOAN-PNDG","Loan pending"),
                ("COLL-DNE","Collections done"),("PRE-SALES","Pre-Sales Registrations"),
                ("COURSE ENROLLED","Course Enrolled"),
                ("TOTAL","Sum of all stage columns = total assigned leads for this caller/team."),
            ]),SP(1,6*_mm),
            Ban("⚠️","SECTION 3 — POTENTIAL BREACHED LEADS"),SP(1,3*_mm),
            lt([
                ("Condition 1","Follow_up_date IS NULL OR Follow_up_date < today (IST)."),
                ("Condition 2","LastCalledDate < today minus 3 days (both conditions required)."),
                ("Stages","CBL · FLW-UP · COUNSELLED · DISCOVERY · ROADMAP only."),
                ("TOTAL","Sum of all five stage columns."),
            ]),SP(1,6*_mm),
            Ban("📞","SECTION 4 — LESS DIALLED LEADS"),SP(1,3*_mm),
            lt([
                ("Filter","Assigned_On_Call_Counter < 11 (fewer than 11 call attempts since assignment)."),
                ("CALL NOT PICKING UP","ContactStage = 'Call Not Picking Up' with <11 attempts."),
                ("CALL NOT CONNECTED","ContactStage = 'Call Not Connected' with <11 attempts."),
                ("TOTAL","Sum of both columns."),
            ]),SP(1,6*_mm),
            Ban("📖","KEY TERMS GLOSSARY",color=GD),SP(1,3*_mm),
            bt([
                ("AssignedOn","Date the lead was assigned. Primary date filter for the dashboard."),
                ("Owner","Caller the lead is assigned to. Mapped via lowercase merge key to Caller Name."),
                ("ContactStage","Current CRM lifecycle stage. Drives all bucketing logic."),
                ("Follow_up_date","Scheduled follow-up date. NULL or past = overdue."),
                ("LastCalledDate","Date last called. Used in breached leads filter."),
                ("Assigned_On_Call_Counter","Call attempts since assignment. Used in less-dialled filter."),
                ("TOTAL row","Bold dark row at bottom summing all numeric columns."),
            ]),SP(1,8*_mm),HR_(),
            _P("Designed by Amit Ray  \u00b7  amitray@lawsikho.in  \u00b7  "
               "For Internal Use of Sales and Operations Team Only. All Rights Reserved.",S['ft']),
        ]
        doc=_SDT(buf,pagesize=_A4,leftMargin=15*_mm,rightMargin=15*_mm,
                 topMargin=14*_mm,bottomMargin=14*_mm,
                 title="Lead Metrics — Logic Reference Guide",author="Amit Ray")
        doc.build(story)
        return buf.getvalue()
 
    st.sidebar.markdown("""
    <div style='padding:.5rem 0 .4rem; text-align:center;'>
        <div style='font-size:.72rem; font-weight:700; text-transform:uppercase;
                    letter-spacing:1px; color:var(--text-muted,#6B7280);'>Report Controls</div>
    </div>
    """, unsafe_allow_html=True)
 
    teams_ld, verts_ld, df_meta_ld = _ld_get_metadata()
    min_dr, max_dr = _ld_available_dates()
    min_date_ld = pd.Timestamp(min_dr).date()
    max_date_ld = pd.Timestamp(max_dr).date()
 
    date_range_ld = st.sidebar.date_input(
        "📅 Date Range", value=(max_date_ld, max_date_ld),
        min_value=min_date_ld, max_value=max_date_ld,
        format="DD-MM-YYYY", key="ld_date_range_cd"
    )
    if isinstance(date_range_ld, tuple) and len(date_range_ld) == 2:
        start_date_ld, end_date_ld = date_range_ld
    else:
        start_date_ld = end_date_ld = (
            date_range_ld if not isinstance(date_range_ld, tuple) else date_range_ld[0])
 
    # ── Role-aware sidebar filters ──────────────────────────
    _role     = st.session_state.get('rf_role', 'admin')
    _rf_teams = st.session_state.get('rf_teams', [])
    _rf_cname = st.session_state.get('rf_caller_name', '')

    if _role == 'admin':
        sel_team_ld = st.sidebar.multiselect("👥 Filter by Team",     options=teams_ld, key="ld_team_cd")
        sel_vert_ld = st.sidebar.multiselect("👑 Filter by Vertical", options=verts_ld, key="ld_vert_cd")
        search_ld   = st.sidebar.text_input("👤 Search Caller Name",                    key="ld_search_cd")
    elif _role == 'vertical_head':
        sel_team_ld = _rf_teams
        sel_vert_ld = []
        search_ld   = st.sidebar.text_input("👤 Search Caller Name", key="ld_search_cd")
        st.sidebar.caption(f"🔒 Showing: {', '.join(_rf_teams)}")
    elif _role in ('tl', 'trainer'):
        sel_team_ld = _rf_teams
        sel_vert_ld = []
        search_ld   = st.sidebar.text_input("👤 Search Caller Name", key="ld_search_cd")
        st.sidebar.caption(f"🔒 Team: {', '.join(_rf_teams)}")
    else:  # caller
        sel_team_ld = []
        sel_vert_ld = []
        search_ld   = _rf_cname
        st.sidebar.caption(f"👤 Viewing: {_rf_cname}")
 
    gen_ld = st.sidebar.button("📊 Generate Leads Report", key="ld_gen_cd")
 
    st.sidebar.download_button(
        label="📖 Metrics Guide (PDF)",
        data=_ld_pdf_bytes(),
        file_name="Lead_Metrics_Logic_Guide.pdf",
        mime="application/pdf", key="dl_ld_pdf_cd"
    )
    
    last_upd_ld  = _ld_last_update()
    disp_start   = start_date_ld.strftime('%d-%m-%Y')
    disp_end     = end_date_ld.strftime('%d-%m-%Y')
 
    st.markdown(f"""
    <div class="ld-header">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:.75rem;">
            <div>
                <div class="ld-title">📊 LEAD METRICS</div>
                <div class="ld-subtitle">ASSIGNMENT PERIOD&nbsp;·&nbsp; {disp_start} to {disp_end}</div>
            </div>
            <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.25rem;">
                <span class="ld-badge"><span class="ld-pulse"></span>LEADSQUARED DATA</span>
                <span class="ld-badge">🕐 UPDATED AT: {last_upd_ld}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
 
    tab_ld1, tab_ld2 = st.tabs(["📊 Assigned Leads Report", "🧠 Insights & Teamwise"])
 
    with tab_ld1:
        if gen_ld:
            with st.spinner("Fetching lead data…"):
                df_raw_ld = _ld_fetch(start_date_ld, end_date_ld)
 
            if df_raw_ld.empty:
                st.warning("No leads found for the selected period.")
            else:
                with st.spinner("Processing…"):
                    df_m_ld = _merge_team_ld(df_raw_ld, df_meta_ld)
                    if sel_team_ld: df_m_ld = df_m_ld[df_m_ld['Team Name'].isin(sel_team_ld)]
                    if sel_vert_ld: df_m_ld = df_m_ld[df_m_ld['Vertical'].isin(sel_vert_ld)]
                    if search_ld:   df_m_ld = df_m_ld[df_m_ld['Owner'].str.contains(search_ld, case=False, na=False)]
 
                if df_m_ld.empty:
                    st.error("No results match the selected filters.")
                else:

                    df_valid_ld = df_m_ld[df_m_ld['Team Name'] != "Others"]
                    _ld_section_header("SUMMARY METRICS")
                    fresh_c = int(df_m_ld['ContactStage'].isin(_STAGE_MAP['FRESH']).sum())
                    enrolled_c = int(df_valid_ld['ContactStage'].eq("Actually Enrolled").sum())
                    discovery_c = int(df_valid_ld['ContactStage'].eq("Discovery Call Done").sum())
                    roadmap_c = int(df_valid_ld['ContactStage'].str.contains('Roadmap', case=False, na=False).sum())
                    followup_c = int(df_valid_ld['ContactStage'].isin(["Follow Up For Closure","Counselled lead"]).sum())
                    kc = st.columns(8)
                    for col, (lbl, val, ico) in zip(kc, [
                        ("Total Assigned Leads", f"{len(df_m_ld):,}", "📋"),
                        ("Fresh Leads", f"{fresh_c:,}", "🌱"),
                        ("Lead Conversions", f"{enrolled_c:,}", "🎓"),
                        ("Discovery", f"{discovery_c:,}", "🔍"),
                        ("Roadmap", f"{roadmap_c:,}", "🗺️"),
                        ("Follow Up / Counselled", f"{followup_c:,}", "📞"),
                        ("Active Callers", df_m_ld['Owner'].nunique(), "👤"),
                        ("Active Teams", df_m_ld['Team Name'].nunique(), "👥"),
                    ]):
                        with col:
                            st.markdown(f"""
                            <div class="ld-metric-card">
                                <div class="ld-metric-label">{ico} {lbl}</div>
                                <div class="ld-metric-value">{val}</div>
                            </div>""", unsafe_allow_html=True)
 
                    st.divider()
 
                    _ld_section_header("ASSIGNED LEADS DISTRIBUTION")
                    df_ac_ld = _proc_assigned_caller(df_m_ld)
                    _show_caller(df_ac_ld, "No assigned lead data found.")

                    if not df_m_ld.empty:
                        col_dl_ac, _ = st.columns([1, 3])
                        with col_dl_ac:
                            st.download_button(
                                label="📥 Download Assigned Leads (.xlsx)",
                                data=_build_leads_xlsx_bytes_ld(df_m_ld),
                                file_name=f"Assigned_Leads_{disp_start}_to_{disp_end}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_ac_leads_ld_xlsx"
                            )

                    st.divider()

                    _ld_section_header("POTENTIAL BREACHED LEADS AFTER ASSIGNMENT")
                    cut_disp = (pd.Timestamp.now().normalize() - pd.Timedelta(days=3)).strftime('%d-%m-%Y')
                    st.caption(f"Leads Breached and dialled before {cut_disp} · "
                               "of stages Call Back Later, Follow Up, Counselled, Discovery & Roadmap.")
                    df_bc_ld = _proc_breached_caller(df_m_ld)
                    _show_caller(df_bc_ld, "No potential breached leads found.")
 
                    df_bc_raw_ld = _get_breached(df_m_ld)
                    if not df_bc_raw_ld.empty:
                        col_dl_bc, _ = st.columns([1, 3])
                        with col_dl_bc:
                            st.download_button(
                                label="📥 Download Breached Leads (.xlsx)",
                                data=_build_leads_xlsx_bytes_ld(df_bc_raw_ld),
                                file_name=f"Breached_Leads_{disp_start}_to_{disp_end}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_bc_leads_ld_xlsx"
                            )
                    st.divider()

                    _ld_section_header("LESS DIALLED LEADS AFTER ASSIGNMENT")
                    st.caption("DNP stages (Call Not Picking Up / Call Not Connected) dialled less than 11 times after assignment to counsellor.")
                    df_ldc_ld = _proc_ld_caller(df_m_ld)
                    _show_caller(df_ldc_ld, "No less-dialled leads found.")

                    df_ld_raw_ld = _get_less_dialled(df_m_ld)
                    if not df_ld_raw_ld.empty:
                        col_dl_ld, _ = st.columns([1, 3])
                        with col_dl_ld:
                            st.download_button(
                                label="📥 Download Less Dialled Leads (.xlsx)",
                                data=_build_leads_xlsx_bytes_ld(df_ld_raw_ld),
                                file_name=f"Less_Dialled_Leads_{disp_start}_to_{disp_end}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_ld_leads_ld_xlsx"
                            )

                    st.session_state.update({
                        'ld_merged_cd'   : df_m_ld.copy(),
                        'ld_ac_cd'       : df_ac_ld,
                        'ld_bc_cd'       : df_bc_ld,
                        'ld_ldc_cd'      : df_ldc_ld,
                        'ld_generated_cd': True,
                    })
        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>📊</div>
                <div style='font-size:.9rem;font-weight:600;'>
                    Select a date range and click <b>Generate Leads Report</b>
                </div>
            </div>""", unsafe_allow_html=True)
 
    # ── TAB 2 ──────────────────────────────────────────────────────────────────
    with tab_ld2:
        if st.session_state.get('ld_generated_cd') and 'ld_merged_cd' in st.session_state:
            df_m_ld  = st.session_state['ld_merged_cd']
            df_ac_ld = st.session_state.get('ld_ac_cd',  pd.DataFrame())
            df_bc_ld = st.session_state.get('ld_bc_cd',  pd.DataFrame())
            df_ld_ld = st.session_state.get('ld_ldc_cd', pd.DataFrame())
 
            df_at_ld  = _proc_assigned_team(df_m_ld)
            df_bt_ld  = _proc_breached_team(df_m_ld)
            df_ldt_ld = _proc_ld_team(df_m_ld)
 
            _ld_section_header("🧠 GENERATED LEAD INSIGHTS")
            insights_ld = _compute_insights(df_ac_ld, df_bc_ld, df_ld_ld, df_at_ld, df_bt_ld, df_ldt_ld)
            if insights_ld:
                ic = st.columns(2)
                for i, ins in enumerate(insights_ld):
                    with ic[i % 2]:
                        st.markdown(f"""
                        <div class="ld-insight-card {ins['type']}">
                            <div style='display:flex;align-items:center;gap:.4rem;'>
                                <span class="insight-icon">{ins['icon']}</span>
                                <span style='font-size:.82rem;font-weight:700;color:var(--text-primary,#111827);'>
                                    {ins['title']}</span>
                            </div>
                            <div style='font-size:.76rem;color:var(--text-muted,#6B7280);line-height:1.5;'>
                                {ins['body']}</div>
                        </div>""", unsafe_allow_html=True)
            else:
                st.info("Not enough data to generate insights.")
 
            st.divider()
            _ld_section_header("TEAMWISE ASSIGNED LEADS DISTRIBUTION")
            _show_team(df_at_ld, "No teamwise data available.")
 
            st.divider()
            _ld_section_header("TEAMWISE POTENTIAL BREACHED LEADS AFTER ASSIGNMENT")
            st.caption("Potential leads not dialled in last 3 days / having a older follow up date.")
            _show_team(df_bt_ld, "No teamwise breached data found.")
 
            st.divider()
            _ld_section_header("TEAMWISE LESS DIALLED LEADS AFTER ASSIGNMENT")
            st.caption("DNP Leads dialled less than 11 times after assignment to counsellor")
            _show_team(df_ldt_ld, "No teamwise less-dialled data found.")
        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🧠</div>
                <div style='font-size:.9rem;font-weight:600;'>
                    Generate a <b>Leads Report</b> first — insights will appear here automatically.
                </div>
            </div>""", unsafe_allow_html=True)


# --- MAIN APP ROUTER ---
if not st.session_state.get('password_correct', False):
    show_homepage_with_login()
else:
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    [data-testid="stStatusWidget"] { display: none !important; }
    header[data-testid="stHeader"] { background: transparent !important; }
    </style>
    """, unsafe_allow_html=True)

    _apply_role_filters()

    ri   = st.session_state.get('auth_role_info', {'role': 'admin'})
    role = ri.get('role', 'admin')
    disp = ri.get('display_name', st.session_state.get('current_user', ''))

    _ROLE_LABELS = {
        'admin'        : '🛡️ Admin',
        'vertical_head': '🏢 Vertical Head',
        'trainer'      : '🎓 Sales Leader',
        'tl'           : '👑 Team Lead',
        'caller'       : '📞 Caller',
    }

    _lc = "#3B82F6"
    _sc = "#1D4ED8"
    _shc = "rgba(59,130,246,.35)"

    st.sidebar.markdown(f"""
    <div style='margin:.4rem 0 .6rem;padding:.55rem .7rem;
                background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);
                border-radius:10px;text-align:center;'>
        <div style='font-size:.6rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:1px;color:{_lc};margin-bottom:.2rem;'>
            {_ROLE_LABELS.get(role, role)}
        </div>
        <div style='font-size:.72rem;color:{_lc};word-break:break-all;'>
            {disp}
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown("""
    <style>
    div[data-testid="stSidebar"] div[data-testid="stButton"]:first-of-type > button {
        display: block !important;
        margin: 0 auto !important;
        width: auto !important;
        min-width: 120px !important;
        padding-left: 1.4rem !important;
        padding-right: 1.4rem !important;
    }
    </style>
    """, unsafe_allow_html=True)
    col_so = st.sidebar.columns([1, 2, 1])
    with col_so[1]:
        _sign_out = st.button("🚪 Sign Out", key="signout_btn")
    if _sign_out:
        try:
            supa.auth.sign_out()
        except Exception:
            pass
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    st.sidebar.markdown(f"""
    <div style='padding:.7rem 0 .5rem;text-align:center;
                border-bottom:1px solid rgba(128,128,128,.15);'>
        <div style='display:flex;align-items:center;justify-content:center;margin-bottom:.25rem;'>
            <span style='font-size:.95rem;font-weight:700;color:{_lc};letter-spacing:-.4px;'>LawSikho</span>
            <div style='width:1px;height:16px;margin:0 .55rem;
                        background:linear-gradient(180deg,transparent,{_sc},transparent);
                        box-shadow:0 0 5px {_shc};'></div>
            <span style='font-size:.95rem;font-weight:700;color:{_lc};letter-spacing:-.4px;'>Skill Arbitrage</span>
        </div>
        <div style='font-size:.6rem;color:{_lc};letter-spacing:1px;font-family:monospace;font-weight:600;'>
            India Learning 📖 India Earning
        </div>
    </div>
    """, unsafe_allow_html=True)

    run_leads_dashboard()
