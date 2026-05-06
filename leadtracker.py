import streamlit as st
import pandas as pd
import requests
import time
import re
import io
import os
from datetime import datetime, timedelta
import pytz

from google.cloud import bigquery
from google.oauth2 import service_account

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from supabase import create_client


# ════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ════════════════════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Realtime Lead Information Tracker — LawSikho",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ════════════════════════════════════════════════════════════════════════════
# BIGQUERY CLIENT
# ════════════════════════════════════════════════════════════════════════════

@st.cache_resource
def get_cached_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        creds = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=creds, project=info["project_id"])
    SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
    if os.path.exists(SERVICE_ACCOUNT_FILE):
        os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
        return bigquery.Client()
    return None


client = get_cached_bq_client()


# ════════════════════════════════════════════════════════════════════════════
# SUPABASE CLIENT
# ════════════════════════════════════════════════════════════════════════════

@st.cache_resource
def get_supabase():
    url = (
        st.secrets.get("SUPABASE_URL")
        or st.secrets.get("supabase", {}).get("url")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_URL")
        or ""
    )
    key = (
        st.secrets.get("SUPABASE_KEY")
        or st.secrets.get("supabase", {}).get("key")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_KEY")
        or ""
    )
    if not url or not key:
        st.error(
            "⚠️ Supabase credentials missing from secrets. "
            "Add SUPABASE_URL and SUPABASE_KEY to your secrets.toml at the TOP LEVEL "
            "(before any [section] header)."
        )
        st.stop()
    return create_client(url, key)


supa = get_supabase()


# ════════════════════════════════════════════════════════════════════════════
# LSQ CREDS
# ════════════════════════════════════════════════════════════════════════════

def get_lsq_creds():
    ak = (
        st.secrets.get("LSQ_ACCESS_KEY")
        or st.secrets.get("lsq", {}).get("access_key", "")
        or os.environ.get("LSQ_ACCESS_KEY", "")
    )
    sk = (
        st.secrets.get("LSQ_SECRET_KEY")
        or st.secrets.get("lsq", {}).get("secret_key", "")
        or os.environ.get("LSQ_SECRET_KEY", "")
    )
    return ak, sk


# ════════════════════════════════════════════════════════════════════════════
# AUTH (copied from main dashboard)
# ════════════════════════════════════════════════════════════════════════════

ADMIN_EMAILS = {
    "admin@lawsikho.in",
    "yash@lawsikho.in",
    "akanksha.patil@lawsikho.in",
    "devki.dt@lawsikho.in",
    "mansi.gupta@lawsikho.in",
    "shivamtejpal@lawsikho.in",
    "pratik.s@lawsikho.in",
    "suraj@lawsikho.in",
    "parul.nagar@lawsikho.in",
    "amitray@lawsikho.in",
    "rinku@lawsikho.in",
    "karunakarareddy@lawsikho.in",
    "priyansh.s@lawsikho.in",
}

VERTICAL_HEAD_TEAMS = {
    "uzair@lawsikho.in"    : ["19th Jan US acc/women ai Closure Batch","CD Closures - Inayat","Contract Drafting","Corporate law - Anas",
                              "Corporate law - Jyoti","Elite","Law firm trainees - Anas","US Acc Closure Specialist - Inayat","US Accounting - Inayat",
                              "US accounting closures - Sana","US accounting trainees","Women ai Trainee - Umme","Women ai/CD closure"],
    "shivya.p@lawsikho.in" : ["ID - 2","ID - 4","ID - 8","ID - 9"],
    "mayur@lawsikho.in"    : ["Changemakers"],
    "deepansi@lawsikho.in" : ["US Accounting","US accounting - Closures"],
    "darshan.c@lawsikho.in": ["ID Closure"],
    "anmol.g@lawsikho.in"  : ["DSV- Aditya","DSV- Shivam","ID Closure - Anmol","Women ai"],
    "abhipsa@lawsikho.in"  : ["CD - Community","CD - Community Manager","Criminal - Community","Criminal - Community Manager","ID - Community","ID - Community Manager"],
}

AUTH_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf"
    "/pub?gid=0&single=true&output=csv"
)

_COL_NAME    = "Caller Name"
_COL_EMAIL   = "Email id"
_COL_DESIG   = "Academic Counselor/TL/ATL"
_COL_TEAM    = "Team Name"
_COL_TRAINER = "Sales Leader"
_TL_VALS     = {"TL", "ATL", "AD", "TEAM LEAD", "TEAM LEADER"}


@st.cache_data(ttl=300, show_spinner=False)
def load_auth_sheet() -> pd.DataFrame:
    df = pd.read_csv(AUTH_SHEET_URL)
    df.columns = df.columns.str.strip()
    if _COL_EMAIL in df.columns:
        df['_email_norm'] = df[_COL_EMAIL].astype(str).str.strip().str.lower()
    return df


@st.cache_data(ttl=300, show_spinner=False)
def load_owner_meta() -> dict:
    """Build a lower-cased Owner-name → {Team Name, Vertical, Analyst, Sales Leader} map
    from the auth sheet so the report Excel can be enriched with team metadata."""
    df = load_auth_sheet()
    if df is None or df.empty or _COL_NAME not in df.columns:
        return {}
    has_team   = 'Team Name'    in df.columns
    has_vert   = 'Vertical'     in df.columns
    has_anal   = 'Analyst'      in df.columns
    has_lead   = 'Sales Leader' in df.columns
    out = {}
    for _, r in df.iterrows():
        name = str(r.get(_COL_NAME, '')).strip()
        if not name:
            continue
        key = name.lower()
        if key in out:
            continue
        def _v(col, present):
            if not present: return ''
            val = r.get(col)
            try:
                if pd.isna(val): return ''
            except (TypeError, ValueError):
                pass
            s = str(val).strip()
            return '' if s.lower() in ('nan', 'none') else s
        out[key] = {
            'Team_Name':    _v('Team Name',    has_team),
            'Vertical':     _v('Vertical',     has_vert),
            'Analyst':      _v('Analyst',      has_anal),
            'Sales_Leader': _v('Sales Leader', has_lead),
        }
    return out


def _extract_trainer_email(cell: str):
    m = re.search(r'\(([^)\s]+@[^)\s]+)\)', str(cell))
    return m.group(1).strip().lower() if m else None


def determine_role(email: str, df: pd.DataFrame):
    el = email.strip().lower()

    if el in {e.lower() for e in ADMIN_EMAILS}:
        return {'role': 'admin', 'teams': None, 'callers': None,
                'caller_name': None, 'display_name': email}

    for vh_mail, vh_teams in VERTICAL_HEAD_TEAMS.items():
        if vh_mail.lower() == el:
            return {'role': 'vertical_head', 'teams': vh_teams, 'callers': None,
                    'caller_name': None, 'display_name': email}

    if '_email_norm' not in df.columns:
        return None

    if _COL_TRAINER in df.columns:
        df2 = df.copy()
        df2['_tr_email'] = df2[_COL_TRAINER].apply(_extract_trainer_email)
        trainer_rows = df2[df2['_tr_email'] == el]
        if not trainer_rows.empty:
            teams   = trainer_rows[_COL_TEAM].dropna().unique().tolist() if _COL_TEAM in trainer_rows.columns else []
            callers = trainer_rows[_COL_NAME].dropna().unique().tolist() if _COL_NAME in trainer_rows.columns else []
            _sl_cell    = str(trainer_rows[_COL_TRAINER].iloc[0])
            _name_match = re.match(r'^([^(]+)\s*\(', _sl_cell)
            _tr_name    = _name_match.group(1).strip() if _name_match else None
            if not _tr_name:
                _own_rows = df[df['_email_norm'] == el]
                _tr_name  = _own_rows.iloc[0][_COL_NAME] if not _own_rows.empty and _COL_NAME in _own_rows.columns else email
            return {'role': 'trainer', 'teams': teams, 'callers': callers,
                    'caller_name': None, 'display_name': _tr_name}

    user_rows = df[df['_email_norm'] == el]
    if user_rows.empty:
        return None

    if _COL_DESIG in user_rows.columns:
        tl_rows = user_rows[
            user_rows[_COL_DESIG].astype(str).str.strip().str.upper().isin(_TL_VALS)
        ]
        if not tl_rows.empty:
            team = tl_rows.iloc[0][_COL_TEAM] if _COL_TEAM in tl_rows.columns else None
            team_callers = (
                df[df[_COL_TEAM] == team][_COL_NAME].dropna().unique().tolist()
                if team else []
            )
            disp = tl_rows.iloc[0][_COL_NAME] if _COL_NAME in tl_rows.columns else email
            return {'role': 'tl', 'teams': [team] if team else [], 'callers': team_callers,
                    'caller_name': None, 'display_name': disp}

    caller_name = user_rows.iloc[0][_COL_NAME] if _COL_NAME in user_rows.columns else email
    return {'role': 'caller', 'teams': None, 'callers': None,
            'caller_name': caller_name, 'display_name': caller_name}


def _complete_login(email: str, session):
    df_auth   = load_auth_sheet()
    role_info = determine_role(email, df_auth)
    if role_info is None:
        st.error("⛔ Your email is not authorised to access this app.")
        return
    st.session_state['password_correct'] = True
    st.session_state['current_user']     = email
    st.session_state['supabase_session'] = session
    st.session_state['auth_role_info']   = role_info
    st.rerun()


def _auth_sign_in_panel():
    email = st.text_input("Email",    key="si_email", placeholder="Your mail ID")
    pwd   = st.text_input("Password", key="si_pwd",   placeholder="Your Password", type="password")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Sign In →", key="si_btn", width='stretch'):
            if not email or not pwd:
                st.error("Enter both email and password."); return
            try:
                resp = supa.auth.sign_in_with_password({"email": email, "password": pwd})
                _complete_login(resp.user.email, resp.session)
            except Exception as ex:
                st.error(f"Login failed: {ex}")
    with c2:
        if st.button("Forgot / Change Password", key="si_otp_switch", width='stretch'):
            st.session_state['auth_tab']           = "otp"
            st.session_state['otp_prefill_email']  = email
            st.session_state['otp_step']           = 1
            st.rerun()


def _auth_otp_panel():
    step = st.session_state.get('otp_step', 1)

    if step == 1:
        email = st.text_input(
            "Email", key="otp_email",
            value=st.session_state.get('otp_prefill_email', ''),
            placeholder="your@lawsikho.in"
        )
        if st.button("Send OTP →", key="otp_send_btn", width='stretch'):
            if not email:
                st.error("Enter your email."); return
            role_check = determine_role(email, load_auth_sheet())
            if role_check is None:
                st.error("⛔ This email is not authorised."); return
            try:
                supa.auth.sign_in_with_otp({
                    "email": email,
                    "options": {"should_create_user": True}
                })
                st.session_state['otp_step']          = 2
                st.session_state['otp_pending_email'] = email
                st.rerun()
            except Exception as ex:
                st.error(f"Could not send OTP: {ex}")

        if st.button("← Back to Sign In", key="otp_back1", width='stretch'):
            st.session_state['auth_tab'] = "signin"
            st.rerun()

    elif step == 2:
        pending_email = st.session_state.get('otp_pending_email', '')
        st.success(f"OTP sent to **{pending_email}** — check your inbox (also spam).")

        otp = st.text_input("OTP code from email", key="otp_code", max_chars=8, placeholder="OTP Received on Email")
        pw1 = st.text_input("Set new password",    key="otp_pw1",  type="password")
        pw2 = st.text_input("Confirm password",    key="otp_pw2",  type="password")

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Verify & Continue →", key="otp_verify_btn", width='stretch'):
                if len(otp) != 8:
                    st.error("Enter the 8-digit code."); return
                if pw1 != pw2:
                    st.error("Passwords don't match."); return
                if len(pw1) < 8:
                    st.error("Password must be ≥ 8 characters."); return
                try:
                    resp = supa.auth.verify_otp({
                        "email": pending_email,
                        "token": otp,
                        "type": "magiclink"
                    })
                    supa.auth.update_user({"password": pw1})
                    st.session_state['otp_step'] = 1
                    _complete_login(resp.user.email, resp.session)
                except Exception as ex:
                    st.error(f"Verification failed — wrong code or it expired: {ex}")
        with c2:
            if st.button("← Back", key="otp_back2", width='stretch'):
                st.session_state['otp_step'] = 1
                st.rerun()


def _apply_role_filters():
    ri = st.session_state.get('auth_role_info', {'role': 'admin'})
    st.session_state['rf_role']         = ri.get('role', 'admin')
    st.session_state['rf_teams']        = ri.get('teams')   or []
    st.session_state['rf_callers']      = ri.get('callers') or []
    st.session_state['rf_caller_name']  = ri.get('caller_name') or ''


# ════════════════════════════════════════════════════════════════════════════
# LOGIN PAGE (matching the existing dashboard style)
# ════════════════════════════════════════════════════════════════════════════

def show_login_page():
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    #MainMenu, header[data-testid="stHeader"] { display: none !important; }
    [data-testid="stToolbar"], [data-testid="stDecoration"], [data-testid="stStatusWidget"], [data-testid="collapsedControl"] { display: none !important; }
    [data-testid="stAppViewContainer"], [data-testid="stMain"], .main { background: #0B1120 !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; margin-bottom: 0 !important; }
    [data-testid="stHorizontalBlock"] { gap: 0 !important; }

    div[data-testid="column"]:nth-child(2),
    div[data-testid="column"]:nth-child(2) > div,
    div[data-testid="column"]:nth-child(2) > div > div,
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"],
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #0f172a !important;
    }
    div[data-testid="column"]:nth-child(2) > div:first-child {
        background-color: #0f172a !important;
        border: 1px solid rgba(255,255,255,.12) !important;
        border-radius: 16px !important;
        padding: 1.6rem 1.8rem 1.8rem !important;
    }
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] {
        background-color: #1e293b !important;
        border-color: rgba(148,163,184,.3) !important;
        border-radius: 10px !important;
    }
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"]:focus-within,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"]:focus-within {
        border-color: #A855F7 !important;
        box-shadow: 0 0 0 2px rgba(168,85,247,.2) !important;
    }
    div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #A855F7 !important;
        border: none !important;
    }
    div[data-testid="column"]:nth-child(2) input::placeholder {
        color: rgba(241,245,249,.32) !important;
    }
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:hover,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:focus {
        -webkit-box-shadow: 0 0 0 1000px #1e293b inset !important;
        -webkit-text-fill-color: #f1f5f9 !important;
    }
    div[data-testid="column"]:nth-child(2) label,
    div[data-testid="column"]:nth-child(2) label p {
        color: rgba(241,245,249,.55) !important;
        font-size: 0.8rem !important;
    }
    div[data-testid="column"]:nth-child(2) .stButton > button {
        width: 100% !important;
        background: linear-gradient(135deg, #A855F7, #7E22CE) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 11px !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
    }
    div[data-testid="column"]:nth-child(2) .stButton > button:hover {
        background: linear-gradient(135deg, #7E22CE, #6B21A8) !important;
        box-shadow: 0 4px 16px rgba(168,85,247,.35) !important;
        transform: translateY(-1px) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    html_hero = """<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;800&family=Plus+Jakarta+Sans:wght@300;400;500&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet"/>
    <style>
    *{box-sizing:border-box;margin:0;padding:0;}
    html,body{font-family:'Plus Jakarta Sans',sans-serif;background:#0B1120;color:#E2E8F0;overflow-x:hidden;}
    body{background:
        radial-gradient(ellipse 80% 50% at 50% -10%,rgba(168,85,247,.14) 0%,transparent 60%),
        radial-gradient(ellipse 60% 40% at 90% 80%,rgba(139,92,246,.1) 0%,transparent 55%),
        #0B1120;}
    body::before{content:"";position:fixed;inset:0;
        background-image:linear-gradient(rgba(255,255,255,.022) 1px,transparent 1px),
            linear-gradient(90deg,rgba(255,255,255,.022) 1px,transparent 1px);
        background-size:48px 48px;pointer-events:none;}
    .hero{display:flex;flex-direction:column;align-items:center;text-align:center;padding:3rem 2rem 2.5rem;}
    .logos{display:flex;align-items:center;margin-bottom:1rem;}
    .lname{font-size:1.25rem;font-weight:700;color:#fff;padding:0 1.6rem;}
    .lsep{width:1px;height:48px;background:linear-gradient(180deg,transparent,rgba(168,85,247,.85),transparent);box-shadow:0 0 8px rgba(168,85,247,.5);}
    .tagline{font-family:'Fira Code',monospace;font-size:.75rem;color:rgba(255,255,255,.32);letter-spacing:1.5px;margin-bottom:2rem;}
    .eyebrow{display:inline-flex;align-items:center;gap:.45rem;font-family:'Fira Code',monospace;font-size:.65rem;
        letter-spacing:2.5px;text-transform:uppercase;color:#A855F7;background:rgba(168,85,247,.08);
        border:1px solid rgba(168,85,247,.18);border-radius:100px;padding:.28rem .95rem;margin-bottom:1.2rem;}
    .dot{width:5px;height:5px;background:#A855F7;border-radius:50%;box-shadow:0 0 5px #A855F7;animation:p 2s ease-in-out infinite;}
    @keyframes p{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.45;transform:scale(1.4);}}
    .headline{font-family:'Playfair Display',serif;font-size:clamp(2rem,5vw,3.2rem);font-weight:800;line-height:1.09;color:#fff;letter-spacing:-1.5px;margin-bottom:.65rem;}
    .accent{background:linear-gradient(125deg,#A855F7,#C084FC 40%,#E9D5FF);-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
    .sub{font-size:1rem;font-weight:300;color:rgba(255,255,255,.38);max-width:560px;}
    </style></head><body>
    <div class="hero">
      <div class="logos">
        <div class="lname">LawSikho</div>
        <div class="lsep"></div>
        <div class="lname">Skill Arbitrage</div>
      </div>
      <div class="tagline">India Learning &nbsp;📖&nbsp; India Earning</div>
      <div class="eyebrow"><span class="dot"></span>Realtime Lead Information Tracker</div>
      <div class="headline">Lead status,<br><span class="accent">in real time</span></div>
      <div class="sub">Upload your Assigned-Lead List — Get the latest LSQ information & Enrollment leads mapped in one Excel Sheet.</div>
    </div>
    </body></html>"""
    st.iframe(html_hero, height=420)

    left, mid, right = st.columns([1, 1, 1])
    with mid:
        auth_tab = st.session_state.get('auth_tab', 'signin')

        st.markdown("""
        <div style="text-align:center;margin-bottom:.8rem;">
            <span style="font-family:'Playfair Display',serif;font-size:1.1rem;font-weight:600;color:#fff;">
                🔐 APP ACCESS
            </span>
        </div>
        """, unsafe_allow_html=True)

        if auth_tab == 'signin':
            _auth_sign_in_panel()
        else:
            _auth_otp_panel()

        if st.session_state.get('password_correct'):
            st.rerun()

    # ── Footer (matching consolidated dashboard) ──
    html_footer = """<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Fira+Code:wght@400;500&display=swap" rel="stylesheet"/>
    <style>
    *{box-sizing:border-box;margin:0;padding:0;}
    html,body{background:#0B1120;color:#E2E8F0;font-family:'Fira Code',monospace;}
    .foot{border-top:1px solid rgba(255,255,255,.06);padding:1.6rem 2rem;text-align:center;margin-top:2rem;}
    .f1{font-size:.64rem;color:rgba(255,255,255,.28);margin-bottom:.4rem;}
    .f2{font-size:.58rem;color:rgba(255,255,255,.14);}
    .fd{display:inline-block;width:3px;height:3px;background:rgba(168,85,247,.4);border-radius:50%;margin:0 .45rem;vertical-align:middle;}
    </style></head><body>
    <div class="foot">
      <div class="f1">For Internal Use of Sales and Operations Team Only<span class="fd"></span>All Rights Reserved</div>
      <div class="f2">Developed and Designed by Amit Ray<span class="fd"></span>Reach out for Support and Queries</div>
    </div>
    </body></html>"""
    st.iframe(html_footer, height=120)


# ════════════════════════════════════════════════════════════════════════════
# CORE HELPERS
# ════════════════════════════════════════════════════════════════════════════

def clean_phone(p) -> str:
    """Strip all non-digits and return last 10 digits (for matching only)."""
    if p is None: return ''
    try:
        if pd.isna(p): return ''
    except Exception:
        pass
    s = str(p).strip()
    if not s or s.lower() in ('nan', 'none'): return ''
    digits = re.sub(r'\D', '', s)
    return digits[-10:] if len(digits) >= 10 else digits


def clean_pid(v) -> str:
    """Normalise a Prospect ID: strip whitespace; if numeric float like 1.234e+18 → int str."""
    if v is None: return ''
    try:
        if pd.isna(v): return ''
    except Exception:
        pass
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'none'): return ''
    # If numeric-only, drop any trailing .0
    if re.match(r'^\d+\.0+$', s):
        s = s.split('.')[0]
    return s


def clean_email(v) -> str:
    if v is None: return ''
    try:
        if pd.isna(v): return ''
    except Exception:
        pass
    s = str(v).strip().lower()
    if not s or s in ('nan', 'none'): return ''
    return s


def detect_fields(columns):
    """Auto-detect which columns map to ProspectID, Email, PhoneNumber, AlternatePhoneNumber."""
    cols_lower = {c: str(c).strip().lower() for c in columns}
    result = {'ProspectID': None, 'Email': None, 'PhoneNumber': None, 'AlternatePhoneNumber': None}

    # Prospect ID
    for c, cl in cols_lower.items():
        if re.search(r'prospect.?id|prospectid|^lead.?id$|^lead_id$', cl):
            result['ProspectID'] = c; break

    # Email
    for c, cl in cols_lower.items():
        if 'email' in cl or re.search(r'e[-_]?mail', cl):
            result['Email'] = c; break

    # Alt phone first (so we don't grab it as primary phone)
    for c, cl in cols_lower.items():
        if re.search(r'(alt|alternat|secondary|other).*(phone|number|mobile|no)|mobile.?2|phone.?2|number.?2', cl):
            result['AlternatePhoneNumber'] = c; break

    # Primary phone (skip if same column was already taken as alt)
    for c, cl in cols_lower.items():
        if c == result['AlternatePhoneNumber']: continue
        if re.search(r'alt|alternat|secondary', cl): continue
        if re.search(r'phone|mobile|^contact.?n|^number$|^contact.?no$|primary', cl):
            result['PhoneNumber'] = c; break

    return result


def get_sample_csv_bytes() -> bytes:
    sample = pd.DataFrame({
        'Prospect ID':      [],
        'Email':            [],
        'Phone Number':     [],
        'Alternate Number': [],
    })
    return sample.to_csv(index=False).encode('utf-8')


# ════════════════════════════════════════════════════════════════════════════
# LSQ FETCH
# ════════════════════════════════════════════════════════════════════════════

LSQ_FIELDS = (
    "ProspectID,FirstName,LastName,Phone,Mobile,EmailAddress,mx_Bootcamp_attended,"
    "mx_Funnel_name,mx_Bootcamp_Attendance,OwnerId,ProspectStage,"
    "mx_Follow_up_date_and_time,mx_Course_Fees,mx_Campaign_Name,mx_Enquired_Course,"
    "mx_Phone_call_counter,mx_Last_Call_new,mx_Report_Type,ModifiedByName,"
    "mx_Assigned_By,mx_Assigned_On,mx_Assigned_On_Call_Counter"
)


def fetch_lsq_users(access_key, secret_key):
    url = "https://api-in21.leadsquared.com/v2/UserManagement.svc/Users.Get"
    r = requests.get(url, params={"accessKey": access_key, "secretKey": secret_key}, timeout=60)
    if r.status_code != 200:
        raise RuntimeError(f"LSQ Users API HTTP {r.status_code}: {r.text[:300]}")
    users = r.json()
    if not isinstance(users, list):
        raise RuntimeError(f"LSQ Users API returned non-list: {str(users)[:300]}")
    return {u.get("ID"): f"{u.get('FirstName','')} {u.get('LastName','')}".strip() for u in users}


def fetch_lsq_leads(access_key, secret_key, days_back, status_callback=None):
    """Fetch all leads from LSQ where mx_Assigned_On > today - days_back."""
    india_tz = pytz.timezone("Asia/Kolkata")
    now = datetime.now(india_tz)
    cutoff = (now - timedelta(days=days_back)).strftime("%Y-%m-%d %H:%M:%S")

    URL = (f"https://api-in21.leadsquared.com/v2/LeadManagement.svc/Leads.Get"
           f"?accessKey={access_key}&secretKey={secret_key}")
    page = 1
    page_size = 1000
    all_data = []
    retry_count = 0
    MAX_RETRIES = 10

    while True:
        payload = {
            "Parameter": {
                "LookupName": "mx_Assigned_On",
                "LookupValue": cutoff,
                "SqlOperator": ">",
            },
            "Columns": {"Include_CSV": LSQ_FIELDS},
            "Paging": {"PageIndex": page, "PageSize": page_size},
        }
        try:
            resp = requests.post(URL, headers={"Content-Type": "application/json"}, json=payload, timeout=120)
            data = resp.json()
        except Exception as ex:
            retry_count += 1
            if retry_count > MAX_RETRIES:
                raise RuntimeError(f"LSQ Leads API failed after retries: {ex}")
            time.sleep(retry_count * 5)
            continue

        if isinstance(data, dict):
            # Either an error or "no more pages" sometimes returns {}
            retry_count += 1
            if retry_count > MAX_RETRIES:
                break
            time.sleep(retry_count * 5)
            continue

        retry_count = 0
        if not data:
            break

        all_data.extend(data)
        if status_callback:
            status_callback(len(all_data), page)
        page += 1
        time.sleep(1)

    return all_data, cutoff


def build_lsq_df(raw_data, owner_map):
    df = pd.DataFrame(raw_data)
    if df.empty:
        return df
    for col in LSQ_FIELDS.split(","):
        if col not in df.columns:
            df[col] = None

    df['OwnerId']             = df['OwnerId'].map(owner_map).fillna(df['OwnerId'])
    df['Phone_clean']         = df['Phone'].apply(clean_phone)
    df['Mobile_clean']        = df['Mobile'].apply(clean_phone)
    df['EmailAddress_clean']  = df['EmailAddress'].apply(clean_email)
    df['ProspectID_clean']    = df['ProspectID'].apply(clean_pid)
    df['Follow_up_date']      = pd.to_datetime(df['mx_Follow_up_date_and_time'], errors='coerce').dt.date
    df['LastCalledDate']      = pd.to_datetime(df['mx_Last_Call_new'], errors='coerce').dt.date
    df['AssignedOn']          = pd.to_datetime(df['mx_Assigned_On'], errors='coerce').dt.date
    return df


# ════════════════════════════════════════════════════════════════════════════
# REVENUE OVERRIDE
# ════════════════════════════════════════════════════════════════════════════

def fetch_revenue_maps():
    if client is None:
        return {}, {}
    q = """
        SELECT Contact_No, Email_Id, Caller_name
        FROM `studious-apex-488820-c3.crm_dashboard.revenue_sheet`
        WHERE Enrollment = 'New enrollment' AND Fee_paid >= 999
    """
    df = client.query(q).to_dataframe()
    if df.empty:
        return {}, {}
    df['Contact_No'] = df['Contact_No'].apply(clean_phone)
    df['Email_Id']   = df['Email_Id'].apply(clean_email)
    phone_map = {p: c for p, c in zip(df['Contact_No'], df['Caller_name']) if p}
    email_map = {e: c for e, c in zip(df['Email_Id'],   df['Caller_name']) if e}
    return phone_map, email_map


# ════════════════════════════════════════════════════════════════════════════
# MAPPING USER → LSQ
# ════════════════════════════════════════════════════════════════════════════

OUTPUT_COLS = [
    'ProspectID', 'FirstName', 'LastName', 'PhoneNumber', 'Alternate_PhoneNumber',
    'Email', 'Funnel_name', 'Bootcamp_Attendance', 'Owner', 'ContactStage',
    'Follow_up_date', 'Campaign_Name', 'Enquired_Course',
    'Phone_call_counter', 'LastCalledDate', 'AssignedBy', 'AssignedOn',
    'Assigned_On_Call_Counter',
    'Team_Name', 'Vertical', 'Analyst', 'Sales_Leader',
]


def map_user_to_lsq(df_user, df_lsq, mapping, phone_map_rev, email_map_rev):
    """For each user row, find matching LSQ row using priority. Apply revenue override."""
    if df_lsq is None or df_lsq.empty:
        df_lsq = pd.DataFrame()

    # Build LSQ index dicts (first occurrence wins)
    lsq_by_pid    = {}
    lsq_by_email  = {}
    lsq_by_phone  = {}
    lsq_by_mobile = {}

    if not df_lsq.empty:
        for idx, row in df_lsq.iterrows():
            pid = row.get('ProspectID_clean', '')
            if pid and pid not in lsq_by_pid:
                lsq_by_pid[pid] = idx
            em = row.get('EmailAddress_clean', '')
            if em and em not in lsq_by_email:
                lsq_by_email[em] = idx
            ph = row.get('Phone_clean', '')
            if ph and ph not in lsq_by_phone:
                lsq_by_phone[ph] = idx
            mo = row.get('Mobile_clean', '')
            if mo and mo not in lsq_by_mobile:
                lsq_by_mobile[mo] = idx

    def find_match(urow):
        # 1. ProspectID
        if mapping.get('ProspectID'):
            v = clean_pid(urow.get(mapping['ProspectID']))
            if v and v in lsq_by_pid:
                return lsq_by_pid[v], 'ProspectID'
        # 2. Email
        if mapping.get('Email'):
            v = clean_email(urow.get(mapping['Email']))
            if v and v in lsq_by_email:
                return lsq_by_email[v], 'Email'
        # 3. Phone (try LSQ Phone first, then Mobile)
        if mapping.get('PhoneNumber'):
            v = clean_phone(urow.get(mapping['PhoneNumber']))
            if v:
                if v in lsq_by_phone:  return lsq_by_phone[v],  'PhoneNumber'
                if v in lsq_by_mobile: return lsq_by_mobile[v], 'PhoneNumber→Mobile'
        # 4. Alternate phone
        if mapping.get('AlternatePhoneNumber'):
            v = clean_phone(urow.get(mapping['AlternatePhoneNumber']))
            if v:
                if v in lsq_by_phone:  return lsq_by_phone[v],  'AltPhone→Phone'
                if v in lsq_by_mobile: return lsq_by_mobile[v], 'AltPhone→Mobile'
        return None, None

    out_rows = []
    for _, urow in df_user.iterrows():
        idx, match_type = find_match(urow)
        out = {c: '' for c in OUTPUT_COLS}
        out['_Match_Type'] = match_type or 'Not Found in LSQ'

        # Carry forward whatever the user provided so even unmatched rows are useful
        if mapping.get('ProspectID'):
            out['ProspectID'] = urow.get(mapping['ProspectID'], '') if pd.notna(urow.get(mapping['ProspectID'], '')) else ''
        if mapping.get('Email'):
            out['Email'] = urow.get(mapping['Email'], '') if pd.notna(urow.get(mapping['Email'], '')) else ''
        if mapping.get('PhoneNumber'):
            out['PhoneNumber'] = urow.get(mapping['PhoneNumber'], '') if pd.notna(urow.get(mapping['PhoneNumber'], '')) else ''
        if mapping.get('AlternatePhoneNumber'):
            out['Alternate_PhoneNumber'] = urow.get(mapping['AlternatePhoneNumber'], '') if pd.notna(urow.get(mapping['AlternatePhoneNumber'], '')) else ''

        if idx is not None:
            lrow = df_lsq.iloc[idx]
            out.update({
                'ProspectID':              lrow.get('ProspectID', '')             or out['ProspectID'],
                'FirstName':               lrow.get('FirstName', '')               or '',
                'LastName':                lrow.get('LastName', '')                or '',
                'PhoneNumber':             lrow.get('Phone', '')                   or out['PhoneNumber'],
                'Alternate_PhoneNumber':   lrow.get('Mobile', '')                  or out['Alternate_PhoneNumber'],
                'Email':                   lrow.get('EmailAddress', '')            or out['Email'],
                'Funnel_name':             lrow.get('mx_Funnel_name', '')          or '',
                'Bootcamp_Attendance':     lrow.get('mx_Bootcamp_Attendance', '')  or '',
                'Owner':                   lrow.get('OwnerId', '')                 or '',
                'ContactStage':            lrow.get('ProspectStage', '')           or '',
                'Follow_up_date':          lrow.get('Follow_up_date', '')          or '',
                'Campaign_Name':           lrow.get('mx_Campaign_Name', '')        or '',
                'Enquired_Course':         lrow.get('mx_Enquired_Course', '')      or '',
                'Phone_call_counter':      lrow.get('mx_Phone_call_counter', '')   or '',
                'LastCalledDate':          lrow.get('LastCalledDate', '')          or '',
                'AssignedBy':              lrow.get('mx_Assigned_By', '')          or '',
                'AssignedOn':              lrow.get('AssignedOn', '')              or '',
                'Assigned_On_Call_Counter': lrow.get('mx_Assigned_On_Call_Counter', '') or '',
            })

        # Revenue override — applies to BOTH matched and unmatched rows
        # so an "Actually Enrolled" lead shows up even if LSQ didn't return it.
        ph_keys  = [clean_phone(out.get('PhoneNumber', '')), clean_phone(out.get('Alternate_PhoneNumber', ''))]
        em_key   = clean_email(out.get('Email', ''))
        overridden = False
        for pk in ph_keys:
            if pk and pk in phone_map_rev:
                out['Owner']        = phone_map_rev[pk]
                out['ContactStage'] = 'Actually Enrolled'
                overridden = True
                break
        if not overridden and em_key and em_key in email_map_rev:
            out['Owner']        = email_map_rev[em_key]
            out['ContactStage'] = 'Actually Enrolled'

        out_rows.append(out)

    return pd.DataFrame(out_rows)


# ════════════════════════════════════════════════════════════════════════════
# OWNER META ENRICHMENT
# ════════════════════════════════════════════════════════════════════════════

def enrich_with_owner_meta(df_out: pd.DataFrame, owner_meta: dict) -> pd.DataFrame:
    """Populate Team_Name, Vertical, Analyst, Sales_Leader by matching Owner against
    the auth-sheet's 'Caller Name' column."""
    if df_out is None or df_out.empty:
        return df_out
    blank = {'Team_Name': '', 'Vertical': '', 'Analyst': '', 'Sales_Leader': ''}

    def _row(owner):
        m = owner_meta.get(str(owner).strip().lower(), blank)
        return pd.Series({
            'Team_Name':    m.get('Team_Name', ''),
            'Vertical':     m.get('Vertical', ''),
            'Analyst':      m.get('Analyst', ''),
            'Sales_Leader': m.get('Sales_Leader', ''),
        })

    meta_df = df_out['Owner'].apply(_row)
    for c in ('Team_Name', 'Vertical', 'Analyst', 'Sales_Leader'):
        df_out[c] = meta_df[c]
    return df_out


# ════════════════════════════════════════════════════════════════════════════
# INSIGHTS
# ════════════════════════════════════════════════════════════════════════════

def compute_lead_insights(df_out: pd.DataFrame):
    """Six summary cards for the Insights tab. Filters out 'Not Found in LSQ' rows first."""
    cards = []
    if df_out is None or df_out.empty:
        return cards
    df = df_out[df_out['_Match_Type'] != 'Not Found in LSQ'].copy()
    if df.empty:
        return cards

    df['_stage'] = df['ContactStage'].astype(str).str.strip()
    df['_owner'] = df['Owner'].astype(str).str.strip()
    df['_aoc']   = pd.to_numeric(df['Assigned_On_Call_Counter'], errors='coerce')

    is_followup = df['_stage'].str.lower() == 'follow up for closure'
    is_callback = df['_stage'].str.lower() == 'call back later'
    is_low_dial = (
        df['_stage'].str.lower().isin(['call not picking up', 'call not connected'])
        & (df['_aoc'] < 4)
    )

    fu_count = int(is_followup.sum())
    cards.append({
        'type': 'info', 'icon': '📞',
        'title': 'Follow Ups Made',
        'body': f"{fu_count} leads with Contact Stage = <b>Follow Up For Closure</b>."
    })

    cb_count = int(is_callback.sum())
    cards.append({
        'type': 'info', 'icon': '⏰',
        'title': 'Call Back Later',
        'body': f"{cb_count} leads with Contact Stage = <b>Call Back Later</b>."
    })

    ld_count = int(is_low_dial.sum())
    cards.append({
        'type': 'warn', 'icon': '⚠️',
        'title': 'Less Dialled CNPU & CNC',
        'body': f"{ld_count} less dialled CNPU and CNC leads dialled less than 4 times after lead assignment."
    })

    fu_top = df.loc[is_followup & (df['_owner'] != ''), '_owner'].value_counts()
    if not fu_top.empty and int(fu_top.iloc[0]) > 0:
        cards.append({
            'type': 'good', 'icon': '🏆',
            'title': f"Top Follow Ups: {fu_top.index[0]}",
            'body': f"{fu_top.index[0]} has <b>{int(fu_top.iloc[0])}</b> Follow Up For Closure leads — highest among all owners."
        })
    else:
        cards.append({
            'type': 'good', 'icon': '🏆',
            'title': "Top Follow Ups",
            'body': "No Follow Up For Closure leads found."
        })

    cb_top = df.loc[is_callback & (df['_owner'] != ''), '_owner'].value_counts()
    if not cb_top.empty and int(cb_top.iloc[0]) > 0:
        cards.append({
            'type': 'good', 'icon': '🥇',
            'title': f"Top Call Back Later: {cb_top.index[0]}",
            'body': f"{cb_top.index[0]} has <b>{int(cb_top.iloc[0])}</b> Call Back Later leads — highest among all owners."
        })
    else:
        cards.append({
            'type': 'good', 'icon': '🥇',
            'title': "Top Call Back Later",
            'body': "No Call Back Later leads found."
        })

    ld_top = df.loc[is_low_dial & (df['_owner'] != ''), '_owner'].value_counts()
    if not ld_top.empty and int(ld_top.iloc[0]) > 0:
        cards.append({
            'type': 'bad', 'icon': '🚨',
            'title': f"Most Less-Dialled: {ld_top.index[0]}",
            'body': f"{ld_top.index[0]} has <b>{int(ld_top.iloc[0])}</b> less dialled CNPU and CNC leads dialled less than 4 times after lead assignment."
        })
    else:
        cards.append({
            'type': 'bad', 'icon': '🚨',
            'title': "Most Less-Dialled Owner",
            'body': "No CNPU/CNC leads with call counter under 4."
        })

    return cards


# ════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ════════════════════════════════════════════════════════════════════════════

def build_output_xlsx(df_out, days_back) -> bytes:
    SUB_FILL  = PatternFill("solid", start_color="6B21A8", end_color="6B21A8")
    ALT_FILL  = PatternFill("solid", start_color="FAF5FF", end_color="FAF5FF")
    WHT_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
    OK_FILL   = PatternFill("solid", start_color="DCFCE7", end_color="DCFCE7")
    NF_FILL   = PatternFill("solid", start_color="FEE2E2", end_color="FEE2E2")
    AE_FILL   = PatternFill("solid", start_color="FEF3C7", end_color="FEF3C7")
    HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    DATA_FONT = Font(name="Calibri", size=10)
    CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    BORDER    = Border(
        left=Side(style='thin', color='E9D5FF'), right=Side(style='thin', color='E9D5FF'),
        top=Side(style='thin',  color='E9D5FF'), bottom=Side(style='thin', color='E9D5FF'),
    )

    cols = OUTPUT_COLS + ['_Match_Type']
    col_headers = [
        'PROSPECT ID', 'FIRST NAME', 'LAST NAME', 'PHONE NUMBER', 'ALTERNATE PHONE',
        'EMAIL', 'FUNNEL NAME', 'BOOTCAMP ATTENDANCE', 'OWNER', 'CONTACT STAGE',
        'FOLLOW UP DATE', 'CAMPAIGN NAME', 'ENQUIRED COURSE',
        'PHONE CALL COUNTER', 'LAST CALLED DATE', 'ASSIGNED BY', 'ASSIGNED ON',
        'ASSIGNED ON CALL COUNTER',
        'TEAM NAME', 'VERTICAL', 'ANALYST', 'SALES LEADER',
        'MATCH TYPE',
    ]
    col_widths = [22, 18, 18, 16, 16, 32, 24, 16, 22, 22, 14, 24, 28, 14, 14, 22, 14, 14,
                  22, 18, 18, 32, 18]

    wb = Workbook()
    ws = wb.active
    ws.title = "Lead Tracker Output"

    # Header row (row 1)
    for ci, h in enumerate(col_headers, 1):
        cell = ws.cell(1, ci, h)
        cell.fill = SUB_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[1].height = 28

    # Data rows (start at row 2)
    for r_idx, (_, row) in enumerate(df_out.iterrows(), 2):
        is_alt = (r_idx % 2 == 0)
        match_type = str(row.get('_Match_Type', ''))
        contact_stage = str(row.get('ContactStage', '')).strip().lower()

        # Pick base fill
        if contact_stage == 'actually enrolled':
            base_fill = AE_FILL
        elif match_type == 'Not Found in LSQ':
            base_fill = NF_FILL
        elif match_type and match_type != 'Not Found in LSQ':
            base_fill = OK_FILL if not is_alt else ALT_FILL
        else:
            base_fill = ALT_FILL if is_alt else WHT_FILL

        for ci, col in enumerate(cols, 1):
            val = row.get(col, '')
            try:
                if pd.isna(val): val = ''
            except (TypeError, ValueError):
                pass
            if hasattr(val, 'strftime'):
                val = val.strftime('%Y-%m-%d')
            cell = ws.cell(r_idx, ci, val)
            cell.fill = base_fill; cell.font = DATA_FONT; cell.border = BORDER
            cell.alignment = LEFT if col in ('FirstName', 'LastName', 'Email', 'Owner', 'ContactStage', 'Funnel_name', 'Campaign_Name', 'Enquired_Course', 'AssignedBy', 'Team_Name', 'Vertical', 'Analyst', 'Sales_Leader', '_Match_Type') else CENTER

    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ════════════════════════════════════════════════════════════════════════════

def run_lead_tracker():
    # CSS
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }
    [data-testid="stMainBlockContainer"] {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    .block-container { max-width: 100% !important; }
    footer { visibility: hidden; }
    [data-testid="stStatusWidget"] { display: none !important; }

    .lt-header {
        background: linear-gradient(135deg, #1e1b4b 0%, #4c1d95 45%, #6b21a8 100%);
        border-radius: 16px;
        padding: 1.5rem 2rem 1.2rem;
        margin-bottom: 1.2rem;
        position: relative; overflow: hidden;
        box-shadow: 0 8px 32px rgba(0,0,0,.14);
    }
    .lt-header::before {
        content: ""; position: absolute; top: -40px; right: -40px;
        width: 200px; height: 200px;
        background: radial-gradient(circle, rgba(168,85,247,.25) 0%, transparent 70%);
        border-radius: 50%;
    }
    .lt-title    { font-size: 1.65rem; font-weight: 700; color: #FFF; letter-spacing: .5px; margin: 0 0 .25rem; }
    .lt-subtitle { font-size: .82rem; color: rgba(255,255,255,.6); margin: 0; font-family: 'DM Mono', monospace; }
    .lt-badge {
        display: inline-flex; align-items: center; gap: 5px;
        background: rgba(255,255,255,.12); border: 1px solid rgba(255,255,255,.18);
        border-radius: 20px; padding: 3px 10px; font-size: .73rem;
        color: rgba(255,255,255,.9); font-family: 'DM Mono', monospace;
    }
    .lt-pulse {
        width: 6px; height: 6px; background: #C084FC; border-radius: 50%;
        display: inline-block; animation: pulse-lt 1.8s ease-in-out infinite;
    }
    @keyframes pulse-lt { 0%,100%{opacity:1;transform:scale(1);} 50%{opacity:.5;transform:scale(1.4);} }

    .lt-step-card {
        background: var(--metric-bg, #fff);
        border: 1px solid rgba(168,85,247,.15);
        border-radius: 12px;
        padding: 1rem 1.2rem;
        margin-bottom: .8rem;
        box-shadow: 0 1px 3px rgba(0,0,0,.05);
    }
    .lt-step-num {
        display:inline-flex; width:28px; height:28px; border-radius:50%;
        background: linear-gradient(135deg, #A855F7, #7E22CE); color:#fff;
        font-weight:700; font-size:.85rem; align-items:center; justify-content:center;
        margin-right:.6rem;
    }

    .stButton > button {
        background: linear-gradient(135deg, #7E22CE, #4C1D95) !important;
        color: #fff !important; border: none !important;
        font-weight: 600 !important; border-radius: 8px !important;
    }
    .stButton > button:hover { opacity: .88; transform: translateY(-1px); }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #7E22CE, #4C1D95) !important;
        color: #fff !important; border: none !important;
        font-weight: 600 !important; border-radius: 8px !important;
        width: 100% !important;
    }

    div[data-testid="stDataFrame"] thead tr th {
        background: linear-gradient(135deg, #4c1d95, #7e22ce) !important;
        color: #fff !important;
        font-size: .72rem !important; font-weight: 700 !important;
        text-transform: uppercase; letter-spacing: .5px;
        text-align: center !important; padding: 8px !important;
    }

    .lt-section-header { display: flex; align-items: center; gap: .6rem; margin: 1.5rem 0 .8rem; }
    .lt-section-line   { flex: 1; height: 1px; background: linear-gradient(90deg, #A855F7, transparent); opacity: .4; }
    .lt-section-line.r { background: linear-gradient(90deg, transparent, #A855F7); }
    .lt-section-title  { font-size: .8rem; font-weight: 700; text-transform: uppercase;
                         letter-spacing: 1.2px; color: #A855F7; white-space: nowrap; text-align: center; }

    .insight-card {
        background: #fff;
        border: 1px solid rgba(168,85,247,.15);
        border-radius: 12px; padding: 1rem 1.1rem;
        margin-bottom: .6rem; box-shadow: 0 1px 3px rgba(0,0,0,.05);
        transition: all .2s ease;
    }
    .insight-card:hover { box-shadow: 0 4px 12px rgba(0,0,0,.08); }
    .insight-card.good  { border-left: 3px solid #16A34A; }
    .insight-card.warn  { border-left: 3px solid #F59E0B; }
    .insight-card.bad   { border-left: 3px solid #EF4444; }
    .insight-card.info  { border-left: 3px solid #A855F7; }
    .insight-icon  { font-size: 1.1rem; }
    .insight-title { font-size: .82rem; font-weight: 700; color: #111827; margin: .2rem 0; text-align: center; }
    .insight-body  { font-size: .76rem; color: #6B7280; line-height: 1.5; text-align: center; }

    [data-testid="stTabs"] [role="tablist"] { gap: .3rem; border-bottom: 1px solid rgba(168,85,247,.15); }
    [data-testid="stTabs"] button[role="tab"] {
        font-family: 'DM Sans', sans-serif !important; font-size: .85rem !important;
        font-weight: 600 !important; letter-spacing: .3px;
        border-radius: 8px 8px 0 0; padding: .55rem 1.1rem !important;
    }
    [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
        color: #6B21A8 !important; border-bottom: 2px solid #A855F7 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Header
    st.markdown(f"""
    <div class="lt-header">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:.75rem;">
            <div>
                <div class="lt-title">🎯 REALTIME LEAD INFORMATION TRACKER</div>
                <div class="lt-subtitle">UPLOAD &middot; FETCH &middot; MAP &middot; DOWNLOAD</div>
            </div>
            <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.25rem;">
                <span class="lt-badge"><span class="lt-pulse"></span>LSQ & REVENUE DATA LIVE</span>
                <span class="lt-badge">🕐 {datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%d-%b-%Y %I:%M %p')}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ─── STEP 1 — File upload + days input + sample CSV ─────────────────────
    st.markdown("""
    <div class="lt-step-card">
        <span class="lt-step-num">1</span>
        <span style="font-weight:700;font-size:1rem;color:#111111;">Upload your assigned-lead list</span>
        <div style="font-size:.78rem;color:#6B7280;margin-top:.3rem;margin-left:2.2rem;">
            Accepted formats: <b>.csv</b> &middot; <b>.xlsx</b>. The file should contain at least one of:
            Prospect ID, Email, Phone Number or Alternate Number field to process.
        </div>
    </div>
    """, unsafe_allow_html=True)

    col_a, col_b = st.columns([2, 1])
    with col_a:
        uploaded = st.file_uploader(
            "Upload CSV or Excel",
            type=['csv', 'xlsx', 'xls'],
            key='lt_upload',
            label_visibility='collapsed'
        )
    with col_b:
        st.download_button(
            "📥 Download Sample CSV",
            data=get_sample_csv_bytes(),
            file_name="sample_leads.csv",
            mime="text/csv",
            key="dl_sample",
            width='stretch',
        )

    days_back = st.number_input(
        "📅 When were these leads assigned to counsellors? (days ago — between 1 and 60)",
        min_value=1, max_value=60, value=1, step=1, key='lt_days'
    )

    # Parse uploaded file
    df_user = None
    if uploaded is not None:
        try:
            if uploaded.name.lower().endswith('.csv'):
                df_user = pd.read_csv(uploaded, dtype=str)
            else:
                df_user = pd.read_excel(uploaded, dtype=str)
            df_user.columns = df_user.columns.str.strip()
        except Exception as ex:
            st.error(f"❌ Could not read the file: {ex}")
            df_user = None

    # ─── STEP 2 — Field mapping ─────────────────────────────────────────────
    mapping = None
    if df_user is not None and not df_user.empty:
        st.markdown("""
        <div class="lt-step-card" style="margin-top:1rem;">
            <span class="lt-step-num">2</span>
            <span style="font-weight:700;font-size:1rem;color:#111111;">Confirm field mapping</span>
            <div style="font-size:.78rem;color:#6B7280;margin-top:.3rem;margin-left:2.2rem;">
                We auto-detect the columns. Adjust below if anything looks wrong.
            </div>
        </div>
        """, unsafe_allow_html=True)

        st.caption(f"📄 File loaded: **{uploaded.name}** &middot; {len(df_user):,} rows &middot; {len(df_user.columns)} columns")

        with st.expander("👁️ Preview first 5 rows", expanded=False):
            st.dataframe(df_user.head(5), width='stretch', hide_index=True)

        detected = detect_fields(list(df_user.columns))
        col_options = ['(none)'] + list(df_user.columns)

        c1, c2, c3, c4 = st.columns(4)

        def _idx(value):
            return col_options.index(value) if value in col_options else 0

        with c1:
            sel_pid = st.selectbox(
                "🆔 Prospect ID",
                col_options,
                index=_idx(detected['ProspectID']),
                key='map_pid',
            )
        with c2:
            sel_email = st.selectbox(
                "📧 Email",
                col_options,
                index=_idx(detected['Email']),
                key='map_email',
            )
        with c3:
            sel_phone = st.selectbox(
                "📱 Phone Number",
                col_options,
                index=_idx(detected['PhoneNumber']),
                key='map_phone',
            )
        with c4:
            sel_alt = st.selectbox(
                "📞 Alternate Number",
                col_options,
                index=_idx(detected['AlternatePhoneNumber']),
                key='map_alt',
            )

        mapping = {
            'ProspectID':           None if sel_pid   == '(none)' else sel_pid,
            'Email':                None if sel_email == '(none)' else sel_email,
            'PhoneNumber':          None if sel_phone == '(none)' else sel_phone,
            'AlternatePhoneNumber': None if sel_alt   == '(none)' else sel_alt,
        }

        if not any(mapping.values()):
            st.error("❌ At least one field must be mapped (Prospect ID, Email, Phone, or Alt Phone).")
            mapping = None

        # Sanity check for non-empty values
        if mapping:
            non_empty = {}
            for k, c in mapping.items():
                if c is None:
                    non_empty[k] = 0
                else:
                    s = df_user[c].astype(str).str.strip().str.lower()
                    mask = (~s.isin(['', 'nan', 'none', 'null', 'na', '<na>'])) & df_user[c].notna()
                    non_empty[k] = int(mask.sum())
            cap = " &middot; ".join(
                f"<span style='font-family:DM Mono,monospace;font-size:.72rem;color:#6B21A8;'>"
                f"{k}: <b>{v:,}</b></span>"
                for k, v in non_empty.items() if mapping[k]
            )
            st.markdown(
                f"<div style='margin-top:.4rem;padding:.5rem .75rem;background:#FAF5FF;"
                f"border-left:3px solid #A855F7;border-radius:6px;color:#111111;'>"
                f"<b style='font-size:.78rem;color:#111111;'>Non-empty values:</b>&nbsp;{cap}</div>",
                unsafe_allow_html=True
            )

    # ─── STEP 3 — Generate ──────────────────────────────────────────────────
    st.markdown("""
    <div class="lt-step-card" style="margin-top:1rem;">
        <span class="lt-step-num">3</span>
        <span style="font-weight:700;font-size:1rem;color:#111111;">Generate the report</span>
    </div>
    """, unsafe_allow_html=True)

    btn_disabled = (df_user is None) or (mapping is None) or (df_user.empty)
    gen = st.button(
        "🚀 Generate Realtime Lead Report",
        disabled=btn_disabled,
        key='lt_gen',
        use_container_width=True,
    )

    # ─── PROCESSING ─────────────────────────────────────────────────────────
    if gen:
        ak, sk = get_lsq_creds()
        if not ak or not sk:
            st.error("❌ LSQ credentials missing. Add `LSQ_ACCESS_KEY` and `LSQ_SECRET_KEY` to secrets.toml.")
            return
        if client is None:
            st.error("❌ BigQuery client not configured. Check `gcp_service_account` in secrets.toml.")
            return

        st.markdown(
            "<div style='text-align:center;padding:.6rem 1rem;background:#FAF5FF;border-radius:8px;"
            "border:1px solid #E9D5FF;font-weight:600;color:#6B21A8;margin:1rem 0;'>"
            "⏳ Your file is getting processed! Please wait…</div>",
            unsafe_allow_html=True
        )

        with st.status("Processing…", expanded=True) as status:
            try:
                # Step A — Fetch users
                st.write("🔹 Fetching users from LSQ…")
                owner_map = fetch_lsq_users(ak, sk)
                st.write(f"   ✓ {len(owner_map):,} users loaded.")

                # Step B — Fetch leads (cutoff = days_back + 5)
                lookup_days = int(days_back) + 5
                st.write(f"🔹 Fetching leads from LSQ (assigned in last {lookup_days} days)…")
                lead_status = st.empty()

                def _cb(total, page):
                    lead_status.write(f"   📥 Fetched **{total:,}** leads so far (page {page})…")

                raw_leads, cutoff_str = fetch_lsq_leads(ak, sk, lookup_days, status_callback=_cb)
                st.write(f"   ✓ Total leads fetched: **{len(raw_leads):,}** (cutoff: {cutoff_str}).")

                # Step C — Build LSQ DataFrame
                st.write("🔹 Cleaning LSQ data…")
                df_lsq = build_lsq_df(raw_leads, owner_map)
                st.write(f"   ✓ {len(df_lsq):,} rows ready for matching.")

                # Step D — Fetch revenue maps
                st.write("🔹 Fetching enrollment data from BigQuery…")
                phone_map_rev, email_map_rev = fetch_revenue_maps()
                st.write(f"   ✓ {len(phone_map_rev):,} phones & {len(email_map_rev):,} emails for enrollment match.")

                # Step E — Map user → LSQ
                st.write("🔹 Mapping your leads to LSQ records…")
                df_out = map_user_to_lsq(df_user, df_lsq, mapping, phone_map_rev, email_map_rev)

                matched = (df_out['_Match_Type'] != 'Not Found in LSQ').sum()
                enrolled = (df_out['ContactStage'].astype(str).str.strip().str.lower() == 'actually enrolled').sum()
                st.write(f"   ✓ Matched **{int(matched):,} of {len(df_out):,}** rows.")
                st.write(f"   ✓ **{int(enrolled):,}** rows flagged 'Actually Enrolled' from revenue data.")

                # Step E.1 — Enrich with Team / Vertical / Analyst / Sales Leader from auth sheet
                st.write("🔹 Mapping owner → Team / Vertical / Analyst / Sales Leader…")
                owner_meta = load_owner_meta()
                df_out = enrich_with_owner_meta(df_out, owner_meta)
                tagged = int((df_out['Team_Name'].astype(str).str.strip() != '').sum())
                st.write(f"   ✓ {tagged:,} of {len(df_out):,} rows enriched with team metadata.")

                # Step F — Build Excel
                st.write("🔹 Building Excel output…")
                xlsx_bytes = build_output_xlsx(df_out, days_back)
                st.write("   ✓ Excel ready.")

                status.update(label="✅ Your processed file is ready to download.", state="complete", expanded=False)

                st.session_state['lt_output_bytes'] = xlsx_bytes
                st.session_state['lt_output_df']    = df_out
                st.session_state['lt_summary']      = {
                    'rows_in':  len(df_user),
                    'matched':  int(matched),
                    'enrolled': int(enrolled),
                }

            except Exception as ex:
                status.update(label=f"❌ Processing failed: {ex}", state="error")
                st.exception(ex)
                return

    # ─── RESULT DOWNLOAD ────────────────────────────────────────────────────
    if 'lt_output_bytes' in st.session_state and 'lt_output_df' in st.session_state:
        df_out  = st.session_state['lt_output_df']
        summary = st.session_state.get('lt_summary', {})

        st.divider()
        st.markdown("#### ✨ Result")
        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric("Rows Uploaded",      f"{summary.get('rows_in', 0):,}")
        with m2:
            st.metric("Matched in LSQ",     f"{summary.get('matched', 0):,}")
        with m3:
            st.metric("Actually Enrolled",  f"{summary.get('enrolled', 0):,}")

        tab_report, tab_insights = st.tabs(["📄 Lead Tracker Report", "🧠 Insights"])

        with tab_report:
            st.download_button(
                "📥 Download Realtime Lead Report (.xlsx)",
                data=st.session_state['lt_output_bytes'],
                file_name=f"Realtime_Lead_Report_{datetime.now(pytz.timezone('Asia/Kolkata')).strftime('%d-%m-%Y_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_final",
                use_container_width=True,
            )

            with st.expander("👁️ Preview first 20 rows of output", expanded=False):
                preview_cols = ['ProspectID', 'FirstName', 'LastName', 'Email', 'PhoneNumber',
                                'Owner', 'ContactStage', 'AssignedOn', 'Phone_call_counter',
                                'Assigned_On_Call_Counter',
                                'Team_Name', 'Vertical', 'Analyst', 'Sales_Leader',
                                '_Match_Type']
                preview_cols = [c for c in preview_cols if c in df_out.columns]
                st.dataframe(df_out[preview_cols].head(20), width='stretch', hide_index=True)

        with tab_insights:
            st.markdown("""
            <div style='text-align:center;margin-bottom:1rem;'>
                <span style='font-size:.72rem;font-weight:600;color:#A855F7;
                             background:rgba(168,85,247,.1);border:1px solid rgba(168,85,247,.2);
                             border-radius:20px;padding:4px 14px;font-family:DM Mono,monospace;'>
                    ⚡ AUTO-GENERATED FROM LEAD TRACKER REPORT
                </span>
            </div>
            <div class="lt-section-header">
                <div class="lt-section-line"></div>
                <span class="lt-section-title">🧠 GENERATED INSIGHTS</span>
                <div class="lt-section-line r"></div>
            </div>
            """, unsafe_allow_html=True)

            insights = compute_lead_insights(df_out)
            if insights:
                cols_ins = st.columns(2)
                for i, ins in enumerate(insights):
                    with cols_ins[i % 2]:
                        st.markdown(f"""
                        <div class="insight-card {ins['type']}">
                            <div style='display:flex;align-items:center;justify-content:center;gap:.4rem;'>
                                <span class="insight-icon">{ins['icon']}</span>
                                <span class="insight-title">{ins['title']}</span>
                            </div>
                            <div class="insight-body">{ins['body']}</div>
                        </div>""", unsafe_allow_html=True)
            else:
                st.info("Not enough matched data to generate insights — every row was 'Not Found in LSQ'.")


# ════════════════════════════════════════════════════════════════════════════
# ROUTER
# ════════════════════════════════════════════════════════════════════════════

if not st.session_state.get('password_correct', False):
    show_login_page()
else:
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    [data-testid="stStatusWidget"] { display: none !important; }
    header[data-testid="stHeader"] { background: transparent !important; }
    </style>
    """, unsafe_allow_html=True)

    _apply_role_filters()

    ri   = st.session_state.get('auth_role_info', {'role': 'admin'})
    role = ri.get('role', 'admin')
    disp = ri.get('display_name', st.session_state.get('current_user', ''))

    _ROLE_LABELS = {
        'admin'        : '🛡️ Admin',
        'vertical_head': '🏢 Vertical Head',
        'trainer'      : '🎓 Sales Leader',
        'tl'           : '👑 Team Lead',
        'caller'       : '📞 Caller',
    }

    st.sidebar.markdown(f"""
    <div style='margin:.4rem 0 .6rem;padding:.55rem .7rem;
                background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);
                border-radius:10px;text-align:center;'>
        <div style='font-size:.6rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:1px;color:#A855F7;margin-bottom:.2rem;'>
            {_ROLE_LABELS.get(role, role)}
        </div>
        <div style='font-size:.72rem;color:#A855F7;word-break:break-all;'>
            {disp}
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown("""
    <style>
    div[data-testid="stSidebar"] div[data-testid="stButton"]:first-of-type > button {
        display: block !important;
        margin: 0 auto !important;
        width: auto !important;
        min-width: 120px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    col_so = st.sidebar.columns([1, 2, 1])
    with col_so[1]:
        _sign_out = st.button("🚪 Sign Out", key="signout_btn")
    if _sign_out:
        try:
            supa.auth.sign_out()
        except Exception:
            pass
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    st.sidebar.markdown(f"""
    <div style='padding:.7rem 0 .5rem;text-align:center;
                border-bottom:1px solid rgba(128,128,128,.15);'>
        <div style='display:flex;align-items:center;justify-content:center;margin-bottom:.25rem;'>
            <span style='font-size:.95rem;font-weight:700;color:#A855F7;letter-spacing:-.4px;'>LawSikho</span>
            <div style='width:1px;height:16px;margin:0 .55rem;
                        background:linear-gradient(180deg,transparent,rgba(168,85,247,.85),transparent);
                        box-shadow:0 0 5px rgba(168,85,247,.5);'></div>
            <span style='font-size:.95rem;font-weight:700;color:#A855F7;letter-spacing:-.4px;'>Skill Arbitrage</span>
        </div>
        <div style='font-size:.6rem;color:#A855F7;letter-spacing:1px;font-family:monospace;font-weight:600;'>
            India Learning 📖 India Earning
        </div>
    </div>
    """, unsafe_allow_html=True)


    run_lead_tracker()
