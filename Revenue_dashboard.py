import streamlit as st
import pandas as pd
import os
import pytz
import io
from datetime import datetime, date, time, timedelta
from google.cloud import bigquery
from google.oauth2 import service_account
import warnings
warnings.filterwarnings(
    "ignore",
    message="Please replace `st.components.v1.html` with `st.iframe`"
)
# ReportLab imports (used by both dashboards)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Flowable
from reportlab.lib.enums import TA_CENTER

# Openpyxl imports (used by Revenue dashboard)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client
import re

# --- GLOBAL CONFIG & CREDENTIALS ---
st.set_page_config(
    page_title="Revenue Metrics — LawSikho",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="collapsed"    
)

def get_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=credentials, project=info["project_id"])
    else:
        SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
        if os.path.exists(SERVICE_ACCOUNT_FILE):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
            return bigquery.Client()
        return None

@st.cache_resource
def get_cached_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=credentials, project=info["project_id"])
    else:
        SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
        if os.path.exists(SERVICE_ACCOUNT_FILE):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
            return bigquery.Client()
        return None

client = get_cached_bq_client()

# --- LOGIN MODULE ---

def _apply_role_filters():
    ri = st.session_state.get('auth_role_info', {'role': 'admin'})
    role    = ri.get('role', 'admin')
    teams   = ri.get('teams')      # list | None
    callers = ri.get('callers')    # list | None
    cname   = ri.get('caller_name')

    st.session_state['rf_role']        = role
    st.session_state['rf_teams']       = teams   or []
    st.session_state['rf_callers']     = callers or []
    st.session_state['rf_caller_name'] = cname   or ''
    # Roles that see team/vertical multiselects in the sidebar
    st.session_state['rf_full_filters']= role == 'admin'

# ── Hardcoded access tiers ─────────────────────────────────────
ADMIN_EMAILS = {
    "admin@lawsikho.in",
    "yash@lawsikho.in",
    "akanksha.patil@lawsikho.in",
    "devki.dt@lawsikho.in",
    "mansi.gupta@lawsikho.in",
    "shivamtejpal@lawsikho.in",
    "pratik.s@lawsikho.in",
    "suraj@lawsikho.in",
    "parul.nagar@lawsikho.in",
    "amitray@lawsikho.in",
    "rinku@lawsikho.in",
    "karunakarareddy@lawsikho.in",
    "priyansh.s@lawsikho.in"
}

VERTICAL_HEAD_TEAMS = {
    # email               : [list of Team Names they can see]
    "uzair@lawsikho.in"       : ["19th Jan US acc/women ai Closure Batch","CD Closures - Inayat","Contract Drafting","Corporate law - Anas",
                                 "Corporate law - Jyoti","Elite","Law firm trainees - Anas","US Acc Closure Specialist - Inayat","US Accounting - Inayat",
                                 "US accounting closures - Sana","US accounting trainees","Women ai Trainee - Umme","Women ai/CD closure"],

    "shivya.p@lawsikho.in"    : ["ID - 2","ID - 4","ID - 8","ID - 9"],
    
    "mayur@lawsikho.in"       : ["Changemakers"],
    "deepansi@lawsikho.in"    : ["US Accounting","US accounting - Closures"],
    "darshan.c@lawsikho.in"   : ["ID Closure"],
    "anmol.g@lawsikho.in"     : ["DSV- Aditya","DSV- Shivam","ID Closure - Anmol","Women ai"],
    "abhipsa@lawsikho.in"     : ["CD - Community","CD - Community Manager","Criminal - Community","Criminal - Community Manager","ID - Community","ID - Community Manager"],
}

AUTH_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf"
    "/pub?gid=0&single=true&output=csv"
)

# Sheet column names — update if your sheet differs
_COL_NAME    = "Caller Name"
_COL_EMAIL   = "Email id"
_COL_DESIG   = "Academic Counselor/TL/ATL"
_COL_TEAM    = "Team Name"
_COL_TRAINER = "Sales Leader"
_TL_VALS     = {"TL", "ATL", "AD", "TEAM LEAD", "TEAM LEADER"}


@st.cache_resource
def get_supabase():
    # Try top-level keys first, then nested under a section
    url = (
        st.secrets.get("SUPABASE_URL")
        or st.secrets.get("supabase", {}).get("url")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_URL")
        or ""
    )
    key = (
        st.secrets.get("SUPABASE_KEY")
        or st.secrets.get("supabase", {}).get("key")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_KEY")
        or ""
    )

    if not url or not key:
        st.error(
            "⚠️ Supabase credentials missing from secrets. "
            "Add SUPABASE_URL and SUPABASE_KEY to your secrets.toml at the TOP LEVEL "
            "(before any [section] header)."
        )
        st.stop()

    return create_client(url, key)

supa = get_supabase()


@st.cache_data(ttl=300, show_spinner=False)
def load_auth_sheet() -> pd.DataFrame:
    df = pd.read_csv(AUTH_SHEET_URL)
    df.columns = df.columns.str.strip()
    if _COL_EMAIL in df.columns:
        df['_email_norm'] = df[_COL_EMAIL].astype(str).str.strip().str.lower()
    return df


def _extract_trainer_email(cell: str) -> str | None:
    """Parses 'Name (email@domain.com)' → 'email@domain.com'"""
    m = re.search(r'\(([^)\s]+@[^)\s]+)\)', str(cell))
    return m.group(1).strip().lower() if m else None


def determine_role(email: str, df: pd.DataFrame) -> dict | None:
   
    el = email.strip().lower()

    # 1 ── Admin
    if el in {e.lower() for e in ADMIN_EMAILS}:
        return {'role': 'admin', 'teams': None, 'callers': None,
                'caller_name': None, 'display_name': email}

    # 2 ── Vertical Head (hardcoded)
    for vh_mail, vh_teams in VERTICAL_HEAD_TEAMS.items():
        if vh_mail.lower() == el:
            return {'role': 'vertical_head', 'teams': vh_teams, 'callers': None,
                    'caller_name': None, 'display_name': email}

    if '_email_norm' not in df.columns:
        return None

    # 3 ── Trainer  (Sales Leader column)
    if _COL_TRAINER in df.columns:
        df2 = df.copy()
        df2['_tr_email'] = df2[_COL_TRAINER].apply(_extract_trainer_email)
        trainer_rows = df2[df2['_tr_email'] == el]
        if not trainer_rows.empty:
            teams   = trainer_rows[_COL_TEAM].dropna().unique().tolist() if _COL_TEAM in trainer_rows.columns else []
            callers = trainer_rows[_COL_NAME].dropna().unique().tolist() if _COL_NAME in trainer_rows.columns else []
        
            # ── Extract trainer's OWN name from "Name (email)" format in Sales Leader cell ──
            _sl_cell   = str(trainer_rows[_COL_TRAINER].iloc[0])
            _name_match = re.match(r'^([^(]+)\s*\(', _sl_cell)
            _tr_name   = _name_match.group(1).strip() if _name_match else None
        
            # ── Fallback: look up trainer's own row by email ──
            if not _tr_name:
                _own_rows = df[df['_email_norm'] == el]
                _tr_name  = _own_rows.iloc[0][_COL_NAME] if not _own_rows.empty and _COL_NAME in _own_rows.columns else email
        
            return {'role': 'trainer', 'teams': teams, 'callers': callers,
                    'caller_name': None, 'display_name': _tr_name}

    user_rows = df[df['_email_norm'] == el]
    if user_rows.empty:
        return None  # email not in sheet → not authorised

    # 4 ── TL / AD / ATL
    if _COL_DESIG in user_rows.columns:
        tl_rows = user_rows[
            user_rows[_COL_DESIG].astype(str).str.strip().str.upper().isin(_TL_VALS)
        ]
        if not tl_rows.empty:
            team = tl_rows.iloc[0][_COL_TEAM] if _COL_TEAM in tl_rows.columns else None
            team_callers = (
                df[df[_COL_TEAM] == team][_COL_NAME].dropna().unique().tolist()
                if team else []
            )
            disp = tl_rows.iloc[0][_COL_NAME] if _COL_NAME in tl_rows.columns else email
            return {'role': 'tl', 'teams': [team] if team else [], 'callers': team_callers,
                    'caller_name': None, 'display_name': disp}

    # 5 ── Regular Caller
    caller_name = user_rows.iloc[0][_COL_NAME] if _COL_NAME in user_rows.columns else email
    return {'role': 'caller', 'teams': None, 'callers': None,
            'caller_name': caller_name, 'display_name': caller_name}


# ──────────────────────────────────────────────────────────────
# AUTH UI HELPERS
# ──────────────────────────────────────────────────────────────

def _complete_login(email: str, session):
    """Resolve role and store everything in session state."""
    df_auth   = load_auth_sheet()
    role_info = determine_role(email, df_auth)
    if role_info is None:
        st.error("⛔ Your email is not authorised to access this dashboard.")
        return
    st.session_state['password_correct'] = True
    st.session_state['current_user']     = email
    st.session_state['supabase_session'] = session
    st.session_state['auth_role_info']   = role_info
    st.rerun()


def _auth_sign_in_panel():
    """Email + Password login panel."""
    email = st.text_input("Email", key="si_email", placeholder="Your mail ID")
    pwd   = st.text_input("Password", type="password", key="si_pwd", placeholder="Your Password")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Sign In →", key="si_btn", width='stretch'):
            if not email or not pwd:
                st.error("Enter both email and password."); return
            try:
                resp = supa.auth.sign_in_with_password({"email": email, "password": pwd})
                _complete_login(resp.user.email, resp.session)
            except Exception as ex:
                st.error(f"Login failed: {ex}")
    with c2:
        if st.button("Forgot / Change Password", key="si_otp_switch", width='stretch'):
            st.session_state['auth_tab']           = "otp"
            st.session_state['otp_prefill_email']  = email
            st.session_state['otp_step']           = 1
            st.rerun()


def _auth_otp_panel():
    """OTP verification + password-set panel (first-time or reset)."""
    step = st.session_state.get('otp_step', 1)

    # ── Step 1: request OTP ──────────────────────────────────
    if step == 1:
        email = st.text_input(
            "Email", key="otp_email",
            value=st.session_state.get('otp_prefill_email', ''),
            placeholder="your@lawsikho.in"
        )
        if st.button("Send OTP →", key="otp_send_btn", width='stretch'):
            if not email:
                st.error("Enter your email."); return
            # Pre-check: is the email in the sheet?
            role_check = determine_role(email, load_auth_sheet())
            if role_check is None:
                st.error("⛔ This email is not authorised."); return
            try:
                supa.auth.sign_in_with_otp({
                    "email": email,
                    "options": {"should_create_user": True}
                })
                st.session_state['otp_step']          = 2
                st.session_state['otp_pending_email'] = email
                st.rerun()
            except Exception as ex:
                st.error(f"Could not send OTP: {ex}")

        if st.button("← Back to Sign In", key="otp_back1", width='stretch'):
            st.session_state['auth_tab'] = "signin"
            st.rerun()

    # ── Step 2: verify OTP + set password ───────────────────
    elif step == 2:
        pending_email = st.session_state.get('otp_pending_email', '')
        st.success(f"OTP sent to **{pending_email}** — check your inbox (also spam).")

        otp = st.text_input("OTP code from email", key="otp_code", max_chars=8, placeholder="OTP Received on Email")
        pw1  = st.text_input("Set new password",    type="password", key="otp_pw1")
        pw2  = st.text_input("Confirm password",    type="password", key="otp_pw2")

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Verify & Continue →", key="otp_verify_btn", width='stretch'):
                if len(otp) != 8:
                    st.error("Enter the 8-digit code."); return
                if pw1 != pw2:
                    st.error("Passwords don't match."); return
                if len(pw1) < 8:
                    st.error("Password must be ≥ 8 characters."); return
                try:
                    resp = supa.auth.verify_otp({
                        "email": pending_email,
                        "token": otp,
                        "type": "magiclink"
                    })
                    # Save the password (user is now logged in to Supabase)
                    supa.auth.update_user({"password": pw1})
                    st.session_state['otp_step'] = 1
                    _complete_login(resp.user.email, resp.session)
                except Exception as ex:
                    st.error(f"Verification failed — wrong code or it expired: {ex}")
        with c2:
            if st.button("← Back", key="otp_back2", width='stretch'):
                st.session_state['otp_step'] = 1
                st.rerun()

def show_homepage_with_login():
    # ── Full-page dark background + input styling ──
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    #MainMenu, header[data-testid="stHeader"] { display: none !important; }
    [data-testid="stToolbar"] { display: none !important; }
    [data-testid="stDecoration"] { display: none !important; }
    [data-testid="stStatusWidget"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"], .main { background: #0B1120 !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; margin-bottom: 0 !important; }
    [data-testid="stHorizontalBlock"] { gap: 0 !important; }

    div[data-testid="column"]:nth-child(2),
    div[data-testid="column"]:nth-child(2) > div,
    div[data-testid="column"]:nth-child(2) > div > div,
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"],
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #0f172a !important;
    }
    div[data-testid="column"]:nth-child(2) > div:first-child {
        background-color: #0f172a !important;
        border: 1px solid rgba(255,255,255,.12) !important;
        border-radius: 16px !important;
        padding: 1.6rem 1.8rem 1.8rem !important;
    }

    div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] {
        background-color: #1e293b !important;
        border-color: rgba(148,163,184,.3) !important;
        border-radius: 10px !important;
    }
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"]:focus-within,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"]:focus-within {
        border-color: #10B981 !important;
        box-shadow: 0 0 0 2px rgba(16,185,129,.2) !important;
    }

    div[data-testid="column"]:nth-child(2) [data-baseweb="input"] input,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] input,
    div[data-testid="column"]:nth-child(2) input[type="text"],
    div[data-testid="column"]:nth-child(2) input[type="password"],
    div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #10B981 !important;
        border: none !important;
    }
    div[data-testid="column"]:nth-child(2) input::placeholder {
        color: rgba(241,245,249,.32) !important;
        -webkit-text-fill-color: rgba(241,245,249,.32) !important;
    }

    div[data-testid="column"]:nth-child(2) input:-webkit-autofill,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:hover,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:focus,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:active {
        -webkit-box-shadow: 0 0 0px 1000px #1e293b inset !important;
        box-shadow: 0 0 0px 1000px #1e293b inset !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #10B981 !important;
    }

    div[data-testid="column"]:nth-child(2) label,
    div[data-testid="column"]:nth-child(2) label p {
        color: rgba(241,245,249,.55) !important;
        font-size: 0.8rem !important;
    }

    div[data-testid="column"]:nth-child(2) [data-baseweb="input"] button,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] button {
        background-color: transparent !important;
        color: rgba(241,245,249,.5) !important;
        border: none !important;
    }

    div[data-testid="column"]:nth-child(2) .stButton > button {
        width: 100% !important;
        background: linear-gradient(135deg, #10B981, #059669) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 11px !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        box-shadow: none !important;
    }
    div[data-testid="column"]:nth-child(2) .stButton > button:hover,
    div[data-testid="column"]:nth-child(2) .stButton > button:active,
    div[data-testid="column"]:nth-child(2) .stButton > button:focus {
        background: linear-gradient(135deg, #059669, #065f46) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        outline: none !important;
        box-shadow: 0 4px 16px rgba(16,185,129,.35) !important;
        transform: translateY(-1px) !important;
    }

    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"],
    [data-theme="light"] div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
    }
    [data-theme="light"] div[data-testid="column"]:nth-child(2),
    [data-theme="light"] div[data-testid="column"]:nth-child(2) > div,
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"] {
        background-color: #0f172a !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── HERO HTML ──
    html_hero = """<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;800&family=Plus+Jakarta+Sans:wght@300;400;500&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet"/>
    <style>
    *{box-sizing:border-box;margin:0;padding:0;}
    html,body{font-family:'Plus Jakarta Sans',sans-serif;background:#0B1120;color:#E2E8F0;overflow-x:hidden;}
    body{background:
        radial-gradient(ellipse 80% 50% at 50% -10%,rgba(59,130,246,.12) 0%,transparent 60%),
        radial-gradient(ellipse 60% 40% at 90% 80%,rgba(16,185,129,.35) 0%,transparent 55%),
        #0B1120;}
    body::before{content:"";position:fixed;inset:0;
        background-image:linear-gradient(rgba(255,255,255,.022) 1px,transparent 1px),
            linear-gradient(90deg,rgba(255,255,255,.022) 1px,transparent 1px);
        background-size:48px 48px;pointer-events:none;}
    .hero{display:flex;flex-direction:column;align-items:center;text-align:center;padding:3rem 2rem 2.5rem;}
    .logos{display:flex;align-items:center;margin-bottom:1rem;}
    .lname{font-size:1.25rem;font-weight:700;color:#fff;padding:0 1.6rem;}
    .lsep{width:1px;height:48px;background:linear-gradient(180deg,transparent,#10B981E0,transparent);box-shadow:0 0 8px rgba(16,185,129,.35);}
    .tagline{font-family:'Fira Code',monospace;font-size:.75rem;color:rgba(255,255,255,.32);letter-spacing:1.5px;margin-bottom:2rem;}
    .eyebrow{display:inline-flex;align-items:center;gap:.45rem;font-family:'Fira Code',monospace;font-size:.65rem;
        letter-spacing:2.5px;text-transform:uppercase;color:#10B981;background:rgba(16,185,129,.2);
        border:1px solid rgba(16,185,129,.2);border-radius:100px;padding:.28rem .95rem;margin-bottom:1.2rem;}
    .dot{width:5px;height:5px;background:#10B981;border-radius:50%;box-shadow:0 0 5px #10B981;
        animation:p 2s ease-in-out infinite;}
    @keyframes p{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.45;transform:scale(1.4);}}
    .headline{font-family:'Playfair Display',serif;font-size:clamp(2rem,5vw,3.5rem);
        font-weight:800;line-height:1.09;color:#fff;letter-spacing:-1.5px;margin-bottom:.65rem;}
    .accent{background:linear-gradient(125deg,#10B981,#34D399 40%,#A7F3D0);
        -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}
    .sub{font-size:1rem;font-weight:300;color:rgba(255,255,255,.38);max-width:560px;}
    </style></head><body>
    <div class="hero">
      <div class="logos">
        <div class="lname">LawSikho</div>
        <div class="lsep"></div>
        <div class="lname">Skill Arbitrage</div>
      </div>
      <div class="tagline">India Learning &nbsp;📖&nbsp; India Earning</div>
      <div class="eyebrow"><span class="dot"></span>Revenue Analytics Hub</div>
      <div class="headline">Revenue Performance,<br><span class="accent">tracked live</span></div>
      <div class="sub">Real-time insights into enrollments, collections &amp; target achievement</div>
    </div>
    </body></html>"""
    st.iframe(html_hero, height=420)

    # ── AUTH PANEL ──
    left, mid, right = st.columns([1, 1, 1])
    with mid:
        auth_tab = st.session_state.get('auth_tab', 'signin')

        st.markdown("""
        <div style="text-align:center;margin-bottom:.8rem;">
            <span style="font-family:'Playfair Display',serif;font-size:1.1rem;font-weight:600;color:#fff;">
                🔐 DASHBOARD ACCESS
            </span>
        </div>
        """, unsafe_allow_html=True)

        if auth_tab == 'signin':
            _auth_sign_in_panel()
        else:
            _auth_otp_panel()

        if st.session_state.get('password_correct'):
            st.rerun()

    # ── Simple footer ──
    st.markdown("""
    <div style='text-align:center;padding:2rem 1rem 1.2rem;border-top:1px solid rgba(255,255,255,.06);margin-top:2.5rem;'>
        <div style='font-family:"Fira Code",monospace;font-size:.64rem;color:rgba(255,255,255,.28);margin-bottom:.4rem;'>
            For Internal Use of Sales and Operations Team Only
            <span style='display:inline-block;width:3px;height:3px;background:#10B98166;border-radius:50%;margin:0 .45rem;vertical-align:middle;'></span>
            All Rights Reserved
        </div>
        <div style='font-family:"Fira Code",monospace;font-size:.58rem;color:rgba(255,255,255,.14);'>
            Developed and Designed by Amit Ray
            <span style='display:inline-block;width:3px;height:3px;background:#10B98166;border-radius:50%;margin:0 .45rem;vertical-align:middle;'></span>
            Reach out for Support and Queries
        </div>
    </div>
    """, unsafe_allow_html=True)

@st.cache_data(ttl=300, show_spinner=False)

def _load_rev_update_team_sheet():
    _url_active = (
        "https://docs.google.com/spreadsheets/d/e/"
        "2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf"
        "/pub?gid=0&single=true&output=csv"
    )
    _url_resigned = (
        "https://docs.google.com/spreadsheets/d/e/"
        "2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf"
        "/pub?gid=973926168&single=true&output=csv"
    )
    try:
        df_active = pd.read_csv(_url_active)
    except Exception:
        df_active = pd.DataFrame()
        
    try:
        df_resigned = pd.read_csv(_url_resigned)
    except Exception:
        df_resigned = pd.DataFrame()
        
    df = pd.concat([df_active, df_resigned], ignore_index=True)
    if df.empty:
        return pd.DataFrame()

    df.columns = df.columns.str.strip()
    mc = next((c for c in df.columns if c.strip().lower() == 'month'), None)
    if mc:
        df[mc] = pd.to_datetime(df[mc], dayfirst=True, errors='coerce')
        df = df.sort_values(mc, ascending=False, na_position='last')
    if 'Caller Name' in df.columns:
        df['merge_key'] = df['Caller Name'].str.strip().str.lower()
        if mc:
            # Sort by month desc before dropping duplicates if month is available
            df = df.sort_values(mc, ascending=False, na_position='last')
        df = df.drop_duplicates(subset='merge_key', keep='first').reset_index(drop=True)

def run_revenue_dashboard():
    # ADD THIS CSS BLOCK FIRST:
    st.markdown("""
    <style>
    [data-testid="stMainBlockContainer"] {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    .block-container {
        max-width: 100% !important;
    }
    </style>
    """, unsafe_allow_html=True)
    # ─────────────────────────────────────────────
    # 1. CREDENTIALS & CONFIG
    # ─────────────────────────────────────────────

    CSV_URL      = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf/pub?gid=973926168&single=true&output=csv"
    REV_TABLE_ID = "studious-apex-488820-c3.crm_dashboard.revenue_sheet"
    
    # Callers that are not real agents
    EXCLUDE_CALLERS = {'direct', 'bootcamp - direct'}

    # ─────────────────────────────────────────────
    # 3. CSS
    # ─────────────────────────────────────────────

    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

:root {
    --accent-primary:   #10B981;
    --accent-secondary: #34D399;
    --accent-gold:      #F59E0B;
    --accent-warn:      #FBBF24;
    --accent-danger:    #F87171;
    --accent-info:      #60A5FA;
    --gold:             #F59E0B;
    --silver:           #9CA3AF;
    --bronze:           #CD7F32;
    --radius-sm:        8px;
    --radius-md:        12px;
    --radius-lg:        16px;
    --shadow-sm:        0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
    --shadow-md:        0 4px 16px rgba(0,0,0,.10);
    --shadow-lg:        0 8px 32px rgba(0,0,0,.14);
    --transition:       all 0.22s cubic-bezier(.4,0,.2,1);
}

[data-testid="stAppViewContainer"]:not([class*="dark"]) {
    --bg-base:      #F0FDF4;
    --bg-surface:   #FFFFFF;
    --bg-elevated:  #FFFFFF;
    --bg-muted:     #DCFCE7;
    --border:       rgba(0,0,0,.08);
    --text-primary: #111827;
    --text-muted:   #6B7280;
    --metric-bg:    #FFFFFF;
}

@media (prefers-color-scheme: dark) {
    :root {
        --bg-base:      #0A0F0D;
        --bg-surface:   #111B16;
        --bg-elevated:  #162119;
        --bg-muted:     #0F1A14;
        --border:       rgba(255,255,255,.07);
        --text-primary: #F1F5F9;
        --text-muted:   #94A3B8;
        --metric-bg:    #162119;
    }
}

[data-theme="dark"] {
    --bg-base:      #0A0F0D !important;
    --bg-surface:   #111B16 !important;
    --bg-elevated:  #162119 !important;
    --bg-muted:     #0F1A14 !important;
    --border:       rgba(255,255,255,.07) !important;
    --text-primary: #F1F5F9 !important;
    --text-muted:   #94A3B8 !important;
    --metric-bg:    #162119 !important;
}

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }

footer { visibility: hidden; }
[data-testid="stStatusWidget"], .stStatusWidget { display: none !important; }
[data-testid="stMainViewContainer"] { padding-top: 1.5rem; }
[data-testid="stSidebar"] { border-right: 1px solid var(--border, rgba(0,0,0,.08)); }

.rv-header {
    background: linear-gradient(135deg, #064e3b 0%, #065f46 45%, #1e3a5f 100%);
    border-radius: var(--radius-lg);
    padding: 1.5rem 2rem 1.2rem;
    margin-bottom: 1.2rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow-lg);
}
.rv-header::before {
    content: "";
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(16,185,129,.2) 0%, transparent 70%);
    border-radius: 50%;
}
.rv-title   { font-size: 1.65rem; font-weight: 700; color: #FFFFFF; letter-spacing: .5px; margin: 0 0 .25rem; }
.rv-subtitle{ font-size: .82rem; color: rgba(255,255,255,.6); font-weight: 400; margin: 0; font-family: 'DM Mono', monospace; }
.rv-badge {
    display: inline-flex; align-items: center; gap: 5px;
    background: rgba(255,255,255,.12); backdrop-filter: blur(8px);
    border: 1px solid rgba(255,255,255,.18); border-radius: 20px;
    padding: 3px 10px; font-size: .73rem;
    color: rgba(255,255,255,.9); font-family: 'DM Mono', monospace;
}
.rv-pulse {
    width: 6px; height: 6px; background: #34D399; border-radius: 50%;
    display: inline-block; animation: pulse-ring 1.8s ease-in-out infinite;
}
@keyframes pulse-ring {
    0%, 100% { opacity: 1; transform: scale(1); }
    50%       { opacity: .5; transform: scale(1.4); }
}

.metric-card {
    background: var(--metric-bg, #fff);
    border: 1px solid var(--border, rgba(0,0,0,.08));
    border-radius: var(--radius-md);
    padding: .9rem 1rem;
    transition: var(--transition);
    box-shadow: var(--shadow-sm);
    position: relative; overflow: hidden; text-align: center;
}
.metric-card::before {
    content: ""; position: absolute; top: 0; left: 0;
    width: 100%; height: 3px;
    background: linear-gradient(90deg, #10B981, #34D399);
    opacity: 0; transition: opacity .2s;
}
.metric-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.metric-card:hover::before { opacity: 1; }
.metric-label  { font-size: .68rem; font-weight: 600; text-transform: uppercase; letter-spacing: .8px; color: var(--text-muted, #6B7280); margin: 0 0 .3rem; }
.metric-value  { font-size: 1.45rem; font-weight: 700; color: var(--text-primary, #111827); line-height: 1; font-family: 'DM Mono', monospace; }
.metric-delta  { font-size: .7rem; color: #10B981; margin-top: .2rem; font-weight: 500; }

.section-header { display: flex; align-items: center; gap: .6rem; margin: 1.5rem 0 .8rem; }
.section-header-line { flex: 1; height: 1px; background: linear-gradient(90deg, #10B981, transparent); opacity: .35; }
.section-title { font-size: .78rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1.2px; color: #10B981; white-space: nowrap; text-align: center; }

.insight-card { background: var(--metric-bg, #fff); border: 1px solid var(--border, rgba(0,0,0,.08)); border-radius: var(--radius-md); padding: 1rem 1.1rem; margin-bottom: .6rem; box-shadow: var(--shadow-sm); transition: var(--transition); }
.insight-card:hover { box-shadow: var(--shadow-md); }
.insight-card.good { border-left: 3px solid #10B981; }
.insight-card.warn { border-left: 3px solid #FBBF24; }
.insight-card.bad  { border-left: 3px solid #F87171; }
.insight-card.info { border-left: 3px solid #60A5FA; }
.insight-icon  { font-size: 1.1rem; }
.insight-title { font-size: .82rem; font-weight: 700; color: var(--text-primary, #111827); margin: .2rem 0; }
.insight-body  { font-size: .76rem; color: var(--text-muted, #6B7280); line-height: 1.5; }

[data-testid="stTabs"] [role="tablist"] { gap: .3rem; border-bottom: 1px solid var(--border, rgba(0,0,0,.08)); }
[data-testid="stTabs"] button[role="tab"] {
    font-family: 'DM Sans', sans-serif !important; font-size: .82rem !important;
    font-weight: 600 !important; letter-spacing: .3px;
    border-radius: var(--radius-sm) var(--radius-sm) 0 0;
    padding: .55rem 1.1rem !important; transition: var(--transition);
}

div[data-testid="stDataFrame"] thead tr th {
    background: linear-gradient(135deg, #064e3b, #065f46) !important;
    color: #fff !important; font-family: 'DM Sans', sans-serif !important;
    font-size: .72rem !important; font-weight: 700 !important; letter-spacing: .6px;
    text-transform: uppercase; white-space: normal !important; word-wrap: break-word !important;
    text-align: center !important; vertical-align: middle !important;
    min-width: 100px !important; padding: 10px !important;
}

[data-testid="stSidebar"] .stButton>button {
    width: 100%; font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important; font-size: .82rem !important;
    border-radius: var(--radius-sm); transition: var(--transition);
}

[data-testid="stSidebar"] .stButton>button {
    background: linear-gradient(135deg, #059669, #065f46) !important;
    color: #fff !important; border: none !important;
}
[data-testid="stSidebar"] .stButton>button:hover {
    opacity: .88 !important; transform: translateY(-1px) !important;
}

hr { border-color: var(--border, rgba(0,0,0,.08)) !important; margin: 1.2rem 0 !important; }

.brand-name {
    font-size: .85rem;
    font-weight: 700;
    letter-spacing: -.3px;
    color: #10B981;
}

.brand-tagline {
    font-size: .58rem;
    letter-spacing: .8px;
    font-family: monospace;
    margin-bottom: .9rem;
    color: #059669;
}

.achieve-bar-wrap { background: var(--bg-muted, #DCFCE7); border-radius: 999px; height: 6px; margin-top: .4rem; overflow: hidden; }
.achieve-bar-fill { height: 100%; border-radius: 999px; background: linear-gradient(90deg, #10B981, #34D399); transition: width .6s ease; }
/* ── Download Buttons — match Generate Revenue Report style ── */
.stDownloadButton > button {
    width: 100% !important;
    background: linear-gradient(135deg, #059669, #065f46) !important;
    color: #fff !important;
    border: none !important;
    border-radius: var(--radius-sm) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: .82rem !important;
    transition: var(--transition) !important;
}
.stDownloadButton > button:hover {
    opacity: .88 !important;
    transform: translateY(-1px) !important;
}
</style>
""", unsafe_allow_html=True)


    # ─────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────

    def fmt_inr(value):
        if pd.isna(value) or value == 0: return "₹0"
        if value >= 1_00_00_000: return f"₹{value/1_00_00_000:.2f}Cr"
        if value >= 1_00_000:    return f"₹{value/1_00_000:.2f}L"
        if value >= 1_000:       return f"₹{value/1_000:.2f}K"
        return f"₹{int(value)}"

    def section_header(label):
        st.markdown(f"""
        <div class="section-header">
            <div class="section-header-line"></div>
            <span class="section-title">{label}</span>
            <div class="section-header-line" style="background:linear-gradient(90deg,transparent,#10B981)"></div>
        </div>""", unsafe_allow_html=True)

    def style_total_rev(row):
        if row.get("CALLER NAME") == "TOTAL":
            return ['font-weight: bold; background-color: #374151; color: #FFFFFF;'] * len(row)
        return [''] * len(row)

    def get_months_in_range(start_date, end_date):
        months, cur = [], date(start_date.year, start_date.month, 1)
        end_month   = date(end_date.year, end_date.month, 1)
        while cur <= end_month:
            months.append(cur)
            cur = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)
        return months

    def count_working_days(start_date, end_date):
        """Count Mon–Fri days between start_date and end_date inclusive."""
        count = 0
        cur   = start_date
        while cur <= end_date:
            if cur.weekday() < 5:   # 0=Mon … 4=Fri
                count += 1
            cur += timedelta(days=1)
        return count


    # ─────────────────────────────────────────────
    # DATA FETCHING
    # ─────────────────────────────────────────────

    @st.cache_data(ttl=120, show_spinner=False)
    def get_metadata():
        df_meta = pd.read_csv(CSV_URL)
        df_meta.columns = df_meta.columns.str.strip().str.replace('\xa0', '', regex=False)

        month_col = next((c for c in df_meta.columns if c.strip().lower() == 'month'), None)
        if month_col and month_col != 'Month':
            df_meta.rename(columns={month_col: 'Month'}, inplace=True)

        if 'Month' in df_meta.columns:
            df_meta['Month'] = pd.to_datetime(df_meta['Month'], dayfirst=True, errors='coerce').dt.date
        else:
            df_meta['Month'] = None

        teams     = sorted(df_meta['Team Name'].dropna().unique()) if 'Team Name' in df_meta.columns else []
        verticals = sorted(df_meta['Vertical'].dropna().unique())  if 'Vertical'  in df_meta.columns else []
        df_meta['merge_key'] = df_meta['Caller Name'].str.strip().str.lower()
        return teams, verticals, df_meta

    @st.cache_data(ttl=600, show_spinner=False)
    def get_last_update():
        query = f"""
            SELECT updated_at_ampm FROM `{REV_TABLE_ID}`
            WHERE updated_at IS NOT NULL
            ORDER BY updated_at DESC LIMIT 1
        """
        try:
            res = client.query(query).to_dataframe()
            return str(res['updated_at_ampm'].iloc[0]) if not res.empty else "N/A"
        except:
            return "N/A"

    @st.cache_data(ttl=600, show_spinner=False)
    def get_available_dates():
        query = f"SELECT MIN(Date) as min_date, MAX(Date) as max_date FROM `{REV_TABLE_ID}`"
        try:
            df = client.query(query).to_dataframe()
            if not df.empty and not pd.isna(df['min_date'].iloc[0]):
                return df['min_date'].iloc[0], df['max_date'].iloc[0]
        except:
            pass
        return date.today(), date.today()

    @st.cache_data(ttl=120, show_spinner=False)
    def fetch_revenue_data(start_date, end_date):
        query = f"""
            SELECT * FROM `{REV_TABLE_ID}`
            WHERE Date BETWEEN '{start_date}' AND '{end_date}'
            AND Fee_paid > 0
        """
        df = client.query(query).to_dataframe()
        if not df.empty:
            df['Caller_name']  = df['Caller_name'].astype(str).str.strip()
            df['merge_key']    = df['Caller_name'].str.lower()
            df['Fee_paid']     = pd.to_numeric(df['Fee_paid'],     errors='coerce').fillna(0)
            df['Course_Price'] = pd.to_numeric(df['Course_Price'], errors='coerce').fillna(0)

            enr       = df['Enrollment'].astype(str).str.strip()
            enr_lower = enr.str.lower()
            src_lower = df['Source'].astype(str).str.lower()

            df['is_new_enrollment']       = enr_lower == 'new enrollment'
            df['is_balance_payment']      = enr_lower == 'new enrollment - balance payment'
            df['is_bootcamp_collection']  = enr_lower == 'bootcamp collections - balance payments'
            df['is_community_collection'] = enr_lower == 'community collections - balance payments'
            df['is_other_revenue']        = enr_lower == 'other revenue'
            df['is_empty_enrollment']     = enr == ''
            df['source_has_community']    = src_lower.str.contains('community', na=False)

            # Legacy compat
            df['is_new'] = df['is_new_enrollment']
        return df


    # ─────────────────────────────────────────────
    # TARGET / DESIGNATION RESOLUTION
    # ─────────────────────────────────────────────

    def _col(df, name):
        name_clean = name.strip().lower()
        match = next((c for c in df.columns if c.strip().lower() == name_clean), None)
        if match is None:
            raise KeyError(f"Column '{name}' not found. Available: {list(df.columns)}")
        return match

    def resolve_targets(df_meta, start_date, end_date):
        months_needed = get_months_in_range(start_date, end_date)
        caller_col    = _col(df_meta, 'Caller Name')
        month_col     = _col(df_meta, 'Month')

        target_col = None
        for candidate in ['Target', 'target', 'TARGET', 'Monthly Target', 'Monthly target', 'Sales Target']:
            try:
                target_col = _col(df_meta, candidate)
                break
            except KeyError:
                continue

        if target_col is None:
            st.warning(f"⚠️ Target column not found. Sheet columns: `{list(df_meta.columns)}`")
            return {}

        months_ym = {(m.year, m.month) for m in months_needed}

        def _ym(val):
            try:
                if pd.isna(val):
                    return None
                d = pd.to_datetime(val, dayfirst=True, errors='coerce')
                if pd.isna(d):
                    return None
                return (d.year, d.month)
            except:
                return None

        relevant = df_meta[df_meta[month_col].apply(_ym).isin(months_ym)].copy()
        if relevant.empty:
            return {}

        dedup = relevant.drop_duplicates(subset=[caller_col, month_col])
        dedup[target_col] = pd.to_numeric(
            dedup[target_col].astype(str).str.replace(',', '', regex=False),
            errors='coerce'
        ).fillna(0)
        return dedup.groupby(caller_col)[target_col].sum().to_dict()

    def resolve_designations(df_meta, start_date, end_date):
        months_needed = get_months_in_range(start_date, end_date)
        caller_col    = _col(df_meta, 'Caller Name')
        month_col     = _col(df_meta, 'Month')

        def safe_col(name):
            try:    return _col(df_meta, name)
            except: return None

        desig_col   = safe_col('Academic Counselor/TL/ATL')
        team_col    = safe_col('Team Name')
        vert_col    = safe_col('Vertical')
        analyst_col = safe_col('Analyst')

        months_ym = {(m.year, m.month) for m in months_needed}

        def _ym(val):
            try:
                if pd.isna(val):
                    return None
                d = pd.to_datetime(val, dayfirst=True, errors='coerce')
                if pd.isna(d):
                    return None
                return (d.year, d.month)
            except:
                return None

        relevant = df_meta[df_meta[month_col].apply(_ym).isin(months_ym)].copy()
        if relevant.empty:
            return {}

        relevant = relevant.sort_values(month_col, ascending=False)
        dedup    = relevant.drop_duplicates(subset=[caller_col], keep='first').set_index(caller_col)

        result = {}
        for caller in dedup.index:
            result[caller] = {
                'Academic Counselor/TL/ATL': dedup.at[caller, desig_col]   if desig_col   else '—',
                'Team Name':                 dedup.at[caller, team_col]     if team_col    else '—',
                'Vertical':                  dedup.at[caller, vert_col]     if vert_col    else '—',
                'Analyst':                   dedup.at[caller, analyst_col]  if analyst_col else '—',
            }
        return result


    # ─────────────────────────────────────────────
    # CALLER CLASSIFICATION + METRICS PROCESSING
    # ─────────────────────────────────────────────

    def classify_and_process(df, df_meta, start_date, end_date):
        target_map       = resolve_targets(df_meta, start_date, end_date)
        desig_map        = resolve_designations(df_meta, start_date, end_date)
        target_map_norm  = {str(k).strip().lower(): v for k, v in target_map.items()}
        desig_map_norm   = {str(k).strip().lower(): v for k, v in desig_map.items()}

        # Pre-compute working days ratio once for all callers
        months_count    = max(len(get_months_in_range(start_date, end_date)), 1)
        working_days    = count_working_days(start_date, end_date)
        till_day_ratio  = min(working_days / (20 * months_count), 1.0)

        calling_rows, collection_rows, both_rows = [], [], []

        for caller, grp in df.groupby('Caller_name'):
            if caller.strip().lower() in EXCLUDE_CALLERS:
                continue

            caller_key = caller.strip().lower()
            info       = desig_map_norm.get(caller_key, {})
            team       = str(info.get('Team Name', '—')).strip()
            target     = float(target_map_norm.get(caller_key, 0) or 0)

            till_day_target = round(target * till_day_ratio)

            enr_rev      = grp[grp['is_new_enrollment']]['Fee_paid'].sum()
            bal_rev      = grp[grp['is_balance_payment']]['Fee_paid'].sum()
            boot_coll    = grp[grp['is_bootcamp_collection']]['Fee_paid'].sum()
            comm_coll    = grp[grp['is_community_collection']]['Fee_paid'].sum()
            enrollments  = int(grp['is_new_enrollment'].sum())

            calling_rev    = enr_rev + bal_rev
            collection_rev = boot_coll + comm_coll
            total_rev      = calling_rev + collection_rev

            is_changemakers = team.lower() == 'changemakers'
            has_calling     = calling_rev > 0
            has_collection  = collection_rev > 0

            row = {
                'DESIGNATION'         : info.get('Academic Counselor/TL/ATL', '—'),
                'CALLER NAME'         : caller,
                'TEAM'                : team,
                'VERTICAL'            : info.get('Vertical', '—'),
                'TOTAL TARGET (₹)'    : target,
                'TILL DAY TARGET (₹)' : till_day_target,
                'ENROLLMENTS'         : enrollments,
                'ENROLLMENT REV'      : enr_rev,
                'BALANCE REV'         : bal_rev,
                'COMMUNITY COLLECTION': comm_coll,
                'BOOTCAMP COLLECTION' : boot_coll,
                'CALLING REVENUE'     : calling_rev,
                'COLLECTION REVENUE'  : collection_rev,
                'TOTAL REVENUE'       : total_rev,
                'raw_revenue'         : total_rev,
                'raw_target'          : target,
                'raw_enrollments'     : enrollments,
                'raw_calling_rev'     : calling_rev,
                'raw_collection_rev'  : collection_rev,
            }

            if is_changemakers:
                row['ACHIEVEMENT %'] = round(collection_rev / target * 100, 1) if target > 0 else 0.0
                collection_rows.append(row)
            elif has_calling and has_collection:
                row['ACHIEVEMENT %'] = round(total_rev / target * 100, 1) if target > 0 else 0.0
                both_rows.append(row)
            elif has_collection:
                row['ACHIEVEMENT %'] = round(collection_rev / target * 100, 1) if target > 0 else 0.0
                collection_rows.append(row)
            else:
                row['ACHIEVEMENT %'] = round(calling_rev / target * 100, 1) if target > 0 else 0.0
                calling_rows.append(row)

        def _to_df(rows, sort_col):
            if not rows:
                return pd.DataFrame()
            d = pd.DataFrame(rows).sort_values(sort_col, ascending=False).reset_index(drop=True)
            return d

        calling_df    = _to_df(calling_rows,    'raw_calling_rev')
        collection_df = _to_df(collection_rows, 'raw_collection_rev')
        both_df       = _to_df(both_rows,       'raw_revenue')

        return calling_df, collection_df, both_df


    # ─────────────────────────────────────────────
    # SUMMARY METRICS COMPUTATION
    # ─────────────────────────────────────────────

    def compute_summary_metrics(df):
        excl_mask = df['Caller_name'].str.strip().str.lower().isin(EXCLUDE_CALLERS)

        calling_rev = (
            df[~excl_mask & df['is_new_enrollment']]['Fee_paid'].sum() +
            df[~excl_mask & df['is_balance_payment']]['Fee_paid'].sum()
        )

        collection_rev = df[df['is_bootcamp_collection']]['Fee_paid'].sum()

        community_rev = (
            df[df['is_community_collection']]['Fee_paid'].sum() +
            df[df['is_other_revenue'] & df['source_has_community']]['Fee_paid'].sum() +
            df[
                df['is_new_enrollment'] &
                df['source_has_community'] &
                (df['Caller_name'].str.strip().str.lower() == 'direct')
            ]['Fee_paid'].sum()
        )

        direct_rev = df[
            (df['Caller_name'].str.strip().str.lower() == 'direct') &
            (~df['source_has_community']) &
            (df['is_other_revenue'] | df['is_new_enrollment'] | df['is_balance_payment'])
        ]['Fee_paid'].sum()

        dna_rev = df[df['is_empty_enrollment']]['Fee_paid'].sum()

        bootcamp_direct_rev = df[
            df['is_new_enrollment'] &
            (df['Caller_name'].str.strip().str.lower() == 'bootcamp - direct')
        ]['Fee_paid'].sum()

        total_rev = calling_rev + bootcamp_direct_rev + collection_rev + community_rev + direct_rev + dna_rev

        return {
            'total_rev'          : total_rev,
            'calling_rev'        : calling_rev,
            'bootcamp_direct_rev': bootcamp_direct_rev,
            'collection_rev'     : collection_rev,
            'community_rev'      : community_rev,
            'direct_rev'         : direct_rev,
            'dna_rev'            : dna_rev,
        }

    # ─────────────────────────────────────────────
    # ENROLLMENT SUMMARY COMPUTATION
    # ─────────────────────────────────────────────

    def compute_enrollment_metrics(df):
        caller_lower = df['Caller_name'].str.strip().str.lower()

        direct_enr    = int(df[
            df['is_new_enrollment'] &
            (caller_lower == 'direct') &
            (~df['source_has_community'])
        ].shape[0])

        bootcamp_enr  = int(df[
            df['is_new_enrollment'] &
            (caller_lower == 'bootcamp - direct')
        ].shape[0])

        caller_enr    = int(df[
            df['is_new_enrollment'] &
            (~caller_lower.isin(['direct', 'bootcamp - direct']))
        ].shape[0])

        community_enr = int(df[
            df['is_new_enrollment'] &
            (caller_lower == 'direct') &
            df['source_has_community']
        ].shape[0])

        total_enr = direct_enr + bootcamp_enr + caller_enr + community_enr

        return {
            'total_enr'    : total_enr,
            'direct_enr'   : direct_enr,
            'bootcamp_enr' : bootcamp_enr,
            'caller_enr'   : caller_enr,
            'community_enr': community_enr,
        }

    # ─────────────────────────────────────────────
    # TABLE RENDERING HELPER
    # ─────────────────────────────────────────────

    def render_perf_table(df, display_cols, total_overrides, sort_col, table_key):
        """Renders a styled performance table with medals and a TOTAL row."""
        if df.empty:
            st.info("No data available for this category in the selected period.")
            return

        d = df.copy().reset_index(drop=True)

        # Medals
        d.insert(0, 'Rank', '')
        for i, medal in enumerate(['🥇', '🥈', '🥉']):
            if i < len(d):
                d.at[i, 'Rank'] = medal

        # Format currency cols
        for col in ['ENROLLMENT REV', 'BALANCE REV', 'COMMUNITY COLLECTION',
                    'BOOTCAMP COLLECTION', 'CALLING REVENUE', 'COLLECTION REVENUE',
                    'TOTAL REVENUE', 'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)']:
            if col in d.columns:
                d[col] = d[col].apply(fmt_inr)

        if 'ACHIEVEMENT %' in d.columns:
            d['ACHIEVEMENT %'] = d['ACHIEVEMENT %'].apply(lambda x: f"{x}%")

        # Build total row
        total_dict = {c: '—' for c in ['Rank'] + display_cols}
        total_dict['Rank']        = ''
        total_dict['CALLER NAME'] = 'TOTAL'
        for k, v in total_overrides.items():
            total_dict[k] = v

        all_cols   = ['Rank'] + [c for c in display_cols if c in d.columns]
        total_row  = pd.DataFrame([{c: total_dict.get(c, '—') for c in all_cols}])
        final      = pd.concat([d[all_cols], total_row], ignore_index=True)

        st.dataframe(
            final.style.apply(style_total_rev, axis=1),
            column_order=all_cols,
            width='stretch',
            hide_index=True
        )


    # ─────────────────────────────────────────────
    # INSIGHTS
    # ─────────────────────────────────────────────

    def compute_revenue_insights(df, calling_df, collection_df, both_df, start_date, end_date):
        insights = []
        if df.empty:
            return insights

        month_label = start_date.strftime("%B %Y")

        _all_list = [d for d in [calling_df, collection_df, both_df] if not d.empty]
        all_agents = pd.concat(_all_list, ignore_index=True) if _all_list else pd.DataFrame()

        # ── 1. Top Calling Revenue Achiever ──
        _calling_pool_list = [d for d in [calling_df, both_df] if not d.empty]
        _calling_pool = pd.concat(_calling_pool_list, ignore_index=True) if _calling_pool_list else pd.DataFrame()
        if not _calling_pool.empty:
            top_c = _calling_pool.sort_values('raw_calling_rev', ascending=False).iloc[0]
            insights.append({
                "type": "good", "icon": "🏆",
                "title": f"Top Calling Revenue — {top_c['CALLER NAME']}",
                "body": (f"Achieved {fmt_inr(top_c['raw_calling_rev'])} in calling revenue with "
                         f"{top_c['ENROLLMENTS']} new enrollment(s) in {month_label}. "
                         f"Target achievement: {top_c['ACHIEVEMENT %']}%.")
            })

        # ── 2. Top Collection Revenue Achiever ──
        _coll_pool_list = [d for d in [collection_df, both_df] if not d.empty]
        _coll_pool = pd.concat(_coll_pool_list, ignore_index=True) if _coll_pool_list else pd.DataFrame()
        if not _coll_pool.empty:
            top_col = _coll_pool.sort_values('raw_collection_rev', ascending=False).iloc[0]
            insights.append({
                "type": "good", "icon": "🏦",
                "title": f"Top Collection Revenue — {top_col['CALLER NAME']}",
                "body": (f"Collected {fmt_inr(top_col['raw_collection_rev'])} in {month_label} "
                         f"across community and bootcamp collections. "
                         f"Team: {top_col['TEAM']}.")
            })

        # ── 3. Most Enrollments ──
        if not all_agents.empty and all_agents['raw_enrollments'].max() > 0:
            top_enr = all_agents.sort_values('raw_enrollments', ascending=False).iloc[0]
            insights.append({
                "type": "good", "icon": "🎓",
                "title": f"Most Enrollments — {top_enr['CALLER NAME']}",
                "body": (f"Closed {top_enr['raw_enrollments']} new enrollment(s) in {month_label}, "
                         f"highest among all callers. "
                         f"Calling revenue generated: {fmt_inr(top_enr['raw_calling_rev'])}.")
            })

        # ── 4. Best Target Achievement ──
        if not all_agents.empty:
            _with_target = all_agents[all_agents['raw_target'] > 0]
            if not _with_target.empty:
                top_ach = _with_target.sort_values('ACHIEVEMENT %', ascending=False).iloc[0]
                insights.append({
                    "type": "good" if top_ach['ACHIEVEMENT %'] >= 80 else "warn", "icon": "🎯",
                    "title": f"Best Target Achievement — {top_ach['CALLER NAME']}",
                    "body": (f"Achieved {top_ach['ACHIEVEMENT %']}% of their "
                             f"{fmt_inr(top_ach['raw_target'])} target in {month_label}. "
                             f"Revenue: {fmt_inr(top_ach['raw_revenue'])}. Team: {top_ach['TEAM']}.")
                })

        # ── 5. Focus Required — worst target achievement ──
        if not all_agents.empty:
            _with_target = all_agents[
                (all_agents['raw_target'] > 0) &
                (all_agents['raw_revenue'] > 0)
            ]
            if len(_with_target) > 1:
                worst = _with_target.sort_values('ACHIEVEMENT %').iloc[0]
                if worst['ACHIEVEMENT %'] < 60:
                    insights.append({
                        "type": "bad", "icon": "⚠️",
                        "title": f"Focus Required — {worst['CALLER NAME']}",
                        "body": (f"Only {worst['ACHIEVEMENT %']}% of "
                                 f"{fmt_inr(worst['raw_target'])} target achieved in {month_label}. "
                                 f"Revenue so far: {fmt_inr(worst['raw_revenue'])}. "
                                 f"Team: {worst['TEAM']}. Needs immediate attention.")
                    })

         # ── 6. Callers with target but zero revenue (calling agents only) ──
        _calling_only_list = [d for d in [calling_df, both_df] if not d.empty]
        _calling_only = pd.concat(_calling_only_list, ignore_index=True) if _calling_only_list else pd.DataFrame()

        if not _calling_only.empty:
            _with_target   = _calling_only[_calling_only['raw_target'] > 0]
            _zero_rev      = _with_target[_with_target['raw_calling_rev'] == 0]
            _zero_count    = len(_zero_rev)
            _total_callers = len(_with_target)

            if _zero_count > 0:
                _zero_names = ', '.join(_zero_rev['CALLER NAME'].head(3).tolist())
                _more       = f" and {_zero_count - 3} more" if _zero_count > 3 else ""
                insights.append({
                    "type": "bad", "icon": "🚨",
                    "title": f"{_zero_count} of {_total_callers} Callers Have Zero Revenue in {month_label}",
                    "body": (f"{_zero_names}{_more} have an assigned target but no enrollment "
                             f"or balance revenue recorded in the selected period. "
                             f"Immediate follow-up recommended to unblock closures.")
                })
            else:
                _below_50  = _with_target[_with_target['raw_calling_rev'] / _with_target['raw_target'] * 100 < 50]
                _b50_count = len(_below_50)
                if _b50_count > 0:
                    _top_team = (
                        _calling_only.groupby('TEAM')['raw_calling_rev']
                        .sum().sort_values(ascending=False).index[0]
                    )
                    insights.append({
                        "type": "warn", "icon": "📊",
                        "title": f"{_b50_count} Caller(s) Below 50% Achievement in {month_label}",
                        "body": (f"All callers have some revenue but {_b50_count} are below 50% of their "
                                 f"calling target. Best performing team is {_top_team}. "
                                 f"Focus support on underperforming callers to improve month-end numbers.")
                    })
                else:
                    _top_team = (
                        _calling_only.groupby('TEAM')['raw_calling_rev']
                        .sum().sort_values(ascending=False).index[0]
                    )
                    _top_team_rev = _calling_only.groupby('TEAM')['raw_calling_rev'].sum().max()
                    insights.append({
                        "type": "good", "icon": "🌟",
                        "title": f"Strong Month — All Callers On Track in {month_label}",
                        "body": (f"Every caller with a target has recorded enrollment revenue in the selected period. "
                                 f"Leading team is {_top_team} with {fmt_inr(_top_team_rev)} in calling revenue. "
                                 f"Keep monitoring till-day targets to sustain momentum through month-end.")
                    })

        return insights[:6]
    @st.cache_data(show_spinner=False)
    def generate_helper_pdf_bytes() -> bytes:
        buffer = io.BytesIO()

        GREEN_DARK  = colors.HexColor("#064e3b")
        GREEN_MID   = colors.HexColor("#065f46")
        GREEN_PALE  = colors.HexColor("#DCFCE7")
        GREEN_ROW   = colors.HexColor("#F0FDF4")
        GREY_DARK   = colors.HexColor("#374151")
        GREY_MID    = colors.HexColor("#6B7280")
        WHITE       = colors.white
        BLACK       = colors.HexColor("#111827")
        W, H        = A4

        def s(name, **kw):
            defaults = dict(fontName='Helvetica', fontSize=9,
                            textColor=BLACK, spaceAfter=3, leading=14)
            defaults.update(kw)
            return ParagraphStyle(name, **defaults)

        S = {
            'body'      : s('body'),
            'label'     : s('label',   fontName='Helvetica-Bold', fontSize=8,
                             textColor=GREEN_DARK, spaceAfter=1),
            'formula'   : s('formula', fontName='Helvetica-Oblique', fontSize=8.5,
                             textColor=colors.HexColor("#065f46"),
                             backColor=GREEN_PALE, leftIndent=8, rightIndent=8),
            'footer'    : s('footer',  fontSize=7.5, textColor=GREY_MID,
                             alignment=TA_CENTER),
        }

        class CoverBlock(Flowable):
            def __init__(self, w):
                Flowable.__init__(self)
                self.w = w
                self.height = 90
            def draw(self):
                c = self.canv
                c.setFillColor(GREEN_DARK);  c.rect(0, 52, self.w, 38, fill=1, stroke=0)
                c.setFillColor(GREEN_MID);   c.rect(0, 22, self.w, 30, fill=1, stroke=0)
                c.setFillColor(colors.HexColor("#065f46")); c.rect(0, 0, self.w, 22, fill=1, stroke=0)
                c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 22)
                c.drawCentredString(self.w/2, 66, "REVENUE METRICS DASHBOARD")
                c.setFillColor(colors.HexColor("#A7F3D0")); c.setFont("Helvetica-Bold", 11)
                c.drawCentredString(self.w/2, 34, "Logic & Metric Reference Guide")
                c.setFillColor(colors.HexColor("#D1FAE5")); c.setFont("Helvetica", 8.5)
                c.drawCentredString(self.w/2, 8,
                    "LawSikho & Skill Arbitrage  \u00b7  Sales & Operations Team  \u00b7  Internal Use Only")

        class SectionBanner(Flowable):
            def __init__(self, icon, title, color=None, w=None):
                Flowable.__init__(self)
                self.icon = icon; self.title = title
                self.color = color or GREEN_DARK
                self.w = w or (W - 30*mm); self.height = 22
            def draw(self):
                c = self.canv
                c.setFillColor(self.color)
                c.roundRect(0, 0, self.w, self.height, 4, fill=1, stroke=0)
                c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 11)
                c.drawString(10, 6, f"{self.icon}  {self.title}")

        def btable(rows, cw=None):
            cw = cw or [42*mm, 118*mm]
            data = [[Paragraph(f"<b>{r[0]}</b>", S['label']),
                     Paragraph(r[1], S['body'])] for r in rows]
            t = Table(data, colWidths=cw, hAlign='LEFT')
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(0,-1), GREEN_PALE),
                ('VALIGN',        (0,0),(-1,-1),'TOP'),
                ('GRID',          (0,0),(-1,-1),0.3, colors.HexColor("#D1FAE5")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[WHITE, GREEN_ROW]),
                ('LEFTPADDING',   (0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',    (0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]))
            return t

        def ltable(rows):
            data = [[Paragraph(f"<b>{r[0]}</b>", S['label']),
                     Paragraph(r[1], S['formula'])] for r in rows]
            t = Table(data, colWidths=[50*mm, 110*mm], hAlign='LEFT')
            t.setStyle(TableStyle([
                ('VALIGN',        (0,0),(-1,-1),'TOP'),
                ('GRID',          (0,0),(-1,-1),0.3, colors.HexColor("#A7F3D0")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[WHITE, GREEN_ROW]),
                ('LEFTPADDING',   (0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',    (0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]))
            return t

        SP  = Spacer
        HR  = lambda: HRFlowable(width="100%", thickness=0.6,
                                  color=colors.HexColor("#A7F3D0"),
                                  spaceAfter=6, spaceBefore=4)
        BAN = SectionBanner
        cw  = W - 30*mm

        story = [
            SP(1,18*mm), CoverBlock(cw), SP(1,10*mm),
            Paragraph("This document explains every metric, table, and highlight card in the "
                      "Revenue Metrics Dashboard — a quick reference for the Sales & Operations team.", S['body']),
            SP(1,4*mm),

            # ── Section 1 ──
            BAN("🏆","SECTION 1 — TOP 3 REVENUE HIGHLIGHTS"), SP(1,3*mm),
            Paragraph("Three cards at the top of the dashboard — each shows the single best performer "
                      "in one dimension for the selected date range.", S['body']), SP(1,2*mm),
            btable([
                ("🥇 Top Revenue — Caller",    "Caller with highest Calling Revenue (Enrollment Rev + Balance Rev). Excludes 'direct' and 'bootcamp-direct'."),
                ("🎓 Most Enrollments",         "Caller with most rows where Enrollment = 'New Enrollment'. Covers all agent types."),
                ("🥇 Top Revenue — Collection", "Caller with highest Collection Revenue (Community Collections + Bootcamp Collections). From Collection table only."),
            ]), SP(1,4*mm),

            # ── Section 2 ──
            BAN("💵","SECTION 2 — REVENUE SUMMARY METRICS"), SP(1,3*mm),
            Paragraph("Full-picture revenue breakdown. Filter: Fee Paid > 0 applied globally.", S['body']), SP(1,2*mm),
            ltable([
                ("Total Revenue\n(EXCL. Services)",    "Calling + Bootcamp-Direct + Bootcamp-Collection + Community + Direct/Other + DNA revenue."),
                ("Calling Revenue\n(INCL. Funnel)",    "Fee Paid where Enrollment = 'New Enrollment' OR 'New Enrollment - Balance Payment', caller NOT in {direct, bootcamp-direct}."),
                ("Bootcamp-Direct\nRevenue",            "Fee Paid where Enrollment = 'New Enrollment' AND Caller = 'bootcamp-direct'."),
                ("Bootcamp-Collection\nRevenue",        "Fee Paid where Enrollment = 'Bootcamp Collections - Balance Payments'."),
                ("Community Revenue\n(Direct+Collection)", "Fee Paid from: (a) Community Collections rows, (b) Other Revenue with community Source, (c) New Enrollment by 'direct' with community Source."),
                ("Direct/Other Revenue\n(INCL. Funnel)","Fee Paid where Caller='direct', Source has no 'community', Enrollment in {Other Revenue, New Enrollment, Balance Payment}."),
                ("DNA / Not Updated Yet",               "Fee Paid where Enrollment column is blank. Not yet categorised in BigQuery."),
            ]), SP(1,4*mm),

            # ── Section 3 ──
            BAN("🎓","SECTION 3 — ENROLLMENT SUMMARY METRICS"), SP(1,3*mm),
            Paragraph("Counts ONLY rows where Enrollment = 'New Enrollment'. Balance payments and collections are excluded.", S['body']), SP(1,2*mm),
            ltable([
                ("Total Enrollments",            "All New Enrollment rows across all callers."),
                ("Caller Enrollments",           "New Enrollment rows where Caller is NOT 'direct' or 'bootcamp-direct'."),
                ("Direct Enrollments",           "New Enrollment by 'direct' caller AND Source has no 'community'."),
                ("Bootcamp-Direct Enrollments",  "New Enrollment where Caller = 'bootcamp-direct'."),
                ("Community-Direct Enrollments", "New Enrollment by 'direct' caller AND Source contains 'community'."),
            ]), SP(1,4*mm),

            # ── Section 4 ──
            BAN("📞","SECTION 4 — CALLER REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Calling agents with Calling Revenue > 0, no collection revenue, not Changemakers team. Sorted by Calling Revenue descending.", S['body']), SP(1,2*mm),
            btable([
                ("DESIGNATION",         "Role from team sheet: Academic Counselor, TL, ATL, etc."),
                ("TOTAL TARGET (₹)",    "Sum of monthly targets from team sheet for all months in selected range."),
                ("TILL DAY TARGET (₹)", "Total Target × (Mon-Fri days elapsed ÷ (20 × months in range)). Capped at 1.0."),
                ("ENROLLMENTS",         "Count of New Enrollment rows for this caller."),
                ("ENROLLMENT REV",      "Fee Paid where Enrollment = 'New Enrollment' for this caller."),
                ("BALANCE REV",         "Fee Paid where Enrollment = 'New Enrollment - Balance Payment' for this caller."),
                ("CALLING REVENUE",     "Enrollment Rev + Balance Rev."),
                ("ACHIEVEMENT %",       "Calling Revenue ÷ Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 5 ──
            BAN("🏦","SECTION 5 — COLLECTION CALLER REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Callers with Collection Revenue > 0 and no calling revenue. Changemakers team always here. Sorted by Collection Revenue descending.", S['body']), SP(1,2*mm),
            btable([
                ("COMMUNITY COLLECTION", "Fee Paid where Enrollment = 'Community Collections - Balance Payments'."),
                ("BOOTCAMP COLLECTION",  "Fee Paid where Enrollment = 'Bootcamp Collections - Balance Payments'."),
                ("COLLECTION REVENUE",   "Community Collection + Bootcamp Collection."),
                ("ACHIEVEMENT %",        "Collection Revenue ÷ Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 6 ──
            BAN("📞🏦","SECTION 6 — CALLING + COLLECTION CALLER REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Callers with BOTH Calling Revenue > 0 AND Collection Revenue > 0, not Changemakers. Sorted by Total Revenue descending.", S['body']), SP(1,2*mm),
            btable([
                ("TOTAL REVENUE",  "Calling Revenue + Collection Revenue."),
                ("ACHIEVEMENT %",  "Total Revenue ÷ Total Target × 100."),
                ("Other columns",  "Same definitions as Sections 4 and 5."),
            ]), SP(1,4*mm),

            # ── Section 7 ──
            BAN("📞","SECTION 7 — CALLER REVENUE TEAM PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Section 4 callers grouped by team. Insights & Leaderboard tab only. Sorted by Calling Revenue descending.", S['body']), SP(1,2*mm),
            ltable([
                ("Callers",         "Count of agents from this team in the Caller table."),
                ("Total Target",    "Sum of all agent targets for this team."),
                ("Till Day Target", "Sum of all agent till-day targets for this team."),
                ("Calling Revenue", "Sum of Calling Revenue across all agents in this team."),
                ("Achievement %",   "Team Calling Revenue ÷ Team Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 8 ──
            BAN("🏦","SECTION 8 — COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Section 5 callers grouped by team. Sorted by Collection Revenue descending.", S['body']), SP(1,2*mm),
            ltable([
                ("Community Collection", "Sum of Community Collection revenue for this team."),
                ("Bootcamp Collection",  "Sum of Bootcamp Collection revenue for this team."),
                ("Collection Revenue",   "Community + Bootcamp Collection total for this team."),
                ("Achievement %",        "Team Collection Revenue ÷ Team Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 9 ──
            BAN("📞🏦","SECTION 9 — CALLING + COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Section 6 callers grouped by team. Sorted by Total Revenue descending.", S['body']), SP(1,2*mm),
            ltable([
                ("Calling Revenue",      "Sum of Calling Revenue for dual-stream agents in this team."),
                ("Community Collection", "Sum of Community Collection revenue."),
                ("Bootcamp Collection",  "Sum of Bootcamp Collection revenue."),
                ("Total Revenue",        "Calling + Community + Bootcamp Collection for this team."),
                ("Achievement %",        "Team Total Revenue ÷ Team Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Glossary ──
            # ── Section 10 — Pending Revenue ──
            BAN("📊","SECTION 10 — CALLERWISE PENDING REVENUE TAB"), SP(1,3*mm),
            Paragraph("Auto-loads on tab open. Always shows current month + previous month. Not affected by sidebar filters.", S['body']), SP(1,2*mm),
            ltable([
                ("Revenue Pool",             "Sum of Course Price for all booking-fee leads (Full_Installment = 'booking fees') with Fee_paid ≥ ₹999 and positive balance. Excludes drop-sheet leads and students who appear in both months (paid balance)."),
                ("Revenue Collected",        "Sum of Fee_paid (booking fee already received) for the same leads in the pool."),
                ("Balance to be Recovered",  "Course Price − Fee_paid for each lead. Only leads with balance > 0 are shown."),
                ("No. of Leads",             "Count of individual pending-balance leads for this caller/team."),
                ("Leads >48 HRS",            "Count of leads whose enrollment Date is ≤ today − 3 days (IST). Today and the two preceding calendar days are excluded so only genuinely overdue leads are flagged."),
                ("Balance >48 HRS",          "Sum of pending balance for leads that qualify as >48 hrs overdue."),
                ("% Pending >48 HRS",        "Balance >48 HRS ÷ Total Balance × 100 for this row."),
                ("Previous Month columns",   "Same balance and lead count computed from the previous calendar month's pending leads pool."),
                ("Grand Balance / Leads",    "Current month balance + previous month balance (and lead counts). Rows with grand balance = 0 are hidden."),
            ]), SP(1,4*mm),
            BAN("🚫","SECTION 11 — CALLERWISE DROPPED LEADS"), SP(1,3*mm),
            Paragraph("Drop leads are sourced from the Drop Leads Google Sheet. Each drop is attributed to the caller who originally enrolled the student (matched by email then phone against the revenue sheet).", S['body']), SP(1,2*mm),
            ltable([
                ("Current Month Drops",  "Drop form submissions with a timestamp falling in the current calendar month."),
                ("Previous Month Drops", "Drop form submissions with a timestamp falling in the previous calendar month."),
                ("Total Drop Cases",     "Current + Previous month drops for this caller."),
                ("Attribution Logic",    "Email match first; if no email match, phone match. If neither matches any New Enrollment row, attributed to 'Unknown'."),
                ("Team / Vertical",      "Pulled from the team sheet using the attributed caller name as the merge key."),
            ]), SP(1,4*mm),

            # ── Glossary ──
            BAN("📖","KEY TERMS GLOSSARY", color=GREY_DARK), SP(1,3*mm),
            btable([
                ("New Enrollment",       "A fresh admission. Counts toward enrollment metrics and Calling Revenue."),
                ("Balance Payment",      "Remaining fee from a prior enrollment. NOT a new enrollment."),
                ("Bootcamp Collections", "Balance payments from bootcamp-enrolled students."),
                ("Community Collections","Balance payments from community-channel students."),
                ("DNA / Not Updated",    "Rows with blank Enrollment. Tracked separately in summary."),
                ("direct",               "Pseudo-caller for organic closures. Excluded from agent tables."),
                ("bootcamp-direct",      "Pseudo-caller for bootcamp direct admissions. Tracked separately."),
                ("Till Day Target",      "Daily-prorated target: Total Target x (Working Days / (20 x Months)). Working Days = Mon-Fri in selected range."),
                ("Changemakers",         "Special team always routed to Collection table regardless of calling revenue."),
            ], cw=[44*mm, 116*mm]),

            SP(1,8*mm), HR(),
            Paragraph("Designed by Amit Ray  \u00b7  amitray@lawsikho.com  \u00b7  "
                      "For Internal Use of Sales and Operations Team Only. All Rights Reserved.", S['footer']),
        ]

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=14*mm,  bottomMargin=14*mm,
            title="Revenue Metrics — Logic Reference Guide",
            author="Amit Ray",
        )
        doc.build(story)
        return buffer.getvalue()


    # ─────────────────────────────────────────────
    # PENDING REVENUE — CONSTANTS & HELPERS
    # ─────────────────────────────────────────────

    DROP_LEADS_URL_P = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRJ9qaFD5Sc9O1JFH7ijR0JI2SEkAgXM8PJZsWslsASLDnTCA_fwIP0fg_PmtdMm_zs3KwTLI45fPog/pub?gid=1082322056&single=true&output=csv"

    def pending_months():
        today   = date.today()
        c_start = date(today.year, today.month, 1)
        c_end   = today
        p_end   = c_start - timedelta(days=1)
        p_start = date(p_end.year, p_end.month, 1)
        return c_start, c_end, p_start, p_end

    @st.cache_data(ttl=300, show_spinner=False)
    def load_drop_leads():
        try:
            df = pd.read_csv(DROP_LEADS_URL_P)
            df.columns = df.columns.str.strip()
            e_col = next((c for c in df.columns if 'email' in c.lower() and ('drop' in c.lower() or 'student' in c.lower())), None)
            p_col = next((c for c in df.columns if 'phone' in c.lower() or ('number' in c.lower() and 'drop' in c.lower())), None)
            t_col = next((c for c in df.columns if 'timestamp' in c.lower()), None)
            df['drop_email'] = df[e_col].astype(str).str.strip().str.lower() if e_col else ''
            df['drop_phone'] = (df[p_col].astype(str).str.replace(r'\D', '', regex=True).str[-10:]) if p_col else ''
            df['drop_date']  = pd.to_datetime(df[t_col], errors='coerce', dayfirst=True) if t_col else pd.NaT
            return df
        except:
            return pd.DataFrame()

    @st.cache_data(ttl=120, show_spinner=False)
    def fetch_both_months_rev(p_start, c_end):
        q = f"SELECT * FROM `{REV_TABLE_ID}` WHERE Date BETWEEN '{p_start}' AND '{c_end}'"
        df = client.query(q).to_dataframe()
        if df.empty:
            return df
        df['Caller_name']     = df['Caller_name'].astype(str).str.strip()
        df['merge_key']       = df['Caller_name'].str.lower()
        df['Fee_paid']        = pd.to_numeric(df['Fee_paid'],     errors='coerce').fillna(0)
        df['Course_Price']    = pd.to_numeric(df['Course_Price'], errors='coerce').fillna(0)
        df['Email_Id_norm']   = df['Email_Id'].astype(str).str.strip().str.lower()
        df['Contact_No_norm'] = df['Contact_No'].astype(str).str.replace(r'\D', '', regex=True).str[-10:]
        enr_l = df['Enrollment'].astype(str).str.strip().str.lower()
        df['is_new'] = enr_l == 'new enrollment'
        df['is_bal'] = enr_l == 'new enrollment - balance payment'
        df['Date']   = pd.to_datetime(df['Date'], errors='coerce').dt.date
        # Normalize Full_Installment field — "booking fees" = partial payment pending
        df['Full_Installment'] = (
            df['Full_Installment'].astype(str).str.strip().str.lower()
            if 'Full_Installment' in df.columns
            else 'unknown'
        )
        return df

    def pending_leads_for_month(df_m, excl_emails, excl_phones, df_combined=None):
        if df_m.empty:
            return pd.DataFrame()

        # ── Step 1: All rows with Fee_paid >= 999 across ALL enrollment types ──
        df_valid = df_m[df_m['Fee_paid'] >= 999].copy()
        if df_valid.empty:
            return pd.DataFrame()

        # ── Step 2: Count email + phone occurrences across COMBINED two-month pool ──
        # Mirrors original: counts built from ALL rows in both months together
        # so a student who enrolled in month A and paid balance in month B
        # appears twice in the combined pool → count=2 → excluded from both months
        if df_combined is not None and not df_combined.empty:
            df_pool = df_combined[df_combined['Fee_paid'] >= 999]
        else:
            df_pool = df_valid  # fallback: single month only
        email_counts = df_pool['Email_Id_norm'].value_counts()
        phone_counts = df_pool['Contact_No_norm'].value_counts()

        # ── Step 3: Filter to "booking fees" rows only (this month's slice) ──
        if 'Full_Installment' in df_valid.columns:
            pending_rows = df_valid[
                df_valid['Full_Installment'].str.lower().str.strip() == 'booking fees'
            ].copy()
        else:
            # Fallback until pipeline pushes Full_Installment to BigQuery
            pending_rows = df_valid[df_valid['is_new']].copy()

        if pending_rows.empty:
            return pd.DataFrame()

        # ── Step 4: count==1 against COMBINED pool — same as original script ──
        pending_rows = pending_rows[
            pending_rows['Email_Id_norm'].map(email_counts) == 1
        ].copy()
        pending_rows = pending_rows[
            pending_rows['Contact_No_norm'].map(phone_counts) == 1
        ].copy()
        if pending_rows.empty:
            return pd.DataFrame()

        # ── Step 5: Exclude drop sheet (email AND phone) ──
        pending_rows = pending_rows[
            ~pending_rows['Email_Id_norm'].isin(excl_emails) &
            ~pending_rows['Contact_No_norm'].isin(excl_phones)
        ].copy()
        if pending_rows.empty:
            return pd.DataFrame()

        # ── Step 6: Pending balance = Course Price - Fee Paid ──
        pending_rows['balance'] = pending_rows['Course_Price'] - pending_rows['Fee_paid']
        p = pending_rows[pending_rows['balance'] > 0].copy()
        if p.empty:
            return pd.DataFrame()

        india        = pytz.timezone("Asia/Kolkata")
        today_ist    = datetime.now(india).date()
        cut48        = today_ist - timedelta(days=3)
        p['over_48'] = p['Date'] <= cut48
        return p

    def _raw_caller_agg(pending_df):
        """Pure groupby aggregation — NO metadata merge. Team info added later in one pass."""
        if pending_df.empty:
            return pd.DataFrame()
        b48 = (pending_df[pending_df['over_48']]
               .groupby('Caller_name')['balance'].sum()
               .rename('bal_48hr').reset_index())
        agg = (pending_df.groupby('Caller_name')
               .agg(pool     =('Course_Price', 'sum'),
                    collected=('Fee_paid',     'sum'),
                    balance  =('balance',      'sum'),
                    leads    =('balance',      'count'),
                    leads_48 =('over_48',      'sum'))
               .reset_index()
               .merge(b48, on='Caller_name', how='left'))
        agg['bal_48hr'] = agg['bal_48hr'].fillna(0)
        agg['leads_48'] = agg['leads_48'].astype(int)
        return agg

    def build_combined_agg(curr_pending, prev_pending, meta_map, df_rev=None):
        """
        Mirrors attribute_drops_to_callers exactly:
          0. (NEW) Re-attribute each pending lead's Caller_name to the ORIGINAL
             NEW ENROLLMENT CALLER via email/phone lookup from df_rev.
             This is why DROPPED LEADS always shows correct teams — it uses the
             new-enrollment caller, not the raw booking-fees row caller.
          1. Raw-aggregate curr and prev separately (no metadata yet)
          2. Outer-join on Caller_name
          3. ONE meta_map merge at the end → correct team for every caller.
        """
        # ── Step 0: remap Caller_name to canonical new-enrollment caller ──
        if df_rev is not None and not df_rev.empty:
            new_enr = df_rev[df_rev['is_new']].copy()
            email_to_caller = (new_enr.drop_duplicates('Email_Id_norm')
                               .set_index('Email_Id_norm')['Caller_name'].to_dict())
            phone_to_caller = (new_enr.drop_duplicates('Contact_No_norm')
                               .set_index('Contact_No_norm')['Caller_name'].to_dict())

            def _resolve(row):
                e = str(row.get('Email_Id_norm', '')).strip().lower()
                p = str(row.get('Contact_No_norm', '')).strip()
                if e and e in email_to_caller: return email_to_caller[e]
                if p and p in phone_to_caller: return phone_to_caller[p]
                return row['Caller_name']  # fallback: keep original

            def _remap(df):
                if df.empty: return df
                df = df.copy()
                df['Caller_name'] = df.apply(_resolve, axis=1)
                return df

            curr_pending = _remap(curr_pending)
            prev_pending = _remap(prev_pending)

        curr_agg = _raw_caller_agg(curr_pending)
        prev_agg = _raw_caller_agg(prev_pending)

        if curr_agg.empty and prev_agg.empty:
            return pd.DataFrame()

        if not prev_agg.empty:
            prev_slim = prev_agg[['Caller_name', 'balance', 'leads']].rename(
                columns={'balance': 'prev_bal', 'leads': 'prev_leads'})
        else:
            prev_slim = pd.DataFrame(columns=['Caller_name', 'prev_bal', 'prev_leads'])

        if curr_agg.empty:
            combined = prev_slim.copy()
            for c in ['pool', 'collected', 'balance', 'leads', 'leads_48', 'bal_48hr']:
                combined[c] = 0
        elif prev_slim.empty:
            combined = curr_agg.copy()
            combined['prev_bal']   = 0
            combined['prev_leads'] = 0
        else:
            combined = curr_agg.merge(prev_slim, on='Caller_name', how='outer')

        for c in ['pool', 'collected', 'balance', 'leads', 'leads_48', 'bal_48hr', 'prev_bal', 'prev_leads']:
            if c not in combined.columns:
                combined[c] = 0
            combined[c] = combined[c].fillna(0)

        combined['grand_bal']   = combined['balance'] + combined['prev_bal']
        combined['grand_leads'] = combined['leads'].astype(int) + combined['prev_leads'].astype(int)

        # Deduplicate meta_map and merge ONCE for ALL callers (curr + prev + both)
        meta_dedup = (meta_map[['merge_key', 'Caller Name', 'Team Name', 'Vertical']]
                      .drop_duplicates(subset=['merge_key'], keep='first'))
        combined['mk'] = combined['Caller_name'].str.strip().str.lower()
        combined = combined.merge(
            meta_dedup.rename(columns={'merge_key': 'mk'}), on='mk', how='left')
        combined['Caller_name'] = combined['Caller Name'].fillna(combined['Caller_name'])
        combined['Team Name']   = combined['Team Name'].fillna('Others')
        combined['Vertical']    = combined['Vertical'].fillna('Others')
        return combined

    def _pct48(b48, bal):
        return f"{round(b48 / bal * 100)}%" if bal and bal > 0 else "0%"

    

    def attribute_drops_to_callers(drop_df, df_rev, meta_map, c_start, c_end, p_start, p_end, curr_label, prev_label):
        if drop_df.empty or df_rev.empty:
            return pd.DataFrame()
        new_enr = df_rev[df_rev['is_new']].copy()
        email_to_caller = new_enr.drop_duplicates('Email_Id_norm').set_index('Email_Id_norm')['Caller_name'].to_dict()
        phone_to_caller = new_enr.drop_duplicates('Contact_No_norm').set_index('Contact_No_norm')['Caller_name'].to_dict()

        def get_caller(row):
            e = str(row.get('drop_email', '')).strip().lower()
            p = str(row.get('drop_phone', '')).strip()
            if e and e in email_to_caller: return email_to_caller[e]
            if p and p in phone_to_caller: return phone_to_caller[p]
            return 'Unknown'

        df = drop_df.copy()
        df['attributed_caller'] = df.apply(get_caller, axis=1)
        df = df[df['drop_date'].notna()].copy()
        df['drop_d'] = df['drop_date'].dt.date

        curr_drops = (df[(df['drop_d'] >= c_start) & (df['drop_d'] <= c_end)]
                      .groupby('attributed_caller').size().rename('curr_drops'))
        prev_drops = (df[(df['drop_d'] >= p_start) & (df['drop_d'] <= p_end)]
                      .groupby('attributed_caller').size().rename('prev_drops'))

        combined = pd.DataFrame({'curr_drops': curr_drops, 'prev_drops': prev_drops}).fillna(0).reset_index()
        combined.columns = ['Caller_name', 'curr_drops', 'prev_drops']
        combined['curr_drops']   = combined['curr_drops'].astype(int)
        combined['prev_drops']   = combined['prev_drops'].astype(int)
        combined['total_drops']  = combined['curr_drops'] + combined['prev_drops']
        combined['mk']           = combined['Caller_name'].str.strip().str.lower()
        combined = combined.merge(meta_map.rename(columns={'merge_key': 'mk'}), on='mk', how='left')
        combined['Caller_name'] = combined['Caller Name'].fillna(combined['Caller_name'])
        combined['Team Name']   = combined['Team Name'].fillna('Others')
        combined['Vertical']    = combined['Vertical'].fillna('Others')
        return combined.sort_values('total_drops', ascending=False).reset_index(drop=True)

    def render_html_pending_table(combined, mode, curr_label, prev_label, title):
        hdr_style  = "background:#064e3b;color:#fff;font-size:.72rem;font-weight:700;text-transform:uppercase;padding:8px 6px;text-align:center;border:1px solid #065f46;"
        sub_style  = "background:#065f46;color:#fff;font-size:.68rem;font-weight:600;padding:6px 4px;text-align:center;border:1px solid #0d9e6e;"
        data_style = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#ffffff;"
        data_style_alt = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#f0fdf4;"
        team_style = "font-weight:700;background:#1f2937;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #374151;"
        vert_style = "font-weight:700;background:#064e3b;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #065f46;"
        grand_style= "font-weight:700;background:#1e3a5f;color:#fff;font-size:.72rem;padding:8px 5px;text-align:center;border:1px solid #1e3a5f;"
        name_col   = "CALLER NAME" if mode == "caller" else "TEAM NAME"
        extra_th   = '<th rowspan="2" style="' + hdr_style + '">TEAM</th>' if mode == "caller" else ""

        html = (
            "<div style='margin:1.5rem 0 .5rem;text-align:center;'>"
            "<div style='font-size:1rem;font-weight:800;text-transform:uppercase;"
            "letter-spacing:1px;color:#10B981;margin-bottom:.5rem;'>"
            + title +
            "</div></div>"
            "<div style='overflow-x:auto;'>"
            "<table style='width:100%;border-collapse:collapse;'>"
            "<thead>"
            "<tr>"
            "<th rowspan='2' style='" + hdr_style + "'>" + name_col + "</th>"
            + extra_th +
            "<th colspan='7' style='" + hdr_style + "background:#065f46;'>" + curr_label.upper() + "</th>"
            "<th colspan='2' style='" + hdr_style + "background:#92400e;'>" + prev_label.upper() + "</th>"
            "<th colspan='2' style='" + hdr_style + "background:#1e3a5f;'>GRAND TOTAL</th>"
            "</tr>"
            "<tr>"
            "<th style='" + sub_style + "'>REVENUE POOL</th>"
            "<th style='" + sub_style + "'>TOTAL REVENUE COLLECTED</th>"
            "<th style='" + sub_style + "'>BALANCE AMOUNT TO BE RECOVERED</th>"
            "<th style='" + sub_style + "'>NO. OF LEADS (BALANCE TO BE RECOVERED)</th>"
            "<th style='" + sub_style + "background:#0f766e;'>NO. OF LEADS PENDING &gt;48 HRS</th>"
            "<th style='" + sub_style + "background:#0f766e;'>BALANCE RECOVERY PENDING &gt;48 HRS</th>"
            "<th style='" + sub_style + "background:#0f766e;'>PERCENTAGE OF REVENUE PENDING &gt;48 HRS</th>"
            "<th style='" + sub_style + "background:#92400e;'>BALANCE AMOUNT TO BE RECOVERED</th>"
            "<th style='" + sub_style + "background:#92400e;'>NO. OF LEADS (BALANCE AMOUNT TO BE RECOVERED)</th>"
            "<th style='" + sub_style + "background:#1e3a5f;'>AMOUNT TO BE RECOVERED</th>"
            "<th style='" + sub_style + "background:#1e3a5f;'>TOTAL LEAD TO BE RECOVERED</th>"
            "</tr>"
            "</thead><tbody>"
        )

        g = {k: 0 for k in ["pool","collected","balance","leads","leads_48","bal_48hr","prev_bal","prev_leads","grand_bal","grand_leads"]}

        num_keys = ["pool","collected","balance","leads","leads_48","bal_48hr","prev_bal","prev_leads","grand_bal","grand_leads"]

        # Sort verticals by total grand_bal descending
        vert_order = (
            combined.copy()
            .assign(_vert=combined["Vertical"].fillna("Unassigned"))
            .groupby("_vert")["grand_bal"]
            .sum()
            .sort_values(ascending=False)
            .index.tolist()
        )

        for vert in vert_order:
            v_df = combined[combined["Vertical"].fillna("Unassigned") == vert].copy()
            if v_df.empty:
                continue
            # Force numeric on all aggregation columns
            for k in num_keys:
                if k in v_df.columns:
                    v_df[k] = pd.to_numeric(v_df[k], errors='coerce').fillna(0)
            v = {k: 0 for k in g}

            # Sort teams within vertical by total grand_bal descending
            team_order = (
                v_df.assign(_team=v_df["Team Name"].fillna("Unassigned"))
                .groupby("_team")["grand_bal"]
                .sum()
                .sort_values(ascending=False)
                .index.tolist()
            )

            for team in team_order:
                t_df = v_df[v_df["Team Name"].fillna("Unassigned") == team].copy()
                t = {k: 0 for k in g}

                if mode == "caller":
                    for idx, r in t_df.sort_values("balance", ascending=False).iterrows():
                        row_style = data_style_alt if idx % 2 == 0 else data_style
                        html += "<tr>"
                        html += "<td style='" + row_style + "'>" + str(r.get("Caller_name","—")) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(r.get("Team Name","—")) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("pool",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("collected",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("balance",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("leads",0))) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("leads_48",0))) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("bal_48hr",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + _pct48(r.get("bal_48hr",0), r.get("balance",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("prev_bal",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("prev_leads",0))) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("grand_bal",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("grand_leads",0))) + "</td>"
                        html += "</tr>"
                        for k in g: t[k] += r.get(k,0)

                else: # Team mode
                    for k in g: t[k] = t_df[k].sum()

                # Team Total
                html += "<tr>"
                html += "<td style='" + team_style + "'>" + (team if mode=="team" else team+" Total") + "</td>"
                if mode == "caller": html += "<td style='" + team_style + "'>—</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["pool"]) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["collected"]) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["balance"]) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["leads"])) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["leads_48"])) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["bal_48hr"]) + "</td>"
                html += "<td style='" + team_style + "'>" + _pct48(t["bal_48hr"], t["balance"]) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["prev_bal"]) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["prev_leads"])) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["grand_bal"]) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["grand_leads"])) + "</td>"
                html += "</tr>"
                for k in g: v[k] += t[k]

            # Vertical Total
            html += "<tr>"
            html += "<td style='" + vert_style + "'>" + vert + " Total</td>"
            if mode == "caller": html += "<td style='" + vert_style + "'>—</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["pool"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["collected"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["balance"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["leads"])) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["leads_48"])) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["bal_48hr"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + _pct48(v["bal_48hr"], v["balance"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["prev_bal"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["prev_leads"])) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["grand_bal"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["grand_leads"])) + "</td>"
            html += "</tr>"
            for k in g: g[k] += v[k]

        # Grand Total
        html += "<tr>"
        html += "<td style='" + grand_style + "'>GRAND TOTAL</td>"
        if mode == "caller": html += "<td style='" + grand_style + "'>—</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["pool"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["collected"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["balance"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["leads"])) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["leads_48"])) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["bal_48hr"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + _pct48(g["bal_48hr"], g["balance"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["prev_bal"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["prev_leads"])) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["grand_bal"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["grand_leads"])) + "</td>"
        html += "</tr>"
        html += "</tbody></table></div>"
        return html


    def render_drop_html(drop_agg, curr_label, prev_label):
        hs = "background:#7c2d12;color:#fff;font-size:.72rem;font-weight:700;text-transform:uppercase;padding:8px 6px;text-align:center;border:1px solid #991b1b;"
        ds = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#ffffff;"
        ds_alt = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#f0fdf4;"
        ts = "font-weight:700;background:#1f2937;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #374151;"
        vs = "font-weight:700;background:#7c2d12;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #991b1b;"
        gs = "font-weight:700;background:#1e3a5f;color:#fff;font-size:.72rem;padding:8px 5px;text-align:center;border:1px solid #1e3a5f;"

        html = (
            "<div style='overflow-x:auto;'>"
            "<table style='width:100%;border-collapse:collapse;'>"
            "<thead><tr>"
            "<th style='" + hs + "'>CALLER NAME</th>"
            "<th style='" + hs + "'>TEAM</th>"
            "<th style='" + hs + "'>VERTICAL</th>"
            "<th style='" + hs + "'>" + curr_label.upper() + " DROP CASES</th>"
            "<th style='" + hs + "'>" + prev_label.upper() + " DROP CASES</th>"
            "<th style='" + hs + "'>TOTAL DROP CASES</th>"
            "</tr></thead><tbody>"
        )

        g_c = g_p = g_t = 0
        for vert in sorted(drop_agg["Vertical"].fillna("Others").unique()):
            v_df = drop_agg[drop_agg["Vertical"].fillna("Others") == vert]
            if v_df.empty:
                continue
            vc = vp = vt = 0
            for team in sorted(v_df["Team Name"].fillna("Others").unique()):
                t_df = v_df[v_df["Team Name"].fillna("Others") == team]
                if t_df.empty:
                    continue
                tc = tp = tt = 0
                drop_row_idx = 0
                for _, r in t_df.sort_values('total_drops', ascending=False).iterrows():
                    drow = ds_alt if drop_row_idx % 2 == 1 else ds
                    drop_row_idx += 1
                    html += f"""<tr>
                        <td style='{drow}text-align:center;'>{r['Caller_name']}</td>
                        <td style='{drow}'>{team}</td>
                        <td style='{drow}'>{vert}</td>
                        <td style='{drow}'>{int(r['curr_drops'])}</td>
                        <td style='{drow}'>{int(r['prev_drops'])}</td>
                        <td style='{drow}font-weight:600;color:#DC2626;'>{int(r['total_drops'])}</td>
                    </tr>"""
                    tc += r['curr_drops']; tp += r['prev_drops']; tt += r['total_drops']
                html += (
                    "<tr>"
                    "<td style='" + ts + "'>" + team + " Total</td>"
                    "<td style='" + ts + "'>—</td>"
                    "<td style='" + ts + "'>" + str(vert) + "</td>"
                    "<td style='" + ts + "'>" + str(int(tc)) + "</td>"
                    "<td style='" + ts + "'>" + str(int(tp)) + "</td>"
                    "<td style='" + ts + "'>" + str(int(tt)) + "</td>"
                    "</tr>"
                )
                vc += tc; vp += tp; vt += tt
            html += (
                "<tr>"
                "<td style='" + vs + "'>" + vert + " Total</td>"
                "<td style='" + vs + "'>—</td>"
                "<td style='" + vs + "'>—</td>"
                "<td style='" + vs + "'>" + str(int(vc)) + "</td>"
                "<td style='" + vs + "'>" + str(int(vp)) + "</td>"
                "<td style='" + vs + "'>" + str(int(vt)) + "</td>"
                "</tr>"
            )
            g_c += vc; g_p += vp; g_t += vt
        html += (
            "<tr>"
            "<td style='" + gs + "'>Grand Total</td>"
            "<td style='" + gs + "'>—</td>"
            "<td style='" + gs + "'>—</td>"
            "<td style='" + gs + "'>" + str(int(g_c)) + "</td>"
            "<td style='" + gs + "'>" + str(int(g_p)) + "</td>"
            "<td style='" + gs + "'>" + str(int(g_t)) + "</td>"
            "</tr></tbody></table></div>"
        )
        return html


    # ─────────────────────────────────────────────
    # EXCEL DOWNLOAD BUILDERS — PENDING TAB
    # ─────────────────────────────────────────────

    def build_pending_excel(combined, pend_curr, pend_prev, meta_map_pending, curr_label, prev_label):
        """
        Returns bytes of an xlsx with 2 tabs:
          Tab 1 — Teamwise Pending Revenue
          Tab 2 — Callerwise Pending Revenue
        All monetary values are raw integers (no K/L formatting).
        """
        HDR_FILL   = PatternFill("solid", start_color="064e3b", end_color="064e3b")
        TEAM_FILL  = PatternFill("solid", start_color="1f2937", end_color="1f2937")
        VERT_FILL  = PatternFill("solid", start_color="064e3b", end_color="064e3b")
        GRAND_FILL = PatternFill("solid", start_color="1e3a5f", end_color="1e3a5f")
        ALT_FILL   = PatternFill("solid", start_color="f0fdf4", end_color="f0fdf4")
        WHITE_FILL = PatternFill("solid", start_color="ffffff", end_color="ffffff")
        HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        DATA_FONT  = Font(name="Arial", size=9)
        # BOLD_WHITE and others used in sub-functions or later
        BOLD_WHITE = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        BORDER     = Border(
            left=Side(style='thin', color='D1FAE5'),
            right=Side(style='thin', color='D1FAE5'),
            top=Side(style='thin', color='D1FAE5'),
            bottom=Side(style='thin', color='D1FAE5'),
        )
        CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        wb = Workbook()

        num_keys = ["pool","collected","balance","leads","leads_48","bal_48hr",
                    "prev_bal","prev_leads","grand_bal","grand_leads"]

        def _safe(v):
            try:
                f = float(v)
                return 0 if pd.isna(f) else f
            except:
                return 0

        def _pct(b48, bal):
            if bal and bal > 0:
                return round(b48 / bal * 100, 1)
            return 0.0

        def write_pending_sheet(ws, mode):
            # ── Header rows ──
            cl = curr_label.upper()
            pl = prev_label.upper()
            name_col_hdr = "CALLER NAME" if mode == "caller" else "TEAM NAME"
            extra = 1 if mode == "caller" else 0   # extra TEAM column for caller mode

            # Row 1 — group headers
            col = 1
            ws.cell(1, col, name_col_hdr).fill = HDR_FILL
            ws.cell(1, col).font = HDR_FONT
            ws.cell(1, col).alignment = CENTER
            ws.cell(1, col).border = BORDER
            col += 1
            if mode == "caller":
                ws.cell(1, col, "TEAM").fill = HDR_FILL
                ws.cell(1, col).font = HDR_FONT
                ws.cell(1, col).alignment = CENTER
                ws.cell(1, col).border = BORDER
                col += 1

            # Current month group (7 cols)
            for i in range(7):
                ws.cell(1, col+i, cl if i == 0 else "").fill = PatternFill("solid", start_color="065f46", end_color="065f46")
                ws.cell(1, col+i).font = HDR_FONT
                ws.cell(1, col+i).alignment = CENTER
                ws.cell(1, col+i).border = BORDER
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+6)
            col += 7

            # Previous month group (2 cols)
            for i in range(2):
                ws.cell(1, col+i, pl if i == 0 else "").fill = PatternFill("solid", start_color="92400e", end_color="92400e")
                ws.cell(1, col+i).font = HDR_FONT
                ws.cell(1, col+i).alignment = CENTER
                ws.cell(1, col+i).border = BORDER
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            col += 2

            # Grand total group (2 cols)
            for i in range(2):
                ws.cell(1, col+i, "GRAND TOTAL" if i == 0 else "").fill = PatternFill("solid", start_color="1e3a5f", end_color="1e3a5f")
                ws.cell(1, col+i).font = HDR_FONT
                ws.cell(1, col+i).alignment = CENTER
                ws.cell(1, col+i).border = BORDER
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)

            # Row 2 — sub-headers
            sub_headers = [name_col_hdr] + (["TEAM"] if mode == "caller" else []) + [
                "REVENUE POOL (₹)", "COLLECTED (₹)", "BALANCE (₹)", "LEADS",
                "LEADS >48HR", "BALANCE >48HR (₹)", "% PENDING >48HR",
                f"BALANCE ({pl}) (₹)", f"LEADS ({pl})",
                "GRAND BALANCE (₹)", "GRAND LEADS"
            ]
            for c_idx, h in enumerate(sub_headers, 1):
                cell = ws.cell(2, c_idx, h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER

            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 28

            # ── Data rows ──
            row_num = 3
            g = {k: 0.0 for k in num_keys}

            vert_order = (
                combined.assign(_v=combined["Vertical"].fillna("Unassigned"))
                .groupby("_v")["grand_bal"].sum()
                .sort_values(ascending=False).index.tolist()
            )

            for vert in vert_order:
                v_df = combined[combined["Vertical"].fillna("Unassigned") == vert].copy()
                if v_df.empty:
                    continue
                for k in num_keys:
                    if k in v_df.columns:
                        v_df[k] = pd.to_numeric(v_df[k], errors='coerce').fillna(0)
                v = {k: 0.0 for k in num_keys}

                team_order = (
                    v_df.assign(_t=v_df["Team Name"].fillna("Unassigned"))
                    .groupby("_t")["grand_bal"].sum()
                    .sort_values(ascending=False).index.tolist()
                )

                for team in team_order:
                    t_df = v_df[v_df["Team Name"].fillna("Unassigned") == team].copy()
                    if t_df.empty:
                        continue
                    t = {k: 0.0 for k in num_keys}

                    if mode == "caller":
                        alt = False
                        for _, r in t_df.sort_values("balance", ascending=False).iterrows():
                            vals = [
                                str(r.get("Caller_name", "—")),
                                str(team),
                                _safe(r.get("pool", 0)),
                                _safe(r.get("collected", 0)),
                                _safe(r.get("balance", 0)),
                                int(_safe(r.get("leads", 0))),
                                int(_safe(r.get("leads_48", 0))),
                                _safe(r.get("bal_48hr", 0)),
                                _pct(_safe(r.get("bal_48hr", 0)), _safe(r.get("balance", 0))),
                                _safe(r.get("prev_bal", 0)),
                                int(_safe(r.get("prev_leads", 0))),
                                _safe(r.get("grand_bal", 0)),
                                int(_safe(r.get("grand_leads", 0))),
                            ]
                            fill = ALT_FILL if alt else WHITE_FILL
                            alt = not alt
                            for c_idx, v_ in enumerate(vals, 1):
                                cell = ws.cell(row_num, c_idx, v_)
                                cell.fill = fill
                                cell.font = DATA_FONT
                                cell.border = BORDER
                                cell.alignment = LEFT if c_idx <= 2 else CENTER
                            row_num += 1
                            for k in num_keys:
                                t[k] += _safe(r.get(k, 0))
                    else:
                        for k in num_keys:
                            t[k] = pd.to_numeric(t_df[k], errors='coerce').fillna(0).sum() if k in t_df.columns else 0

                    # Team total row
                    t_vals = ([f"{team} Total"] + (["—"] if mode == "caller" else []) +
                              [t["pool"], t["collected"], t["balance"], int(t["leads"]),
                               int(t["leads_48"]), t["bal_48hr"], _pct(t["bal_48hr"], t["balance"]),
                               t["prev_bal"], int(t["prev_leads"]), t["grand_bal"], int(t["grand_leads"])])
                    for c_idx, v_ in enumerate(t_vals, 1):
                        cell = ws.cell(row_num, c_idx, v_)
                        cell.fill = TEAM_FILL
                        cell.font = BOLD_WHITE
                        cell.border = BORDER
                        cell.alignment = LEFT if c_idx == 1 else CENTER
                    row_num += 1

                    for k in num_keys:
                        v[k] += t[k]

                # Vertical total row
                v_vals = ([f"{vert} Total"] + (["—"] if mode == "caller" else []) +
                          [v["pool"], v["collected"], v["balance"], int(v["leads"]),
                           int(v["leads_48"]), v["bal_48hr"], _pct(v["bal_48hr"], v["balance"]),
                           v["prev_bal"], int(v["prev_leads"]), v["grand_bal"], int(v["grand_leads"])])
                for c_idx, v_ in enumerate(v_vals, 1):
                    cell = ws.cell(row_num, c_idx, v_)
                    cell.fill = VERT_FILL
                    cell.font = BOLD_WHITE
                    cell.border = BORDER
                    cell.alignment = LEFT if c_idx == 1 else CENTER
                row_num += 1

                for k in num_keys:
                    g[k] += v[k]

            # Grand total row
            g_vals = (["Grand Total"] + (["—"] if mode == "caller" else []) +
                      [g["pool"], g["collected"], g["balance"], int(g["leads"]),
                       int(g["leads_48"]), g["bal_48hr"], _pct(g["bal_48hr"], g["balance"]),
                       g["prev_bal"], int(g["prev_leads"]), g["grand_bal"], int(g["grand_leads"])])
            for c_idx, v_ in enumerate(g_vals, 1):
                cell = ws.cell(row_num, c_idx, v_)
                cell.fill = GRAND_FILL
                cell.font = BOLD_WHITE
                cell.border = BORDER
                cell.alignment = LEFT if c_idx == 1 else CENTER

            # Column widths
            col_widths = [28] + ([18] if mode == "caller" else []) + [16, 16, 16, 8, 10, 16, 14, 16, 10, 16, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # Tab 1 — Teamwise
        ws1 = wb.active
        ws1.title = f"Teamwise {curr_label[:3]}{curr_label[-4:]}+{prev_label[:3]}{prev_label[-4:]}"[:31]
        ws1.freeze_panes = "A3"
        write_pending_sheet(ws1, "team")

        # Tab 2 — Callerwise
        ws2 = wb.create_sheet(f"Callerwise {curr_label[:3]}{curr_label[-4:]}+{prev_label[:3]}{prev_label[-4:]}"[:31])
        ws2.freeze_panes = "A3"
        write_pending_sheet(ws2, "caller")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    def build_pending_leads_excel(pend_curr, pend_prev, meta_map_pending, curr_label, prev_label):
        """
        Returns bytes of an xlsx with 2 tabs — one per month — with full lead details
        including Caller Name, Team Name, Vertical, Course Price, Revenue Collected, Balance.
        """
        HDR_FILL  = PatternFill("solid", start_color="064e3b", end_color="064e3b")
        ALT_FILL  = PatternFill("solid", start_color="f0fdf4", end_color="f0fdf4")
        WHITE_FILL= PatternFill("solid", start_color="ffffff", end_color="ffffff")
        HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        DATA_FONT = Font(name="Arial", size=9)
        CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        BORDER    = Border(
            left=Side(style='thin', color='D1FAE5'),
            right=Side(style='thin', color='D1FAE5'),
            top=Side(style='thin', color='D1FAE5'),
            bottom=Side(style='thin', color='D1FAE5'),
        )

        COLS = ["DATE", "NAME", "CONTACT NO", "EMAIL ID", "COURSE",
                "CALLER NAME", "TEAM NAME", "VERTICAL",
                "COURSE PRICE (₹)", "REVENUE COLLECTED (₹)", "BALANCE (₹)"]
        COL_WIDTHS = [12, 28, 14, 32, 32, 22, 22, 18, 16, 20, 14]

        # Build meta lookup: merge_key → {Team Name, Vertical, Caller Name}
        meta_lkp = {}
        if not meta_map_pending.empty:
            for _, row in meta_map_pending.iterrows():
                k = str(row.get('merge_key', '')).strip().lower()
                meta_lkp[k] = {
                    'Caller Name': row.get('Caller Name', '—'),
                    'Team Name':   row.get('Team Name',   '—'),
                    'Vertical':    row.get('Vertical',    '—'),
                }

        def enrich(df):
            if df.empty:
                return df
            d = df.copy()
            d['_mk'] = d['Caller_name'].astype(str).str.strip().str.lower()
            d['_cn']   = d['_mk'].map(lambda k: meta_lkp.get(k, {}).get('Caller Name', d.loc[d['_mk']==k, 'Caller_name'].iloc[0] if not d[d['_mk']==k].empty else '—'))
            d['_team'] = d['_mk'].map(lambda k: meta_lkp.get(k, {}).get('Team Name', '—'))
            d['_vert'] = d['_mk'].map(lambda k: meta_lkp.get(k, {}).get('Vertical',  '—'))
            return d

        def write_leads_sheet(ws, df, label):
            ws.row_dimensions[1].height = 28
            for c_idx, h in enumerate(COLS, 1):
                cell = ws.cell(1, c_idx, h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER
            for i, w in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            if df.empty:
                ws.cell(2, 1, f"No pending leads found for {label}")
                return

            df = enrich(df)
            alt = False
            for r_idx, (_, r) in enumerate(df.sort_values('Date').iterrows(), 2):
                fill = ALT_FILL if alt else WHITE_FILL
                alt  = not alt
                _raw_phone = str(r.get('Contact_No', '') or '').strip().replace(' ', '')
                try:
                    _phone_val = int(_raw_phone) if _raw_phone.lstrip('+').isdigit() else _raw_phone
                except:
                    _phone_val = _raw_phone
                vals = [
                    str(r.get('Date', '')),
                    str(r.get('Name', '')),
                    _phone_val,
                    str(r.get('Email_Id', '')),
                    str(r.get('Course', '')),
                    str(r.get('_cn', r.get('Caller_name', '—'))),
                    str(r.get('_team', '—')),
                    str(r.get('_vert', '—')),
                    float(r.get('Course_Price', 0) or 0),
                    float(r.get('Fee_paid', 0) or 0),
                    float(r.get('balance', 0) or 0),
                ]
                for c_idx, v_ in enumerate(vals, 1):
                    cell = ws.cell(r_idx, c_idx, v_)
                    cell.fill = fill
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    cell.alignment = LEFT if c_idx <= 6 else CENTER

        wb = Workbook()
        ws1 = wb.active
        ws1.title = f"Pending {curr_label[:3]} {curr_label[-4:]}"[:31]
        write_leads_sheet(ws1, pend_curr, curr_label)

        ws2 = wb.create_sheet(f"Pending {prev_label[:3]} {prev_label[-4:]}"[:31])
        write_leads_sheet(ws2, pend_prev, prev_label)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    def build_drop_leads_excel(drop_df, c_start, c_end, p_start, p_end, curr_label, prev_label, meta_map_pending=None):
        """
        Returns bytes of an xlsx with 2 tabs:
          Tab 1 — Current month dropped leads
          Tab 2 — Previous month dropped leads
        Columns: DATE, EMAIL, PHONE NUMBER, CALLER NAME, TEAM, VERTICAL
        """
        if meta_map_pending is None:
            meta_map_pending = pd.DataFrame()

        HDR_FILL   = PatternFill("solid", start_color="7c2d12", end_color="7c2d12")
        ALT_FILL   = PatternFill("solid", start_color="fff7ed", end_color="fff7ed")
        WHITE_FILL = PatternFill("solid", start_color="ffffff", end_color="ffffff")
        HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        DATA_FONT  = Font(name="Arial", size=9)
        CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT       = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        BORDER     = Border(
            left=Side(style='thin', color='FED7AA'),
            right=Side(style='thin', color='FED7AA'),
            top=Side(style='thin', color='FED7AA'),
            bottom=Side(style='thin', color='FED7AA'),
        )

        COLS       = ["DATE", "EMAIL", "PHONE NUMBER", "CALLER NAME", "TEAM", "VERTICAL"]
        COL_WIDTHS = [14, 36, 16, 28, 22, 18]

        e_col = next((c for c in drop_df.columns if 'email' in c.lower() and ('drop' in c.lower() or 'student' in c.lower())), None)
        p_col = next((c for c in drop_df.columns if 'phone' in c.lower() or ('number' in c.lower() and 'drop' in c.lower())), None)
        t_col = next((c for c in drop_df.columns if 'timestamp' in c.lower()), None)

        def write_drop_sheet(ws, df_slice, label):
            ws.row_dimensions[1].height = 28
            for c_idx, h in enumerate(COLS, 1):
                cell = ws.cell(1, c_idx, h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER
            for i, w in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            if df_slice.empty:
                ws.cell(2, 1, f"No dropped leads for {label}")
                return

            alt = False
            for r_idx, (_, r) in enumerate(df_slice.iterrows(), 2):
                fill = ALT_FILL if alt else WHITE_FILL
                alt  = not alt
                drop_date = r.get('drop_date', '')
                if hasattr(drop_date, 'date'):
                    drop_date = drop_date.date()
                _dp = str(r.get('drop_phone', '') or '').strip().replace(' ', '')
                try:
                    _dp_val = int(_dp) if _dp.lstrip('+').isdigit() else _dp
                except:
                    _dp_val = _dp
                vals = [
                    str(drop_date),
                    str(r.get('drop_email', '')),
                    _dp_val,
                    str(r.get('attributed_caller', '—')),
                    str(r.get('_team', '—')),
                    str(r.get('_vert', '—')),
                ]
                for c_idx, v_ in enumerate(vals, 1):
                    cell = ws.cell(r_idx, c_idx, v_)
                    cell.fill = fill
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    cell.alignment = LEFT

        wb  = Workbook()
        ws1 = wb.active
        ws1.title = f"{curr_label[:3]} {curr_label[-4:]} Drops"[:31]

        # Attribute caller to each drop row using revenue data
        df_work = drop_df.copy()
        df_work['drop_d'] = pd.to_datetime(df_work['drop_date'], errors='coerce').dt.date

        # Build email/phone → caller attribution from the revenue table
        try:
            _rev_attr = fetch_both_months_rev(p_start, c_end)
            if not _rev_attr.empty:
                _new_enr_attr = _rev_attr[_rev_attr['is_new']].copy()
                _email_to_caller = _new_enr_attr.drop_duplicates('Email_Id_norm').set_index('Email_Id_norm')['Caller_name'].to_dict()
                _phone_to_caller = _new_enr_attr.drop_duplicates('Contact_No_norm').set_index('Contact_No_norm')['Caller_name'].to_dict()

                def _get_caller(row):
                    e = str(row.get('drop_email', '')).strip().lower()
                    p = str(row.get('drop_phone', '')).strip()
                    if e and e in _email_to_caller: return _email_to_caller[e]
                    if p and p in _phone_to_caller: return _phone_to_caller[p]
                    return 'Unknown'

                df_work['attributed_caller'] = df_work.apply(_get_caller, axis=1)

                # Build caller → team/vertical lookup from meta
                _meta_lkp_drop = {}
                if not meta_map_pending.empty:
                    for _, _mr in meta_map_pending.iterrows():
                        _mk = str(_mr.get('merge_key', '')).strip().lower()
                        _meta_lkp_drop[_mk] = {
                            'Team Name': _mr.get('Team Name', '—'),
                            'Vertical':  _mr.get('Vertical',  '—'),
                        }
                df_work['_team'] = df_work['attributed_caller'].apply(
                    lambda c: _meta_lkp_drop.get(str(c).strip().lower(), {}).get('Team Name', '—'))
                df_work['_vert'] = df_work['attributed_caller'].apply(
                    lambda c: _meta_lkp_drop.get(str(c).strip().lower(), {}).get('Vertical', '—'))
            else:
                df_work['attributed_caller'] = 'Unknown'
                df_work['_team'] = '—'
                df_work['_vert'] = '—'
        except Exception:
            df_work['attributed_caller'] = 'Unknown'
            df_work['_team'] = '—'
            df_work['_vert'] = '—'

        curr_drops = df_work[(df_work['drop_d'] >= c_start) & (df_work['drop_d'] <= c_end)].copy()
        prev_drops = df_work[(df_work['drop_d'] >= p_start) & (df_work['drop_d'] <= p_end)].copy()

        write_drop_sheet(ws1, curr_drops.sort_values('drop_d'), curr_label)

        ws2 = wb.create_sheet(f"{prev_label[:3]} {prev_label[-4:]} Drops"[:31])
        write_drop_sheet(ws2, prev_drops.sort_values('drop_d'), prev_label)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    # ─────────────────────────────────────────────
    # SIDEBAR
    # ─────────────────────────────────────────────

    min_d, max_d = get_available_dates()

    st.sidebar.markdown("""
    <div style='padding:.5rem 0 .4rem; text-align:center;'>
        <div style='font-size:.72rem; font-weight:700; text-transform:uppercase;
                    letter-spacing:1px; color:var(--text-muted,#6B7280);'>Report Controls</div>
    </div>
    """, unsafe_allow_html=True)

    min_d = pd.Timestamp(min_d).date()
    max_d = pd.Timestamp(max_d).date()

    _s            = date(max_d.year, max_d.month, 1)
    _next_m       = (_s.replace(day=28) + timedelta(days=4)).replace(day=1)
    _month_end    = _next_m - timedelta(days=1)
    default_start = max(_s, min_d)
    default_end   = min(_month_end, max_d)
    if default_start > default_end:
        default_start = default_end

    selected_dates = st.sidebar.date_input(
        "📅 Date Range",
        value=(default_start, default_end),
        min_value=min_d, max_value=max_d, format="DD-MM-YYYY",
        key="rev_date_input"
    )

    if isinstance(selected_dates, tuple) and len(selected_dates) == 2:
        start_date, end_date = selected_dates
    else:
        start_date = end_date = selected_dates if not isinstance(selected_dates, tuple) else selected_dates[0]

    teams_meta, verticals_meta, df_team_mapping = get_metadata()
    # ── Role-aware sidebar filters ──────────────────────────
    _role     = st.session_state.get('rf_role', 'admin')
    _rf_teams = st.session_state.get('rf_teams', [])
    _rf_cname = st.session_state.get('rf_caller_name', '')

    if _role == 'admin':
        selected_vertical = st.sidebar.multiselect("👑 Filter by Vertical", options=verticals_meta, key="rev_vert_multiselect")
        selected_team     = st.sidebar.multiselect("👥 Filter by Team",     options=teams_meta,    key="rev_team_multiselect")
        search_query      = st.sidebar.text_input("👤 Search Caller Name",                         key="rev_search_input")
    elif _role == 'vertical_head':
        selected_team     = _rf_teams
        selected_vertical = []
        search_query      = st.sidebar.text_input("👤 Search Caller Name", key="rev_search_input")
        st.sidebar.caption(f"🔒 Showing: {', '.join(_rf_teams)}")
    elif _role in ('tl', 'trainer'):
        selected_team     = _rf_teams
        selected_vertical = []
        search_query      = st.sidebar.text_input("👤 Search Caller Name", key="rev_search_input")
        st.sidebar.caption(f"🔒 Team: {', '.join(_rf_teams)}")
    else:  # caller
        selected_team     = []
        selected_vertical = []
        search_query      = _rf_cname
        st.sidebar.caption(f"👤 Viewing: {_rf_cname}")

    gen_report = st.sidebar.button("💰 Generate Revenue Report", key="rev_gen_btn")
    st.sidebar.download_button(
        label="📖 Metrics Guide (PDF)",
        data=generate_helper_pdf_bytes(),
        file_name="Revenue_Metrics_Logic_Guide.pdf",
        mime="application/pdf",
        key="dl_helper_pdf"
    )
  

    # ─────────────────────────────────────────────
    # HEADER BANNER
    # ─────────────────────────────────────────────

    last_update_str = get_last_update()
    display_start   = start_date.strftime('%d-%m-%Y')
    display_end     = end_date.strftime('%d-%m-%Y')

    st.markdown(f"""
    <div class="rv-header">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:.75rem;">
            <div>
                <div class="rv-title">💰 REVENUE METRICS</div>
                <div class="rv-subtitle">REVENUE PERIOD&nbsp;·&nbsp; {display_start} to {display_end}</div>
            </div>
            <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.25rem;">
                <span class="rv-badge"><span class="rv-pulse"></span>LIVE REVENUE DATA</span>
                <span class="rv-badge">🕐 UPDATED AT: {last_update_str}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "💰 Revenue Dashboard",
    "🧠 Insights & Leaderboard",
    "📊 Callerwise Pending Revenue",
    "📋 Revenue Update",
    "🎯 Target Vs Achievement",
])

    with tab1:
        if gen_report:
            with st.spinner("Fetching revenue data…"):
                df_raw = fetch_revenue_data(start_date, end_date)

                if df_raw.empty:
                    st.warning("No revenue records found for the selected period (Fee Paid > 0).")
                else:
                    # Merge team info
                    df = pd.merge(
                        df_raw,
                        df_team_mapping[['merge_key','Caller Name','Team Name','Vertical']].drop_duplicates('merge_key'),
                        on='merge_key', how='left'
                    )
                    df['Caller_name'] = df['Caller Name'].fillna(df['Caller_name'])

                    # Apply filters
                    if selected_team:
                        df = df[df['Team Name'].isin(selected_team)]
                    if selected_vertical:
                        df = df[df['Vertical'].isin(selected_vertical)]
                    if search_query:
                        df = df[df['Caller_name'].str.contains(search_query, case=False, na=False)]

                    if df.empty:
                        st.error("No records match the selected filters.")
                    else:
                        # ── Classify callers ──
                        calling_df, collection_df, both_df = classify_and_process(
                            df, df_team_mapping, start_date, end_date
                        )

                        # ── Summary metrics ──
                        metrics = compute_summary_metrics(df)

                        # ── TOP 3 HIGHLIGHTS ──
                        section_header("🏆 TOP 3 REVENUE HIGHLIGHTS")
                        top_cols = st.columns(3)

                        # Top Revenue Caller — from calling_df
                        with top_cols[0]:
                            if not calling_df.empty:
                                top_c = calling_df.iloc[0]
                                st.markdown(f"""
                                <div class="metric-card" style="border-top:3px solid var(--accent-primary);">
                                    <div class="metric-label">🥇 TOP REVENUE — CALLER</div>
                                    <div class="metric-value" style="font-size:1.1rem;">{top_c['CALLER NAME']}</div>
                                    <div class="metric-delta">{fmt_inr(top_c['raw_calling_rev'])} Calling Revenue</div>
                                </div>""", unsafe_allow_html=True)
                            else:
                                st.markdown("""<div class="metric-card" style="border-top:3px solid var(--gold);">
                                    <div class="metric-label">🥇 TOP REVENUE — CALLER</div>
                                    <div class="metric-value" style="font-size:1rem;">No Data</div>
                                </div>""", unsafe_allow_html=True)

                        # Most Enrollments — from all agent types combined
                        with top_cols[1]:
                            _agent_list = [d for d in [calling_df, collection_df, both_df] if not d.empty]
                            all_agents = pd.concat(_agent_list, ignore_index=True) if _agent_list else pd.DataFrame()
                            if not all_agents.empty and all_agents['raw_enrollments'].max() > 0:
                                top_enr = all_agents.sort_values('raw_enrollments', ascending=False).iloc[0]
                                st.markdown(f"""
                                <div class="metric-card" style="border-top:3px solid var(--accent-primary);">
                                    <div class="metric-label">🎓 MOST ENROLLMENTS</div>
                                    <div class="metric-value" style="font-size:1.1rem;">{top_enr['CALLER NAME']}</div>
                                    <div class="metric-delta">{top_enr['raw_enrollments']} New Enrollments</div>
                                </div>""", unsafe_allow_html=True)
                            else:
                                st.markdown("""<div class="metric-card" style="border-top:3px solid var(--silver);">
                                    <div class="metric-label">🎓 MOST ENROLLMENTS</div>
                                    <div class="metric-value" style="font-size:1rem;">No Data</div>
                                </div>""", unsafe_allow_html=True)

                        # Top Revenue Collection Caller — from collection_df
                        with top_cols[2]:
                            _coll_list = [d for d in [collection_df] if not d.empty]
                            coll_combined = pd.concat(_coll_list, ignore_index=True) if _coll_list else pd.DataFrame()
                            if not coll_combined.empty:
                                top_coll = coll_combined.sort_values('raw_collection_rev', ascending=False).iloc[0]
                                st.markdown(f"""
                                <div class="metric-card" style="border-top:3px solid var(--accent-primary);">
                                    <div class="metric-label">🥇 TOP REVENUE — COLLECTION</div>
                                    <div class="metric-value" style="font-size:1.1rem;">{top_coll['CALLER NAME']}</div>
                                    <div class="metric-delta">{fmt_inr(top_coll['raw_collection_rev'])} Collection Revenue</div>
                                </div>""", unsafe_allow_html=True)
                            else:
                                st.markdown("""<div class="metric-card" style="border-top:3px solid var(--bronze);">
                                    <div class="metric-label">🥇 TOP REVENUE — COLLECTION</div>
                                    <div class="metric-value" style="font-size:1rem;">No Data</div>
                                </div>""", unsafe_allow_html=True)

                        # ── SUMMARY KPI CARDS ──
                        section_header("💵 REVENUE SUMMARY METRICS")

                        kpis = [
                            ("Total Revenue (EXCL. Services)",                   fmt_inr(metrics['total_rev']),      "💰"),
                            ("Calling Revenue (INCL. Funnel)",                  fmt_inr(metrics['calling_rev']),    "📞"),
                            ("Bootcamp-Direct Revenue",         fmt_inr(metrics['bootcamp_direct_rev']),     "🎓"),
                            ("Bootcamp-Collection Revenue",              fmt_inr(metrics['collection_rev']), "🏦"),
                            ("Community Revenue (Direct + Collection)",               fmt_inr(metrics['community_rev']),  "🌐"),
                            ("Direct/Other Revenue (INCL. Funnel)",                  fmt_inr(metrics['direct_rev']),     "🎯"),
                            ("Details Not Available/ Not Updated Yet",           fmt_inr(metrics['dna_rev']),        "❓"),
                        ]

                        cols = st.columns(len(kpis))
                        for col, (label, val, icon) in zip(cols, kpis):
                            with col:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">{icon} {label}</div>
                                    <div class="metric-value">{val}</div>
                                </div>""", unsafe_allow_html=True)

                        st.divider()

                        # ══════════════════════════════════════
                        # ENROLLMENT SUMMARY METRICS
                        # ══════════════════════════════════════
                        section_header("🎓 ENROLLMENT SUMMARY METRICS")

                        enr_metrics = compute_enrollment_metrics(df)

                        enr_kpis = [
                            ("Total Enrollments",             enr_metrics['total_enr'],     "🎯"),
                            ("Caller Enrollments",            enr_metrics['caller_enr'],    "📞"),
                            ("Direct Enrollments",            enr_metrics['direct_enr'],    "🎯"),
                            ("Bootcamp-Direct Enrollments",   enr_metrics['bootcamp_enr'],  "🎓"),
                            ("Community-Direct Enrollments",  enr_metrics['community_enr'], "🌐"),
                        ]

                        enr_cols = st.columns(len(enr_kpis))
                        for col, (label, val, icon) in zip(enr_cols, enr_kpis):
                            with col:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">{icon} {label}</div>
                                    <div class="metric-value">{val}</div>
                                </div>""", unsafe_allow_html=True)

                        st.divider()

                        # ── Download Summary Cards ──
                        _rev_summary = pd.DataFrame([{
                            'Metric'  : label,
                            'Amount'  : val
                        } for label, val, _ in [
                            ("Total Revenue (EXCL. Services)",          metrics['total_rev'],           ""),
                            ("Calling Revenue (INCL. Funnel)",          metrics['calling_rev'],         ""),
                            ("Bootcamp-Direct Revenue",                 metrics['bootcamp_direct_rev'], ""),
                            ("Bootcamp-Collection Revenue",             metrics['collection_rev'],      ""),
                            ("Community Revenue (Direct+Collection)",   metrics['community_rev'],       ""),
                            ("Direct/Other Revenue (INCL. Funnel)",     metrics['direct_rev'],          ""),
                            ("Details Not Available / Not Updated Yet", metrics['dna_rev'],             ""),
                        ]])

                        _enr_summary = pd.DataFrame([{
                            'Metric': label,
                            'Count' : val
                        } for label, val, _ in [
                            ("Total Enrollments",            enr_metrics['total_enr'],     ""),
                            ("Caller Enrollments",           enr_metrics['caller_enr'],    ""),
                            ("Direct Enrollments",           enr_metrics['direct_enr'],    ""),
                            ("Bootcamp-Direct Enrollments",  enr_metrics['bootcamp_enr'],  ""),
                            ("Community-Direct Enrollments", enr_metrics['community_enr'], ""),
                        ]])

                        _combined = pd.concat([
                            _rev_summary.rename(columns={'Amount': 'Value'}),
                            _enr_summary.rename(columns={'Count': 'Value'})
                        ], ignore_index=True)

                        st.download_button(
                            label="📥 Download Revenue & Enrollment Summary",
                            data=_combined.to_csv(index=False).encode('utf-8'),
                            file_name=f"Summary_{display_start}_to_{display_end}.csv",
                            mime='text/csv',
                            key='dl_summary'
                        )

                        st.divider()

                        # ══════════════════════════════════════
                        # TABLE 1 — CALLER REVENUE PERFORMANCE
                        # ══════════════════════════════════════
                        section_header("📞 CALLER REVENUE PERFORMANCE TABLE")

                        calling_display_cols = [
                            'DESIGNATION', 'CALLER NAME', 'TEAM', 'VERTICAL',
                            'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)', 'ENROLLMENTS',
                            'ENROLLMENT REV', 'BALANCE REV',
                            'CALLING REVENUE', 'ACHIEVEMENT %'
                        ]

                        if not calling_df.empty:
                            calling_totals = {
                                'TOTAL TARGET (₹)'    : fmt_inr(calling_df['raw_target'].sum()),
                                'TILL DAY TARGET (₹)' : fmt_inr(calling_df['TILL DAY TARGET (₹)'].sum()),
                                'ENROLLMENTS'         : int(calling_df['raw_enrollments'].sum()),
                                'ENROLLMENT REV'      : fmt_inr(calling_df['ENROLLMENT REV'].sum()),
                                'BALANCE REV'         : fmt_inr(calling_df['BALANCE REV'].sum()),
                                'CALLING REVENUE'     : fmt_inr(calling_df['CALLING REVENUE'].sum()),
                                'ACHIEVEMENT %'       : f"{round(calling_df['CALLING REVENUE'].sum() / calling_df['raw_target'].sum() * 100, 1) if calling_df['raw_target'].sum() > 0 else 0}%",
                            }
                        else:
                            calling_totals = {}

                        render_perf_table(
                            calling_df, calling_display_cols,
                            calling_totals, 'raw_calling_rev', 'calling'
                        )
                        if not calling_df.empty:
                            st.download_button(
                                label="📥 Download Caller Revenue CSV",
                                data=calling_df[calling_display_cols].to_csv(index=False).encode('utf-8'),
                                file_name=f"Caller_Revenue_{display_start}_to_{display_end}.csv",
                                mime='text/csv', key='dl_calling'
                            )

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE 2
                        # ══════════════════════════════════════════════
                        section_header("🏦 COLLECTION CALLER REVENUE PERFORMANCE TABLE")

                        collection_display_cols = [
                            'DESIGNATION', 'CALLER NAME', 'TEAM', 'VERTICAL',
                            'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)', 'ENROLLMENTS',
                            'CALLING REVENUE',
                            'COMMUNITY COLLECTION', 'BOOTCAMP COLLECTION',
                            'COLLECTION REVENUE'
                        ]

                        if not collection_df.empty:
                            collection_totals = {
                                'TOTAL TARGET (₹)'    : fmt_inr(collection_df['raw_target'].sum()),
                                'TILL DAY TARGET (₹)' : fmt_inr(collection_df['TILL DAY TARGET (₹)'].sum()),
                                'ENROLLMENTS'         : int(collection_df['raw_enrollments'].sum()),
                                'CALLING REVENUE'     : fmt_inr(collection_df['CALLING REVENUE'].sum()),
                                'COMMUNITY COLLECTION': fmt_inr(collection_df['COMMUNITY COLLECTION'].sum()),
                                'BOOTCAMP COLLECTION' : fmt_inr(collection_df['BOOTCAMP COLLECTION'].sum()),
                                'COLLECTION REVENUE'  : fmt_inr(collection_df['COLLECTION REVENUE'].sum()),
                            }
                        else:
                            collection_totals = {}

                        render_perf_table(
                            collection_df, collection_display_cols,
                            collection_totals, 'raw_collection_rev', 'collection'
                        )
                        if not collection_df.empty:
                            st.download_button(
                                label="📥 Download Collection Revenue CSV",
                                data=collection_df[collection_display_cols].to_csv(index=False).encode('utf-8'),
                                file_name=f"Collection_Revenue_{display_start}_to_{display_end}.csv",
                                mime='text/csv', key='dl_collection'
                            )

                        st.divider()

                        # ══════════════════════════════════════════════════════════
                        # TABLE 3
                        # ══════════════════════════════════════════════════════════
                        section_header("📞🏦 CALLING + COLLECTION CALLER REVENUE PERFORMANCE TABLE")

                        both_display_cols = [
                            'DESIGNATION', 'CALLER NAME', 'TEAM', 'VERTICAL',
                            'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)', 'ENROLLMENTS',
                            'CALLING REVENUE', 'COMMUNITY COLLECTION', 'BOOTCAMP COLLECTION',
                            'TOTAL REVENUE', 'ACHIEVEMENT %'
                        ]

                        if not both_df.empty:
                            _both_target     = both_df['raw_target'].sum()
                            _both_till       = both_df['TILL DAY TARGET (₹)'].sum()
                            _both_enr        = int(both_df['raw_enrollments'].sum())
                            _both_comm       = both_df['COMMUNITY COLLECTION'].sum()
                            _both_boot       = both_df['BOOTCAMP COLLECTION'].sum()
                            _both_calling    = both_df['CALLING REVENUE'].sum()
                            _both_collection = both_df['COLLECTION REVENUE'].sum()
                            _both_total      = both_df['raw_revenue'].sum()
                            _both_ach        = round(_both_total / _both_target * 100, 1) if _both_target > 0 else 0
                            both_totals = {
                                'TOTAL TARGET (₹)'    : fmt_inr(_both_target),
                                'TILL DAY TARGET (₹)' : fmt_inr(_both_till),
                                'ENROLLMENTS'         : _both_enr,
                                'COMMUNITY COLLECTION': fmt_inr(_both_comm),
                                'BOOTCAMP COLLECTION' : fmt_inr(_both_boot),
                                'CALLING REVENUE'     : fmt_inr(_both_calling),
                                'COLLECTION REVENUE'  : fmt_inr(_both_collection),
                                'TOTAL REVENUE'       : fmt_inr(_both_total),
                                'ACHIEVEMENT %'       : f"{_both_ach}%",
                            }
                        else:
                            both_totals = {}
                        render_perf_table(
                            both_df, both_display_cols,
                            both_totals, 'raw_revenue', 'both'
                        )
                        if not both_df.empty:
                            st.download_button(
                                label="📥 Download Calling+Collection Revenue CSV",
                                data=both_df[both_display_cols].to_csv(index=False).encode('utf-8'),
                                file_name=f"Both_Revenue_{display_start}_to_{display_end}.csv",
                                mime='text/csv', key='dl_both'
                            )

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>💰</div>
                <div style='font-size:.9rem;font-weight:600;'>Select a date range and click <b>Generate Revenue Report</b></div>
            </div>""", unsafe_allow_html=True)


        # ══════════════════════════════════════════════
        # TAB 2 — INSIGHTS & LEADERBOARD
        # ══════════════════════════════════════════════

    with tab2:
        if gen_report:
            with st.spinner("Analysing revenue patterns…"):
                df_raw = fetch_revenue_data(start_date, end_date)
                if df_raw.empty:
                    st.warning("No revenue data found for the selected period.")
                else:
                    df_ins = pd.merge(
                        df_raw,
                        df_team_mapping[['merge_key','Caller Name','Team Name','Vertical']].drop_duplicates('merge_key'),
                        on='merge_key', how='left'
                    )
                    df_ins['Caller_name'] = df_ins['Caller Name'].fillna(df_ins['Caller_name'])

                    if selected_team:
                        df_ins = df_ins[df_ins['Team Name'].isin(selected_team)]
                    if selected_vertical:
                        df_ins = df_ins[df_ins['Vertical'].isin(selected_vertical)]
                    if search_query:
                        df_ins = df_ins[df_ins['Caller_name'].str.contains(search_query, case=False, na=False)]

                    calling_ins, collection_ins, both_ins = classify_and_process(
                        df_ins, df_team_mapping, start_date, end_date
                    )

                    _ins_list = [d for d in [calling_ins, collection_ins, both_ins] if not d.empty]
                    all_agents_ins = pd.concat(_ins_list, ignore_index=True) if _ins_list else pd.DataFrame()

                    if all_agents_ins.empty:
                        st.warning("Not enough data for insights with current filters.")
                    else:
                        # ── 6 INSIGHTS ──
                        section_header("🧠 REVENUE INSIGHTS")
                        insights = compute_revenue_insights(
                            df_ins, calling_ins, collection_ins, both_ins,
                            start_date, end_date
                        )

                        # ── Pending Revenue Insights (unfiltered, loaded fresh) ──
                        try:
                            _c_start, _c_end, _p_start, _p_end = pending_months()
                            _curr_label_ins = _c_start.strftime("%B %Y")
                            _prev_label_ins = _p_start.strftime("%B %Y")
                            _, _, _df_meta_ins = get_metadata()
                            _meta_map_ins = (_df_meta_ins.sort_values('Month', ascending=False)
                                             .drop_duplicates(subset=['merge_key'], keep='first')
                                             [['merge_key', 'Caller Name', 'Team Name', 'Vertical']].copy())
                            _drop_df_ins  = load_drop_leads()
                            _excl_e = set(_drop_df_ins['drop_email'].dropna().astype(str).unique()) if not _drop_df_ins.empty else set()
                            _excl_p = set(_drop_df_ins['drop_phone'].dropna().astype(str).unique()) if not _drop_df_ins.empty else set()
                            _df_both_ins  = fetch_both_months_rev(_p_start, _c_end)

                            _pending_insight = None
                            _drop_insight    = None

                            if not _df_both_ins.empty:
                                _df_curr_ins = _df_both_ins[_df_both_ins['Date'] >= _c_start].copy()
                                _df_prev_ins = _df_both_ins[(_df_both_ins['Date'] >= _p_start) & (_df_both_ins['Date'] <= _p_end)].copy()
                                _pc = pending_leads_for_month(_df_curr_ins, _excl_e, _excl_p, _df_both_ins)
                                _pp = pending_leads_for_month(_df_prev_ins, _excl_e, _excl_p, _df_both_ins)
                                _cc = build_combined_agg(_pc, _pp, _meta_map_ins, _df_both_ins)

                                if not _cc.empty:
                                    _comb_ins = _cc[_cc['grand_bal'] > 0].copy()
                                    if not _comb_ins.empty:
                                        _top_bal = _comb_ins.sort_values('grand_bal', ascending=False).iloc[0]
                                        _team_bal = (_comb_ins.groupby('Team Name')['grand_bal'].sum()
                                                     .sort_values(ascending=False))
                                        _top_team_bal = _team_bal.index[0] if not _team_bal.empty else '—'
                                        _top_team_bal_amt = _team_bal.iloc[0] if not _team_bal.empty else 0
                                        _pending_body = (
                                            f"{_top_bal['Caller_name']} has the highest combined pending balance of "
                                            f"{fmt_inr(_top_bal['grand_bal'])} across {int(_top_bal['grand_leads'])} lead(s) "
                                            f"({_curr_label_ins} + {_prev_label_ins}). "
                                            f"Top team by pending balance is {_top_team_bal} "
                                            f"with {fmt_inr(_top_team_bal_amt)} outstanding."
                                        )
                                        _pending_insight = {
                                            "type": "warn", "icon": "💰",
                                            "title": (f"Highest Pending Balance — {_top_bal['Caller_name']} "
                                                      f"({_top_bal.get('Team Name','—')})"),
                                            "body": _pending_body
                                        }

                            if not _drop_df_ins.empty and not _df_both_ins.empty:
                                _drop_agg_ins = attribute_drops_to_callers(
                                    _drop_df_ins, _df_both_ins, _meta_map_ins,
                                    _c_start, _c_end, _p_start, _p_end,
                                    _curr_label_ins, _prev_label_ins
                                )
                                if not _drop_agg_ins.empty:
                                    _drop_agg_ins = _drop_agg_ins[_drop_agg_ins['Team Name'] != 'Others'].copy()
                                    if not _drop_agg_ins.empty:
                                        _top_drop = _drop_agg_ins.sort_values('total_drops', ascending=False).iloc[0]
                                        _team_drops = (_drop_agg_ins.groupby('Team Name')['total_drops'].sum()
                                                       .sort_values(ascending=False))
                                        _top_team_drop = _team_drops.index[0] if not _team_drops.empty else '—'
                                        _top_team_drop_cnt = int(_team_drops.iloc[0]) if not _team_drops.empty else 0
                                        _drop_body = (
                                            f"{_top_drop['Caller_name']} has the highest combined drop cases: "
                                            f"{int(_top_drop['curr_drops'])} in {_curr_label_ins} + "
                                            f"{int(_top_drop['prev_drops'])} in {_prev_label_ins} = "
                                            f"{int(_top_drop['total_drops'])} total. "
                                            f"Top team by drops is {_top_team_drop} "
                                            f"with {_top_team_drop_cnt} cases."
                                        )
                                        _drop_insight = {
                                            "type": "bad", "icon": "🚫",
                                            "title": (f"Most Drop Cases — {_top_drop['Caller_name']} "
                                                      f"({_top_drop.get('Team Name','—')})"),
                                            "body": _drop_body
                                        }

                            if _pending_insight:
                                insights.append(_pending_insight)
                            if _drop_insight:
                                insights.append(_drop_insight)
                        except Exception:
                            pass

                        if insights:
                            cols_ins = st.columns(2)
                            for i, ins in enumerate(insights):
                                with cols_ins[i % 2]:
                                    st.markdown(f"""
                                    <div class="insight-card {ins['type']}">
                                        <div style='display:flex;align-items:center;gap:.4rem;'>
                                            <span class="insight-icon">{ins['icon']}</span>
                                            <span class="insight-title">{ins['title']}</span>
                                        </div>
                                        <div class="insight-body">{ins['body']}</div>
                                    </div>""", unsafe_allow_html=True)
                        else:
                            st.info("Not enough variation in data to generate insights.")

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE A — CALLER TEAM LEADERBOARD
                        # ══════════════════════════════════════════════
                        section_header("📞 CALLER REVENUE TEAM PERFORMANCE TABLE")

                        if not calling_ins.empty:
                            lb_calling = (
                                calling_ins.groupby('TEAM')
                                .agg(
                                    Callers   = ('CALLER NAME',       'count'),
                                    Enrollments=('raw_enrollments',   'sum'),
                                    Target    = ('raw_target',        'sum'),
                                    Till_Day  = ('TILL DAY TARGET (₹)','sum'),
                                    Enr_Rev   = ('ENROLLMENT REV',    'sum'),
                                    Bal_Rev   = ('BALANCE REV',       'sum'),
                                    Revenue   = ('raw_calling_rev',   'sum'),
                                )
                                .reset_index()
                                .sort_values('Revenue', ascending=False)
                            )
                            lb_calling['Achievement %'] = lb_calling.apply(
                                lambda r: f"{round(r['Revenue']/r['Target']*100,1)}%" if r['Target'] > 0 else "—", axis=1
                            )
                            lb_calling['Target']   = lb_calling['Target'].apply(fmt_inr)
                            lb_calling['Till_Day'] = lb_calling['Till_Day'].apply(fmt_inr)
                            lb_calling['Enr_Rev']  = lb_calling['Enr_Rev'].apply(fmt_inr)
                            lb_calling['Bal_Rev']  = lb_calling['Bal_Rev'].apply(fmt_inr)
                            lb_calling['Revenue']  = lb_calling['Revenue'].apply(fmt_inr)
                            medals = (["🥇", "🥈", "🥉"] + [""] * len(lb_calling))[:len(lb_calling)]
                            lb_calling.insert(0, "🏅", medals)
                            lb_calling.columns = ['🏅', 'Team', 'Callers', 'Enrollments',
                                                  'Total Target', 'Till Day Target',
                                                  'Enrollment Rev', 'Balance Rev',
                                                  'Calling Revenue', 'Achievement %']
                            st.dataframe(lb_calling.reset_index(drop=True), width='stretch', hide_index=True)
                        else:
                            st.info("No calling agent data available.")

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE B — COLLECTION TEAM LEADERBOARD
                        # ══════════════════════════════════════════════
                        section_header("🏦 COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE")

                        if not collection_ins.empty:
                            lb_coll = (
                                collection_ins.groupby('TEAM')
                                .agg(
                                    Callers     = ('CALLER NAME',          'count'),
                                    Enrollments = ('raw_enrollments',      'sum'),
                                    Target      = ('raw_target',           'sum'),
                                    Till_Day    = ('TILL DAY TARGET (₹)',  'sum'),
                                    Calling_Rev = ('CALLING REVENUE',      'sum'),
                                    Comm_Coll   = ('COMMUNITY COLLECTION', 'sum'),
                                    Boot_Coll   = ('BOOTCAMP COLLECTION',  'sum'),
                                    Revenue     = ('raw_collection_rev',   'sum'),
                                )
                                .reset_index()
                                .sort_values('Revenue', ascending=False)
                            )
                            lb_coll['Achievement %'] = lb_coll.apply(
                                lambda r: f"{round(r['Revenue']/r['Target']*100,1)}%" if r['Target'] > 0 else "—", axis=1
                            )
                            lb_coll['Target']      = lb_coll['Target'].apply(fmt_inr)
                            lb_coll['Till_Day']    = lb_coll['Till_Day'].apply(fmt_inr)
                            lb_coll['Calling_Rev'] = lb_coll['Calling_Rev'].apply(fmt_inr)
                            lb_coll['Comm_Coll']   = lb_coll['Comm_Coll'].apply(fmt_inr)
                            lb_coll['Boot_Coll']   = lb_coll['Boot_Coll'].apply(fmt_inr)
                            lb_coll['Revenue']     = lb_coll['Revenue'].apply(fmt_inr)
                            medals = (["🥇", "🥈", "🥉"] + [""] * len(lb_coll))[:len(lb_coll)]
                            lb_coll.insert(0, "🏅", medals)
                            lb_coll.columns = ['🏅', 'Team', 'Callers', 'Enrollments',
                                               'Total Target', 'Till Day Target',
                                               'Calling Revenue', 'Community Collection',
                                               'Bootcamp Collection', 'Collection Revenue', 'Achievement %']
                            st.dataframe(lb_coll.reset_index(drop=True), width='stretch', hide_index=True)
                        else:
                            st.info("No collection agent data available.")

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE C — BOTH TEAM LEADERBOARD
                        # ══════════════════════════════════════════════
                        section_header("📞🏦 CALLING + COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE")

                        if not both_ins.empty:
                            lb_both = (
                                both_ins.groupby('TEAM')
                                .agg(
                                    Callers     = ('CALLER NAME',          'count'),
                                    Enrollments = ('raw_enrollments',      'sum'),
                                    Target      = ('raw_target',           'sum'),
                                    Till_Day    = ('TILL DAY TARGET (₹)',  'sum'),
                                    Calling_Rev = ('CALLING REVENUE',      'sum'),
                                    Comm_Coll   = ('COMMUNITY COLLECTION', 'sum'),
                                    Boot_Coll   = ('BOOTCAMP COLLECTION',  'sum'),
                                    Total_Rev   = ('raw_revenue',          'sum'),
                                )
                                .reset_index()
                                .sort_values('Total_Rev', ascending=False)
                            )
                            lb_both['Achievement %'] = lb_both.apply(
                                lambda r: f"{round(r['Total_Rev']/r['Target']*100,1)}%" if r['Target'] > 0 else "—", axis=1
                            )
                            lb_both['Target']      = lb_both['Target'].apply(fmt_inr)
                            lb_both['Till_Day']    = lb_both['Till_Day'].apply(fmt_inr)
                            lb_both['Calling_Rev'] = lb_both['Calling_Rev'].apply(fmt_inr)
                            lb_both['Comm_Coll']   = lb_both['Comm_Coll'].apply(fmt_inr)
                            lb_both['Boot_Coll']   = lb_both['Boot_Coll'].apply(fmt_inr)
                            lb_both['Total_Rev']   = lb_both['Total_Rev'].apply(fmt_inr)
                            medals = (["🥇", "🥈", "🥉"] + [""] * len(lb_both))[:len(lb_both)]
                            lb_both.insert(0, "🏅", medals)
                            lb_both.columns = ['🏅', 'Team', 'Callers', 'Enrollments',
                                               'Total Target', 'Till Day Target',
                                               'Calling Revenue', 'Community Collection',
                                               'Bootcamp Collection', 'Total Revenue', 'Achievement %']
                            st.dataframe(lb_both.reset_index(drop=True), width='stretch', hide_index=True)
                        else:
                            st.info("No calling+collection agent data available.")

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🧠</div>
                <div style='font-size:.95rem;font-weight:600;'>Click <b>Generate Revenue Report</b> to load insights</div>
            </div>""", unsafe_allow_html=True)


    with tab3:
        c_start, c_end, p_start, p_end = pending_months()
        curr_label = c_start.strftime("%B %Y")
        prev_label = p_start.strftime("%B %Y")

        gen_pending = st.button("📊 Generate Callerwise Pending", key="rev_pending_btn", width='content')

        if gen_pending:
            st.session_state['pending_revenue_loaded'] = True

        if not st.session_state.get('pending_revenue_loaded', False):
            st.markdown("""
            <div style='text-align:center;padding:4rem 1rem;opacity:.6;'>
                <div style='font-size:3rem;margin-bottom:1rem;'>📊</div>
                <div style='font-size:.9rem;font-weight:600;'>Click <b>📊 Generate Callerwise Pending</b> in the sidebar to load this tab</div>
            </div>""", unsafe_allow_html=True)

        else:
            with st.spinner("Loading pending revenue data…"):
                _, _, df_meta_all = get_metadata()
                meta_map_pending  = (df_meta_all.sort_values('Month', ascending=False)
                                     .drop_duplicates(subset=['merge_key'], keep='first')
                                     [['merge_key', 'Caller Name', 'Team Name', 'Vertical']].copy())
    
                drop_df     = load_drop_leads()
                excl_emails = set(drop_df['drop_email'].dropna().astype(str).unique()) if not drop_df.empty else set()
                excl_phones = set(drop_df['drop_phone'].dropna().astype(str).unique()) if not drop_df.empty else set()
                df_both     = fetch_both_months_rev(p_start, c_end)
    
            if df_both.empty:
                st.warning("No revenue data found.")
            else:
                df_curr   = df_both[df_both['Date'] >= c_start].copy()
                df_prev   = df_both[(df_both['Date'] >= p_start) & (df_both['Date'] <= p_end)].copy()
    
                pend_curr = pending_leads_for_month(df_curr, excl_emails, excl_phones, df_both)
                pend_prev = pending_leads_for_month(df_prev, excl_emails, excl_phones, df_both)
    
                combined  = build_combined_agg(pend_curr, pend_prev, meta_map_pending, df_both)
    
                if not combined.empty:
                    combined = combined[~(
                        (combined['Team Name'] == 'Others') &
                        (combined['Vertical'] == 'Others') &
                        (combined['grand_bal'] == 0)
                    )].copy()
                    combined = combined[combined['grand_bal'] > 0].copy()
    
                    if selected_vertical:
                        combined = combined[combined['Vertical'].isin(selected_vertical)].copy()
                    if selected_team:
                        combined = combined[combined['Team Name'].isin(selected_team)].copy()
    
                if combined.empty:
                    st.info("No pending leads found.")
                else:
                    st.markdown(render_html_pending_table(
                        combined, 'team', curr_label, prev_label,
                        f"TEAMWISE PENDING REVENUE {prev_label.upper()} + {curr_label.upper()}"
                    ), unsafe_allow_html=True)
    
                    st.markdown("<div style='margin:2rem 0;'></div>", unsafe_allow_html=True)
    
                    st.markdown(render_html_pending_table(
                        combined, 'caller', curr_label, prev_label,
                        f"CALLERWISE PENDING REVENUE {prev_label.upper()} + {curr_label.upper()}"
                    ), unsafe_allow_html=True)
    
                    st.markdown("<div style='margin:1rem 0;'></div>", unsafe_allow_html=True)
                    dl_col1, dl_col2 = st.columns(2)
    
                    with dl_col1:
                        _pending_xlsx = build_pending_excel(
                            combined, pend_curr, pend_prev,
                            meta_map_pending, curr_label, prev_label
                        )
                        st.download_button(
                            label="📥 Download Teamwise + Callerwise Pending Revenue",
                            data=_pending_xlsx,
                            file_name=f"Pending_Revenue_{prev_label.replace(' ','_')}_{curr_label.replace(' ','_')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_pending_revenue_xlsx'
                        )
    
                    with dl_col2:
                        _leads_xlsx = build_pending_leads_excel(
                            pend_curr, pend_prev,
                            meta_map_pending, curr_label, prev_label
                        )
                        st.download_button(
                            label=f"📥 Download {prev_label} + {curr_label} Pending Leads",
                            data=_leads_xlsx,
                            file_name=f"Pending_Leads_{prev_label.replace(' ','_')}_{curr_label.replace(' ','_')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_pending_leads_xlsx'
                        )
    
                # ══ DROPPED LEADS — inside else: so drop_df and df_both are always defined ══
                st.markdown("<div style='margin:2rem 0;'></div>", unsafe_allow_html=True)
                section_header(f"🚫 CALLERWISE DROPPED LEADS — {prev_label} + {curr_label}")
    
                if not drop_df.empty and not df_both.empty:
                    drop_agg = attribute_drops_to_callers(
                        drop_df, df_both, meta_map_pending,
                        c_start, c_end, p_start, p_end,
                        curr_label, prev_label
                    )
                    if not drop_agg.empty:
                        drop_agg = drop_agg[drop_agg['Team Name'] != 'Others'].copy()
    
                        if selected_vertical:
                            drop_agg = drop_agg[drop_agg['Vertical'].isin(selected_vertical)].copy()
                        if selected_team:
                            drop_agg = drop_agg[drop_agg['Team Name'].isin(selected_team)].copy()
    
                        st.markdown(
                            render_drop_html(drop_agg, curr_label, prev_label),
                            unsafe_allow_html=True
                        )
                    else:
                        st.info("No dropped leads found for current or previous month.")
    
                    if not drop_agg.empty:
                        _drop_xlsx = build_drop_leads_excel(
                            drop_df, c_start, c_end, p_start, p_end,
                            curr_label, prev_label, meta_map_pending
                        )
                        st.download_button(
                            label=f"📥 Download Dropped Leads — {prev_label} + {curr_label}",
                            data=_drop_xlsx,
                            file_name=f"Dropped_Leads_{prev_label.replace(' ','_')}_{curr_label.replace(' ','_')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_drop_leads_xlsx'
                        )
                else:
                    st.info("Drop leads sheet could not be loaded.")
    
        with tab4:
            # ── Services Input ──────────────────────────────────────────────────────
            gen_update = st.button("📋 Generate Revenue Update", key="rev_update_btn", width='content')
            st.markdown("""
            <div style='background:rgba(16,185,129,.07);border:1px solid rgba(16,185,129,.2);
                        border-radius:10px;padding:.9rem 1.2rem;margin-bottom:.9rem;'>
                <div style='font-size:.82rem;font-weight:700;color:#10B981;margin-bottom:.3rem;'>
                    💼 SERVICES REVENUE INPUT
                </div>
                <div style='font-size:.76rem;color:var(--text-muted,#6B7280);'>
                    Take This Value From Manager's Group Report by Finance and input below
                    before clicking Generate Revenue Update.
                </div>
            </div>
            """, unsafe_allow_html=True)
         
            # Parse & validate Services
            services_raw = st.text_input(
                "Services Revenue (₹)",
                placeholder="e.g. 431211  or  4,31,211  or  4,31,211.50",
                key="ru_services_input",
                help="Numbers, commas and dots only. Commas are stripped; decimals are rounded.",
            )

            # Parse & validate
            _ru_services_val  = 0.0
            _ru_services_ok   = True
            if services_raw and services_raw.strip():
                _cleaned = services_raw.strip().replace(',', '')
                try:
                    _ru_services_val = round(float(_cleaned))
                except ValueError:
                    st.error("⚠️  Invalid input — please enter a plain number (e.g. 431211 or 4,31,211).")
                    _ru_services_ok = False
            st.caption(f"Services value: ₹{int(_ru_services_val):,}")

            if not gen_update:
                st.markdown("""
                <div style='text-align:center;padding:5rem 1rem;opacity:.6;'>
                    <div style='font-size:4rem;margin-bottom:1rem;'>📋</div>
                    <div style='font-size:.9rem;font-weight:600;'>
                        Enter both Services Revenue and Total Revenue Till Today above, then click
                        <b>📋 Generate Revenue Update</b> in the sidebar.
                    </div>
                </div>""", unsafe_allow_html=True)
         
            else:
                if not _ru_services_ok:
                    st.error("Please enter a valid Services revenue value before generating the report.")
                else:
                    with st.spinner("Building Revenue Update…"):
         
                        # ── heading ──────────────────────────────────────────────────
                        if start_date == end_date:
                            _hd_str = start_date.strftime('%d/%m/%Y')
                        else:
                            _hd_str = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
                        _ru_heading = f"REVENUE UPDATE — ({_hd_str})"

                        # ── fetch data ────────────────────────────────────────────────
                        _df_ru  = fetch_revenue_data(start_date, end_date)
                        _df_tru = _load_rev_update_team_sheet()

                    if _df_ru.empty:
                        st.warning("No revenue data found for the selected period.")
                    else:
                        # ── normalise revenue sheet ───────────────────────────────────
                        _dr = _df_ru.copy()

                        # ── Role-based + sidebar filters ──────────────────────────────
                        _ru_role   = st.session_state.get('rf_role', 'admin')
                        _ru_cname  = st.session_state.get('rf_caller_name', '')
                        _ru_teams  = st.session_state.get('rf_teams', [])
                        # Merge team info temporarily for filtering
                        if not _df_tru.empty and 'merge_key' in _df_tru.columns:
                            _tru_slim = _df_tru[['merge_key','Team Name','Vertical']].drop_duplicates('merge_key')
                            _dr = _dr.merge(_tru_slim, on='merge_key', how='left')
                        else:
                            _dr['Team Name'] = ''
                            _dr['Vertical']  = ''
                        if _ru_role == 'caller' and _ru_cname:
                            _dr = _dr[_dr['Caller_name'].str.strip().str.lower() == _ru_cname.strip().lower()]
                        elif _ru_role in ('tl', 'trainer', 'vertical_head') and _ru_teams:
                            _dr = _dr[_dr['Team Name'].isin(_ru_teams)]
                        if selected_team:
                            _dr = _dr[_dr['Team Name'].isin(selected_team)]
                        if selected_vertical:
                            _dr = _dr[_dr['Vertical'].isin(selected_vertical)]
                        if search_query:
                            _dr = _dr[_dr['Caller_name'].str.contains(search_query, case=False, na=False)]

                        if _dr.empty:
                            st.warning("No data found for your access level and selected filters.")
                            st.stop()

                        _enr_l    = _dr['Enrollment'].astype(str).str.strip().str.lower()

                        # ── Apply role-based filters (same as tab1/tab2) ──────────────
                        # Merge team info first so we can filter by Team Name / Vertical
                        _dr = pd.merge(
                            _dr,
                            _df_tru[['merge_key', 'Team Name', 'Vertical']].drop_duplicates('merge_key')
                            if 'merge_key' in _df_tru.columns and not _df_tru.empty
                            else pd.DataFrame(columns=['merge_key', 'Team Name', 'Vertical']),
                            on='merge_key', how='left'
                        )
                        if 'Team Name' not in _dr.columns:
                            _dr['Team Name'] = ''
                        if 'Vertical' not in _dr.columns:
                            _dr['Vertical'] = ''

                        # Role filter: caller sees only their own rows
                        _ru_role = st.session_state.get('rf_role', 'admin')
                        _ru_cname = st.session_state.get('rf_caller_name', '')
                        if _ru_role == 'caller' and _ru_cname:
                            _dr = _dr[_dr['Caller_name'].str.strip().str.lower() == _ru_cname.strip().lower()]
                        elif _ru_role in ('tl', 'trainer', 'vertical_head'):
                            _ru_rf_teams = st.session_state.get('rf_teams', [])
                            if _ru_rf_teams:
                                _dr = _dr[_dr['Team Name'].isin(_ru_rf_teams)]

                        # Sidebar filter overrides (admin uses multiselects)
                        if selected_team:
                            _dr = _dr[_dr['Team Name'].isin(selected_team)]
                        if selected_vertical:
                            _dr = _dr[_dr['Vertical'].isin(selected_vertical)]
                        if search_query:
                            _dr = _dr[_dr['Caller_name'].str.contains(search_query, case=False, na=False)]

                        if _dr.empty:
                            st.warning("No revenue data found for the selected filters.")
                        else:
                            pass

                        _enr_l    = _dr['Enrollment'].astype(str).str.strip().str.lower()
                        _src_l    = _dr['Source'].astype(str).str.lower()
                        _caller_l = _dr['Caller_name'].str.strip().str.lower()
                        _eotm     = (
                            _dr['Enrollment_of_this_month'].astype(str).str.strip().str.lower()
                            if 'Enrollment_of_this_month' in _dr.columns
                            else pd.Series([''] * len(_dr), index=_dr.index)
                        )
                        _ch_l = (
                            _dr['community_head'].astype(str).str.strip().str.lower()
                            if 'community_head' in _dr.columns
                            else pd.Series([''] * len(_dr), index=_dr.index)
                        )
         
                        # ── build lookup from gid=0 team sheet ────────────────────────
                        _dt = _df_tru.copy()
                        # Smart dedup: prefer rows that have a non-empty Team Name for each caller.
                        # This ensures resigned callers whose most-recent sheet row has a blank/different
                        # team (post-resignation update) still get their original team mapping.
                        if 'merge_key' in _dt.columns and 'Team Name' in _dt.columns:
                            _dt['_clean_team'] = _dt['Team Name'].astype(str).str.strip().str.lower()
                            _dt_has_team = _dt[
                                (_dt['_clean_team'] != '') &
                                (_dt['_clean_team'] != 'nan') &
                                (_dt['_clean_team'] != 'none')
                            ].drop_duplicates(subset='merge_key', keep='first')
                            _dt_fallback = _dt[
                                ~_dt['merge_key'].isin(_dt_has_team['merge_key'])
                            ].drop_duplicates(subset='merge_key', keep='first')
                            _dedup_dt = pd.concat([_dt_has_team, _dt_fallback], ignore_index=True)
                            if '_clean_team' in _dedup_dt.columns:
                                _dedup_dt = _dedup_dt.drop(columns=['_clean_team'])
                        else:
                            _dedup_dt = (
                                _dt.drop_duplicates(subset='merge_key', keep='first')
                                if 'merge_key' in _dt.columns else _dt
                            )
         
                        def _mk_map(col):
                            if col in _dedup_dt.columns and 'merge_key' in _dedup_dt.columns:
                                return _dedup_dt.dropna(subset=['merge_key']).set_index('merge_key')[col].to_dict()
                            return {}
         
                        _vert_map  = _mk_map('Vertical')
                        _team_map  = _mk_map('Team Name')
                        _desig_col = next(
                            (c for c in _dedup_dt.columns if 'academic' in c.lower() or 'counselor' in c.lower()),
                            None
                        )
                        _desig_map = _mk_map(_desig_col) if _desig_col else {}
         
                        _dr['_vert'] = _caller_l.map(_vert_map).fillna('').astype(str)
                        _dr['_team'] = _caller_l.map(_team_map).fillna('').astype(str)
         
                        # community manager callers (for Abhipsha extra)
                        _cm_callers = {
                            k for k, v in _desig_map.items()
                            if str(v).strip().lower() == 'community manager'
                        }.union({
                            k for k, v in _team_map.items()
                            if 'community manager' in str(v).strip().lower()
                        })
         
                        # ── tiny helpers ──────────────────────────────────────────────
                        def _sf(m):
                            return float(_dr.loc[m, 'Fee_paid'].sum())
         
                        def _cnew(m):
                            return int((m & (_enr_l == 'new enrollment')).sum())
         
                        def _vc(kw):
                            return _dr['_vert'].str.contains(kw, case=False, na=False)
         
                        def _fnum(v):
                            if v is None:
                                return ""
                            return f"{int(round(v)):,}"
         
                        def _fenr(v):
                            try:
                                return str(int(v)) if int(v) > 0 else ""
                            except Exception:
                                return ""
         
                        # ─── OLD SALES TEAM ────────────────────────────────────────────
                        _uzair_m    = _vc('Uzair')
                        _old_enr_f  = _enr_l.isin(['new enrollment', 'new enrollment - balance payment'])
                        _old_rev    = _sf(_uzair_m & _old_enr_f)
                        _old_enr    = _cnew(_uzair_m)
         
                        _uz_teams = (
                            _dedup_dt[
                                _dedup_dt['Vertical'].astype(str).str.contains('Uzair', case=False, na=False)
                            ]['Team Name'].dropna().unique().tolist()
                            if 'Vertical' in _dedup_dt.columns and 'Team Name' in _dedup_dt.columns
                            else []
                        )
                        _corp_t = sorted([t for t in _uz_teams if 'corporate' in t.lower()])
                        _ncorp_t = sorted([t for t in _uz_teams if 'corporate' not in t.lower()])
         
                        _old_sub = {}
                        if _corp_t:
                            _cm2 = _dr['_team'].isin(_corp_t)
                            _cr  = _sf(_cm2 & _old_enr_f)
                            _ce  = _cnew(_cm2)
                            _cfr = _sf(_cm2 & _src_l.str.contains('funnel', na=False))
                            _old_sub['Corporate Law'] = {
                                'rev': _cr, 'enr': _ce,
                                'funnel_rev': _cfr, 'has_funnel': _cfr > 0,
                            }
                        for _t in _ncorp_t:
                            _tm = _dr['_team'] == _t
                            _old_sub[_t] = {
                                'rev': _sf(_tm & _old_enr_f),
                                'enr': _cnew(_tm),
                                'funnel_rev': 0, 'has_funnel': False,
                            }
                        # filter zero rows, sort by revenue desc
                        _old_sub = dict(
                            sorted(
                                {k: v for k, v in _old_sub.items() if v['rev'] > 0 or v['enr'] > 0}.items(),
                                key=lambda x: x[1]['rev'], reverse=True,
                            )
                        )
         
                        # ─── NEW SALES TEAM ────────────────────────────────────────────
                        _ns_kws   = ['Deepanshi', 'Darshan', 'Shivya', 'Anmol']
                        _ns_m     = pd.Series(False, index=_dr.index)
                        for _kw in _ns_kws:
                            _ns_m |= _vc(_kw)
                        _ns_excl  = _enr_l.isin([
                            'bootcamp collections - balance payments',
                            'community collections - balance payments',
                        ])
                        _ns_rev   = _sf(_ns_m & ~_ns_excl)
                        _ns_enr   = _cnew(_ns_m)
         
                        _ns_teams = (
                            _dedup_dt[
                                _dedup_dt['Vertical'].astype(str).str.contains(
                                    '|'.join(_ns_kws), case=False, na=False
                                )
                            ]['Team Name'].dropna().unique().tolist()
                            if 'Vertical' in _dedup_dt.columns and 'Team Name' in _dedup_dt.columns
                            else []
                        )
                        # US Accounting: ONLY teams under Deepanshi's vertical (excludes Uzair's US teams)
                        if 'Vertical' in _dedup_dt.columns and 'Team Name' in _dedup_dt.columns:
                            _deepanshi_teams = (
                                _dedup_dt[
                                    _dedup_dt['Vertical'].astype(str).str.contains('Deepanshi', case=False, na=False)
                                ]['Team Name'].dropna().unique().tolist()
                            )
                        else:
                            _deepanshi_teams = []
                        _us_t = [
                            t for t in _deepanshi_teams
                            if 'us accounting' in t.lower() or 'us acc' in t.lower()
                        ]

                        # Split ID teams: those whose Vertical contains "Anmol" vs those that don't
                        _all_id_t = [t for t in _ns_teams
                                     if re.match(r'^id[\s\-]', t.strip().lower()) or t.strip().lower() == 'id']
                        if 'Vertical' in _dedup_dt.columns and 'Team Name' in _dedup_dt.columns:
                            _anmol_vert_teams = (
                                _dedup_dt[
                                    _dedup_dt['Vertical'].astype(str).str.contains('Anmol', case=False, na=False)
                                ]['Team Name'].dropna().unique().tolist()
                            )
                        else:
                            _anmol_vert_teams = []
                        _id_anmol_t = [t for t in _all_id_t if t in _anmol_vert_teams]
                        _id_t       = [t for t in _all_id_t if t not in _anmol_vert_teams]

                        _dsv_t = [t for t in _ns_teams if t.strip().lower().startswith('dsv')]
                        _oth_t = [t for t in _ns_teams
                                  if t not in _us_t and t not in _all_id_t and t not in _dsv_t]

                        _ns_sub = {}
                        if _us_t:
                            _um = _dr['_team'].isin(_us_t)
                            _ns_sub['US Accounting'] = {'rev': _sf(_um & ~_ns_excl), 'enr': _cnew(_um)}
                        # ID-Anmol appears before ID
                        if _id_anmol_t:
                            _iam = _dr['_team'].isin(_id_anmol_t)
                            _ns_sub['ID-Anmol'] = {'rev': _sf(_iam & ~_ns_excl), 'enr': _cnew(_iam)}
                        if _id_t:
                            _im = _dr['_team'].isin(_id_t)
                            _ns_sub['ID'] = {'rev': _sf(_im & ~_ns_excl), 'enr': _cnew(_im)}
                        if _dsv_t:
                            _dvm = _dr['_team'].isin(_dsv_t)
                            _ns_sub['DSV'] = {'rev': _sf(_dvm & ~_ns_excl), 'enr': _cnew(_dvm)}
                        for _t in _oth_t:
                            _tm = _dr['_team'] == _t
                            _ns_sub[_t] = {'rev': _sf(_tm & ~_ns_excl), 'enr': _cnew(_tm)}
                        # Keep ID-Anmol and ID always in fixed order (don't sort them away)
                        _fixed_order = ['ID-Anmol', 'ID']
                        _fixed_entries = {k: _ns_sub[k] for k in _fixed_order if k in _ns_sub and (_ns_sub[k]['rev'] > 0 or _ns_sub[k]['enr'] > 0)}
                        _other_entries = dict(
                            sorted(
                                {k: v for k, v in _ns_sub.items()
                                 if k not in _fixed_order and (v['rev'] > 0 or v['enr'] > 0)}.items(),
                                key=lambda x: x[1]['rev'], reverse=True,
                            )
                        )
                        _ns_sub = {**_other_entries, **_fixed_entries}
         
                        # ─── BOOTCAMP BOOKING FEES ─────────────────────────────────────
                        _boot_m   = (_caller_l == 'bootcamp - direct') & (_enr_l == 'new enrollment')
                        _boot_rev = _sf(_boot_m)
                        _boot_enr = int(_boot_m.sum())
         
                        # ─── MAYUR (CHANGEMAKERS) ──────────────────────────────────────
                        _chm          = _dr['_team'].str.contains('Changemakers', case=False, na=False)
                        _mayur_rev    = _sf(_chm & (_enr_l == 'bootcamp collections - balance payments'))
                        _mayur_enr    = _cnew(_chm)
                        _mayur_call_r = _sf(_chm & _enr_l.isin(['new enrollment', 'new enrollment - balance payment']))
                        _mayur_call_e = _cnew(_chm)
                        _mayur_curr   = _sf(_chm & (_enr_l == 'bootcamp collections - balance payments') & (_eotm == 'yes'))
                        _mayur_prev   = _sf(_chm & (_enr_l == 'bootcamp collections - balance payments') & (_eotm == 'no'))
         
                        # ─── ANMOL ─────────────────────────────────────────────────────
                        _anm_m      = _vc('Anmol')
                        _anmol_rev  = _sf(_anm_m & (_enr_l == 'bootcamp collections - balance payments'))
                        _anmol_curr = _sf(_anm_m & (_enr_l == 'bootcamp collections - balance payments') & (_eotm == 'yes'))
                        _anmol_prev = _sf(_anm_m & (_enr_l == 'bootcamp collections - balance payments') & (_eotm == 'no'))
         
                        # ─── DEEPANSHI (Previous balances) ─────────────────────────────
                        _dep_m      = _vc('Deepanshi')
                        _dep_rev    = _sf(_dep_m & (_enr_l == 'bootcamp collections - balance payments'))
         
                        # ─── COLLECTIONS ───────────────────────────────────────────────
                        _coll_rev = _boot_rev + _mayur_rev + _anmol_rev + _dep_rev
         
                        # ─── COMMUNITY (total) ──────────────────────────────────────────
                        _src_comm     = _src_l.str.contains('community', na=False)
                        _comm_all     = _sf(_src_comm)
                        _comm_nd_new  = _sf(
                            _src_comm
                            & ~_caller_l.isin(['direct', 'bootcamp - direct'])
                            & _enr_l.isin(['new enrollment', 'new enrollment - balance payment'])
                        )
                        _cm_extra_rev = _sf(_caller_l.isin(_cm_callers))
                        _comm_rev     = _comm_all - _comm_nd_new + _cm_extra_rev
         
                        # ─── COMMUNITY HEADS ────────────────────────────────────────────
                        _comm_heads = {}
                        for _hkw, _hfull in [
                            ('komal',     'Komal Shah'),
                            ('jayantika', 'Jayantika Ganguly'),
                            ('garima',    'Garima'),
                            ('abhipsha',  'Abhipsha'),
                        ]:
                            _hm      = _ch_l.str.contains(_hkw, case=False, na=False)
                            # Other revenue: source=community + head matches + enrollment=other revenue
                            _oth_r   = _sf(_hm & _src_comm & (_enr_l == 'other revenue'))
                            # Community-Direct: caller=Direct + source=community + head matches + enrollment=new enrollment
                            _dir_m   = (_caller_l == 'direct') & _hm & _src_comm & (_enr_l == 'new enrollment')
                            _dir_r   = _sf(_dir_m)
                            _dir_e   = int(_dir_m.sum())
                            # Collections split by eotm
                            _curr_h  = _sf(_hm & _src_comm & (_enr_l == 'community collections - balance payments') & (_eotm == 'yes'))
                            _prev_h  = _sf(_hm & _src_comm & (_enr_l == 'community collections - balance payments') & (_eotm == 'no'))
                            # Subtract: source=community + head matches + new/balance enrollment + caller != Direct
                            _nd_new  = _sf(
                                _hm & _src_comm
                                & _enr_l.isin(['new enrollment', 'new enrollment - balance payment'])
                                & (_caller_l != 'direct')
                            )
                            # Abhipsha only: add ALL revenue for callers designated Community Manager in team sheet
                            _extra   = _sf(_caller_l.isin(_cm_callers)) if _hkw == 'abhipsha' else 0.0
                            _htotal  = _oth_r + _dir_r + _curr_h + _prev_h + _extra - _nd_new
                            _comm_heads[_hkw] = {
                                'full': _hfull, 'total': _htotal,
                                'dir_r': _dir_r, 'dir_e': _dir_e,
                                'curr': _curr_h, 'prev': _prev_h,
                            }
         
                        # ─── DIRECT-FUNNEL ──────────────────────────────────────────────
                        _dfunnel_rev = _sf(
                            _src_l.str.contains('funnel', na=False)
                            & (_caller_l == 'direct')
                            & _enr_l.isin(['new enrollment', 'new enrollment - balance payment', 'other revenue'])
                        )
         
                        # ─── DIRECT ────────────────────────────────────────────────────
                        _direct_rev = _sf(
                            ~_src_l.str.contains('funnel', na=False)
                            & ~_src_l.str.contains('community', na=False)
                            & (_caller_l == 'direct')
                            & _enr_l.isin(['new enrollment', 'new enrollment - balance payment', 'other revenue'])
                        )
         
                        # ─── LEAD DETAILS NOT AVAILABLE ─────────────────────────────────
                        _dna_m = (
                            _dr['Caller_name'].astype(str).str.strip().isin(['', 'nan', 'None', 'NaN'])
                            | _dr['Enrollment'].astype(str).str.strip().isin(['', 'nan', 'None', 'NaN'])
                            | _dr['Source'].astype(str).str.strip().isin(['', 'nan', 'None', 'NaN'])
                        )
                        _dna_rev = _sf(_dna_m)
         
                        # ─── TOTAL REVENUE ───────────────────────────────────────────────
                        _total_rev = (
                            _old_rev + _ns_rev
                            + _coll_rev + _comm_rev
                            + _dfunnel_rev + _direct_rev
                            + _ru_services_val + _dna_rev
                        )
                        _total_enr = _old_enr + _ns_enr   # matches image (121+84=205)
         
                        # ── Render HTML table ─────────────────────────────────────────
                        _TITLE  = "#1c3a5f"
                        _COLHDR = "#1e4d6b"
                        _BOLD   = "#dbeafe"
                        _SUB    = "#ffffff"
                        _SUB2   = "#f8faff"
                        _TOTAL  = "#1c3a5f"
                        _TBOLD  = "#0d2137"
        
                        def _hs(extra=""):
                            return (
                                f"background:{_COLHDR};color:#fff;font-size:.72rem;"
                                f"font-weight:700;padding:5px 10px;border:1px solid #163d5a;{extra}"
                            )
        
                        def _bs(extra=""):
                            return (
                                f"background:{_BOLD};color:{_TBOLD};font-weight:700;"
                                f"font-size:.75rem;padding:4px 8px;border:1px solid #bfdbfe;{extra}"
                            )
        
                        def _ss(extra=""):
                            return (
                                f"background:{_SUB};color:#1a1a2e;font-size:.73rem;"
                                f"padding:3px 8px 3px 22px;border:1px solid #e8f0fe;{extra}"
                            )
        
                        def _ss2(extra=""):
                            return (
                                f"background:{_SUB2};color:#1a1a2e;font-size:.71rem;"
                                f"padding:3px 8px 3px 36px;border:1px solid #e8f0fe;{extra}"
                            )
        
                        def _ts(extra=""):
                            return (
                                f"background:{_TOTAL};color:#fff;font-weight:700;"
                                f"font-size:.76rem;padding:5px 8px;border:1px solid #0d2137;{extra}"
                            )
        
                        _NR = "text-align:right;font-family:'DM Mono',monospace;min-width:90px;"
        
                        def _tr(lbl_css, num_css, label, rev=None, enr=None):
                            rv = _fnum(rev) if rev is not None else ""
                            en = _fenr(enr) if enr is not None else ""
                            return (
                                f"<tr>"
                                f"<td style='{lbl_css}'>{label}</td>"
                                f"<td style='{num_css}{_NR}'>{rv}</td>"
                                f"<td style='{num_css}{_NR}'>{en}</td>"
                                f"</tr>"
                            )
        
                        def _show(rev, enr=0):
                            return (rev or 0) != 0 or (enr or 0) != 0

                        _html = f"""
    <div style='overflow-x:auto;margin:.5rem auto;max-width:720px;'>
    <table style='width:100%;border-collapse:collapse;font-family:"DM Sans",sans-serif;'>
    <thead>
    <tr>
      <th colspan='3' style='background:{_TITLE};color:#fff;text-align:center;
          padding:8px 12px;font-size:.82rem;font-weight:700;letter-spacing:.4px;
          border:2px solid {_TITLE};'>{_ru_heading}</th>
    </tr>
    <tr>
      <th style='{_hs("text-align:left;width:58%;")}'>&nbsp;</th>
      <th style='{_hs("text-align:right;width:26%;")}'> Revenue Split</th>
      <th style='{_hs("text-align:right;width:16%;")}'> Enrollments</th>
    </tr>
    </thead>
    <tbody>
    """

                        # Old Sales
                        _html += _tr(_bs(), _bs(), "Old Sales team", _old_rev, _old_enr)
                        for _tn, _td in _old_sub.items():
                            if not _show(_td['rev'], _td['enr']): continue
                            _lbl = _tn
                            if _td['has_funnel']:
                                _lbl = f"{_tn} ({_fnum(_td['funnel_rev'])} - Funnel included)"
                            _html += _tr(_ss(), _ss(), _lbl, _td['rev'], _td['enr'])

                        # New Sales
                        _html += _tr(_bs(), _bs(), "New Sales team", _ns_rev, _ns_enr)
                        for _tn, _td in _ns_sub.items():
                            if not _show(_td['rev'], _td['enr']): continue
                            _html += _tr(_ss(), _ss(), _tn, _td['rev'], _td['enr'])

                        # Bootcamp
                        if _show(_boot_rev, _boot_enr):
                            _html += _tr(_bs(), _bs(), "Bootcamp booking fees", _boot_rev, _boot_enr)

                        # Mayur
                        _html += _tr(_bs(), _bs(), "Mayur", _mayur_rev)
                        if _show(_mayur_call_r, _mayur_call_e):
                            _html += _tr(_ss(), _ss(), "Calling Revenue", _mayur_call_r, _mayur_call_e)
                        if _show(_mayur_curr):
                            _html += _tr(_ss(), _ss(), "Current month collections", _mayur_curr)
                        if _show(_mayur_prev):
                            _html += _tr(_ss(), _ss(), "Previous month collections", _mayur_prev)

                        # Anmol
                        _html += _tr(_bs(), _bs(), "Anmol", _anmol_rev)
                        if _show(_anmol_curr):
                            _html += _tr(_ss(), _ss(), "Current month collections", _anmol_curr)
                        if _show(_anmol_prev):
                            _html += _tr(_ss(), _ss(), "Previous month collections", _anmol_prev)

                        # Deepanshi
                        if _show(_dep_rev):
                            _html += _tr(_bs(), _bs(), "Deepanshi (Previous balances)", _dep_rev)

                        _html += _tr(_bs(), _bs(), "Collections", _coll_rev)

                        # Community
                        _html += _tr(_bs(), _bs(), "Community", _comm_rev)
                        for _hk, _hd in _comm_heads.items():
                            if not _show(_hd['total'], _hd['dir_e']): continue
                            _html += _tr(_ss(), _ss(), _hd['full'], _hd['total'])
                            if _show(_hd['dir_r'], _hd['dir_e']):
                                _html += _tr(_ss2(), _ss2(), f"{_hd['full']} Community-Direct", _hd['dir_r'], _hd['dir_e'])
                            if _show(_hd['curr']):
                                _html += _tr(_ss2(), _ss2(), "Current month collections", _hd['curr'])
                            if _show(_hd['prev']):
                                _html += _tr(_ss2(), _ss2(), "Previous month collections", _hd['prev'])

                        # Direct-Funnel & Direct
                        if _show(_dfunnel_rev):
                            _html += _tr(_bs(), _bs(), "Direct-Funnel", _dfunnel_rev)
                        if _show(_direct_rev):
                            _html += _tr(_bs(), _bs(), "Direct", _direct_rev)

                        # Services
                        _html += _tr(_bs(), _bs(), "Services", _ru_services_val)

                        # DNA
                        if _show(_dna_rev):
                            _html += _tr(_bs(), _bs(), "Lead Details not Available", _dna_rev)

                        # Total
                        _html += _tr(_ts(), _ts(), "Total Revenue", _total_rev, _total_enr)
         
                        _html += "</tbody></table></div>"
         
                        section_header(f"📋 {_ru_heading}")
                        st.markdown(_html, unsafe_allow_html=True)
         
                        # ── CSV download ──────────────────────────────────────────────
                        _rows_dl = []
                        _rows_dl.append({"Category": "Old Sales team", "Sub-category": "",
                                          "Revenue Split": int(round(_old_rev)), "Enrollments": _old_enr or ""})
                        for _tn, _td in _old_sub.items():
                            _lbl = f"{_tn} ({_fnum(_td['funnel_rev'])} - Funnel included)" if _td['has_funnel'] else _tn
                            _rows_dl.append({"Category": "", "Sub-category": _lbl,
                                             "Revenue Split": int(round(_td['rev'])), "Enrollments": _td['enr'] or ""})
                        _rows_dl.append({"Category": "New Sales team", "Sub-category": "",
                                          "Revenue Split": int(round(_ns_rev)), "Enrollments": _ns_enr or ""})
                        for _tn, _td in _ns_sub.items():
                            _rows_dl.append({"Category": "", "Sub-category": _tn,
                                             "Revenue Split": int(round(_td['rev'])), "Enrollments": _td['enr'] or ""})
                        _rows_dl.append({"Category": "Bootcamp booking fees", "Sub-category": "",
                                          "Revenue Split": int(round(_boot_rev)), "Enrollments": _boot_enr or ""})
                        _rows_dl.append({"Category": "Mayur", "Sub-category": "",
                                          "Revenue Split": int(round(_mayur_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "", "Sub-category": "Calling Revenue",
                                          "Revenue Split": int(round(_mayur_call_r)), "Enrollments": _mayur_call_e or ""})
                        _rows_dl.append({"Category": "", "Sub-category": "Current month collections",
                                          "Revenue Split": int(round(_mayur_curr)), "Enrollments": ""})
                        _rows_dl.append({"Category": "", "Sub-category": "Previous month collections",
                                          "Revenue Split": int(round(_mayur_prev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Anmol", "Sub-category": "",
                                          "Revenue Split": int(round(_anmol_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "", "Sub-category": "Current month collections",
                                          "Revenue Split": int(round(_anmol_curr)), "Enrollments": ""})
                        _rows_dl.append({"Category": "", "Sub-category": "Previous month collections",
                                          "Revenue Split": int(round(_anmol_prev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Deepanshi (Previous balances)", "Sub-category": "",
                                          "Revenue Split": int(round(_dep_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Collections", "Sub-category": "",
                                          "Revenue Split": int(round(_coll_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Community", "Sub-category": "",
                                          "Revenue Split": int(round(_comm_rev)), "Enrollments": ""})
                        for _hk, _hd in _comm_heads.items():
                            _rows_dl.append({"Category": "", "Sub-category": _hd['full'],
                                             "Revenue Split": int(round(_hd['total'])), "Enrollments": ""})
                            _rows_dl.append({"Category": "", "Sub-category": f"  {_hd['full']} Community-Direct",
                                             "Revenue Split": int(round(_hd['dir_r'])), "Enrollments": _hd['dir_e'] or ""})
                            _rows_dl.append({"Category": "", "Sub-category": "  Current month collections",
                                             "Revenue Split": int(round(_hd['curr'])), "Enrollments": ""})
                            _rows_dl.append({"Category": "", "Sub-category": "  Previous month collections",
                                             "Revenue Split": int(round(_hd['prev'])), "Enrollments": ""})
                        _rows_dl.append({"Category": "Direct-Funnel", "Sub-category": "",
                                          "Revenue Split": int(round(_dfunnel_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Direct", "Sub-category": "",
                                          "Revenue Split": int(round(_direct_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Services", "Sub-category": "",
                                          "Revenue Split": int(round(_ru_services_val)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Lead Details not Available", "Sub-category": "",
                                          "Revenue Split": int(round(_dna_rev)), "Enrollments": ""})
                        _rows_dl.append({"Category": "Total Revenue", "Sub-category": "",
                                          "Revenue Split": int(round(_total_rev)), "Enrollments": _total_enr or ""})
         
                        def _build_ru_excel(rows, heading):
                            TITLE_FILL = PatternFill("solid", start_color="1c3a5f", end_color="1c3a5f")
                            HDR_FILL   = PatternFill("solid", start_color="1e4d6b", end_color="1e4d6b")
                            SEC_FILL   = PatternFill("solid", start_color="dbeafe", end_color="dbeafe")
                            SUB_FILL   = PatternFill("solid", start_color="ffffff", end_color="ffffff")
                            SUB2_FILL  = PatternFill("solid", start_color="f8faff", end_color="f8faff")
                            TOT_FILL   = PatternFill("solid", start_color="1c3a5f", end_color="1c3a5f")
                            W_FONT     = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                            S_FONT     = Font(bold=True, color="0d2137", name="Arial", size=9)
                            D_FONT     = Font(name="Arial", size=9, color="1a1a2e")
                            CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            LEFT       = Alignment(horizontal='left',   vertical='center')
                            RIGHT      = Alignment(horizontal='right',  vertical='center')
                            BDR        = Border(
                                left=Side(style='thin', color='bfdbfe'),
                                right=Side(style='thin', color='bfdbfe'),
                                top=Side(style='thin', color='bfdbfe'),
                                bottom=Side(style='thin', color='bfdbfe'),
                            )
                            wb = Workbook(); ws = wb.active; ws.title = "Revenue Update"
                            ws.merge_cells("A1:C1")
                            tc = ws["A1"]
                            tc.value = heading
                            tc.fill  = TITLE_FILL
                            tc.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                            tc.alignment = CENTER; tc.border = BDR
                            ws.row_dimensions[1].height = 26
                            for ci, h in enumerate(["CATEGORY / SUB-CATEGORY", "REVENUE SPLIT (₹)", "ENROLLMENTS"], 1):
                                c = ws.cell(2, ci, h)
                                c.fill = HDR_FILL; c.font = W_FONT; c.border = BDR
                                c.alignment = LEFT if ci == 1 else CENTER
                            ws.row_dimensions[2].height = 20
                            for ri, row in enumerate(rows, 3):
                                cat  = str(row.get("Category", "")).strip()
                                sub  = str(row.get("Sub-category", "")).strip()
                                rev  = row.get("Revenue Split", "")
                                enr  = row.get("Enrollments", "")
                                is_total   = cat == "Total Revenue"
                                is_section = cat != "" and sub == ""
                                is_sub2    = sub.startswith("  ")
                                label      = cat if (is_section or is_total) else sub.strip()
                                indent     = "" if (is_section or is_total) else ("      " if is_sub2 else "   ")
                                if is_total:
                                    fill, lfont, nfont = TOT_FILL, W_FONT, W_FONT
                                elif is_section:
                                    fill, lfont, nfont = SEC_FILL, S_FONT, S_FONT
                                elif is_sub2:
                                    fill, lfont, nfont = SUB2_FILL, D_FONT, D_FONT
                                else:
                                    fill, lfont, nfont = SUB_FILL, D_FONT, D_FONT
                                lc = ws.cell(ri, 1, indent + label)
                                lc.fill = fill; lc.font = lfont; lc.alignment = LEFT; lc.border = BDR
                                try:
                                    rev_v = int(str(rev).replace(",","")) if str(rev).strip() else ""
                                except Exception:
                                    rev_v = str(rev)
                                rc = ws.cell(ri, 2, rev_v)
                                rc.fill = fill; rc.font = nfont; rc.alignment = RIGHT; rc.border = BDR
                                if isinstance(rev_v, int): rc.number_format = '#,##0'
                                try:
                                    enr_v = int(str(enr)) if str(enr).strip() not in ("", "0") else ""
                                except Exception:
                                    enr_v = ""
                                ec = ws.cell(ri, 3, enr_v)
                                ec.fill = fill; ec.font = nfont; ec.alignment = RIGHT; ec.border = BDR
                            ws.column_dimensions["A"].width = 44
                            ws.column_dimensions["B"].width = 22
                            ws.column_dimensions["C"].width = 14
                            ws.freeze_panes = "A3"
                            buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

                        # ── Raw source data download ──────────────────────────────
                        # ── Segment classification ────────────────────────────────
                        _seg_team = pd.Series('', index=_dr.index, dtype=str)
                        _seg_sub  = pd.Series('', index=_dr.index, dtype=str)

                        # Bootcamp booking fees
                        _seg_team[_boot_m] = 'Bootcamp booking fees'

                        # Mayur — bootcamp collections + calling
                        _m_bc = _chm & (_enr_l == 'bootcamp collections - balance payments')
                        _seg_team[_m_bc] = 'Mayur'
                        _seg_sub[_m_bc & (_eotm == 'yes')] = 'Current month collections'
                        _seg_sub[_m_bc & (_eotm == 'no')]  = 'Previous month collections'
                        _m_call = _chm & _enr_l.isin(['new enrollment', 'new enrollment - balance payment'])
                        _seg_team[_m_call] = 'Mayur'
                        _seg_sub[_m_call]  = 'Calling Revenue'

                        # Anmol — bootcamp collections
                        _a_bc = _anm_m & (_enr_l == 'bootcamp collections - balance payments') & ~_chm
                        _seg_team[_a_bc] = 'Anmol'
                        _seg_sub[_a_bc & (_eotm == 'yes')] = 'Current month collections'
                        _seg_sub[_a_bc & (_eotm == 'no')]  = 'Previous month collections'

                        # Deepanshi — bootcamp collections
                        _d_bc = _dep_m & (_enr_l == 'bootcamp collections - balance payments') & ~_chm
                        _seg_team[_d_bc] = 'Deepanshi (Previous balances)'

                        # Old Sales team
                        _old_seg = _uzair_m & _old_enr_f
                        _seg_team[_old_seg] = 'Old Sales team'
                        for _st in _corp_t:
                            _seg_sub[_old_seg & (_dr['_team'] == _st)] = 'Corporate Law'
                        for _st in _ncorp_t:
                            _seg_sub[_old_seg & (_dr['_team'] == _st)] = _st

                        # New Sales team
                        _ns_seg = _ns_m & ~_ns_excl & ~_chm
                        _seg_team[_ns_seg] = 'New Sales team'
                        if _us_t:
                            _seg_sub[_ns_seg & _dr['_team'].isin(_us_t)] = 'US Accounting'
                        if _id_anmol_t:
                            _seg_sub[_ns_seg & _dr['_team'].isin(_id_anmol_t)] = 'ID-Anmol'
                        if _id_t:
                            _seg_sub[_ns_seg & _dr['_team'].isin(_id_t)] = 'ID'
                        if _dsv_t:
                            _seg_sub[_ns_seg & _dr['_team'].isin(_dsv_t)] = 'DSV'
                        for _st in _oth_t:
                            _seg_sub[_ns_seg & (_dr['_team'] == _st)] = _st

                        # Community Manager callers
                        _cm_seg = _caller_l.isin(_cm_callers)
                        _seg_team[_cm_seg] = 'Community'
                        _seg_sub[_cm_seg]  = 'Abhipsha — Community Managers'

                        # Community source rows (excluding nd_new already in NS/OS)
                        _src_c_seg = _src_l.str.contains('community', na=False)
                        _nd_excl   = _src_c_seg & ~_caller_l.isin(['direct', 'bootcamp - direct']) & _enr_l.isin(['new enrollment', 'new enrollment - balance payment'])
                        _comm_seg  = _src_c_seg & ~_nd_excl & ~_cm_seg
                        _seg_team[_comm_seg] = 'Community'
                        for _hkw2, _hfull2 in [('komal','Komal Shah'),('jayantika','Jayantika Ganguly'),('garima','Garima'),('abhipsha','Abhipsha')]:
                            _hm2 = _ch_l.str.contains(_hkw2, case=False, na=False)
                            _seg_sub[_comm_seg & _hm2 & (_enr_l == 'other revenue')] = _hfull2
                            _seg_sub[_comm_seg & _hm2 & (_caller_l == 'direct') & (_enr_l == 'new enrollment')] = f'{_hfull2} — Community-Direct'
                            _seg_sub[_comm_seg & _hm2 & (_enr_l == 'community collections - balance payments') & (_eotm == 'yes')] = f'{_hfull2} — Current month collections'
                            _seg_sub[_comm_seg & _hm2 & (_enr_l == 'community collections - balance payments') & (_eotm == 'no')] = f'{_hfull2} — Previous month collections'

                        # Direct-Funnel
                        _df_seg = _src_l.str.contains('funnel', na=False) & (_caller_l == 'direct') & _enr_l.isin(['new enrollment', 'new enrollment - balance payment', 'other revenue'])
                        _seg_team[_df_seg] = 'Direct-Funnel'

                        # Direct
                        _dir_seg = ~_src_l.str.contains('funnel', na=False) & ~_src_l.str.contains('community', na=False) & (_caller_l == 'direct') & _enr_l.isin(['new enrollment', 'new enrollment - balance payment', 'other revenue'])
                        _seg_team[_dir_seg] = 'Direct'

                        # DNA / Lead Details not Available
                        _seg_team[_dna_m] = 'Lead Details not Available'

                        _dr['Considered under Team']     = _seg_team.values
                        _dr['Considered under Sub-Team'] = _seg_sub.values

                        # ── Raw source data download ──────────────────────────────
                        _EXPORT_COLS = [
                            'Date', 'Name', 'Contact_No', 'Email_Id', 'Course',
                            'Fee_paid', 'Caller_name', 'Enrollment', 'Source',
                            'Course_Price', 'LawSikho_Skill_Arbitrage',
                            'Full_Installment', 'Enrollment_of_this_month',
                            '_vert', '_team',
                            'Considered under Team', 'Considered under Sub-Team',
                        ]
                        _dr_export = _dr.copy()

                        # Force Contact_No to numeric (strip non-digits, convert to int where possible)
                        if 'Contact_No' in _dr_export.columns:
                            _dr_export['Contact_No'] = (
                                _dr_export['Contact_No']
                                .astype(str)
                                .str.replace(r'\D', '', regex=True)
                                .str.strip()
                            )
                            _dr_export['Contact_No'] = pd.to_numeric(
                                _dr_export['Contact_No'], errors='coerce'
                            ).astype('Int64')

                        _export_existing = [c for c in _EXPORT_COLS if c in _dr_export.columns]
                        _dr_export = _dr_export[_export_existing].reset_index(drop=True)

                        def _build_ru_raw_excel(df, heading):
                            HDR_FILL  = PatternFill("solid", start_color="1c3a5f", end_color="1c3a5f")
                            ALT_FILL  = PatternFill("solid", start_color="f0f7ff", end_color="f0f7ff")
                            WHT_FILL  = PatternFill("solid", start_color="ffffff", end_color="ffffff")
                            HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                            DATA_FONT = Font(name="Arial", size=9)
                            CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            LEFT      = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            BORDER    = Border(
                                left=Side(style='thin', color='bfdbfe'),
                                right=Side(style='thin', color='bfdbfe'),
                                top=Side(style='thin', color='bfdbfe'),
                                bottom=Side(style='thin', color='bfdbfe'),
                            )
                            COL_WIDTHS = {
                                'Date': 12, 'Name': 28, 'Contact_No': 16,
                                'Email_Id': 32, 'Course': 28, 'Fee_paid': 12,
                                'Caller_name': 22, 'Enrollment': 28, 'Source': 20,
                                'Course_Price': 14, 'LawSikho_Skill_Arbitrage': 18,
                                'Full_Installment': 16, 'Enrollment_of_this_month': 20,
                                '_vert': 18, '_team': 22,
                                'Considered under Team': 26, 'Considered under Sub-Team': 32,
                            }
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Revenue Source Data"

                            # Title row
                            ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
                            tc = ws["A1"]
                            tc.value = heading
                            tc.fill = HDR_FILL
                            tc.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                            tc.alignment = CENTER
                            tc.border = BORDER
                            ws.row_dimensions[1].height = 24

                            # Header row
                            for ci, col in enumerate(df.columns, 1):
                                c = ws.cell(2, ci, col.upper().replace('_', ' '))
                                c.fill = PatternFill("solid", start_color="1e4d6b", end_color="1e4d6b")
                                c.font = HDR_FONT
                                c.alignment = CENTER
                                c.border = BORDER
                            ws.row_dimensions[2].height = 20

                            # Data rows
                            for ri, (_, row) in enumerate(df.iterrows(), 3):
                                fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
                                for ci, col in enumerate(df.columns, 1):
                                    val = row[col]
                                    # Handle NA/NaT
                                    try:
                                        if pd.isna(val):
                                            val = ''
                                    except (TypeError, ValueError):
                                        pass
                                    if hasattr(val, 'strftime'):
                                        val = val.strftime('%Y-%m-%d')
                                    # Keep Contact_No as integer in Excel
                                    if col == 'Contact_No' and val != '':
                                        try:
                                            val = int(val)
                                        except (ValueError, TypeError):
                                            pass
                                    cell = ws.cell(ri, ci, val)
                                    cell.fill = fill
                                    cell.font = DATA_FONT
                                    cell.border = BORDER
                                    cell.alignment = LEFT if col in ('Name', 'Course', 'Email_Id', 'Caller_name', 'Enrollment', 'Source', '_vert', '_team', 'Considered under Team', 'Considered under Sub-Team') else CENTER

                            # Column widths
                            for ci, col in enumerate(df.columns, 1):
                                ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 16)

                            ws.freeze_panes = "A3"
                            buf = io.BytesIO()
                            wb.save(buf)
                            return buf.getvalue()

                        _total_fetched = float(_dr['Fee_paid'].sum()) + _ru_services_val
                        _diff          = _total_fetched - _total_rev
                        _diff_color    = "#10B981" if _diff >= 0 else "#F87171"
                        _diff_sign     = "+" if _diff >= 0 else ""
                        _diff_html     = f"""
                        <div style='text-align:center;padding:.55rem .8rem;
                                    background:rgba(255,255,255,.04);
                                    border:1px solid rgba(255,255,255,.10);
                                    border-radius:10px;line-height:1.6;'>
                            <span style='font-size:.72rem;color:rgba(255,255,255,.5);'>
                                Difference in Revenue&nbsp;
                            </span>
                            <span style='font-size:.88rem;font-weight:700;
                                         color:{_diff_color};font-family:DM Mono,monospace;'>
                                {_diff_sign}₹{int(abs(round(_diff))):,}
                            </span><br>
                            <span style='font-size:.65rem;color:rgba(255,255,255,.3);
                                         font-family:DM Mono,monospace;'>
                                Today: ₹{int(round(_total_fetched)):,}
                                &nbsp;−&nbsp;
                                Report: ₹{int(round(_total_rev)):,}
                            </span>
                        </div>"""

                        _dl_col1, _diff_col, _dl_col2 = st.columns([1, 1.4, 1])
                        with _dl_col1:
                            st.download_button(
                                label="📥 Download Revenue Update",
                                data=_build_ru_excel(_rows_dl, _ru_heading),
                                file_name=f"Revenue_Update_{display_start}_to_{display_end}.xlsx",
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                key='dl_rev_update_xlsx',
                                width='stretch',
                            )
                        with _diff_col:
                            st.markdown(_diff_html, unsafe_allow_html=True)
                        with _dl_col2:
                            st.download_button(
                                label="📥 Download Revenue Sheet",
                                data=_build_ru_raw_excel(_dr_export, _ru_heading),
                                file_name=f"Revenue_Source_{display_start}_to_{display_end}.xlsx",
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                key='dl_rev_source_xlsx',
                                width='stretch',
                            )

    with tab5:
        gen_tva = st.button("🎯 Generate Target Vs Achievement", key="rev_tva_btn", width='content')

        st.markdown("""
        <div style='background:rgba(16,185,129,.07);border:1px solid rgba(16,185,129,.2);
                    border-radius:10px;padding:.9rem 1.2rem;margin-bottom:.9rem;'>
            <div style='font-size:.82rem;font-weight:700;color:#10B981;margin-bottom:.3rem;'>
                📅 WORKING DAYS INPUT
            </div>
            <div style='font-size:.76rem;color:var(--text-muted,#6B7280);'>
                Please enter the number of working days of this month.
            </div>
        </div>
        """, unsafe_allow_html=True)

        _tva_wd_raw = st.text_input(
            "Number of Working Days",
            placeholder="e.g. 20",
            key="tva_working_days_input",
            help="Enter a number between 0 and 20.",
        )

        _tva_wd_val = None
        _tva_wd_ok  = True
        if _tva_wd_raw and _tva_wd_raw.strip():
            try:
                _tva_wd_val = int(_tva_wd_raw.strip())
                if not (0 <= _tva_wd_val <= 20):
                    st.error("⚠️ Please enter a valid number of working days (0–20).")
                    _tva_wd_ok = False
            except ValueError:
                st.error("⚠️ Please enter a valid number of working days (0–20).")
                _tva_wd_ok = False

        if _tva_wd_val is not None and _tva_wd_ok:
            st.caption(f"Working days: {_tva_wd_val}")

        if not gen_tva:
            st.markdown("""
            <div style='text-align:center;padding:5rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🎯</div>
                <div style='font-size:.9rem;font-weight:600;'>
                    Enter the number of working days and click
                    <b>🎯 Generate Target Vs Achievement</b>.
                </div>
            </div>""", unsafe_allow_html=True)

        else:
            if start_date.year != end_date.year or start_date.month != end_date.month:
                st.error("⚠️ Please input valid dates and Generate Report Again. The date range must be within a single month.")
            elif _tva_wd_val is None:
                st.error("⚠️ Please enter the number of working days before generating.")
            elif not _tva_wd_ok:
                pass
            else:
                with st.spinner("Building Target Vs Achievement…"):

                    # ── Load teamsheet for the specific month ──────────────────
                    _, _, _df_meta_tva = get_metadata()
                    _tva_month_start = date(start_date.year, start_date.month, 1)
                    _month_col_tva  = next((c for c in _df_meta_tva.columns if c.strip().lower() == 'month'), None)
                    _status_col_tva = next((c for c in _df_meta_tva.columns if c.strip().lower() == 'status'), None)

                    def _tva_ym(val):
                        try:
                            if pd.isna(val): return False
                            d = pd.to_datetime(val, dayfirst=True, errors='coerce')
                            if pd.isna(d): return False
                            return d.year == _tva_month_start.year and d.month == _tva_month_start.month
                        except: return False

                    _tva_sheet = (
                        _df_meta_tva[_df_meta_tva[_month_col_tva].apply(_tva_ym)].copy()
                        if _month_col_tva else _df_meta_tva.copy()
                    )
                    if _status_col_tva:
                        _tva_sheet = _tva_sheet[
                            _tva_sheet[_status_col_tva].astype(str).str.strip().str.lower() == 'active'
                        ].copy()

                    if _tva_sheet.empty:
                        st.warning("No active callers found for the selected month in the team sheet.")
                        st.stop()

                    # Target column
                    _tgt_col_tva = None
                    for _tc in ['Target','target','TARGET','Monthly Target','Monthly target','Sales Target']:
                        if _tc in _tva_sheet.columns:
                            _tgt_col_tva = _tc; break

                    # ── Revenue data ───────────────────────────────────────────
                    _tva_rev_full = fetch_revenue_data(start_date, end_date)
                    _yesterday_tva = end_date - timedelta(days=1)
                    _tva_rev_yst  = fetch_revenue_data(_yesterday_tva, _yesterday_tva)

                    # ── Pending revenue per caller ─────────────────────────────
                    _pending_map_tva = {}
                    try:
                        _c2, _ce2, _p2, _pe2 = pending_months()
                        _meta_tva2 = (
                            _tva_sheet[['merge_key','Caller Name','Team Name','Vertical']]
                            .drop_duplicates('merge_key').copy()
                            if 'merge_key' in _tva_sheet.columns
                            else pd.DataFrame(columns=['merge_key','Caller Name','Team Name','Vertical'])
                        )
                        _db2 = fetch_both_months_rev(_p2, _ce2)
                        _dd2 = load_drop_leads()
                        _ee2 = set(_dd2['drop_email'].dropna().astype(str).unique()) if not _dd2.empty else set()
                        _ep2 = set(_dd2['drop_phone'].dropna().astype(str).unique()) if not _dd2.empty else set()
                        if not _db2.empty:
                            _dfc2 = _db2[_db2['Date'] >= _c2].copy()
                            _dfp2 = _db2[(_db2['Date'] >= _p2) & (_db2['Date'] <= _pe2)].copy()
                            _pc2  = pending_leads_for_month(_dfc2, _ee2, _ep2, _db2)
                            _pp2  = pending_leads_for_month(_dfp2, _ee2, _ep2, _db2)
                            _cb2  = build_combined_agg(_pc2, _pp2, _meta_tva2, _db2)
                            if not _cb2.empty:
                                for _, _pr in _cb2.iterrows():
                                    _pending_map_tva[str(_pr.get('Caller_name','')).strip().lower()] = float(_pr.get('grand_bal', 0) or 0)
                    except Exception:
                        pass

                    # ── Exclusions & role filters ──────────────────────────────
                    _EXCL_TVA = {'direct', 'bootcamp - direct', 'bootcamp direct', 'lead details not available'}
                    _ru_role_tva  = st.session_state.get('rf_role', 'admin')
                    _ru_teams_tva = st.session_state.get('rf_teams', [])
                    _ru_cname_tva = st.session_state.get('rf_caller_name', '')

                    _all_tva_teams = sorted([
                        t for t in _tva_sheet['Team Name'].dropna().unique()
                        if t.strip().lower() != 'changemakers'
                        and 'community manager' not in t.strip().lower()
                    ])
                    if _ru_role_tva in ('tl','trainer','vertical_head') and _ru_teams_tva:
                        _all_tva_teams = [t for t in _all_tva_teams if t in _ru_teams_tva]
                    if selected_team:
                        _all_tva_teams = [t for t in _all_tva_teams if t in selected_team]
                    _end_disp_tva = end_date.strftime('%d-%b-%Y')

                    def _tf(v):
                        if v is None or (isinstance(v, float) and pd.isna(v)): return "₹0"
                        return f"₹{int(round(v)):,}"

                    def _def_fmt(v):
                        if v < 0: return f"-₹{int(abs(round(v))):,}"
                        return f"₹{int(round(v)):,}"

                    # ── Build per-team data ────────────────────────────────────
                    _tva_all_data = {}

                    for _tm in _all_tva_teams:
                        _ts = _tva_sheet[_tva_sheet['Team Name'] == _tm].copy()
                        _rows_tm = []

                        for _, _cr in _ts.iterrows():
                            _cn = str(_cr.get('Caller Name', '')).strip()
                            if not _cn or _cn.lower() in _EXCL_TVA:
                                continue
                            if _ru_role_tva == 'caller' and _ru_cname_tva:
                                if _cn.strip().lower() != _ru_cname_tva.strip().lower():
                                    continue
                            if search_query and search_query.lower() not in _cn.lower():
                                continue

                            _tgt_v = 0.0
                            if _tgt_col_tva:
                                try: _tgt_v = float(str(_cr.get(_tgt_col_tva, 0) or 0).replace(',',''))
                                except: _tgt_v = 0.0
                            if pd.isna(_tgt_v): _tgt_v = 0.0

                            _cnl = _cn.strip().lower()

                            _rev_mtd_v = 0.0; _enr_mtd_v = 0
                            if not _tva_rev_full.empty:
                                _mf = _tva_rev_full['Caller_name'].str.strip().str.lower() == _cnl
                                _rev_mtd_v = float(_tva_rev_full.loc[_mf, 'Fee_paid'].sum())
                                _enr_mtd_v = int((_tva_rev_full.loc[_mf, 'Enrollment'].astype(str).str.strip().str.lower() == 'new enrollment').sum())

                            _rev_yst_v = 0.0; _enr_yst_v = 0
                            if not _tva_rev_yst.empty:
                                _yf = _tva_rev_yst['Caller_name'].str.strip().str.lower() == _cnl
                                _rev_yst_v = float(_tva_rev_yst.loc[_yf, 'Fee_paid'].sum())
                                _enr_yst_v = int((_tva_rev_yst.loc[_yf, 'Enrollment'].astype(str).str.strip().str.lower() == 'new enrollment').sum())

                            # Exclude if both target<=0 and revenue<=0
                            if _tgt_v <= 0 and _rev_mtd_v <= 0:
                                continue

                            _till_v    = round((_tgt_v / 20) * _tva_wd_val) if _tgt_v > 0 else 0
                            _pct_v     = round(_rev_mtd_v / _tgt_v * 100) if _tgt_v > 0 else (100 if _rev_mtd_v > 0 else 0)
                            _deficit_v = _till_v - _rev_mtd_v
                            _pend_v    = _pending_map_tva.get(_cnl, 0)

                            _rows_tm.append({
                                'cn': _cn, 'tgt': _tgt_v, 'till': _till_v,
                                'yst': _rev_yst_v, 'mtd': _rev_mtd_v, 'pct': _pct_v,
                                'deficit': _deficit_v, 'ey': _enr_yst_v,
                                'em': _enr_mtd_v, 'pend': _pend_v,
                            })

                        if _rows_tm:
                            _tva_all_data[_tm] = _rows_tm

                    if not _tva_all_data:
                        st.warning("No data found for the selected filters.")
                        st.stop()

                    # ── Render HTML per team ───────────────────────────────────
                    _HS = ("background:#1c3a5f;color:#fff;font-size:.68rem;font-weight:700;"
                           "padding:7px 8px;text-align:center;border:1px solid #163d5a;"
                           "white-space:normal;word-wrap:break-word;")
                    _DS = ("font-size:.72rem;padding:5px 7px;text-align:center;"
                           "border:1px solid #dbeafe;color:#111827;background:#ffffff;")
                    _DA = _DS.replace('#ffffff','#f0f7ff')
                    _TS = ("background:#1c3a5f;color:#fff;font-weight:700;font-size:.72rem;"
                           "padding:6px 7px;text-align:center;border:1px solid #0d2137;")
                    _COLS_H = [
                        "COUNSELLORS","TARGET FOR<br>THE MONTH","TARGET<br>(TILL DATE)",
                        "YESTERDAY","REVENUE<br>(MTD)","CURRENT<br>DEFICIT",
                        "ENROLLMENTS<br>YESTERDAY","ENROLLMENTS<br>(MTD)",
                        "Pending Revenue<br>(Current + Prev Month)"
                    ]

                    for _tm, _rows_tm in _tva_all_data.items():
                        _hdg = f"Team {_tm} (Target Vs Achievement) {_end_disp_tva}"
                        _html_tva = f"""
<div style='margin:1.5rem 0 .3rem;'>
<div style='text-align:center;font-size:.88rem;font-weight:800;letter-spacing:.4px;
color:#fff;background:#1c3a5f;padding:10px 14px;border-radius:8px 8px 0 0;
border:2px solid #0d2137;'>{_hdg}</div>
<div style='overflow-x:auto;'>
<table style='width:100%;border-collapse:collapse;font-family:"DM Sans",sans-serif;'>
<thead><tr>{''.join(f"<th style='{_HS}'>{h}</th>" for h in _COLS_H)}</tr></thead>
<tbody>"""
                        _tt = _tl = _ty = _tm2 = _td = _tey = _tem = _tp = 0.0

                        for _i, _r in enumerate(_rows_tm):
                            _ds = _DA if _i % 2 == 1 else _DS
                            _mtds = f"{_tf(_r['mtd'])} ({_r['pct']}%)"
                            _defs = _def_fmt(_r['deficit'])
                            _html_tva += (
                                f"<tr>"
                                f"<td style='{_ds}text-align:left;font-weight:500;padding-left:10px;'>{_r['cn']}</td>"
                                f"<td style='{_ds}'>{_tf(_r['tgt'])}</td>"
                                f"<td style='{_ds}'>{_tf(_r['till'])}</td>"
                                f"<td style='{_ds}'>{_tf(_r['yst'])}</td>"
                                f"<td style='{_ds}'>{_mtds}</td>"
                                f"<td style='{_ds}'>{_defs}</td>"
                                f"<td style='{_ds}'>{_r['ey']}</td>"
                                f"<td style='{_ds}'>{_r['em']}</td>"
                                f"<td style='{_ds}'>{_tf(_r['pend'])}</td>"
                                f"</tr>"
                            )
                            _tt += _r['tgt']; _tl += _r['till']; _ty += _r['yst']
                            _tm2 += _r['mtd']; _td += _r['deficit']
                            _tey += _r['ey'];  _tem += _r['em'];  _tp  += _r['pend']

                        _tpct = round(_tm2 / _tt * 100) if _tt > 0 else 0
                        _html_tva += (
                            f"<tr>"
                            f"<td style='{_TS}text-align:left;padding-left:10px;'>TOTAL</td>"
                            f"<td style='{_TS}'>{_tf(_tt)}</td>"
                            f"<td style='{_TS}'>{_tf(_tl)}</td>"
                            f"<td style='{_TS}'>{_tf(_ty)}</td>"
                            f"<td style='{_TS}'>{_tf(_tm2)} ({_tpct}%)</td>"
                            f"<td style='{_TS}'>{_def_fmt(_td)}</td>"
                            f"<td style='{_TS}'>{int(_tey)}</td>"
                            f"<td style='{_TS}'>{int(_tem)}</td>"
                            f"<td style='{_TS}'>{_tf(_tp)}</td>"
                            f"</tr>"
                        )
                        _html_tva += "</tbody></table></div></div>"
                        st.markdown(_html_tva, unsafe_allow_html=True)

                    st.markdown("<div style='margin:1.5rem 0;'></div>", unsafe_allow_html=True)

                    # ── Single xlsx download ───────────────────────────────────
                    def _build_tva_xlsx(all_data, end_disp):
                        H_FILL  = PatternFill("solid", start_color="1c3a5f", end_color="1c3a5f")
                        S_FILL  = PatternFill("solid", start_color="1e4d6b", end_color="1e4d6b")
                        T_FILL  = PatternFill("solid", start_color="1c3a5f", end_color="1c3a5f")
                        A_FILL  = PatternFill("solid", start_color="f0f7ff", end_color="f0f7ff")
                        W_FILL  = PatternFill("solid", start_color="ffffff", end_color="ffffff")
                        H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                        D_FONT  = Font(name="Arial", size=9)
                        BW_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                        TH_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
                        CENTER  = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        LEFT    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
                        BDR     = Border(
                            left=Side(style='thin',  color='BFDBFE'),
                            right=Side(style='thin', color='BFDBFE'),
                            top=Side(style='thin',   color='BFDBFE'),
                            bottom=Side(style='thin',color='BFDBFE'),
                        )
                        COLS_XL = [
                            "COUNSELLORS","TARGET FOR THE MONTH (₹)","TARGET (TILL DATE) (₹)",
                            "YESTERDAY (₹)","REVENUE MTD",
                            "CURRENT DEFICIT (₹)","ENROLLMENTS YESTERDAY",
                            "ENROLLMENTS (MTD)","PENDING REVENUE (₹)"
                        ]
                        CW = [28,22,20,16,22,20,20,16,24]

                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Target Vs Achievement"

                        rn = 1
                        for _tm, _rows in all_data.items():
                            # Team heading
                            ws.merge_cells(start_row=rn, start_column=1,
                                           end_row=rn, end_column=len(COLS_XL))
                            _hc = ws.cell(rn, 1, f"Team {_tm} (Target Vs Achievement) {end_disp}")
                            _hc.fill = H_FILL; _hc.font = TH_FONT
                            _hc.alignment = CENTER; _hc.border = BDR
                            ws.row_dimensions[rn].height = 22; rn += 1

                            # Column headers
                            for ci, ch in enumerate(COLS_XL, 1):
                                c = ws.cell(rn, ci, ch)
                                c.fill = S_FILL; c.font = H_FONT
                                c.alignment = CENTER; c.border = BDR
                                ws.column_dimensions[get_column_letter(ci)].width = CW[ci-1]
                            ws.row_dimensions[rn].height = 24; rn += 1

                            _tt=_tl=_ty=_tm2=_td=_tey=_tem=_tp = 0.0
                            for ri, _r in enumerate(_rows):
                                fill = A_FILL if ri % 2 == 1 else W_FILL
                                _mtd_v = int(round(pd.Series(_r['mtd']).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                vals = [
                                    _r['cn'], 
                                    int(round(pd.Series(_r['tgt']).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0])), 
                                    int(round(pd.Series(_r['till']).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0])), 
                                    int(round(pd.Series(_r['yst']).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0])), 
                                    f"₹{_mtd_v:,} ({_r['pct']}%)", 
                                    int(round(pd.Series(_r['deficit']).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0])), 
                                    _r['ey'], 
                                    _r['em'], 
                                    int(round(pd.Series(_r['pend']).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0])), 
                                ]
                                for ci, v in enumerate(vals, 1):
                                    c = ws.cell(rn, ci, v)
                                    c.fill = fill; c.font = D_FONT
                                    c.border = BDR
                                    c.alignment = LEFT if ci == 1 else CENTER
                                rn += 1
                                _tt+=_r['tgt']; _tl+=_r['till']; _ty+=_r['yst']
                                _tm2+=_r['mtd']; _td+=_r['deficit']
                                _tey+=_r['ey']; _tem+=_r['em']; _tp+=_r['pend']

                                _tt_final = int(round(pd.Series([_tt]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _tl_final = int(round(pd.Series([_tl]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _ty_final = int(round(pd.Series([_ty]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _tm2_final = int(round(pd.Series([_tm2]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _td_final = int(round(pd.Series([_td]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _tp_final = int(round(pd.Series([_tp]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _tey_final = int(round(pd.Series([_tey]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                                _tem_final = int(round(pd.Series([_tem]).apply(pd.to_numeric, errors='coerce').fillna(0).iloc[0]))
                        
                                _tpct_xl = f"{round(_tm2_final / _tt_final * 100)}%" if _tt_final > 0 else "0%"
                        
                                tot = [
                                    "TOTAL",
                                    _tt_final,
                                    _tl_final,
                                    _ty_final,
                                    f"₹{_tm2_final:,} ({_tpct_xl})",
                                    _td_final,
                                    _tey_final,
                                    _tem_final,
                                    _tp_final,
                                ]    
                            for ci, v in enumerate(tot, 1):
                                c = ws.cell(rn, ci, v)
                                c.fill = T_FILL; c.font = BW_FONT
                                c.border = BDR
                                c.alignment = LEFT if ci == 1 else CENTER
                            rn += 2  # +1 blank row between teams

                        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

                    _dl_tva_col, _ = st.columns([1, 2])
                    with _dl_tva_col:
                        st.download_button(
                            label="📥 Download Target Vs Achievement (.xlsx)",
                            data=_build_tva_xlsx(_tva_all_data, _end_disp_tva),
                            file_name=f"Target_Vs_Achievement_{display_start}_to_{display_end}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_tva_xlsx',
                        )


# --- MAIN APP ROUTER ---
if not st.session_state.get('password_correct', False):
    show_homepage_with_login()
else:
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    [data-testid="stStatusWidget"] { display: none !important; }
    header[data-testid="stHeader"] { background: transparent !important; }
    </style>
    """, unsafe_allow_html=True)

    _apply_role_filters()

    ri   = st.session_state.get('auth_role_info', {'role': 'admin'})
    role = ri.get('role', 'admin')
    disp = ri.get('display_name', st.session_state.get('current_user', ''))

    _ROLE_LABELS = {
        'admin'        : '🛡️ Admin',
        'vertical_head': '🏢 Vertical Head',
        'trainer'      : '🎓 Sales Leader',
        'tl'           : '👑 Team Lead',
        'caller'       : '📞 Caller',
    }

    _lc = "#10B981"
    _sc = "#059669"
    _shc = "rgba(16,185,129,.35)"

    st.sidebar.markdown(f"""
    <div style='margin:.4rem 0 .6rem;padding:.55rem .7rem;
                background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);
                border-radius:10px;text-align:center;'>
        <div style='font-size:.6rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:1px;color:{_lc};margin-bottom:.2rem;'>
            {_ROLE_LABELS.get(role, role)}
        </div>
        <div style='font-size:.72rem;color:{_lc};word-break:break-all;'>
            {disp}
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.sidebar.markdown("""
    <style>
    div[data-testid="stSidebar"] div[data-testid="stButton"]:first-of-type > button {
        display: block !important;
        margin: 0 auto !important;
        width: auto !important;
        min-width: 120px !important;
        padding-left: 1.4rem !important;
        padding-right: 1.4rem !important;
    }
    </style>
    """, unsafe_allow_html=True)
    col_so = st.sidebar.columns([1, 2, 1])
    with col_so[1]:
        _sign_out = st.button("🚪 Sign Out", key="signout_btn")
    if _sign_out:
        try:
            supa.auth.sign_out()
        except Exception:
            pass
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    st.sidebar.markdown(f"""
    <div style='padding:.7rem 0 .5rem;text-align:center;
                border-bottom:1px solid rgba(128,128,128,.15);'>
        <div style='display:flex;align-items:center;justify-content:center;margin-bottom:.25rem;'>
            <span style='font-size:.95rem;font-weight:700;color:{_lc};letter-spacing:-.4px;'>LawSikho</span>
            <div style='width:1px;height:16px;margin:0 .55rem;
                        background:linear-gradient(180deg,transparent,{_sc},transparent);
                        box-shadow:0 0 5px {_shc};'></div>
            <span style='font-size:.95rem;font-weight:700;color:{_lc};letter-spacing:-.4px;'>Skill Arbitrage</span>
        </div>
        <div style='font-size:.6rem;color:{_lc};letter-spacing:1px;font-family:monospace;font-weight:600;'>
            India Learning 📖 India Earning
        </div>
    </div>
    """, unsafe_allow_html=True)

    run_revenue_dashboard()
