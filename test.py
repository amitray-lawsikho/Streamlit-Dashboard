import streamlit as st
import pandas as pd
import os
import pytz
import io
from datetime import datetime, date, time, timedelta
from google.cloud import bigquery
from google.oauth2 import service_account
import warnings
warnings.filterwarnings(
    "ignore",
    message="Please replace `st.components.v1.html` with `st.iframe`"
)
# ReportLab imports (used by both dashboards)
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Flowable
from reportlab.lib.enums import TA_CENTER

# Openpyxl imports (used by Revenue dashboard)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client
import re

# --- GLOBAL CONFIG & CREDENTIALS ---
import streamlit.components.v1 as components
st.set_page_config(
    page_title="Analytics Dashboard — LawSikho",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

def get_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=credentials, project=info["project_id"])
    else:
        SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
        if os.path.exists(SERVICE_ACCOUNT_FILE):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
            return bigquery.Client()
        return None

@st.cache_resource
def get_cached_bq_client():
    if "gcp_service_account" in st.secrets:
        info = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(info)
        return bigquery.Client(credentials=credentials, project=info["project_id"])
    else:
        SERVICE_ACCOUNT_FILE = "C:\\Users\\AMIT GAMING\\.gemini\\antigravity\\secrets\\bigquery_key.json"
        if os.path.exists(SERVICE_ACCOUNT_FILE):
            os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = SERVICE_ACCOUNT_FILE
            return bigquery.Client()
        return None

client = get_cached_bq_client()

# --- LOGIN MODULE ---

def _apply_role_filters():
    ri = st.session_state.get('auth_role_info', {'role': 'admin'})
    role    = ri.get('role', 'admin')
    teams   = ri.get('teams')      # list | None
    callers = ri.get('callers')    # list | None
    cname   = ri.get('caller_name')

    st.session_state['rf_role']        = role
    st.session_state['rf_teams']       = teams   or []
    st.session_state['rf_callers']     = callers or []
    st.session_state['rf_caller_name'] = cname   or ''
    # Roles that see team/vertical multiselects in the sidebar
    st.session_state['rf_full_filters']= role == 'admin'

# ── Hardcoded access tiers ─────────────────────────────────────
ADMIN_EMAILS = {
    "amitray@lawsikho.in",
    "rinku@lawsikho.in",
}

VERTICAL_HEAD_TEAMS = {
    # email               : [list of Team Names they can see]
    "uzair@lawsikho.in"   : ["Elite", "Corporate law - Jyoti"],

}

AUTH_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf"
    "/pub?gid=0&single=true&output=csv"
)

# Sheet column names — update if your sheet differs
_COL_NAME    = "Caller Name"
_COL_EMAIL   = "Email id"
_COL_DESIG   = "Academic Counselor/TL/ATL"
_COL_TEAM    = "Team Name"
_COL_TRAINER = "Sales Leader"
_TL_VALS     = {"TL", "ATL", "AD", "TEAM LEAD", "TEAM LEADER"}


@st.cache_resource
def get_supabase():
    # Try top-level keys first, then nested under a section
    url = (
        st.secrets.get("SUPABASE_URL")
        or st.secrets.get("supabase", {}).get("url")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_URL")
        or ""
    )
    key = (
        st.secrets.get("SUPABASE_KEY")
        or st.secrets.get("supabase", {}).get("key")
        or st.secrets.get("gcp_service_account", {}).get("SUPABASE_KEY")
        or ""
    )

    if not url or not key:
        st.error(
            "⚠️ Supabase credentials missing from secrets. "
            "Add SUPABASE_URL and SUPABASE_KEY to your secrets.toml at the TOP LEVEL "
            "(before any [section] header)."
        )
        st.stop()

    return create_client(url, key)

supa = get_supabase()


@st.cache_data(ttl=300, show_spinner=False)
def load_auth_sheet() -> pd.DataFrame:
    df = pd.read_csv(AUTH_SHEET_URL)
    df.columns = df.columns.str.strip()
    if _COL_EMAIL in df.columns:
        df['_email_norm'] = df[_COL_EMAIL].astype(str).str.strip().str.lower()
    return df


def _extract_trainer_email(cell: str) -> str | None:
    """Parses 'Name (email@domain.com)' → 'email@domain.com'"""
    m = re.search(r'\(([^)\s]+@[^)\s]+)\)', str(cell))
    return m.group(1).strip().lower() if m else None


def determine_role(email: str, df: pd.DataFrame) -> dict | None:
   
    el = email.strip().lower()

    # 1 ── Admin
    if el in {e.lower() for e in ADMIN_EMAILS}:
        return {'role': 'admin', 'teams': None, 'callers': None,
                'caller_name': None, 'display_name': email}

    # 2 ── Vertical Head (hardcoded)
    for vh_mail, vh_teams in VERTICAL_HEAD_TEAMS.items():
        if vh_mail.lower() == el:
            return {'role': 'vertical_head', 'teams': vh_teams, 'callers': None,
                    'caller_name': None, 'display_name': email}

    if '_email_norm' not in df.columns:
        return None

    # 3 ── Trainer  (Sales Leader column)
    if _COL_TRAINER in df.columns:
        df2 = df.copy()
        df2['_tr_email'] = df2[_COL_TRAINER].apply(_extract_trainer_email)
        trainer_rows = df2[df2['_tr_email'] == el]
        if not trainer_rows.empty:
            teams   = trainer_rows[_COL_TEAM].dropna().unique().tolist()  if _COL_TEAM in trainer_rows else []
            callers = trainer_rows[_COL_NAME].dropna().unique().tolist()  if _COL_NAME in trainer_rows else []
            name    = trainer_rows[_COL_NAME].iloc[0] if _COL_NAME in trainer_rows.columns and len(trainer_rows) else email
            return {'role': 'trainer', 'teams': teams, 'callers': callers,
                    'caller_name': None, 'display_name': name}

    user_rows = df[df['_email_norm'] == el]
    if user_rows.empty:
        return None  # email not in sheet → not authorised

    # 4 ── TL / AD / ATL
    if _COL_DESIG in user_rows.columns:
        tl_rows = user_rows[
            user_rows[_COL_DESIG].astype(str).str.strip().str.upper().isin(_TL_VALS)
        ]
        if not tl_rows.empty:
            team = tl_rows.iloc[0][_COL_TEAM] if _COL_TEAM in tl_rows.columns else None
            team_callers = (
                df[df[_COL_TEAM] == team][_COL_NAME].dropna().unique().tolist()
                if team else []
            )
            disp = tl_rows.iloc[0][_COL_NAME] if _COL_NAME in tl_rows.columns else email
            return {'role': 'tl', 'teams': [team] if team else [], 'callers': team_callers,
                    'caller_name': None, 'display_name': disp}

    # 5 ── Regular Caller
    caller_name = user_rows.iloc[0][_COL_NAME] if _COL_NAME in user_rows.columns else email
    return {'role': 'caller', 'teams': None, 'callers': None,
            'caller_name': caller_name, 'display_name': caller_name}


# ──────────────────────────────────────────────────────────────
# AUTH UI HELPERS
# ──────────────────────────────────────────────────────────────

def _complete_login(email: str, session):
    """Resolve role and store everything in session state."""
    df_auth   = load_auth_sheet()
    role_info = determine_role(email, df_auth)
    if role_info is None:
        st.error("⛔ Your email is not authorised to access this dashboard.")
        return
    st.session_state['password_correct'] = True
    st.session_state['current_user']     = email
    st.session_state['supabase_session'] = session
    st.session_state['auth_role_info']   = role_info
    st.rerun()


def _auth_sign_in_panel():
    """Email + Password login panel."""
    email = st.text_input("Email", key="si_email", placeholder="your@lawsikho.in")
    pwd   = st.text_input("Password", type="password", key="si_pwd", placeholder="Your password")

    c1, c2 = st.columns(2)
    with c1:
        if st.button("Sign In →", key="si_btn", use_container_width=True):
            if not email or not pwd:
                st.error("Enter both email and password."); return
            try:
                resp = supa.auth.sign_in_with_password({"email": email, "password": pwd})
                _complete_login(resp.user.email, resp.session)
            except Exception as ex:
                st.error(f"Login failed: {ex}")
    with c2:
        if st.button("Forgot / First time?", key="si_otp_switch", use_container_width=True):
            st.session_state['auth_tab']           = "otp"
            st.session_state['otp_prefill_email']  = email
            st.session_state['otp_step']           = 1
            st.rerun()


def _auth_otp_panel():
    """OTP verification + password-set panel (first-time or reset)."""
    step = st.session_state.get('otp_step', 1)

    # ── Step 1: request OTP ──────────────────────────────────
    if step == 1:
        email = st.text_input(
            "Email", key="otp_email",
            value=st.session_state.get('otp_prefill_email', ''),
            placeholder="your@lawsikho.in"
        )
        if st.button("Send OTP →", key="otp_send_btn", use_container_width=True):
            if not email:
                st.error("Enter your email."); return
            # Pre-check: is the email in the sheet?
            role_check = determine_role(email, load_auth_sheet())
            if role_check is None:
                st.error("⛔ This email is not authorised."); return
            try:
                supa.auth.sign_in_with_otp({
                    "email": email,
                    "options": {"should_create_user": True}
                })
                st.session_state['otp_step']          = 2
                st.session_state['otp_pending_email'] = email
                st.rerun()
            except Exception as ex:
                st.error(f"Could not send OTP: {ex}")

        if st.button("← Back to Sign In", key="otp_back1", use_container_width=True):
            st.session_state['auth_tab'] = "signin"
            st.rerun()

    # ── Step 2: verify OTP + set password ───────────────────
    elif step == 2:
        pending_email = st.session_state.get('otp_pending_email', '')
        st.success(f"OTP sent to **{pending_email}** — check your inbox (also spam).")

        otp  = st.text_input("6-digit OTP code", key="otp_code", max_chars=6, placeholder="123456")
        pw1  = st.text_input("Set new password",    type="password", key="otp_pw1")
        pw2  = st.text_input("Confirm password",    type="password", key="otp_pw2")

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Verify & Continue →", key="otp_verify_btn", use_container_width=True):
                if len(otp) != 6:
                    st.error("Enter the 6-digit code."); return
                if pw1 != pw2:
                    st.error("Passwords don't match."); return
                if len(pw1) < 8:
                    st.error("Password must be ≥ 8 characters."); return
                try:
                    resp = supa.auth.verify_otp({
                        "email": pending_email,
                        "token": otp,
                        "type": "email"
                    })
                    # Save the password (user is now logged in to Supabase)
                    supa.auth.update_user({"password": pw1})
                    st.session_state['otp_step'] = 1
                    _complete_login(resp.user.email, resp.session)
                except Exception as ex:
                    st.error(f"Verification failed — wrong code or it expired: {ex}")
        with c2:
            if st.button("← Back", key="otp_back2", use_container_width=True):
                st.session_state['otp_step'] = 1
                st.rerun()

# --- DASHBOARD FUNCTIONS ---
@st.cache_data(ttl=300, show_spinner=False)
def get_stats():
    try:
        r1 = client.query("""
            SELECT updated_at_ampm FROM (
                SELECT updated_at_ampm FROM `studious-apex-488820-c3.crm_dashboard.acefone_calls`
                UNION ALL
                SELECT updated_at_ampm FROM `studious-apex-488820-c3.crm_dashboard.ozonetel_calls`
            ) WHERE updated_at_ampm IS NOT NULL ORDER BY 1 DESC LIMIT 1
        """).to_dataframe()
        call_time = str(r1["updated_at_ampm"].iloc[0]) if not r1.empty else "N/A"

        r2 = client.query("""
            SELECT SUM(c) AS t FROM (
                SELECT COUNT(*) AS c FROM `studious-apex-488820-c3.crm_dashboard.acefone_calls`
                UNION ALL
                SELECT COUNT(*) AS c FROM `studious-apex-488820-c3.crm_dashboard.ozonetel_calls`
            )
        """).to_dataframe()
        call_cnt = "{:,}".format(int(r2["t"].iloc[0])) if not r2.empty else "—"

        r3 = client.query("""
             SELECT MAX(updated_at_ampm) AS last_updated, COUNT(*) AS cnt
             FROM `studious-apex-488820-c3.crm_dashboard.revenue_sheet`
             """).to_dataframe()
        rev_time = str(r3["last_updated"].iloc[0]) if not r3.empty and r3["last_updated"].iloc[0] else "N/A"
        rev_cnt  = "{:,}".format(int(r3["cnt"].iloc[0])) if not r3.empty else "0"

        r4 = client.query("""
            SELECT MAX(updated_at_ampm) AS last_updated, COUNT(*) AS cnt
            FROM `studious-apex-488820-c3.crm_dashboard.lsq_leads`
            """).to_dataframe()
        lead_time = str(r4["last_updated"].iloc[0]) if not r4.empty and r4["last_updated"].iloc[0] else "N/A"
        lead_cnt  = "{:,}".format(int(r4["cnt"].iloc[0])) if not r4.empty else "0"

        return call_time, call_cnt, rev_time, rev_cnt, lead_time, lead_cnt
    except:
        return "N/A", "—", "N/A", "—", "N/A", "—"

# ADD THIS FUNCTION AFTER get_stats() and BEFORE run_calling_dashboard()
# ===========================================================================

def show_homepage_with_login():
    # ── Full-page dark background + input styling ──
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    #MainMenu, header[data-testid="stHeader"] { display: none !important; }
    [data-testid="stStatusWidget"],
    [data-testid="collapsedControl"],
    [data-testid="stSidebarCollapsedControl"] { display: none !important; }
    [data-testid="stAppViewContainer"],
    [data-testid="stMain"], .main { background: #0B1120 !important; }
    .block-container { padding: 0 !important; max-width: 100% !important; margin-bottom: 0 !important; }
    [data-testid="stHorizontalBlock"] { gap: 0 !important; }

    /* ── Force every layer of the login column dark ── */
    div[data-testid="column"]:nth-child(2),
    div[data-testid="column"]:nth-child(2) > div,
    div[data-testid="column"]:nth-child(2) > div > div,
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"],
    div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlockBorderWrapper"] {
        background-color: #0f172a !important;
    }
    div[data-testid="column"]:nth-child(2) > div:first-child {
        background-color: #0f172a !important;
        border: 1px solid rgba(255,255,255,.12) !important;
        border-radius: 16px !important;
        padding: 1.6rem 1.8rem 1.8rem !important;
    }

    /* ── BaseUI input container (what Streamlit's theme actually targets) ── */
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] {
        background-color: #1e293b !important;
        border-color: rgba(148,163,184,.3) !important;
        border-radius: 10px !important;
    }
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"]:focus-within,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"]:focus-within {
        border-color: #F97316 !important;
        box-shadow: 0 0 0 2px rgba(249,115,22,.2) !important;
    }

    /* ── The actual input element — every state ── */
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"] input,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] input,
    div[data-testid="column"]:nth-child(2) input[type="text"],
    div[data-testid="column"]:nth-child(2) input[type="password"],
    div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #F97316 !important;
        border: none !important;
    }
    div[data-testid="column"]:nth-child(2) input::placeholder {
        color: rgba(241,245,249,.32) !important;
        -webkit-text-fill-color: rgba(241,245,249,.32) !important;
    }

    /* ── Chrome/Safari autofill override (the inset shadow trick) ── */
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:hover,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:focus,
    div[data-testid="column"]:nth-child(2) input:-webkit-autofill:active {
        -webkit-box-shadow: 0 0 0px 1000px #1e293b inset !important;
        box-shadow: 0 0 0px 1000px #1e293b inset !important;
        -webkit-text-fill-color: #f1f5f9 !important;
        caret-color: #F97316 !important;
    }

    /* ── Labels ── */
    div[data-testid="column"]:nth-child(2) label,
    div[data-testid="column"]:nth-child(2) label p {
        color: rgba(241,245,249,.55) !important;
        font-size: 0.8rem !important;
    }

    /* ── Eye icon button in password field ── */
    div[data-testid="column"]:nth-child(2) [data-baseweb="input"] button,
    div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"] button {
        background-color: transparent !important;
        color: rgba(241,245,249,.5) !important;
        border: none !important;
    }

    /* ── Sign In button — orange always, white text always ── */
    div[data-testid="column"]:nth-child(2) .stButton > button {
        width: 100% !important;
        background: linear-gradient(135deg, #F97316, #EA580C) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 11px !important;
        font-size: 0.9rem !important;
        font-weight: 600 !important;
        box-shadow: none !important;
    }
    div[data-testid="column"]:nth-child(2) .stButton > button:hover,
    div[data-testid="column"]:nth-child(2) .stButton > button:active,
    div[data-testid="column"]:nth-child(2) .stButton > button:focus {
        background: linear-gradient(135deg, #EA580C, #C2410C) !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        border: none !important;
        outline: none !important;
        box-shadow: 0 4px 16px rgba(249,115,22,.35) !important;
        transform: translateY(-1px) !important;
    }

    /* ── Streamlit light-theme specific override (belt and braces) ── */
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-baseweb="input"],
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-baseweb="base-input"],
    [data-theme="light"] div[data-testid="column"]:nth-child(2) input {
        background-color: #1e293b !important;
        color: #f1f5f9 !important;
        -webkit-text-fill-color: #f1f5f9 !important;
    }
    [data-theme="light"] div[data-testid="column"]:nth-child(2),
    [data-theme="light"] div[data-testid="column"]:nth-child(2) > div,
    [data-theme="light"] div[data-testid="column"]:nth-child(2) [data-testid="stVerticalBlock"] {
        background-color: #0f172a !important;
    }
    </style>
    """, unsafe_allow_html=True)

    call_time, call_cnt, rev_time, rev_cnt, lead_time, lead_cnt = get_stats()

    # ── HERO HTML ──
    html_hero = f"""<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700;800&family=Plus+Jakarta+Sans:wght@300;400;500&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet"/>
    <style>
    *{{box-sizing:border-box;margin:0;padding:0;}}
    html,body{{font-family:'Plus Jakarta Sans',sans-serif;background:#0B1120;color:#E2E8F0;overflow-x:hidden;}}
    body{{background:
        radial-gradient(ellipse 80% 50% at 50% -10%,rgba(59,130,246,.12) 0%,transparent 60%),
        radial-gradient(ellipse 60% 40% at 90% 80%,rgba(249,115,22,.08) 0%,transparent 55%),
        #0B1120;}}
    body::before{{content:"";position:fixed;inset:0;
        background-image:linear-gradient(rgba(255,255,255,.022) 1px,transparent 1px),
            linear-gradient(90deg,rgba(255,255,255,.022) 1px,transparent 1px);
        background-size:48px 48px;pointer-events:none;}}
    .hero{{display:flex;flex-direction:column;align-items:center;text-align:center;padding:3rem 2rem 2.5rem;}}
    .logos{{display:flex;align-items:center;margin-bottom:1rem;}}
    .lname{{font-size:1.25rem;font-weight:700;color:#fff;padding:0 1.6rem;}}
    .lsep{{width:1px;height:48px;background:linear-gradient(180deg,transparent,rgba(249,115,22,.85),transparent);box-shadow:0 0 8px rgba(249,115,22,.5);}}
    .tagline{{font-family:'Fira Code',monospace;font-size:.75rem;color:rgba(255,255,255,.32);letter-spacing:1.5px;margin-bottom:2rem;}}
    .eyebrow{{display:inline-flex;align-items:center;gap:.45rem;font-family:'Fira Code',monospace;font-size:.65rem;
        letter-spacing:2.5px;text-transform:uppercase;color:#F97316;background:rgba(249,115,22,.08);
        border:1px solid rgba(249,115,22,.18);border-radius:100px;padding:.28rem .95rem;margin-bottom:1.2rem;}}
    .dot{{width:5px;height:5px;background:#F97316;border-radius:50%;box-shadow:0 0 5px #F97316;
        animation:p 2s ease-in-out infinite;}}
    @keyframes p{{0%,100%{{opacity:1;transform:scale(1);}}50%{{opacity:.45;transform:scale(1.4);}}}}
    .headline{{font-family:'Playfair Display',serif;font-size:clamp(2rem,5vw,3.5rem);
        font-weight:800;line-height:1.09;color:#fff;letter-spacing:-1.5px;margin-bottom:.65rem;}}
    .accent{{background:linear-gradient(125deg,#F97316,#FB923C 40%,#FBBF24);
        -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;}}
    .sub{{font-size:1rem;font-weight:300;color:rgba(255,255,255,.38);max-width:500px;}}
    </style></head><body>
    <div class="hero">
      <div class="logos">
        <div class="lname">LawSikho</div>
        <div class="lsep"></div>
        <div class="lname">Skill Arbitrage</div>
      </div>
      <div class="tagline">India Learning &nbsp;📖&nbsp; India Earning</div>
      <div class="eyebrow"><span class="dot"></span>Internal Analytics Hub</div>
      <div class="headline">All your dashboards,<br><span class="accent">at one place</span></div>
      <div class="sub">Real-time insights across Leads, Revenue &amp; Calling</div>
    </div>
    </body></html>"""
    components.html(html_hero, height=420, scrolling=False)

    # ── AUTH PANEL ─────────────────────────────────────────────
    left, mid, right = st.columns([1, 1, 1])
    with mid:
        auth_tab = st.session_state.get('auth_tab', 'signin')

        st.markdown("""
        <div style="text-align:center;margin-bottom:.8rem;">
            <span style="font-family:'Playfair Display',serif;font-size:1.1rem;font-weight:600;color:#fff;">
                🔐 DASHBOARD ACCESS
            </span>
        </div>
        """, unsafe_allow_html=True)

        if auth_tab == 'signin':
            _auth_sign_in_panel()
        else:
            _auth_otp_panel()

        if st.session_state.get('password_correct'):
            st.rerun()
            
    # ── STATS + CARDS + FOOTER HTML ──
    html_bottom = f"""<!DOCTYPE html><html><head>
    <link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Plus+Jakarta+Sans:wght@300;400;500&family=Fira+Code:wght@400;500&display=swap" rel="stylesheet"/>
    <style>
    *{{box-sizing:border-box;margin:0;padding:0;}}
    html,body{{font-family:'Plus Jakarta Sans',sans-serif;background:#0B1120;color:#E2E8F0;}}
    .stats{{display:flex;justify-content:center;gap:1rem;flex-wrap:wrap;padding:2.5rem 2rem 1.5rem;}}
    .sc{{display:flex;align-items:center;gap:.8rem;background:rgba(255,255,255,.04);
        border:1px solid rgba(255,255,255,.08);border-radius:14px;padding:.85rem 1.3rem;
        min-width:230px;flex:1;max-width:310px;transition:all .2s;}}
    .sc:hover{{transform:translateY(-2px);}}
    .sc.co:hover{{border-color:rgba(249,115,22,.22);background:rgba(249,115,22,.04);}}
    .sc.cg:hover{{border-color:rgba(52,211,153,.22);background:rgba(52,211,153,.04);}}
    .sb{{background:rgba(59,130,246,.14);}}
    .cb:hover{{border-color:rgba(59,130,246,.22);background:rgba(59,130,246,.04);}}
    .si{{width:34px;height:34px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:.88rem;flex-shrink:0;}}
    .so{{background:rgba(249,115,22,.14);}} .sg{{background:rgba(52,211,153,.12);}} .sp{{background:rgba(139,92,246,.12);}}
    .sinfo{{display:flex;flex-direction:column;gap:2px;}}
    .slbl{{font-family:'Fira Code',monospace;font-size:.55rem;text-transform:uppercase;color:rgba(255,255,255,.28);}}
    .sval{{font-family:'Fira Code',monospace;font-size:.76rem;color:rgba(255,255,255,.8);white-space:nowrap;}}
    .ssub{{font-family:'Fira Code',monospace;font-size:.54rem;color:rgba(255,255,255,.18);}}
    .pl,.pw{{margin-left:auto;font-family:'Fira Code',monospace;font-size:.52rem;letter-spacing:.8px;text-transform:uppercase;border-radius:20px;padding:2px 8px;}}
    .pl{{color:#34D399;background:rgba(52,211,153,.1);border:1px solid rgba(52,211,153,.18);}}
    .pw{{color:#FBBF24;background:rgba(251,191,36,.1);border:1px solid rgba(251,191,36,.18);}}
    .dsec{{padding:0 2rem 3.5rem;max-width:1080px;margin:0 auto;}}
    .sh{{display:flex;align-items:center;gap:1rem;margin-bottom:1.8rem;}}
    .sl{{flex:1;height:1px;background:rgba(255,255,255,.07);}}
    .slb{{font-family:'Fira Code',monospace;font-size:.62rem;letter-spacing:2.5px;text-transform:uppercase;color:rgba(255,255,255,.2);}}
    .grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:1.2rem;}}
    @media(max-width:900px){{.grid{{grid-template-columns:1fr;}}}}
    @media(max-width:1200px) and (min-width:901px){{.grid{{grid-template-columns:repeat(2,1fr);}}}}
    .dc{{background:rgba(255,255,255,.033);border:1px solid rgba(255,255,255,.08);border-radius:16px;
        padding:1.6rem 1.5rem 1.3rem;display:flex;flex-direction:column;gap:.65rem;
        transition:transform .25s,box-shadow .25s,border-color .25s;}}
    .dc:hover{{transform:translateY(-5px);}}
    .dc.wip{{opacity:.5;}}
    .dc-o{{border-top:2px solid rgba(249,115,22,.4);}}
    .dc-g{{border-top:2px solid rgba(52,211,153,.35);}}
    .dc-p{{border-top:2px solid rgba(139,92,246,.3);}}
    .dc-o:hover{{border-color:#F97316;background:rgba(249,115,22,.04);box-shadow:0 18px 50px rgba(249,115,22,.09);}}
    .dc-g:hover{{border-color:#34D399;background:rgba(52,211,153,.04);box-shadow:0 18px 50px rgba(52,211,153,.08);}}
    .dc-b{{border-top:2px solid #3B82F6;}}
    .dc-b:hover{{border-color:#3B82F6;background:rgba(59,130,246,.04);box-shadow:0 18px 50px rgba(59,130,246,.09);}}
    .dh{{display:flex;align-items:flex-start;justify-content:space-between;}}
    .di{{font-size:1.65rem;}}
    .wb{{font-family:'Fira Code',monospace;font-size:.5rem;text-transform:uppercase;color:#FBBF24;
        background:rgba(251,191,36,.08);border:1px solid rgba(251,191,36,.18);border-radius:6px;padding:3px 7px;}}
    .dt{{font-family:'Playfair Display',serif;font-size:1.1rem;font-weight:600;color:#fff;}}
    .dd{{font-size:.77rem;color:rgba(255,255,255,.38);line-height:1.7;}}
    .tags{{display:flex;flex-wrap:wrap;gap:.32rem;}}
    .tag{{font-family:'Fira Code',monospace;font-size:.54rem;color:rgba(255,255,255,.26);
        background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.07);border-radius:5px;padding:2px 7px;text-transform:uppercase;}}
    .foot{{border-top:1px solid rgba(255,255,255,.06);padding:1.6rem 2rem;text-align:center;}}
    .f1{{font-family:'Fira Code',monospace;font-size:.64rem;color:rgba(255,255,255,.28);margin-bottom:.4rem;}}
    .f2{{font-family:'Fira Code',monospace;font-size:.58rem;color:rgba(255,255,255,.14);}}
    .fd{{display:inline-block;width:3px;height:3px;background:rgba(249,115,22,.4);border-radius:50%;margin:0 .45rem;vertical-align:middle;}}
    </style></head><body>
    <div class="stats">
      <div class="sc co"><div class="si so">🔔</div>
        <div class="sinfo"><span class="slbl">Calling Data</span><span class="sval">{call_time}</span><span class="ssub">{call_cnt} records</span></div>
        <span class="pl">● Live</span></div>
      <div class="sc cg"><div class="si sg">💰</div>
        <div class="sinfo"><span class="slbl">Revenue Data</span><span class="sval">{rev_time}</span><span class="ssub">{rev_cnt} records</span></div>
        <span class="pl">● Live</span></div>
      <div class="sc cb"><div class="si sb">📊</div>
         <div class="sinfo"><span class="slbl">Leads Data</span>
         <span class="sval">{lead_time}</span>
         <span class="ssub">{lead_cnt} records</span></div>
         <span class="pl">● Live</span></div>
    </div>
    <div class="dsec">
      <div class="sh"><div class="sl"></div><span class="slb">Dashboards</span><div class="sl"></div></div>
      <div class="grid">
        <div class="dc dc-o">
          <div class="dh"><div class="di">🔔</div></div>
          <div class="dt">Calling Metrics</div>
          <div class="dd">Full CDR analysis across Ozonetel, Acefone &amp; Manual. Agent performance, break tracking, productive hours &amp; leaderboards.</div>
          <div class="tags"><span class="tag">Ozonetel</span><span class="tag">Acefone</span><span class="tag">Manual</span><span class="tag">Teams</span></div>
        </div>
        <div class="dc dc-g">
          <div class="dh"><div class="di">💰</div></div>
          <div class="dt">Revenue Metrics</div>
          <div class="dd">Enrollment revenue, target achievement &amp; caller-level breakdown. Source mix &amp; team leaderboards.</div>
          <div class="tags"><span class="tag">Enrollments</span><span class="tag">Targets</span><span class="tag">Achievement</span><span class="tag">Teams</span></div>
        </div>
        <div class="dc dc-b">
           <div class="dh"><div class="di">📊</div></div>
           <div class="dt">Lead Metrics</div>
           <div class="dd">Assigned lead distribution, potential breached leads & less-dialled leads analysis. Stage-by-stage callerwise and teamwise breakdown.</div>
           <div class="tags"><span class="tag">Assigned</span><span class="tag">Breached</span><span class="tag">Less Dialled</span><span class="tag">Stages</span></div>
       </div>
      </div>
    </div>
    <div class="foot">
      <div class="f1">For Internal Use of Sales and Operations Team Only<span class="fd"></span>All Rights Reserved</div>
      <div class="f2">Developed and Designed by Amit Ray<span class="fd"></span>Reach out for Support and Queries</div>
    </div>
    </body></html>"""
    components.html(html_bottom, height=640, scrolling=False)

def run_calling_dashboard():
    # ADD THIS CSS BLOCK FIRST:
    st.markdown("""
    <style>
    [data-testid="stMainBlockContainer"] {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    .block-container {
        max-width: 100% !important;
    }
    </style>
    """, unsafe_allow_html=True)
    CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf/pub?gid=0&single=true&output=csv"

    # --- PROFESSIONAL WARM THEME ---
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

    :root {
        --accent-primary:   #F97316;
        --accent-secondary: #EF4444;
        --accent-success:   #EAB308;
        --accent-warn:      #FBBF24;
        --accent-danger:    #DC2626;
        --gold:             #F59E0B;
        --silver:           #9CA3AF;
        --bronze:           #CD7F32;
        --radius-sm:        8px;
        --radius-md:        12px;
        --radius-lg:        16px;
        --shadow-sm:        0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
        --shadow-md:        0 4px 16px rgba(0,0,0,.10);
        --shadow-lg:        0 8px 32px rgba(0,0,0,.14);
        --transition:       all 0.22s cubic-bezier(.4,0,.2,1);
    }

    [data-testid="stAppViewContainer"]:not([class*="dark"]) {
        --bg-base:      #FFF8F3;
        --bg-surface:   #FFFFFF;
        --bg-elevated:  #FFFFFF;
        --bg-muted:     #FEF3E8;
        --border:       rgba(249,115,22,.12);
        --text-primary: #111827;
        --text-muted:   #6B7280;
        --metric-bg:    #FFFFFF;
    }

    @media (prefers-color-scheme: dark) {
        :root {
            --bg-base:      #0F0A05;
            --bg-surface:   #1A1006;
            --bg-elevated:  #231508;
            --bg-muted:     #1E1207;
            --border:       rgba(249,115,22,.10);
            --text-primary: #FEF3E8;
            --text-muted:   #D1A67A;
            --metric-bg:    #231508;
        }
    }

    [data-theme="dark"] {
        --bg-base:      #0F0A05 !important;
        --bg-surface:   #1A1006 !important;
        --bg-elevated:  #231508 !important;
        --bg-muted:     #1E1207 !important;
        --border:       rgba(249,115,22,.10) !important;
        --text-primary: #FEF3E8 !important;
        --text-muted:   #D1A67A !important;
        --metric-bg:    #231508 !important;
    }

    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }

    footer { visibility: hidden; }
    [data-testid="stStatusWidget"], .stStatusWidget { display: none !important; }
    [data-testid="stMainViewContainer"] { padding-top: 1.5rem; }
    [data-testid="stSidebar"] { border-right: 1px solid var(--border, rgba(249,115,22,.12)); }

    .cw-header {
        background: linear-gradient(135deg, #1c0700 0%, #7c2d12 50%, #431407 100%);
        border-radius: var(--radius-lg);
        padding: 1.5rem 2rem 1.2rem;
        margin-bottom: 1.2rem;
        position: relative;
        overflow: hidden;
        box-shadow: var(--shadow-lg);
    }
    .cw-header::before, .cw-header::after { display: none; }
    .cw-title { font-size: 1.65rem; font-weight: 700; color: #FFFFFF; letter-spacing: .5px; margin: 0 0 .25rem; }
    .cw-subtitle { font-size: .82rem; color: rgba(255,255,255,.6); font-weight: 400; margin: 0; font-family: 'DM Mono', monospace; }
    .cw-badge {
        display: inline-flex; align-items: center; gap: 5px;
        background: var(--bg-muted, #FEF3E8);
        border: 1px solid var(--border, rgba(249,115,22,.12));
        border-radius: 20px; padding: 3px 10px; font-size: .73rem;
        color: var(--text-primary, #111827); font-family: 'DM Mono', monospace;
    }
    .cw-pulse {
        width: 6px; height: 6px; background: #EAB308; border-radius: 50%;
        display: inline-block; animation: pulse-ring 1.8s ease-in-out infinite;
    }
    @keyframes pulse-ring {
        0%, 100% { opacity: 1; transform: scale(1); }
        50%       { opacity: .5; transform: scale(1.4); }
    }

    .metric-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); gap: .75rem; margin: .5rem 0 1rem; }
    .metric-card {
        background: var(--metric-bg, #fff);
        border: 1px solid var(--border, rgba(249,115,22,.12));
        border-radius: var(--radius-md); padding: .9rem 1rem;
        transition: var(--transition); box-shadow: var(--shadow-sm);
        position: relative; overflow: hidden; text-align: center;
    }
    .metric-card::before {
        content: ""; position: absolute; top: 0; left: 0;
        width: 100%; height: 3px;
        background: linear-gradient(90deg, #F97316, #EF4444);
        opacity: 0; transition: opacity .2s;
    }
    .metric-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
    .metric-card:hover::before { opacity: 1; }
    .metric-label { font-size: .68rem; font-weight: 600; text-transform: uppercase; letter-spacing: .8px; color: var(--text-muted, #6B7280); margin: 0 0 .3rem; }
    .metric-value { font-size: 1.45rem; font-weight: 700; color: var(--text-primary, #111827); line-height: 1; font-family: 'DM Mono', monospace; }
    .metric-delta { font-size: .7rem; color: #EAB308; margin-top: .2rem; font-weight: 500; }

    .section-header { display: flex; align-items: center; gap: .6rem; margin: 1.5rem 0 .8rem; }
    .section-header-line { flex: 1; height: 1px; background: linear-gradient(90deg, #F97316, transparent); opacity: .35; }
    .section-title { font-size: .78rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1.2px; color: #F97316; white-space: nowrap; text-align: center; }

    .static-team-header {
        text-align: center; margin: 2rem 0 .6rem; font-size: 1rem; font-weight: 700;
        text-transform: uppercase; letter-spacing: 1px; color: #F97316;
        display: flex; align-items: center; justify-content: center; gap: .75rem;
    }
    .static-team-header::before, .static-team-header::after {
        content: ""; flex: 1; max-width: 120px; height: 1px;
        background: linear-gradient(90deg, transparent, #F97316); opacity: .4;
    }
    .static-team-header::after { background: linear-gradient(90deg, #F97316, transparent); }

    .insight-card {
        background: var(--metric-bg, #fff);
        border: 1px solid var(--border, rgba(249,115,22,.12));
        border-radius: var(--radius-md); padding: 1rem 1.1rem;
        margin-bottom: .6rem; box-shadow: var(--shadow-sm); transition: var(--transition);
    }
    .insight-card:hover { box-shadow: var(--shadow-md); }
    .insight-card.good  { border-left: 3px solid #EAB308; }
    .insight-card.warn  { border-left: 3px solid #FBBF24; }
    .insight-card.bad   { border-left: 3px solid #EF4444; }
    .insight-card.info  { border-left: 3px solid #F97316; }
    .insight-icon { font-size: 1.1rem; }
    .insight-title { font-size: .82rem; font-weight: 700; color: var(--text-primary, #111827); margin: .2rem 0; text-align: center; }
    .insight-body { font-size: .76rem; color: var(--text-muted, #6B7280); line-height: 1.5; text-align: center; }

    [data-testid="stTabs"] [role="tablist"] { gap: .3rem; border-bottom: 1px solid var(--border, rgba(249,115,22,.12)); padding-bottom: 0; }
    [data-testid="stTabs"] button[role="tab"] {
        font-family: 'DM Sans', sans-serif !important; font-size: .82rem !important;
        font-weight: 600 !important; letter-spacing: .3px;
        border-radius: var(--radius-sm) var(--radius-sm) 0 0;
        padding: .55rem 1.1rem !important; transition: var(--transition);
    }

    div[data-testid="stDataFrame"] thead tr th {
        background: linear-gradient(135deg, #431407, #7c1d1d) !important;
        color: #fff !important; font-family: 'DM Sans', sans-serif !important;
        font-size: .72rem !important; font-weight: 700 !important; letter-spacing: .6px;
        text-transform: uppercase; white-space: normal !important; word-wrap: break-word !important;
        text-align: center !important; vertical-align: middle !important;
        min-width: 100px !important; padding: 10px !important;
    }

    [data-testid="stSidebar"] .stButton>button {
        width: 100%; font-family: 'DM Sans', sans-serif !important;
        font-weight: 600 !important; font-size: .82rem !important;
        border-radius: var(--radius-sm); transition: var(--transition);
        background: linear-gradient(135deg, #431407, #7c1d1d) !important;
        color: #fff !important; border: none !important;
    }
    [data-testid="stSidebar"] .stButton>button:hover {
        opacity: .88 !important; transform: translateY(-1px) !important;
    }

    .stDownloadButton>button {
        background: linear-gradient(135deg, #431407, #7c1d1d) !important;
        color: #fff !important; border: none !important;
        border-radius: var(--radius-sm) !important;
        font-family: 'DM Sans', sans-serif !important;
        font-size: .78rem !important; font-weight: 600 !important;
        transition: var(--transition) !important;
    }
    .stDownloadButton>button:hover { opacity: .88; transform: translateY(-1px); }

    hr { border-color: var(--border, rgba(249,115,22,.12)) !important; margin: 1.2rem 0 !important; }

    .brand-name {
        font-size: .85rem;
        font-weight: 700;
        letter-spacing: -.3px;
        color: #F97316;
    }

    .brand-tagline {
        font-size: .58rem;
        letter-spacing: .8px;
        font-family: monospace;
        margin-bottom: .9rem;
        color: #EA580C;
    }
    .js-plotly-plot { border-radius: var(--radius-md); overflow: hidden; }

    .kpi-pill { display: inline-flex; align-items: center; gap: 4px; padding: 2px 9px; border-radius: 20px; font-size: .7rem; font-weight: 600; font-family: 'DM Mono', monospace; }
    .kpi-pill.green  { background: rgba(234,179,8,.15);   color: #CA8A04; }
    .kpi-pill.amber  { background: rgba(251,191,36,.15);  color: #D97706; }
    .kpi-pill.red    { background: rgba(239,68,68,.15);   color: #DC2626; }
    .kpi-pill.blue   { background: rgba(249,115,22,.15);  color: #EA580C; }
    </style>
    """, unsafe_allow_html=True)


    # ─────────────────────────────────────────────
    # GLOBAL HELPER FUNCTIONS
    # ─────────────────────────────────────────────

    def style_total(row):
        if row["CALLER"] == "TOTAL":
            return ['font-weight: bold; background-color: #374151; color: #FFFFFF;'] * len(row)
        return [''] * len(row)

    def style_static(row):
        if row["CALLER"] == "TOTAL":
            return ['font-weight: bold; background-color: #374151; color: #FFFFFF;'] * len(row)
        return [''] * len(row)

    def format_dur_hm(total_seconds):
        if pd.isna(total_seconds) or total_seconds <= 0: return "0h 0m"
        tm = int(round(total_seconds / 60))
        return f"{tm // 60}h {tm % 60}m"

    def get_display_gap_seconds(start_time, end_time):
        if pd.isna(start_time) or pd.isna(end_time): return 0
        s = start_time.replace(second=0, microsecond=0)
        e = end_time.replace(second=0, microsecond=0)
        return (e - s).total_seconds()

    def section_header(label):
        st.markdown(f"""
        <div class="section-header">
            <div class="section-header-line"></div>
            <span class="section-title">{label}</span>
            <div class="section-header-line" style="background:linear-gradient(90deg,transparent,#F97316)"></div>
        </div>""", unsafe_allow_html=True)

    def _unique_approvals(series):
        seen = {}
        for v in series.dropna().astype(str):
            v = v.strip()
            k = v.lower()
            if k and k not in seen:
                seen[k] = v
        return ", ".join(seen.values()) if seen else "—"

    def style_team_manual_total(row):
        if row.get('TEAM') == 'TOTAL':
            return ['font-weight:bold;background-color:#374151;color:#FFFFFF;'] * len(row)
        return [''] * len(row)
    # ─────────────────────────────────────────────
    # DATA FETCHING
    # ─────────────────────────────────────────────

    @st.cache_data(ttl=120, show_spinner=False)
    def get_metadata():
        df_meta = pd.read_csv(CSV_URL)
        df_meta.columns = df_meta.columns.str.strip()
        df_meta['merge_key'] = df_meta['Caller Name'].str.strip().str.lower()
        teams = sorted(df_meta['Team Name'].dropna().unique())
        verticals = sorted(df_meta['Vertical'].dropna().unique())
        return teams, verticals, df_meta

    @st.cache_data(ttl=600, show_spinner=False)
    def get_global_last_update():
        query = """
        WITH combined AS (
            SELECT updated_at, updated_at_ampm FROM `studious-apex-488820-c3.crm_dashboard.acefone_calls`
            UNION ALL
            SELECT StartTime as updated_at, updated_at_ampm FROM `studious-apex-488820-c3.crm_dashboard.ozonetel_calls`
        )
        SELECT updated_at_ampm FROM combined WHERE updated_at IS NOT NULL ORDER BY updated_at DESC LIMIT 1
        """
        try:
            res = client.query(query).to_dataframe()
            return str(res['updated_at_ampm'].iloc[0]) if not res.empty else "N/A"
        except:
            return "N/A"

    @st.cache_data(ttl=600, show_spinner=False)
    def get_available_dates():
        query = """
        SELECT MIN(min_d) as min_date, MAX(max_d) as max_date FROM (
            SELECT MIN(`Call Date`) as min_d, MAX(`Call Date`) as max_d
            FROM `studious-apex-488820-c3.crm_dashboard.acefone_calls`
            UNION ALL
            SELECT MIN(CallDate) as min_d, MAX(CallDate) as max_d
            FROM `studious-apex-488820-c3.crm_dashboard.ozonetel_calls`
        )
        """
        df_dates = client.query(query).to_dataframe()
        if not df_dates.empty and not pd.isna(df_dates['min_date'].iloc[0]):
            return df_dates['min_date'].iloc[0], df_dates['max_date'].iloc[0]
        return date.today(), date.today()

    @st.cache_data(ttl=120, show_spinner=False)
    def fetch_call_data(start_date, end_date):
        q_ace = f"SELECT * FROM `studious-apex-488820-c3.crm_dashboard.acefone_calls` WHERE `Call Date` BETWEEN '{start_date}' AND '{end_date}'"
        df_ace = client.query(q_ace).to_dataframe()
        if not df_ace.empty:
            df_ace['source'] = 'Acefone'
            df_ace['unique_lead_id'] = df_ace['client_number']

        q_ozo = f"SELECT * FROM `studious-apex-488820-c3.crm_dashboard.ozonetel_calls` WHERE CallDate BETWEEN '{start_date}' AND '{end_date}'"
        df_ozo = client.query(q_ozo).to_dataframe()
        if not df_ozo.empty:
            df_ozo['unique_lead_id'] = df_ozo['phone_number']
            df_ozo = df_ozo.rename(columns={
                'CallID': 'call_id', 'AgentName': 'call_owner', 'phone_number': 'client_number',
                'StartTime': 'call_datetime', 'CallDate': 'Call Date', 'duration_sec': 'call_duration',
                'Status': 'status', 'Type': 'direction', 'Disposition': 'reason'
            })
            df_ozo['status'] = df_ozo['status'].str.lower().replace({'unanswered': 'missed'})
            df_ozo['direction'] = df_ozo['direction'].str.lower().replace({'manual': 'outbound'})
            df_ozo['source'] = 'Ozonetel'

        q_man = f"SELECT * FROM `studious-apex-488820-c3.crm_dashboard.manual_calls` WHERE Call_Date BETWEEN '{start_date}' AND '{end_date}'"
        df_man = client.query(q_man).to_dataframe()
        if not df_man.empty:
            df_man['unique_lead_id'] = df_man['client_number']
            df_man = df_man.rename(columns={'Call_Date': 'Call Date', 'Approved_By': 'reason'})
            df_man['status'] = 'answered'
            df_man['direction'] = 'outbound'
            df_man['source'] = 'Manual'
            df_man['call_datetime'] = pd.NaT

        df = pd.concat([df_ace, df_ozo, df_man], ignore_index=True)
        if not df.empty:
            df['call_endtime'] = pd.to_datetime(df['call_datetime'], utc=True).dt.tz_convert('Asia/Kolkata')
            df['call_duration'] = pd.to_numeric(df['call_duration'], errors='coerce').fillna(0)
            df['call_starttime'] = df['call_endtime'] - pd.to_timedelta(df['call_duration'], unit='s')

            ozo_mask = df['source'] == 'Ozonetel'
            df.loc[ozo_mask, 'call_starttime'] = df.loc[ozo_mask, 'call_endtime']
            df.loc[ozo_mask, 'call_endtime']   = (
                df.loc[ozo_mask, 'call_starttime']
                + pd.to_timedelta(df.loc[ozo_mask, 'call_duration'], unit='s')
            )

            df['call_starttime_clean'] = df['call_starttime'].dt.tz_localize(None)
            df['call_endtime_clean']   = df['call_endtime'].dt.tz_localize(None)
        return df


    # ─────────────────────────────────────────────
    # CORE METRICS PROCESSING
    # ─────────────────────────────────────────────

    def process_metrics_logic(df_filtered):
        agents_list = []
        total_duration_agg = 0
        ist_tz = pytz.timezone("Asia/Kolkata")

        for owner, agent_group in df_filtered.groupby('call_owner'):
            total_ans, total_miss, total_calls = 0, 0, 0
            total_above_3min, total_mid_calls, total_long_calls, agent_valid_dur = 0, 0, 0, 0
            total_break_sec_all_days, total_active_days = 0, 0
            daily_io_list, daily_break_list, all_issues = [], [], []

            for c_date, day_group in agent_group.groupby('Call Date'):
                timed_group = day_group[day_group['call_starttime'].notna()].sort_values('call_starttime')
                total_active_days += 1
                ans  = len(day_group[day_group['status'].str.lower() == 'answered'])
                miss = len(day_group[day_group['status'].str.lower() == 'missed'])
                total_ans  += ans
                total_miss += miss
                total_calls += len(day_group)

                total_above_3min  += len(day_group[day_group['call_duration'] >= 180])
                total_mid_calls   += len(day_group[(day_group['call_duration'] >= 900) & (day_group['call_duration'] < 1200)])
                total_long_calls  += len(day_group[day_group['call_duration'] >= 1200])
                day_dur = day_group.loc[day_group['call_duration'] >= 180, 'call_duration'].sum()
                agent_valid_dur += day_dur

                if timed_group.empty: continue

                first_call_start = timed_group['call_starttime'].min()
                last_call_end    = timed_group['call_endtime'].max()
                daily_io_list.append(
                    f"{c_date.strftime('%d/%m')}: In {first_call_start.strftime('%I:%M %p')} · Out {last_call_end.strftime('%I:%M %p')}"
                )

                start_office = ist_tz.localize(datetime.combine(c_date, time(10, 0)))
                end_office   = ist_tz.localize(datetime.combine(c_date, time(20, 0)))

                if first_call_start > ist_tz.localize(datetime.combine(c_date, time(10, 15))): all_issues.append("Late Check-In")
                if last_call_end   < end_office: all_issues.append("Early Check-Out")

                day_breaks, day_break_sec = [], 0

                if first_call_start > start_office:
                    g = get_display_gap_seconds(start_office, first_call_start)
                    if g >= 900:
                        day_breaks.append({'s': start_office, 'e': first_call_start, 'dur': g})
                        day_break_sec += g

                if len(timed_group) > 1:
                    for i in range(len(timed_group) - 1):
                        cur_end   = timed_group['call_endtime'].iloc[i]
                        nxt_start = timed_group['call_starttime'].iloc[i + 1]
                        act_s = max(cur_end,   start_office)
                        act_e = min(nxt_start, end_office)
                        if act_e > act_s:
                            g = get_display_gap_seconds(act_s, act_e)
                            if g >= 900:
                                day_breaks.append({'s': act_s, 'e': act_e, 'dur': g})
                                day_break_sec += g

                if last_call_end < end_office:
                    g = get_display_gap_seconds(last_call_end, end_office)
                    if g >= 900:
                        day_breaks.append({'s': last_call_end, 'e': end_office, 'dur': g})
                        day_break_sec += g

                total_break_sec_all_days += day_break_sec
                if day_breaks:
                    b_str = f"{c_date.strftime('%d/%m')}: {len(day_breaks)} breaks : {format_dur_hm(day_break_sec)}"
                    for b in day_breaks:
                        b_str += f"\n  {b['s'].strftime('%I:%M %p')}→{b['e'].strftime('%I:%M %p')} ({format_dur_hm(b['dur'])})"
                    daily_break_list.append(b_str)

                day_prod_sec = 36000 - day_break_sec
                if len(day_group[day_group['call_duration'] >= 180]) < 40: all_issues.append("Low Calls")
                if day_dur < 11700:    all_issues.append("Low Duration")
                if len(day_breaks) > 2: all_issues.append("Excessive Breaks")
                if day_prod_sec < 18000: all_issues.append("Less Productive")

            total_duration_agg += agent_valid_dur
            prod_sec_total = (36000 * total_active_days) - total_break_sec_all_days

            agents_list.append({
                "IN/OUT TIME": "\n".join(daily_io_list),
                "CALLER": owner,
                "TEAM": agent_group['Team Name'].iloc[0] if not pd.isna(agent_group['Team Name'].iloc[0]) else "Others",
                "TOTAL CALLS": int(total_calls),
                "CALL STATUS": f"{total_ans} Ans / {total_miss} Unans",
                "PICK UP RATIO %": f"{round((total_ans / total_calls * 100)) if total_calls > 0 else 0}%",
                "CALLS > 3 MINS": int(total_above_3min),
                "CALLS 15-20 MINS": int(total_mid_calls),
                "20+ MIN CALLS": int(total_long_calls),
                "CALL DURATION > 3 MINS": format_dur_hm(agent_valid_dur),
                "PRODUCTIVE HOURS": format_dur_hm(prod_sec_total),
                "BREAKS (>=15 MINS)": "\n---\n".join(daily_break_list) if daily_break_list else "0",
                "REMARKS": ", ".join(sorted(set(all_issues))) if all_issues else "None",
                "raw_prod_sec": prod_sec_total,
                "raw_dur_sec": agent_valid_dur,
            })

        return pd.DataFrame(agents_list), total_duration_agg


    # ─────────────────────────────────────────────
    # INSIGHTS COMPUTATION
    # ─────────────────────────────────────────────

    def compute_team_insights(df_merged, report_df):
        insights = []
        if df_merged.empty or report_df.empty:
            return insights

        team_dur = report_df.groupby("TEAM")["raw_dur_sec"].mean().sort_values(ascending=False)
        if len(team_dur) >= 1:
            top_team = team_dur.index[0]
            top_val  = format_dur_hm(team_dur.iloc[0])
            insights.append({
                "type": "good", "icon": "🏆",
                "title": f"Top Team by Avg Call Duration: {top_team}",
                "body": f"Averaging {top_val} of qualifying call duration per agent — highest across all teams."
            })

        exclude_teams = ['Others', 'CD - Community Manager', 'CD - Community', 'Criminal - Community Manager',
                         'Criminal - Community', 'ID - Community Manager', 'ID - Community',
                         'Clerkship community', 'Women ai - Community']

        manual_df = df_merged[(df_merged['source'] == 'Manual') & (~df_merged['Team Name'].isin(exclude_teams))]
        if not manual_df.empty:
            man_counts = manual_df.groupby('Team Name').agg(
                total_manual=('source', 'count'),
                unique_agents=('call_owner', 'nunique')
            ).sort_values('total_manual', ascending=False)
            if not man_counts.empty:
                top_man_team = man_counts.index[0]
                insights.append({
                    "type": "bad", "icon": "⚠️",
                    "title": f"Focus Required: {top_man_team} (Highest manual calls)",
                    "body": f"Total {int(man_counts.iloc[0]['total_manual'])} Manual Calls are getting dialled by {int(man_counts.iloc[0]['unique_agents'])} agents."
                })

        df_merged['_ans'] = df_merged['status'].str.lower() == 'answered'
        pur = df_merged.groupby('Team Name')['_ans'].mean().mul(100).round(1)
        best_pur  = pur.idxmax()
        worst_pur = pur.idxmin()
        if best_pur != worst_pur:
            insights.append({
                "type": "info", "icon": "🔔",
                "title": f"Pick-Up Ratio Spread: {best_pur} vs {worst_pur}",
                "body": (f"{best_pur} leads at {pur[best_pur]}% answer rate. "
                         f"{worst_pur} trails at {pur[worst_pur]}%. "
                         f"Gap of {round(pur[best_pur]-pur[worst_pur],1)} pp — review missed-call handling in {worst_pur}.")
            })

        long_rate = report_df.groupby("TEAM").apply(
            lambda g: g["20+ MIN CALLS"].sum() / g["TOTAL CALLS"].sum() * 100
            if g["TOTAL CALLS"].sum() > 0 else 0
        ).round(2)
        if not long_rate.empty:
            best_long = long_rate.idxmax()
            insights.append({
                "type": "good", "icon": "💬",
                "title": f"Highest Deep-Engagement Rate: {best_long}",
                "body": (f"{long_rate[best_long]}% of calls in {best_long} exceed 20 minutes — "
                         f"a strong signal of qualified prospect conversations. Replicate best practices across other teams.")
            })

        break_df = report_df[~report_df["TEAM"].isin(exclude_teams)]
        remarks_series = break_df["REMARKS"].str.contains("Excessive Breaks", na=False)
        if remarks_series.sum() > 0:
            b_teams = break_df.loc[remarks_series, "TEAM"].value_counts().idxmax()
            b_count = remarks_series.sum()
            insights.append({
                "type": "warn", "icon": "⏸️",
                "title": f"Break Discipline Alert — {b_teams}",
                "body": f"{b_count} agent(s) flagged for excessive breaks (>2 breaks ≥15 min/day). Heaviest cluster in {b_teams}."
            })

        prod_df = report_df[~report_df["TEAM"].isin(exclude_teams)]
        if not prod_df.empty:
            team_avg_prod = prod_df.groupby("TEAM")["raw_prod_sec"].mean().sort_values()
            if not team_avg_prod.empty:
                worst_prod_team = team_avg_prod.index[0]
                agent_count = len(prod_df[prod_df["TEAM"] == worst_prod_team])
                insights.append({
                    "type": "bad", "icon": "⏱️",
                    "title": f"Focus Required: Lowest Productive Hours: {worst_prod_team}",
                    "body": f"{agent_count} agents on {worst_prod_team} team have the least average productive hours as compared to other teams."
                })

        return insights
    @st.cache_data(show_spinner=False)
    def generate_calling_helper_pdf_bytes() -> bytes:
        
        buffer = io.BytesIO()
        ORANGE_DARK=colors.HexColor("#7c2d12"); ORANGE_MID=colors.HexColor("#431407")
        ORANGE_PALE=colors.HexColor("#FEF3E8"); ORANGE_ROW=colors.HexColor("#FFF8F3")
        GREY_DARK=colors.HexColor("#374151"); GREY_MID=colors.HexColor("#6B7280")
        WHITE=colors.white; BLACK=colors.HexColor("#111827"); W,H=A4

        def s(name,**kw):
            d=dict(fontName='Helvetica',fontSize=9,textColor=BLACK,spaceAfter=3,leading=14)
            d.update(kw); return ParagraphStyle(name,**d)

        S={'body':s('body'),
           'label':s('label',fontName='Helvetica-Bold',fontSize=8,textColor=ORANGE_DARK,spaceAfter=1),
           'formula':s('formula',fontName='Helvetica-Oblique',fontSize=8.5,textColor=colors.HexColor("#7c2d12"),backColor=ORANGE_PALE,leftIndent=8,rightIndent=8),
           'footer':s('footer',fontSize=7.5,textColor=GREY_MID,alignment=TA_CENTER)}

        class CoverBlock(Flowable):
            def __init__(self,w): Flowable.__init__(self); self.w=w; self.height=90
            def draw(self):
                c=self.canv
                c.setFillColor(ORANGE_DARK); c.rect(0,52,self.w,38,fill=1,stroke=0)
                c.setFillColor(ORANGE_MID);  c.rect(0,22,self.w,30,fill=1,stroke=0)
                c.setFillColor(colors.HexColor("#1c0700")); c.rect(0,0,self.w,22,fill=1,stroke=0)
                c.setFillColor(WHITE); c.setFont("Helvetica-Bold",22)
                c.drawCentredString(self.w/2,66,"CALLING METRICS DASHBOARD")
                c.setFillColor(colors.HexColor("#FBBF24")); c.setFont("Helvetica-Bold",11)
                c.drawCentredString(self.w/2,34,"Logic & Metric Reference Guide")
                c.setFillColor(colors.HexColor("#FDE68A")); c.setFont("Helvetica",8.5)
                c.drawCentredString(self.w/2,8,"LawSikho & Skill Arbitrage  \u00b7  Sales & Operations Team  \u00b7  Internal Use Only")

        class SectionBanner(Flowable):
            def __init__(self,icon,title,color=None,w=None):
                Flowable.__init__(self); self.icon=icon; self.title=title
                self.color=color or ORANGE_DARK; self.w=w or (W-30*mm); self.height=22
            def draw(self):
                c=self.canv; c.setFillColor(self.color)
                c.roundRect(0,0,self.w,self.height,4,fill=1,stroke=0)
                c.setFillColor(WHITE); c.setFont("Helvetica-Bold",11)
                c.drawString(10,6,f"{self.icon}  {self.title}")

        def btable(rows,cw=None):
            cw=cw or [44*mm,116*mm]
            data=[[Paragraph(f"<b>{r[0]}</b>",S['label']),Paragraph(r[1],S['body'])] for r in rows]
            t=Table(data,colWidths=cw,hAlign='LEFT')
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(0,-1),ORANGE_PALE),('VALIGN',(0,0),(-1,-1),'TOP'),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor("#FDE68A")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[WHITE,ORANGE_ROW]),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            return t

        def ltable(rows):
            data=[[Paragraph(f"<b>{r[0]}</b>",S['label']),Paragraph(r[1],S['formula'])] for r in rows]
            t=Table(data,colWidths=[52*mm,108*mm],hAlign='LEFT')
            t.setStyle(TableStyle([
                ('VALIGN',(0,0),(-1,-1),'TOP'),('GRID',(0,0),(-1,-1),0.3,colors.HexColor("#FDE68A")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[WHITE,ORANGE_ROW]),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            return t

        SP=Spacer; BAN=SectionBanner; cw=W-30*mm
        HR=lambda: HRFlowable(width="100%",thickness=0.6,color=colors.HexColor("#FBBF24"),spaceAfter=6,spaceBefore=4)

        story=[
            SP(1,18*mm),CoverBlock(cw),SP(1,10*mm),
            Paragraph("This document explains every metric, table, tab, and column in the Calling Metrics Dashboard. Use it as a quick reference to understand how each number is calculated and what each section means.",S['body']),
            SP(1,6*mm),
            BAN("📋","SECTION 1 — DASHBOARD TABS OVERVIEW"),SP(1,3*mm),
            Paragraph("The dashboard has three tabs. Each serves a different purpose.",S['body']),SP(1,2*mm),
            btable([
                ("🚀 Dynamic Dashboard","The main report tab. Shows agent-level performance for the selected date range — Top 3 highlights, Summary KPIs, and the full Agent Performance Table. Click 'Generate Dynamic Report' in the sidebar to load it."),
                ("📅 Duration Report","A team-by-team breakdown showing call duration metrics only. Each team gets its own table sorted by call duration. TLs and ATLs are grouped separately at the bottom. Useful for sharing per-team performance without exposing cross-team data. Click 'Generate Duration Report' in the sidebar."),
                ("🧠 Insights & Leaderboard","Auto-populated from whichever report was generated last. Shows 5-6 auto-generated team insights and a Team Leaderboard table. No separate button needed — switch to this tab after generating any report."),
            ]),SP(1,6*mm),
            BAN("🏆","SECTION 2 — TOP 3 PERFORMANCE HIGHLIGHTS"),SP(1,3*mm),
            Paragraph("Three highlight cards at the top of the Dynamic Dashboard. Each picks the single best agent in one dimension.",S['body']),SP(1,2*mm),
            btable([
                ("🥇 Top Performer","Agent with the highest total Call Duration for calls above 3 minutes (raw_dur_sec). Agents are sorted by qualifying call duration descending — the top row wins this card."),
                ("✆ Highest Calls","Agent with the highest Total Calls count across all statuses (answered + missed). Sorted separately — a different agent may win this card."),
                ("🗣️ Deep Engagement","Agent with the most calls lasting 20 minutes or more (duration >= 1200 seconds). Signals high-quality prospect conversations."),
            ]),SP(1,6*mm),
            BAN("📊","SECTION 3 — SUMMARY METRICS (KPI CARDS)"),SP(1,3*mm),
            Paragraph("Eight KPI cards shown below the Top 3 highlights. Aggregate numbers across ALL agents and sources for the selected date range.",S['body']),SP(1,2*mm),
            ltable([
                ("Total Calls","Count of all call rows across Acefone + Ozonetel + Manual after filters applied."),
                ("Acefone Calls","Count of rows where source = 'Acefone'."),
                ("Ozonetel Calls","Count of rows where source = 'Ozonetel'."),
                ("Manual Calls","Count of rows where source = 'Manual'. Calls logged manually by agents and approved — not system-generated CDR."),
                ("Unique Leads","Count of distinct phone numbers across all sources. One lead dialled multiple times still counts as 1 unique lead."),
                ("Pick-Up Ratio","Answered Calls / Total Calls x 100, rounded to nearest whole percent. Answered = rows where status = 'answered'."),
                ("Active Callers","Count of distinct agents in the report after all filters. Only agents with at least one call in the date range are counted."),
                ("Avg Prod Hrs","Average Productive Hours across all active agents. Productive Hours = (10 hrs x active days) - total break time >= 15 mins. Shown as Xh Ym."),
            ]),SP(1,6*mm),
            BAN("📋","SECTION 4 — AGENT PERFORMANCE TABLE"),SP(1,3*mm),
            Paragraph("Main data table in the Dynamic Dashboard. One row per active agent, sorted by Call Duration > 3 Mins descending. Top 3 rows get medal emojis. A bold TOTAL row is appended at the bottom.",S['body']),SP(1,2*mm),
            btable([
                ("Rank","Medal emoji for top 3 agents by call duration: Gold, Silver, Bronze."),
                ("IN/OUT TIME","First call start (In) and last call end (Out) per day. Format: DD/MM: In HH:MM AM/PM . Out HH:MM AM/PM. Derived from call_starttime and call_endtime after IST conversion."),
                ("CALLER","Agent name from the team sheet, normalised via lowercase merge key. Falls back to raw system name if not matched."),
                ("TEAM","Team name from the team sheet. Shows 'Others' if the agent is unmatched."),
                ("TOTAL CALLS","All calls for this agent in the date range — all statuses combined."),
                ("CALL STATUS","'X Ans / Y Unans' — count of answered and missed/unanswered calls for this agent."),
                ("PICK UP RATIO %","Answered / Total Calls x 100 for this agent, rounded to nearest whole percent."),
                ("CALLS > 3 MINS","Count of calls where duration >= 180 seconds. Qualifying threshold for duration metrics."),
                ("CALLS 15-20 MINS","Count of calls where duration >= 900 seconds AND < 1200 seconds. Mid-range engagement."),
                ("20+ MIN CALLS","Count of calls where duration >= 1200 seconds. Deep engagement indicator."),
                ("CALL DURATION > 3 MINS","Sum of duration for calls >= 3 minutes, shown as Xh Ym."),
                ("PRODUCTIVE HOURS","(10 hrs x active days) - total break time >= 15 mins. Shown as Xh Ym."),
                ("BREAKS (>=15 MINS)","Gaps between calls of 15+ minutes shown per day with time ranges. Gaps < 15 mins are ignored."),
                ("REMARKS","Auto-flagged issues: Late Check-In (first call after 10:15 AM), Early Check-Out (last call before 8 PM), Low Calls (<40 qualifying/day), Low Duration (<3h 15m/day), Excessive Breaks (>2 breaks >= 15 mins/day), Less Productive (<5 hrs productive/day)."),
            ],cw=[46*mm,114*mm]),SP(1,6*mm),
            BAN("📅","SECTION 5 — DURATION REPORT TAB"),SP(1,3*mm),
            Paragraph("Generates a simplified shareable table per team showing duration columns only — no break details or remarks.",S['body']),SP(1,2*mm),
            btable([
                ("Team separation","Each team gets its own section and table. Teams with zero qualifying duration are skipped entirely."),
                ("TL / ATL separation","Agents flagged as TL, ATL, AD, Team Lead, or Team Leader in the team sheet are shown in a single 'TL Duration Report' section at the bottom. Only TLs with > 5 mins qualifying duration are included."),
                ("Columns shown","CALLER, TOTAL CALLS, CALL STATUS, PICK UP RATIO %, CALLS > 3 MINS, CALLS 15-20 MINS, 20+ MIN CALLS, CALL DURATION > 3 MINS. Same definitions as Section 4."),
                ("Sorting","Within each team table, agents are sorted by Call Duration > 3 Mins descending."),
                ("TOTAL row","Each team table has a TOTAL row summing calls and duration."),
                ("CDR per team","Each team section has its own Download CDR button exporting only that team's raw records."),
            ]),SP(1,6*mm),
            BAN("🏅","SECTION 6 — TEAM LEADERBOARD (INSIGHTS TAB)"),SP(1,3*mm),
            Paragraph("Appears in the Insights tab. Aggregates all agents by team, ranked by total call duration descending. Only shown when no Team or Name filter is active.",S['body']),SP(1,2*mm),
            ltable([
                ("Team","Team name from the team sheet."),
                ("Agents","Count of distinct agents from this team in the report."),
                ("Total Calls","Sum of all calls across all agents in this team."),
                ("Total Dur (h)","Sum of qualifying call duration (>3 min) in hours, 1 decimal place."),
                ("Avg Dur/Agent (h)","Total Duration divided by agent count for this team, in hours."),
                ("Avg Prod Hrs (h)","Average productive hours per agent in this team, in hours."),
                ("20+ Min Calls","Sum of 20+ minute calls across all agents in this team."),
                ("Medal","Gold, Silver, Bronze for the top 3 teams by Total Duration."),
            ]),SP(1,6*mm),
            BAN("📥","SECTION 7 — CDR DOWNLOAD COLUMNS"),SP(1,3*mm),
            Paragraph("The Download CDR button exports a CSV of raw call detail records. Each row is one call.",S['body']),SP(1,2*mm),
            btable([
                ("client_number","Lead phone number. Unique lead identifier."),
                ("call_datetime","Original timestamp from the source system (UTC). For Ozonetel: call start. For Acefone: call end."),
                ("call_starttime_clean","Call start time in IST, timezone stripped. Used for break and productive hours calculations."),
                ("call_endtime_clean","Call end time in IST, timezone stripped. For Acefone: call_datetime. For Ozonetel: start + duration."),
                ("call_duration","Call duration in seconds. Zero-duration answered calls from Ozonetel are excluded before ingestion."),
                ("status","Call outcome: 'answered' or 'missed'. Ozonetel 'unanswered' mapped to 'missed'."),
                ("direction","Call direction: 'outbound' or 'inbound'. Ozonetel 'manual' mapped to 'outbound'."),
                ("service","Service or campaign name (Acefone only)."),
                ("reason","Disposition or reason code. For Manual calls this is the Approver name."),
                ("call_owner","Agent name after team sheet normalisation. Raw system name replaced by canonical Caller Name."),
                ("Call Date","Date of the call only. Used for day-level grouping."),
                ("updated_at_ampm","Timestamp when the record was last written to BigQuery, in AM/PM format."),
                ("Team Name","Team name from the team sheet, merged on lowercase agent name."),
                ("Vertical","Business vertical from the team sheet (e.g. Lawsikho, Skill Arbitrage)."),
                ("Analyst","Analyst name from the team sheet, if populated."),
                ("source","Data source: 'Acefone', 'Ozonetel', or 'Manual'."),
            ],cw=[46*mm,114*mm]),SP(1,6*mm),
            BAN("⚠️","SECTION 8 — HIGHEST MANUAL CALLS TABLE (DYNAMIC DASHBOARD)"),SP(1,3*mm),
            Paragraph("Appears below the CDR download button in the Dynamic Dashboard when manual calls exist in the selected date range. Shows agent-level manual call activity sorted by count descending. No medal ranking — manual calls are a flag, not an achievement.",S['body']),SP(1,2*mm),
            btable([
                ("CALLER",                "Agent name from the team sheet, same as the main performance table."),
                ("VERTICAL",              "Business vertical from the team sheet (e.g. Lawsikho, Skill Arbitrage)."),
                ("TEAM",                  "Team name from the team sheet."),
                ("MANUAL CALLS COUNT",    "Total number of manual call entries for this agent in the selected date range. Sorted descending — highest manual caller appears at top."),
                ("MANUAL CALLS DURATION", "Total duration of manual calls for this agent in Xh Ym format. Sourced from the call_duration column of the manual_calls BigQuery table."),
                ("APPROVED BY",           "Unique approver names from the Approved_By field, deduplicated case-insensitively (e.g. 'John', 'john', 'JOHN' collapse to one entry). Multiple unique approvers are separated by ', '."),
                ("TOTAL row",             "Bottom row summing Manual Calls Count and Duration across all agents shown."),
            ]),SP(1,6*mm),

            BAN("⚠️","SECTION 9 — TEAM MANUAL CALLS TABLE (INSIGHTS TAB)"),SP(1,3*mm),
            Paragraph("Appears at the bottom of the Insights & Leaderboard tab after generating any report. Shows the same manual call data aggregated by team instead of by agent. Sorted by Manual Calls Count descending.",S['body']),SP(1,2*mm),
            btable([
                ("VERTICAL",              "Business vertical from the team sheet."),
                ("TEAM",                  "Team name. The TOTAL row at the bottom sums across all teams."),
                ("MANUAL CALLS COUNT",    "Total manual calls across all agents in this team."),
                ("MANUAL CALLS DURATION", "Total manual call duration across all agents in this team, in Xh Ym format."),
                ("APPROVALS BY",          "All unique approver names across all agents in this team, deduplicated case-insensitively and joined by ', '."),
                ("Why no medals?",        "Manual calls indicate reliance on manual logging rather than system-dialled calls. High manual call counts may signal data quality gaps or agents bypassing the dialler. This table exists for monitoring, not ranking."),
            ]),SP(1,6*mm),

            BAN("📖","KEY TERMS GLOSSARY",color=GREY_DARK),SP(1,3*mm),
            btable([
                ("Qualifying Call","Any call with duration >= 180 seconds (3 minutes). Used for all duration metrics and performance flags."),
                ("Office Hours","10:00 AM to 8:00 PM IST (10 hours = 36,000 seconds). Reference window for breaks and productive hours."),
                ("Break","A gap between consecutive calls >= 900 seconds (15 minutes). Shorter gaps are ignored."),
                ("Productive Hours","(10 hrs x active days) - total break seconds. Expressed as Xh Ym."),
                ("Late Check-In","First call of the day starts after 10:15 AM IST."),
                ("Early Check-Out","Last call of the day ends before 8:00 PM IST."),
                ("Low Calls","Fewer than 40 qualifying calls (>3 min) in a single day."),
                ("Low Duration","Total qualifying duration < 3h 15m (11,700 seconds) in a single day."),
                ("Excessive Breaks","More than 2 breaks >= 15 minutes in a single day."),
                ("Less Productive","Productive seconds < 5 hours (18,000 seconds) in a single day."),
                ("Merge Key","Lowercase-stripped agent name used to join call data with the team sheet."),
                ("IST","Indian Standard Time (UTC+5:30). All timestamps are converted to IST for display and calculations."),
            ],cw=[46*mm,114*mm]),
            SP(1,8*mm),HR(),
            Paragraph("Designed by Amit Ray  \u00b7  amitray@lawsikho.com  \u00b7  For Internal Use of Sales and Operations Team Only. All Rights Reserved.",S['footer']),
        ]

        doc=SimpleDocTemplate(buffer,pagesize=A4,leftMargin=15*mm,rightMargin=15*mm,topMargin=14*mm,bottomMargin=14*mm,title="Calling Metrics — Logic Reference Guide",author="Amit Ray")
        doc.build(story)
        return buffer.getvalue()

    # ─────────────────────────────────────────────
    # SIDEBAR & UI
    # ─────────────────────────────────────────────

    # ─────────────────────────────────────────────
    # SIDEBAR — DATA + CONTROLS
    # ─────────────────────────────────────────────

    teams, verticals, df_team_mapping = get_metadata()
    min_date_raw, max_date_raw = get_available_dates()
    min_date = pd.Timestamp(min_date_raw).date()
    max_date = pd.Timestamp(max_date_raw).date()

    st.sidebar.markdown("""
    <div style='padding:.5rem 0 .4rem; text-align:center;'>
        <div style='font-size:.72rem; font-weight:700; text-transform:uppercase;
                    letter-spacing:1px; color:var(--text-muted,#6B7280);'>Report Controls</div>
    </div>
    """, unsafe_allow_html=True)

    date_range = st.sidebar.date_input(
        "📅 Date Range",
        value=(max_date, max_date),
        min_value=min_date,
        max_value=max_date,
        format="DD-MM-YYYY",
        key="call_date_range"
    )
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start_date, end_date = date_range
    else:
        start_date = end_date = date_range if not isinstance(date_range, tuple) else date_range[0]

    # ── Role-aware sidebar filters ──────────────────────────
    _role       = st.session_state.get('rf_role', 'admin')
    _rf_teams   = st.session_state.get('rf_teams', [])
    _rf_cname   = st.session_state.get('rf_caller_name', '')

    if _role == 'admin':
        selected_team     = st.sidebar.multiselect("👥 Filter by Team",     options=teams,     key="call_team_filter")
        selected_vertical = st.sidebar.multiselect("👑 Filter by Vertical", options=verticals, key="call_vert_filter")
        search_query      = st.sidebar.text_input("👤 Search Caller Name",                     key="call_search")
    elif _role == 'vertical_head':
        selected_team     = _rf_teams
        selected_vertical = []
        search_query      = st.sidebar.text_input("👤 Search Caller Name", key="call_search")
        st.sidebar.caption(f"🔒 Showing: {', '.join(_rf_teams)}")
    elif _role in ('tl', 'trainer'):
        selected_team     = _rf_teams
        selected_vertical = []
        search_query      = st.sidebar.text_input("👤 Search Caller Name", key="call_search")
        st.sidebar.caption(f"🔒 Team: {', '.join(_rf_teams)}")
    else:  # caller — only sees their own data
        selected_team     = []
        selected_vertical = []
        search_query      = _rf_cname
        st.sidebar.caption(f"👤 Viewing: {_rf_cname}")

    gen_dynamic = st.sidebar.button("🚀 Generate Dynamic Report",  key="call_gen_dynamic")
    gen_static  = st.sidebar.button("📅 Generate Duration Report", key="call_gen_static")

    st.sidebar.download_button(
        label="📖 Metrics Guide (PDF)",
        data=generate_calling_helper_pdf_bytes(),
        file_name="Calling_Metrics_Logic_Guide.pdf",
        mime="application/pdf",
        key="dl_calling_helper_pdf"
    )
    


    # ─────────────────────────────────────────────
    # HEADER BANNER
    # ─────────────────────────────────────────────

    last_update_str = get_global_last_update()
    display_start   = start_date.strftime('%d-%m-%Y')
    display_end     = end_date.strftime('%d-%m-%Y')

    st.markdown(f"""
    <div class="cw-header">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:.75rem;">
            <div>
                <div class="cw-title">🔔 CALLING METRICS</div>
                <div class="cw-subtitle">DURATION PERIOD&nbsp;·&nbsp; {display_start} to {display_end}</div>
            </div>
            <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.25rem;">
                <span class="cw-badge"><span class="cw-pulse"></span>OZONETEL &amp; ACEFONE</span>
                <span class="cw-badge">🕐 UPDATED AT: {last_update_str}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


    # ─────────────────────────────────────────────
    # TABS
    # ─────────────────────────────────────────────

    tab1, tab2, tab3 = st.tabs([
        "🚀 Dynamic Dashboard",
        "📅 Duration Report",
        "🧠 Insights & Leaderboard"
    ])


    # ══════════════════════════════════════════════
    # TAB 1 — DYNAMIC DASHBOARD
    # ══════════════════════════════════════════════

    with tab1:
        if gen_dynamic:
            with st.spinner("Calculating metrics…"):
                df_raw = fetch_call_data(start_date, end_date)
                if df_raw.empty:
                    st.warning("No data found for the selected period.")
                else:
                    df_raw['merge_key'] = df_raw['call_owner'].str.strip().str.lower()
                    df = pd.merge(df_raw, df_team_mapping, on='merge_key', how='left')
                    df['call_owner'] = df['Caller Name'].fillna(df['call_owner'])
                    df = df[df['call_owner'].notna() & (df['call_owner'] != '')]

                    if selected_team:     df = df[df['Team Name'].isin(selected_team)]
                    if selected_vertical: df = df[df['Vertical'].isin(selected_vertical)]
                    if search_query:      df = df[df['call_owner'].str.contains(search_query, case=False, na=False)]
                    if df.empty:
                        st.error("No results match the selected filters.")
                    else:
                        report_df, total_duration_agg = process_metrics_logic(df)
                        report_df = report_df.sort_values(by="raw_dur_sec", ascending=False)
                        report_df['Rank'] = ""
                        if len(report_df) > 0: report_df.iloc[0, report_df.columns.get_loc('Rank')] = "🥇"
                        if len(report_df) > 1: report_df.iloc[1, report_df.columns.get_loc('Rank')] = "🥈"
                        if len(report_df) > 2: report_df.iloc[2, report_df.columns.get_loc('Rank')] = "🥉"

                        # ── Store for Insights tab ──
                        st.session_state['insights_df']     = df.copy()
                        st.session_state['insights_report'] = report_df.copy()
                        st.session_state['insights_source'] = "Dynamic Report"

                        section_header("🏆 TOP 3 PERFORMANCE HIGHLIGHTS")
                        top_cols = st.columns(3)

                        top_dur = report_df.iloc[0]
                        with top_cols[0]:
                            st.markdown(f"""
                            <div class="metric-card" style="border-top: 3px solid var(--gold);">
                                <div class="metric-label">🥇 TOP PERFORMER</div>
                                <div class="metric-value" style="font-size:1.1rem;">{top_dur['CALLER']}</div>
                                <div class="metric-delta">{top_dur['CALL DURATION > 3 MINS']} Duration</div>
                            </div>""", unsafe_allow_html=True)

                        top_calls = report_df.sort_values('TOTAL CALLS', ascending=False).iloc[0]
                        with top_cols[1]:
                            st.markdown(f"""
                            <div class="metric-card" style="border-top: 3px solid #F97316;">
                                <div class="metric-label">✆ HIGHEST CALLS</div>
                                <div class="metric-value" style="font-size:1.1rem;">{top_calls['CALLER']}</div>
                                <div class="metric-delta">{top_calls['TOTAL CALLS']} Total Calls</div>
                            </div>""", unsafe_allow_html=True)

                        top_long = report_df.sort_values('20+ MIN CALLS', ascending=False).iloc[0]
                        with top_cols[2]:
                            st.markdown(f"""
                            <div class="metric-card" style="border-top: 3px solid var(--bronze);">
                                <div class="metric-label">🗣️ DEEP ENGAGEMENT</div>
                                <div class="metric-value" style="font-size:1.1rem;">{top_long['CALLER']}</div>
                                <div class="metric-delta">{top_long['20+ MIN CALLS']} Long Calls</div>
                            </div>""", unsafe_allow_html=True)

                        section_header("SUMMARY METRICS")
                        ans_t = len(df[df['status'].str.lower() == 'answered'])
                        pur_val = f"{round(ans_t / len(df) * 100)}%" if len(df) > 0 else "0%"
                        kpis = [
                            ("Total Calls",    len(df),                                         "📲"),
                            ("Acefone Calls",  len(df[df['source'] == 'Acefone']),              "🔵"),
                            ("Ozonetel Calls", len(df[df['source'] == 'Ozonetel']),             "🟠"),
                            ("Manual Calls",   len(df[df['source'] == 'Manual']),               "✏️"),
                            ("Unique Leads",   df['unique_lead_id'].nunique(),                  "👤"),
                            ("Pick-Up Ratio",  pur_val,                                         "✅"),
                            ("Active Callers", len(report_df),                                  "🎙️"),
                            ("Avg Prod Hrs",   format_dur_hm(report_df["raw_prod_sec"].mean()), "⏱"),
                        ]

                        cols = st.columns(len(kpis))
                        for col, (label, val, icon) in zip(cols, kpis):
                            with col:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">{icon} {label}</div>
                                    <div class="metric-value">{val}</div>
                                </div>""", unsafe_allow_html=True)

                        st.divider()
                        section_header("AGENT PERFORMANCE TABLE")

                        total_row = pd.DataFrame([{
                            "Rank": "",
                            "IN/OUT TIME": "-", "CALLER": "TOTAL", "TEAM": "-",
                            "TOTAL CALLS": int(report_df["TOTAL CALLS"].sum()),
                            "CALL STATUS": "-", "PICK UP RATIO %": "-",
                            "CALLS > 3 MINS": int(report_df["CALLS > 3 MINS"].sum()),
                            "CALLS 15-20 MINS": int(report_df["CALLS 15-20 MINS"].sum()),
                            "20+ MIN CALLS": int(report_df["20+ MIN CALLS"].sum()),
                            "CALL DURATION > 3 MINS": format_dur_hm(total_duration_agg),
                            "PRODUCTIVE HOURS": format_dur_hm(report_df["raw_prod_sec"].sum()),
                            "BREAKS (>=15 MINS)": "-", "REMARKS": "-"
                        }])

                        display_cols = [
                            "Rank", "IN/OUT TIME", "CALLER", "TEAM", "TOTAL CALLS", "CALL STATUS",
                            "PICK UP RATIO %", "CALLS > 3 MINS", "CALLS 15-20 MINS",
                            "20+ MIN CALLS", "CALL DURATION > 3 MINS",
                            "PRODUCTIVE HOURS", "BREAKS (>=15 MINS)", "REMARKS"
                        ]

                        final_df = pd.concat([report_df, total_row], ignore_index=True)
                        st.dataframe(
                            final_df.style.apply(style_total, axis=1)
                                          .set_properties(**{'white-space': 'pre-wrap'}),
                            column_order=display_cols,
                            use_container_width=True, hide_index=True
                        )

                        st.divider()
                        target_cols = [
                            "client_number", "call_datetime", "call_starttime_clean",
                            "call_endtime_clean", "call_duration", "status", "direction",
                            "service", "reason", "call_owner", "Call Date",
                            "updated_at_ampm", "Team Name", "Vertical", "Analyst", "source"
                        ]
                        existing_cols = [c for c in target_cols if c in df.columns]
                        st.download_button(
                            label="📥 Download CDR",
                            data=df[existing_cols].to_csv(index=False).encode('utf-8'),
                            file_name="CDR_LOG.csv", mime='text/csv'
                        )

                        # ── Manual Calls Table ──
                        manual_df_view = df[df['source'] == 'Manual'].copy()
                        if not manual_df_view.empty:
                            st.divider()
                            section_header("⚠️ HIGHEST MANUAL CALLS")

                            man_agg = (
                                manual_df_view.groupby('call_owner', sort=False)
                                .agg(
                                    Vertical  = ('Vertical',      'first'),
                                    Team      = ('Team Name',     'first'),
                                    Count     = ('source',        'count'),
                                    DurSec    = ('call_duration', 'sum'),
                                    Approvals = ('reason',        _unique_approvals),
                                )
                                .reset_index()
                                .sort_values('Count', ascending=False)
                                .reset_index(drop=True)
                            )

                            man_display = pd.DataFrame({
                                'CALLER'               : man_agg['call_owner'],
                                'VERTICAL'             : man_agg['Vertical'].fillna('—'),
                                'TEAM'                 : man_agg['Team'].fillna('—'),
                                'MANUAL CALLS COUNT'   : man_agg['Count'],
                                'MANUAL CALLS DURATION': man_agg['DurSec'].apply(format_dur_hm),
                                'APPROVED BY'          : man_agg['Approvals'],
                            })

                            total_man_row = pd.DataFrame([{
                                'CALLER'               : 'TOTAL',
                                'VERTICAL'             : '—',
                                'TEAM'                 : '—',
                                'MANUAL CALLS COUNT'   : int(man_agg['Count'].sum()),
                                'MANUAL CALLS DURATION': format_dur_hm(man_agg['DurSec'].sum()),
                                'APPROVED BY'          : '—',
                            }])

                            man_final = pd.concat([man_display, total_man_row], ignore_index=True)
                            st.dataframe(
                                man_final.style.apply(style_total, axis=1),
                                use_container_width=True, hide_index=True
                            )

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🚀</div>
                <div style='font-size:.9rem;font-weight:600;'>Select a date range and click <b>Generate Dynamic Report</b></div>
            </div>""", unsafe_allow_html=True)


    # ══════════════════════════════════════════════
    # TAB 2 — DURATION REPORT
    # ══════════════════════════════════════════════

    with tab2:
        if gen_static:
            with st.spinner("Building static layouts…"):
                df_raw = fetch_call_data(start_date, end_date)
                if df_raw.empty:
                    st.warning("No data found.")
                else:
                    df_raw['merge_key'] = df_raw['call_owner'].str.strip().str.lower()
                    df_static_master = pd.merge(df_raw, df_team_mapping, on='merge_key', how='left')
                    df_static_master['call_owner'] = df_static_master['Caller Name'].fillna(df_static_master['call_owner'])

                    if selected_team:     df_static_master = df_static_master[df_static_master['Team Name'].isin(selected_team)]
                    if selected_vertical: df_static_master = df_static_master[df_static_master['Vertical'].isin(selected_vertical)]
                    if search_query:      df_static_master = df_static_master[df_static_master['call_owner'].str.contains(search_query, case=False, na=False)]

                    if df_static_master.empty:
                        st.error("No results match filters.")
                    else:
                        tl_ad_mask = pd.Series(False, index=df_static_master.index)
                        for col in df_team_mapping.columns:
                            if col in df_static_master.columns:
                                clean_col = df_static_master[col].fillna('').astype(str).str.strip().str.upper()
                                tl_ad_mask |= clean_col.isin(['TL', 'ATL', 'AD', 'TEAM LEAD', 'TEAM LEADER'])

                        static_display_cols = [
                            "CALLER", "TOTAL CALLS", "CALL STATUS", "PICK UP RATIO %",
                            "CALLS > 3 MINS", "CALLS 15-20 MINS", "20+ MIN CALLS",
                            "CALL DURATION > 3 MINS"
                        ]

                        normal_team_data = df_static_master[~tl_ad_mask]
                        normal_teams     = sorted(normal_team_data['Team Name'].dropna().unique())

                        for team in normal_teams:
                            team_df = normal_team_data[normal_team_data['Team Name'] == team]
                            report_df, team_dur_agg_sec = process_metrics_logic(team_df)
                            if team_dur_agg_sec > 0:
                                report_df = report_df.sort_values(by="raw_dur_sec", ascending=False)
                                st.markdown(f"""
                                <div class="static-team-header">
                                    DURATION REPORT — {team.upper()} &nbsp;({display_start} to {display_end})
                                </div>""", unsafe_allow_html=True)

                                total_row = pd.DataFrame([{
                                    "CALLER": "TOTAL",
                                    "TOTAL CALLS": int(report_df["TOTAL CALLS"].sum()),
                                    "CALL STATUS": "-", "PICK UP RATIO %": "-",
                                    "CALLS > 3 MINS": int(report_df["CALLS > 3 MINS"].sum()),
                                    "CALLS 15-20 MINS": int(report_df["CALLS 15-20 MINS"].sum()),
                                    "20+ MIN CALLS": int(report_df["20+ MIN CALLS"].sum()),
                                    "CALL DURATION > 3 MINS": format_dur_hm(team_dur_agg_sec)
                                }])
                                final_team_df = pd.concat([report_df[static_display_cols], total_row], ignore_index=True)
                                calc_h = (len(final_team_df) + 1) * 35 + 20
                                st.dataframe(
                                    final_team_df.style.apply(style_static, axis=1)
                                                       .set_properties(**{'white-space': 'pre-wrap'}),
                                    column_order=static_display_cols,
                                    use_container_width=True, hide_index=True, height=calc_h
                                )

                                target_cols = [
                                    "client_number", "call_datetime", "call_starttime_clean",
                                    "call_endtime_clean", "call_duration", "status", "direction",
                                    "service", "reason", "call_owner", "Call Date",
                                    "updated_at_ampm", "Team Name", "Vertical", "Analyst", "source"
                                ]
                                existing_cols = [c for c in target_cols if c in team_df.columns]
                                st.download_button(
                                    label=f"📥 Download CDR — {team}",
                                    data=team_df[existing_cols].to_csv(index=False).encode('utf-8'),
                                    file_name=f"CDR_{team}.csv", mime='text/csv',
                                    key=f"dl_team_{team}"
                                )
                                st.divider()

                        tl_ad_pool = df_static_master[tl_ad_mask]
                        if not tl_ad_pool.empty:
                            report_df_tl, tl_dur_agg_sec = process_metrics_logic(tl_ad_pool)
                            active_tl = report_df_tl[report_df_tl['raw_dur_sec'] > 300].sort_values(by="raw_dur_sec", ascending=False)
                            if not active_tl.empty:
                                st.markdown(f"""
                                <div class="static-team-header">
                                    TL'S DURATION REPORT &nbsp;({display_start} to {display_end})
                                </div>""", unsafe_allow_html=True)
                                total_row_tl = pd.DataFrame([{
                                    "CALLER": "TOTAL",
                                    "TOTAL CALLS": int(active_tl["TOTAL CALLS"].sum()),
                                    "CALL STATUS": "-", "PICK UP RATIO %": "-",
                                    "CALLS > 3 MINS": int(active_tl["CALLS > 3 MINS"].sum()),
                                    "CALLS 15-20 MINS": int(active_tl["CALLS 15-20 MINS"].sum()),
                                    "20+ MIN CALLS": int(active_tl["20+ MIN CALLS"].sum()),
                                    "CALL DURATION > 3 MINS": format_dur_hm(active_tl["raw_dur_sec"].sum())
                                }])
                                final_tl_df = pd.concat([active_tl[static_display_cols], total_row_tl], ignore_index=True)
                                calc_h_tl = (len(final_tl_df) + 1) * 35 + 20
                                st.dataframe(
                                    final_tl_df.style.apply(style_static, axis=1)
                                                     .set_properties(**{'white-space': 'pre-wrap'}),
                                    column_order=static_display_cols,
                                    use_container_width=True, hide_index=True, height=calc_h_tl
                                )
                                valid_tls    = active_tl['CALLER'].unique()
                                final_tl_cdr = tl_ad_pool[tl_ad_pool['call_owner'].isin(valid_tls)]
                                target_cols  = [
                                    "client_number", "call_datetime", "call_starttime_clean",
                                    "call_endtime_clean", "call_duration", "status", "direction",
                                    "service", "reason", "call_owner", "Call Date",
                                    "updated_at_ampm", "Team Name", "Vertical", "Analyst", "source"
                                ]
                                existing_cols = [c for c in target_cols if c in final_tl_cdr.columns]
                                st.download_button(
                                    label="📥 Download TL CDR",
                                    data=final_tl_cdr[existing_cols].to_csv(index=False).encode('utf-8'),
                                    file_name="CDR_TL_AD.csv", mime='text/csv',
                                    key="dl_tl_ad_final_last"
                                )

                        # ── Store full dataset for Insights tab ──
                        report_all, _ = process_metrics_logic(
                            df_static_master[df_static_master['call_owner'].notna() & (df_static_master['call_owner'] != '')]
                        )
                        st.session_state['insights_df']     = df_static_master.copy()
                        st.session_state['insights_report'] = report_all.copy()
                        st.session_state['insights_source'] = "Duration Report"

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>📅</div>
                <div style='font-size:.9rem;font-weight:600;'>Click <b>Generate Duration Report</b> in the sidebar</div>
            </div>""", unsafe_allow_html=True)


    # ══════════════════════════════════════════════
    # TAB 3 — INSIGHTS (auto-populated from session state)
    # ══════════════════════════════════════════════

    with tab3:
        if 'insights_df' in st.session_state and 'insights_report' in st.session_state:
            df_ins       = st.session_state['insights_df']
            report_df_all = st.session_state['insights_report']
            source_label  = st.session_state.get('insights_source', 'Report')

            st.markdown(f"""
            <div style='text-align:center;margin-bottom:1rem;'>
                <span style='font-size:.75rem;font-weight:600;color:#F97316;
                             background:rgba(249,115,22,.1);border:1px solid rgba(249,115,22,.2);
                             border-radius:20px;padding:4px 14px;font-family:DM Mono,monospace;'>
                    ⚡ AUTO-GENERATED FROM {source_label.upper()}
                </span>
            </div>""", unsafe_allow_html=True)

            section_header("🧠 GENERATED TEAM INSIGHTS")
            insights = compute_team_insights(df_ins, report_df_all)

            if insights:
                cols_ins = st.columns(2)
                for i, ins in enumerate(insights):
                    with cols_ins[i % 2]:
                        st.markdown(f"""
                        <div class="insight-card {ins['type']}">
                            <div style='display:flex;align-items:center;justify-content:center;gap:.4rem;'>
                                <span class="insight-icon">{ins['icon']}</span>
                                <span class="insight-title">{ins['title']}</span>
                            </div>
                            <div class="insight-body">{ins['body']}</div>
                        </div>""", unsafe_allow_html=True)
            else:
                st.info("Not enough data to generate comparative insights.")

            st.divider()

            if not selected_team and not search_query:
                section_header("🏅 TEAM LEADERBOARD")
                lb = (
                    report_df_all.groupby("TEAM")
                    .agg(
                        agents=("CALLER", "count"),
                        total_calls=("TOTAL CALLS", "sum"),
                        total_dur_h=("raw_dur_sec", lambda x: round(x.sum() / 3600, 1)),
                        avg_dur_h=("raw_dur_sec", lambda x: round(x.mean() / 3600, 1)),
                        avg_prod_h=("raw_prod_sec", lambda x: round(x.mean() / 3600, 1)),
                        long_calls=("20+ MIN CALLS", "sum"),
                    )
                    .reset_index().sort_values("total_dur_h", ascending=False)
                    .rename(columns={
                        "TEAM": "Team", "agents": "Agents", "total_calls": "Total Calls",
                        "total_dur_h": "Total Dur (h)", "avg_dur_h": "Avg Dur/Agent (h)",
                        "avg_prod_h": "Avg Prod Hrs (h)", "long_calls": "20+ Min Calls"
                    })
                )
                medals = (["🥇", "🥈", "🥉"] + [""] * max(0, len(lb) - 3))[:len(lb)]
                lb.insert(0, "🏅", medals)
                lb = lb.reset_index(drop=True)
                st.dataframe(lb, use_container_width=True, hide_index=True)

        # ── Team Manual Calls ──
            manual_team_df = df_ins[df_ins['source'] == 'Manual'].copy()
            if not manual_team_df.empty:
                st.divider()
                section_header("⚠️ TEAM MANUAL CALLS")

                team_man_agg = (
                    manual_team_df.groupby('Team Name', sort=False)
                    .agg(
                        Vertical  = ('Vertical',      'first'),
                        Count     = ('source',        'count'),
                        DurSec    = ('call_duration', 'sum'),
                        Approvals = ('reason',        _unique_approvals),
                    )
                    .reset_index()
                    .sort_values('Count', ascending=False)
                    .reset_index(drop=True)
                )

                team_man_display = pd.DataFrame({
                    'VERTICAL'             : team_man_agg['Vertical'].fillna('—'),
                    'TEAM'                 : team_man_agg['Team Name'],
                    'MANUAL CALLS COUNT'   : team_man_agg['Count'],
                    'MANUAL CALLS DURATION': team_man_agg['DurSec'].apply(format_dur_hm),
                    'APPROVALS BY'         : team_man_agg['Approvals'],
                })

                total_team_man = pd.DataFrame([{
                    'VERTICAL'             : '—',
                    'TEAM'                 : 'TOTAL',
                    'MANUAL CALLS COUNT'   : int(team_man_agg['Count'].sum()),
                    'MANUAL CALLS DURATION': format_dur_hm(team_man_agg['DurSec'].sum()),
                    'APPROVALS BY'         : '—',
                }])

                team_man_final = pd.concat([team_man_display, total_team_man], ignore_index=True)
                st.dataframe(
                    team_man_final.style.apply(style_team_manual_total, axis=1),
                    use_container_width=True, hide_index=True
                )

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🧠</div>
                <div style='font-size:.9rem;font-weight:600;'>
                    Generate a <b>Dynamic Report</b> or <b>Duration Report</b> first —<br>
                    Insights will appear here automatically.
                </div>
            </div>""", unsafe_allow_html=True)


def run_revenue_dashboard():
    # ADD THIS CSS BLOCK FIRST:
    st.markdown("""
    <style>
    [data-testid="stMainBlockContainer"] {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    .block-container {
        max-width: 100% !important;
    }
    </style>
    """, unsafe_allow_html=True)
    # ─────────────────────────────────────────────
    # 1. CREDENTIALS & CONFIG
    # ─────────────────────────────────────────────

    CSV_URL      = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf/pub?gid=973926168&single=true&output=csv"
    REV_TABLE_ID = "studious-apex-488820-c3.crm_dashboard.revenue_sheet"
    
    # Callers that are not real agents
    EXCLUDE_CALLERS = {'direct', 'bootcamp - direct'}

    # ─────────────────────────────────────────────
    # 3. CSS
    # ─────────────────────────────────────────────

    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

:root {
    --accent-primary:   #10B981;
    --accent-secondary: #34D399;
    --accent-gold:      #F59E0B;
    --accent-warn:      #FBBF24;
    --accent-danger:    #F87171;
    --accent-info:      #60A5FA;
    --gold:             #F59E0B;
    --silver:           #9CA3AF;
    --bronze:           #CD7F32;
    --radius-sm:        8px;
    --radius-md:        12px;
    --radius-lg:        16px;
    --shadow-sm:        0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
    --shadow-md:        0 4px 16px rgba(0,0,0,.10);
    --shadow-lg:        0 8px 32px rgba(0,0,0,.14);
    --transition:       all 0.22s cubic-bezier(.4,0,.2,1);
}

[data-testid="stAppViewContainer"]:not([class*="dark"]) {
    --bg-base:      #F0FDF4;
    --bg-surface:   #FFFFFF;
    --bg-elevated:  #FFFFFF;
    --bg-muted:     #DCFCE7;
    --border:       rgba(0,0,0,.08);
    --text-primary: #111827;
    --text-muted:   #6B7280;
    --metric-bg:    #FFFFFF;
}

@media (prefers-color-scheme: dark) {
    :root {
        --bg-base:      #0A0F0D;
        --bg-surface:   #111B16;
        --bg-elevated:  #162119;
        --bg-muted:     #0F1A14;
        --border:       rgba(255,255,255,.07);
        --text-primary: #F1F5F9;
        --text-muted:   #94A3B8;
        --metric-bg:    #162119;
    }
}

[data-theme="dark"] {
    --bg-base:      #0A0F0D !important;
    --bg-surface:   #111B16 !important;
    --bg-elevated:  #162119 !important;
    --bg-muted:     #0F1A14 !important;
    --border:       rgba(255,255,255,.07) !important;
    --text-primary: #F1F5F9 !important;
    --text-muted:   #94A3B8 !important;
    --metric-bg:    #162119 !important;
}

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }

footer { visibility: hidden; }
[data-testid="stStatusWidget"], .stStatusWidget { display: none !important; }
[data-testid="stMainViewContainer"] { padding-top: 1.5rem; }
[data-testid="stSidebar"] { border-right: 1px solid var(--border, rgba(0,0,0,.08)); }

.rv-header {
    background: linear-gradient(135deg, #064e3b 0%, #065f46 45%, #1e3a5f 100%);
    border-radius: var(--radius-lg);
    padding: 1.5rem 2rem 1.2rem;
    margin-bottom: 1.2rem;
    position: relative;
    overflow: hidden;
    box-shadow: var(--shadow-lg);
}
.rv-header::before {
    content: "";
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(16,185,129,.2) 0%, transparent 70%);
    border-radius: 50%;
}
.rv-title   { font-size: 1.65rem; font-weight: 700; color: #FFFFFF; letter-spacing: .5px; margin: 0 0 .25rem; }
.rv-subtitle{ font-size: .82rem; color: rgba(255,255,255,.6); font-weight: 400; margin: 0; font-family: 'DM Mono', monospace; }
.rv-badge {
    display: inline-flex; align-items: center; gap: 5px;
    background: rgba(255,255,255,.12); backdrop-filter: blur(8px);
    border: 1px solid rgba(255,255,255,.18); border-radius: 20px;
    padding: 3px 10px; font-size: .73rem;
    color: rgba(255,255,255,.9); font-family: 'DM Mono', monospace;
}
.rv-pulse {
    width: 6px; height: 6px; background: #34D399; border-radius: 50%;
    display: inline-block; animation: pulse-ring 1.8s ease-in-out infinite;
}
@keyframes pulse-ring {
    0%, 100% { opacity: 1; transform: scale(1); }
    50%       { opacity: .5; transform: scale(1.4); }
}

.metric-card {
    background: var(--metric-bg, #fff);
    border: 1px solid var(--border, rgba(0,0,0,.08));
    border-radius: var(--radius-md);
    padding: .9rem 1rem;
    transition: var(--transition);
    box-shadow: var(--shadow-sm);
    position: relative; overflow: hidden; text-align: center;
}
.metric-card::before {
    content: ""; position: absolute; top: 0; left: 0;
    width: 100%; height: 3px;
    background: linear-gradient(90deg, #10B981, #34D399);
    opacity: 0; transition: opacity .2s;
}
.metric-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.metric-card:hover::before { opacity: 1; }
.metric-label  { font-size: .68rem; font-weight: 600; text-transform: uppercase; letter-spacing: .8px; color: var(--text-muted, #6B7280); margin: 0 0 .3rem; }
.metric-value  { font-size: 1.45rem; font-weight: 700; color: var(--text-primary, #111827); line-height: 1; font-family: 'DM Mono', monospace; }
.metric-delta  { font-size: .7rem; color: #10B981; margin-top: .2rem; font-weight: 500; }

.section-header { display: flex; align-items: center; gap: .6rem; margin: 1.5rem 0 .8rem; }
.section-header-line { flex: 1; height: 1px; background: linear-gradient(90deg, #10B981, transparent); opacity: .35; }
.section-title { font-size: .78rem; font-weight: 700; text-transform: uppercase; letter-spacing: 1.2px; color: #10B981; white-space: nowrap; text-align: center; }

.insight-card { background: var(--metric-bg, #fff); border: 1px solid var(--border, rgba(0,0,0,.08)); border-radius: var(--radius-md); padding: 1rem 1.1rem; margin-bottom: .6rem; box-shadow: var(--shadow-sm); transition: var(--transition); }
.insight-card:hover { box-shadow: var(--shadow-md); }
.insight-card.good { border-left: 3px solid #10B981; }
.insight-card.warn { border-left: 3px solid #FBBF24; }
.insight-card.bad  { border-left: 3px solid #F87171; }
.insight-card.info { border-left: 3px solid #60A5FA; }
.insight-icon  { font-size: 1.1rem; }
.insight-title { font-size: .82rem; font-weight: 700; color: var(--text-primary, #111827); margin: .2rem 0; }
.insight-body  { font-size: .76rem; color: var(--text-muted, #6B7280); line-height: 1.5; }

[data-testid="stTabs"] [role="tablist"] { gap: .3rem; border-bottom: 1px solid var(--border, rgba(0,0,0,.08)); }
[data-testid="stTabs"] button[role="tab"] {
    font-family: 'DM Sans', sans-serif !important; font-size: .82rem !important;
    font-weight: 600 !important; letter-spacing: .3px;
    border-radius: var(--radius-sm) var(--radius-sm) 0 0;
    padding: .55rem 1.1rem !important; transition: var(--transition);
}

div[data-testid="stDataFrame"] thead tr th {
    background: linear-gradient(135deg, #064e3b, #065f46) !important;
    color: #fff !important; font-family: 'DM Sans', sans-serif !important;
    font-size: .72rem !important; font-weight: 700 !important; letter-spacing: .6px;
    text-transform: uppercase; white-space: normal !important; word-wrap: break-word !important;
    text-align: center !important; vertical-align: middle !important;
    min-width: 100px !important; padding: 10px !important;
}

[data-testid="stSidebar"] .stButton>button {
    width: 100%; font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important; font-size: .82rem !important;
    border-radius: var(--radius-sm); transition: var(--transition);
}

[data-testid="stSidebar"] .stButton>button {
    background: linear-gradient(135deg, #059669, #065f46) !important;
    color: #fff !important; border: none !important;
}
[data-testid="stSidebar"] .stButton>button:hover {
    opacity: .88 !important; transform: translateY(-1px) !important;
}

hr { border-color: var(--border, rgba(0,0,0,.08)) !important; margin: 1.2rem 0 !important; }

.brand-name {
    font-size: .85rem;
    font-weight: 700;
    letter-spacing: -.3px;
    color: #10B981;
}

.brand-tagline {
    font-size: .58rem;
    letter-spacing: .8px;
    font-family: monospace;
    margin-bottom: .9rem;
    color: #059669;
}

.achieve-bar-wrap { background: var(--bg-muted, #DCFCE7); border-radius: 999px; height: 6px; margin-top: .4rem; overflow: hidden; }
.achieve-bar-fill { height: 100%; border-radius: 999px; background: linear-gradient(90deg, #10B981, #34D399); transition: width .6s ease; }
/* ── Download Buttons — match Generate Revenue Report style ── */
.stDownloadButton > button {
    width: 100% !important;
    background: linear-gradient(135deg, #059669, #065f46) !important;
    color: #fff !important;
    border: none !important;
    border-radius: var(--radius-sm) !important;
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 600 !important;
    font-size: .82rem !important;
    transition: var(--transition) !important;
}
.stDownloadButton > button:hover {
    opacity: .88 !important;
    transform: translateY(-1px) !important;
}
</style>
""", unsafe_allow_html=True)


    # ─────────────────────────────────────────────
    # HELPERS
    # ─────────────────────────────────────────────

    def fmt_inr(value):
        if pd.isna(value) or value == 0: return "₹0"
        if value >= 1_00_00_000: return f"₹{value/1_00_00_000:.2f}Cr"
        if value >= 1_00_000:    return f"₹{value/1_00_000:.2f}L"
        if value >= 1_000:       return f"₹{value/1_000:.2f}K"
        return f"₹{int(value)}"

    def section_header(label):
        st.markdown(f"""
        <div class="section-header">
            <div class="section-header-line"></div>
            <span class="section-title">{label}</span>
            <div class="section-header-line" style="background:linear-gradient(90deg,transparent,#10B981)"></div>
        </div>""", unsafe_allow_html=True)

    def style_total_rev(row):
        if row.get("CALLER NAME") == "TOTAL":
            return ['font-weight: bold; background-color: #374151; color: #FFFFFF;'] * len(row)
        return [''] * len(row)

    def get_months_in_range(start_date, end_date):
        months, cur = [], date(start_date.year, start_date.month, 1)
        end_month   = date(end_date.year, end_date.month, 1)
        while cur <= end_month:
            months.append(cur)
            cur = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)
        return months

    def count_working_days(start_date, end_date):
        """Count Mon–Fri days between start_date and end_date inclusive."""
        count = 0
        cur   = start_date
        while cur <= end_date:
            if cur.weekday() < 5:   # 0=Mon … 4=Fri
                count += 1
            cur += timedelta(days=1)
        return count


    # ─────────────────────────────────────────────
    # DATA FETCHING
    # ─────────────────────────────────────────────

    @st.cache_data(ttl=120, show_spinner=False)
    def get_metadata():
        df_meta = pd.read_csv(CSV_URL)
        df_meta.columns = df_meta.columns.str.strip().str.replace('\xa0', '', regex=False)

        month_col = next((c for c in df_meta.columns if c.strip().lower() == 'month'), None)
        if month_col and month_col != 'Month':
            df_meta.rename(columns={month_col: 'Month'}, inplace=True)

        if 'Month' in df_meta.columns:
            df_meta['Month'] = pd.to_datetime(df_meta['Month'], dayfirst=True, errors='coerce').dt.date
        else:
            df_meta['Month'] = None

        teams     = sorted(df_meta['Team Name'].dropna().unique()) if 'Team Name' in df_meta.columns else []
        verticals = sorted(df_meta['Vertical'].dropna().unique())  if 'Vertical'  in df_meta.columns else []
        df_meta['merge_key'] = df_meta['Caller Name'].str.strip().str.lower()
        return teams, verticals, df_meta

    @st.cache_data(ttl=600, show_spinner=False)
    def get_last_update():
        query = f"""
            SELECT updated_at_ampm FROM `{REV_TABLE_ID}`
            WHERE updated_at IS NOT NULL
            ORDER BY updated_at DESC LIMIT 1
        """
        try:
            res = client.query(query).to_dataframe()
            return str(res['updated_at_ampm'].iloc[0]) if not res.empty else "N/A"
        except:
            return "N/A"

    @st.cache_data(ttl=600, show_spinner=False)
    def get_available_dates():
        query = f"SELECT MIN(Date) as min_date, MAX(Date) as max_date FROM `{REV_TABLE_ID}`"
        try:
            df = client.query(query).to_dataframe()
            if not df.empty and not pd.isna(df['min_date'].iloc[0]):
                return df['min_date'].iloc[0], df['max_date'].iloc[0]
        except:
            pass
        return date.today(), date.today()

    @st.cache_data(ttl=120, show_spinner=False)
    def fetch_revenue_data(start_date, end_date):
        query = f"""
            SELECT * FROM `{REV_TABLE_ID}`
            WHERE Date BETWEEN '{start_date}' AND '{end_date}'
            AND Fee_paid > 0
        """
        df = client.query(query).to_dataframe()
        if not df.empty:
            df['Caller_name']  = df['Caller_name'].astype(str).str.strip()
            df['merge_key']    = df['Caller_name'].str.lower()
            df['Fee_paid']     = pd.to_numeric(df['Fee_paid'],     errors='coerce').fillna(0)
            df['Course_Price'] = pd.to_numeric(df['Course_Price'], errors='coerce').fillna(0)

            enr       = df['Enrollment'].astype(str).str.strip()
            enr_lower = enr.str.lower()
            src_lower = df['Source'].astype(str).str.lower()

            df['is_new_enrollment']       = enr_lower == 'new enrollment'
            df['is_balance_payment']      = enr_lower == 'new enrollment - balance payment'
            df['is_bootcamp_collection']  = enr_lower == 'bootcamp collections - balance payments'
            df['is_community_collection'] = enr_lower == 'community collections - balance payments'
            df['is_other_revenue']        = enr_lower == 'other revenue'
            df['is_empty_enrollment']     = enr == ''
            df['source_has_community']    = src_lower.str.contains('community', na=False)

            # Legacy compat
            df['is_new'] = df['is_new_enrollment']
        return df


    # ─────────────────────────────────────────────
    # TARGET / DESIGNATION RESOLUTION
    # ─────────────────────────────────────────────

    def _col(df, name):
        name_clean = name.strip().lower()
        match = next((c for c in df.columns if c.strip().lower() == name_clean), None)
        if match is None:
            raise KeyError(f"Column '{name}' not found. Available: {list(df.columns)}")
        return match

    def resolve_targets(df_meta, start_date, end_date):
        months_needed = get_months_in_range(start_date, end_date)
        caller_col    = _col(df_meta, 'Caller Name')
        month_col     = _col(df_meta, 'Month')

        target_col = None
        for candidate in ['Target', 'target', 'TARGET', 'Monthly Target', 'Monthly target', 'Sales Target']:
            try:
                target_col = _col(df_meta, candidate)
                break
            except KeyError:
                continue

        if target_col is None:
            st.warning(f"⚠️ Target column not found. Sheet columns: `{list(df_meta.columns)}`")
            return {}

        months_ym = {(m.year, m.month) for m in months_needed}

        def _ym(val):
            try:
                if pd.isna(val):
                    return None
                d = pd.to_datetime(val, dayfirst=True, errors='coerce')
                if pd.isna(d):
                    return None
                return (d.year, d.month)
            except:
                return None

        relevant = df_meta[df_meta[month_col].apply(_ym).isin(months_ym)].copy()
        if relevant.empty:
            return {}

        dedup = relevant.drop_duplicates(subset=[caller_col, month_col])
        dedup[target_col] = pd.to_numeric(
            dedup[target_col].astype(str).str.replace(',', '', regex=False),
            errors='coerce'
        ).fillna(0)
        return dedup.groupby(caller_col)[target_col].sum().to_dict()

    def resolve_designations(df_meta, start_date, end_date):
        months_needed = get_months_in_range(start_date, end_date)
        caller_col    = _col(df_meta, 'Caller Name')
        month_col     = _col(df_meta, 'Month')

        def safe_col(name):
            try:    return _col(df_meta, name)
            except: return None

        desig_col   = safe_col('Academic Counselor/TL/ATL')
        team_col    = safe_col('Team Name')
        vert_col    = safe_col('Vertical')
        analyst_col = safe_col('Analyst')

        months_ym = {(m.year, m.month) for m in months_needed}

        def _ym(val):
            try:
                if pd.isna(val):
                    return None
                d = pd.to_datetime(val, dayfirst=True, errors='coerce')
                if pd.isna(d):
                    return None
                return (d.year, d.month)
            except:
                return None

        relevant = df_meta[df_meta[month_col].apply(_ym).isin(months_ym)].copy()
        if relevant.empty:
            return {}

        relevant = relevant.sort_values(month_col, ascending=False)
        dedup    = relevant.drop_duplicates(subset=[caller_col], keep='first').set_index(caller_col)

        result = {}
        for caller in dedup.index:
            result[caller] = {
                'Academic Counselor/TL/ATL': dedup.at[caller, desig_col]   if desig_col   else '—',
                'Team Name':                 dedup.at[caller, team_col]     if team_col    else '—',
                'Vertical':                  dedup.at[caller, vert_col]     if vert_col    else '—',
                'Analyst':                   dedup.at[caller, analyst_col]  if analyst_col else '—',
            }
        return result


    # ─────────────────────────────────────────────
    # CALLER CLASSIFICATION + METRICS PROCESSING
    # ─────────────────────────────────────────────

    def classify_and_process(df, df_meta, start_date, end_date):
        target_map       = resolve_targets(df_meta, start_date, end_date)
        desig_map        = resolve_designations(df_meta, start_date, end_date)
        target_map_norm  = {str(k).strip().lower(): v for k, v in target_map.items()}
        desig_map_norm   = {str(k).strip().lower(): v for k, v in desig_map.items()}

        # Pre-compute working days ratio once for all callers
        months_count    = max(len(get_months_in_range(start_date, end_date)), 1)
        working_days    = count_working_days(start_date, end_date)
        till_day_ratio  = min(working_days / (20 * months_count), 1.0)

        calling_rows, collection_rows, both_rows = [], [], []

        for caller, grp in df.groupby('Caller_name'):
            if caller.strip().lower() in EXCLUDE_CALLERS:
                continue

            caller_key = caller.strip().lower()
            info       = desig_map_norm.get(caller_key, {})
            team       = str(info.get('Team Name', '—')).strip()
            target     = float(target_map_norm.get(caller_key, 0) or 0)

            till_day_target = round(target * till_day_ratio)

            enr_rev      = grp[grp['is_new_enrollment']]['Fee_paid'].sum()
            bal_rev      = grp[grp['is_balance_payment']]['Fee_paid'].sum()
            boot_coll    = grp[grp['is_bootcamp_collection']]['Fee_paid'].sum()
            comm_coll    = grp[grp['is_community_collection']]['Fee_paid'].sum()
            enrollments  = int(grp['is_new_enrollment'].sum())

            calling_rev    = enr_rev + bal_rev
            collection_rev = boot_coll + comm_coll
            total_rev      = calling_rev + collection_rev

            is_changemakers = team.lower() == 'changemakers'
            has_calling     = calling_rev > 0
            has_collection  = collection_rev > 0

            row = {
                'DESIGNATION'         : info.get('Academic Counselor/TL/ATL', '—'),
                'CALLER NAME'         : caller,
                'TEAM'                : team,
                'VERTICAL'            : info.get('Vertical', '—'),
                'TOTAL TARGET (₹)'    : target,
                'TILL DAY TARGET (₹)' : till_day_target,
                'ENROLLMENTS'         : enrollments,
                'ENROLLMENT REV'      : enr_rev,
                'BALANCE REV'         : bal_rev,
                'COMMUNITY COLLECTION': comm_coll,
                'BOOTCAMP COLLECTION' : boot_coll,
                'CALLING REVENUE'     : calling_rev,
                'COLLECTION REVENUE'  : collection_rev,
                'TOTAL REVENUE'       : total_rev,
                'raw_revenue'         : total_rev,
                'raw_target'          : target,
                'raw_enrollments'     : enrollments,
                'raw_calling_rev'     : calling_rev,
                'raw_collection_rev'  : collection_rev,
            }

            if is_changemakers:
                row['ACHIEVEMENT %'] = round(collection_rev / target * 100, 1) if target > 0 else 0.0
                collection_rows.append(row)
            elif has_calling and has_collection:
                row['ACHIEVEMENT %'] = round(total_rev / target * 100, 1) if target > 0 else 0.0
                both_rows.append(row)
            elif has_collection:
                row['ACHIEVEMENT %'] = round(collection_rev / target * 100, 1) if target > 0 else 0.0
                collection_rows.append(row)
            else:
                row['ACHIEVEMENT %'] = round(calling_rev / target * 100, 1) if target > 0 else 0.0
                calling_rows.append(row)

        def _to_df(rows, sort_col):
            if not rows:
                return pd.DataFrame()
            d = pd.DataFrame(rows).sort_values(sort_col, ascending=False).reset_index(drop=True)
            return d

        calling_df    = _to_df(calling_rows,    'raw_calling_rev')
        collection_df = _to_df(collection_rows, 'raw_collection_rev')
        both_df       = _to_df(both_rows,       'raw_revenue')

        return calling_df, collection_df, both_df


    # ─────────────────────────────────────────────
    # SUMMARY METRICS COMPUTATION
    # ─────────────────────────────────────────────

    def compute_summary_metrics(df):
        excl_mask = df['Caller_name'].str.strip().str.lower().isin(EXCLUDE_CALLERS)

        calling_rev = (
            df[~excl_mask & df['is_new_enrollment']]['Fee_paid'].sum() +
            df[~excl_mask & df['is_balance_payment']]['Fee_paid'].sum()
        )

        collection_rev = df[df['is_bootcamp_collection']]['Fee_paid'].sum()

        community_rev = (
            df[df['is_community_collection']]['Fee_paid'].sum() +
            df[df['is_other_revenue'] & df['source_has_community']]['Fee_paid'].sum() +
            df[
                df['is_new_enrollment'] &
                df['source_has_community'] &
                (df['Caller_name'].str.strip().str.lower() == 'direct')
            ]['Fee_paid'].sum()
        )

        direct_rev = df[
            (df['Caller_name'].str.strip().str.lower() == 'direct') &
            (~df['source_has_community']) &
            (df['is_other_revenue'] | df['is_new_enrollment'] | df['is_balance_payment'])
        ]['Fee_paid'].sum()

        dna_rev = df[df['is_empty_enrollment']]['Fee_paid'].sum()

        bootcamp_direct_rev = df[
            df['is_new_enrollment'] &
            (df['Caller_name'].str.strip().str.lower() == 'bootcamp - direct')
        ]['Fee_paid'].sum()

        total_rev = calling_rev + bootcamp_direct_rev + collection_rev + community_rev + direct_rev + dna_rev

        return {
            'total_rev'          : total_rev,
            'calling_rev'        : calling_rev,
            'bootcamp_direct_rev': bootcamp_direct_rev,
            'collection_rev'     : collection_rev,
            'community_rev'      : community_rev,
            'direct_rev'         : direct_rev,
            'dna_rev'            : dna_rev,
        }

    # ─────────────────────────────────────────────
    # ENROLLMENT SUMMARY COMPUTATION
    # ─────────────────────────────────────────────

    def compute_enrollment_metrics(df):
        caller_lower = df['Caller_name'].str.strip().str.lower()

        direct_enr    = int(df[
            df['is_new_enrollment'] &
            (caller_lower == 'direct') &
            (~df['source_has_community'])
        ].shape[0])

        bootcamp_enr  = int(df[
            df['is_new_enrollment'] &
            (caller_lower == 'bootcamp - direct')
        ].shape[0])

        caller_enr    = int(df[
            df['is_new_enrollment'] &
            (~caller_lower.isin(['direct', 'bootcamp - direct']))
        ].shape[0])

        community_enr = int(df[
            df['is_new_enrollment'] &
            (caller_lower == 'direct') &
            df['source_has_community']
        ].shape[0])

        total_enr = direct_enr + bootcamp_enr + caller_enr + community_enr

        return {
            'total_enr'    : total_enr,
            'direct_enr'   : direct_enr,
            'bootcamp_enr' : bootcamp_enr,
            'caller_enr'   : caller_enr,
            'community_enr': community_enr,
        }

    # ─────────────────────────────────────────────
    # TABLE RENDERING HELPER
    # ─────────────────────────────────────────────

    def render_perf_table(df, display_cols, total_overrides, sort_col, table_key):
        """Renders a styled performance table with medals and a TOTAL row."""
        if df.empty:
            st.info("No data available for this category in the selected period.")
            return

        d = df.copy().reset_index(drop=True)

        # Medals
        d.insert(0, 'Rank', '')
        for i, medal in enumerate(['🥇', '🥈', '🥉']):
            if i < len(d):
                d.at[i, 'Rank'] = medal

        # Format currency cols
        for col in ['ENROLLMENT REV', 'BALANCE REV', 'COMMUNITY COLLECTION',
                    'BOOTCAMP COLLECTION', 'CALLING REVENUE', 'COLLECTION REVENUE',
                    'TOTAL REVENUE', 'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)']:
            if col in d.columns:
                d[col] = d[col].apply(fmt_inr)

        if 'ACHIEVEMENT %' in d.columns:
            d['ACHIEVEMENT %'] = d['ACHIEVEMENT %'].apply(lambda x: f"{x}%")

        # Build total row
        total_dict = {c: '—' for c in ['Rank'] + display_cols}
        total_dict['Rank']        = ''
        total_dict['CALLER NAME'] = 'TOTAL'
        for k, v in total_overrides.items():
            total_dict[k] = v

        all_cols   = ['Rank'] + [c for c in display_cols if c in d.columns]
        total_row  = pd.DataFrame([{c: total_dict.get(c, '—') for c in all_cols}])
        final      = pd.concat([d[all_cols], total_row], ignore_index=True)

        st.dataframe(
            final.style.apply(style_total_rev, axis=1),
            column_order=all_cols,
            use_container_width=True,
            hide_index=True
        )


    # ─────────────────────────────────────────────
    # INSIGHTS
    # ─────────────────────────────────────────────

    def compute_revenue_insights(df, calling_df, collection_df, both_df, start_date, end_date):
        insights = []
        if df.empty:
            return insights

        month_label = start_date.strftime("%B %Y")

        _all_list = [d for d in [calling_df, collection_df, both_df] if not d.empty]
        all_agents = pd.concat(_all_list, ignore_index=True) if _all_list else pd.DataFrame()

        # ── 1. Top Calling Revenue Achiever ──
        _calling_pool_list = [d for d in [calling_df, both_df] if not d.empty]
        _calling_pool = pd.concat(_calling_pool_list, ignore_index=True) if _calling_pool_list else pd.DataFrame()
        if not _calling_pool.empty:
            top_c = _calling_pool.sort_values('raw_calling_rev', ascending=False).iloc[0]
            insights.append({
                "type": "good", "icon": "🏆",
                "title": f"Top Calling Revenue — {top_c['CALLER NAME']}",
                "body": (f"Achieved {fmt_inr(top_c['raw_calling_rev'])} in calling revenue with "
                         f"{top_c['ENROLLMENTS']} new enrollment(s) in {month_label}. "
                         f"Target achievement: {top_c['ACHIEVEMENT %']}%.")
            })

        # ── 2. Top Collection Revenue Achiever ──
        _coll_pool_list = [d for d in [collection_df, both_df] if not d.empty]
        _coll_pool = pd.concat(_coll_pool_list, ignore_index=True) if _coll_pool_list else pd.DataFrame()
        if not _coll_pool.empty:
            top_col = _coll_pool.sort_values('raw_collection_rev', ascending=False).iloc[0]
            insights.append({
                "type": "good", "icon": "🏦",
                "title": f"Top Collection Revenue — {top_col['CALLER NAME']}",
                "body": (f"Collected {fmt_inr(top_col['raw_collection_rev'])} in {month_label} "
                         f"across community and bootcamp collections. "
                         f"Team: {top_col['TEAM']}.")
            })

        # ── 3. Most Enrollments ──
        if not all_agents.empty and all_agents['raw_enrollments'].max() > 0:
            top_enr = all_agents.sort_values('raw_enrollments', ascending=False).iloc[0]
            insights.append({
                "type": "good", "icon": "🎓",
                "title": f"Most Enrollments — {top_enr['CALLER NAME']}",
                "body": (f"Closed {top_enr['raw_enrollments']} new enrollment(s) in {month_label}, "
                         f"highest among all callers. "
                         f"Calling revenue generated: {fmt_inr(top_enr['raw_calling_rev'])}.")
            })

        # ── 4. Best Target Achievement ──
        if not all_agents.empty:
            _with_target = all_agents[all_agents['raw_target'] > 0]
            if not _with_target.empty:
                top_ach = _with_target.sort_values('ACHIEVEMENT %', ascending=False).iloc[0]
                insights.append({
                    "type": "good" if top_ach['ACHIEVEMENT %'] >= 80 else "warn", "icon": "🎯",
                    "title": f"Best Target Achievement — {top_ach['CALLER NAME']}",
                    "body": (f"Achieved {top_ach['ACHIEVEMENT %']}% of their "
                             f"{fmt_inr(top_ach['raw_target'])} target in {month_label}. "
                             f"Revenue: {fmt_inr(top_ach['raw_revenue'])}. Team: {top_ach['TEAM']}.")
                })

        # ── 5. Focus Required — worst target achievement ──
        if not all_agents.empty:
            _with_target = all_agents[
                (all_agents['raw_target'] > 0) &
                (all_agents['raw_revenue'] > 0)
            ]
            if len(_with_target) > 1:
                worst = _with_target.sort_values('ACHIEVEMENT %').iloc[0]
                if worst['ACHIEVEMENT %'] < 60:
                    insights.append({
                        "type": "bad", "icon": "⚠️",
                        "title": f"Focus Required — {worst['CALLER NAME']}",
                        "body": (f"Only {worst['ACHIEVEMENT %']}% of "
                                 f"{fmt_inr(worst['raw_target'])} target achieved in {month_label}. "
                                 f"Revenue so far: {fmt_inr(worst['raw_revenue'])}. "
                                 f"Team: {worst['TEAM']}. Needs immediate attention.")
                    })

         # ── 6. Callers with target but zero revenue (calling agents only) ──
        _calling_only_list = [d for d in [calling_df, both_df] if not d.empty]
        _calling_only = pd.concat(_calling_only_list, ignore_index=True) if _calling_only_list else pd.DataFrame()

        if not _calling_only.empty:
            _with_target   = _calling_only[_calling_only['raw_target'] > 0]
            _zero_rev      = _with_target[_with_target['raw_calling_rev'] == 0]
            _zero_count    = len(_zero_rev)
            _total_callers = len(_with_target)

            if _zero_count > 0:
                _zero_names = ', '.join(_zero_rev['CALLER NAME'].head(3).tolist())
                _more       = f" and {_zero_count - 3} more" if _zero_count > 3 else ""
                insights.append({
                    "type": "bad", "icon": "🚨",
                    "title": f"{_zero_count} of {_total_callers} Callers Have Zero Revenue in {month_label}",
                    "body": (f"{_zero_names}{_more} have an assigned target but no enrollment "
                             f"or balance revenue recorded in the selected period. "
                             f"Immediate follow-up recommended to unblock closures.")
                })
            else:
                _below_50  = _with_target[_with_target['raw_calling_rev'] / _with_target['raw_target'] * 100 < 50]
                _b50_count = len(_below_50)
                if _b50_count > 0:
                    _top_team = (
                        _calling_only.groupby('TEAM')['raw_calling_rev']
                        .sum().sort_values(ascending=False).index[0]
                    )
                    insights.append({
                        "type": "warn", "icon": "📊",
                        "title": f"{_b50_count} Caller(s) Below 50% Achievement in {month_label}",
                        "body": (f"All callers have some revenue but {_b50_count} are below 50% of their "
                                 f"calling target. Best performing team is {_top_team}. "
                                 f"Focus support on underperforming callers to improve month-end numbers.")
                    })
                else:
                    _top_team = (
                        _calling_only.groupby('TEAM')['raw_calling_rev']
                        .sum().sort_values(ascending=False).index[0]
                    )
                    _top_team_rev = _calling_only.groupby('TEAM')['raw_calling_rev'].sum().max()
                    insights.append({
                        "type": "good", "icon": "🌟",
                        "title": f"Strong Month — All Callers On Track in {month_label}",
                        "body": (f"Every caller with a target has recorded enrollment revenue in the selected period. "
                                 f"Leading team is {_top_team} with {fmt_inr(_top_team_rev)} in calling revenue. "
                                 f"Keep monitoring till-day targets to sustain momentum through month-end.")
                    })

        return insights[:6]
    @st.cache_data(show_spinner=False)
    def generate_helper_pdf_bytes() -> bytes:
        buffer = io.BytesIO()

        GREEN_DARK  = colors.HexColor("#064e3b")
        GREEN_MID   = colors.HexColor("#065f46")
        GREEN_PALE  = colors.HexColor("#DCFCE7")
        GREEN_ROW   = colors.HexColor("#F0FDF4")
        GREY_DARK   = colors.HexColor("#374151")
        GREY_MID    = colors.HexColor("#6B7280")
        WHITE       = colors.white
        BLACK       = colors.HexColor("#111827")
        W, H        = A4

        def s(name, **kw):
            defaults = dict(fontName='Helvetica', fontSize=9,
                            textColor=BLACK, spaceAfter=3, leading=14)
            defaults.update(kw)
            return ParagraphStyle(name, **defaults)

        S = {
            'body'      : s('body'),
            'label'     : s('label',   fontName='Helvetica-Bold', fontSize=8,
                             textColor=GREEN_DARK, spaceAfter=1),
            'formula'   : s('formula', fontName='Helvetica-Oblique', fontSize=8.5,
                             textColor=colors.HexColor("#065f46"),
                             backColor=GREEN_PALE, leftIndent=8, rightIndent=8),
            'footer'    : s('footer',  fontSize=7.5, textColor=GREY_MID,
                             alignment=TA_CENTER),
        }

        class CoverBlock(Flowable):
            def __init__(self, w):
                Flowable.__init__(self)
                self.w = w
                self.height = 90
            def draw(self):
                c = self.canv
                c.setFillColor(GREEN_DARK);  c.rect(0, 52, self.w, 38, fill=1, stroke=0)
                c.setFillColor(GREEN_MID);   c.rect(0, 22, self.w, 30, fill=1, stroke=0)
                c.setFillColor(colors.HexColor("#065f46")); c.rect(0, 0, self.w, 22, fill=1, stroke=0)
                c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 22)
                c.drawCentredString(self.w/2, 66, "REVENUE METRICS DASHBOARD")
                c.setFillColor(colors.HexColor("#A7F3D0")); c.setFont("Helvetica-Bold", 11)
                c.drawCentredString(self.w/2, 34, "Logic & Metric Reference Guide")
                c.setFillColor(colors.HexColor("#D1FAE5")); c.setFont("Helvetica", 8.5)
                c.drawCentredString(self.w/2, 8,
                    "LawSikho & Skill Arbitrage  \u00b7  Sales & Operations Team  \u00b7  Internal Use Only")

        class SectionBanner(Flowable):
            def __init__(self, icon, title, color=None, w=None):
                Flowable.__init__(self)
                self.icon = icon; self.title = title
                self.color = color or GREEN_DARK
                self.w = w or (W - 30*mm); self.height = 22
            def draw(self):
                c = self.canv
                c.setFillColor(self.color)
                c.roundRect(0, 0, self.w, self.height, 4, fill=1, stroke=0)
                c.setFillColor(WHITE); c.setFont("Helvetica-Bold", 11)
                c.drawString(10, 6, f"{self.icon}  {self.title}")

        def btable(rows, cw=None):
            cw = cw or [42*mm, 118*mm]
            data = [[Paragraph(f"<b>{r[0]}</b>", S['label']),
                     Paragraph(r[1], S['body'])] for r in rows]
            t = Table(data, colWidths=cw, hAlign='LEFT')
            t.setStyle(TableStyle([
                ('BACKGROUND',    (0,0),(0,-1), GREEN_PALE),
                ('VALIGN',        (0,0),(-1,-1),'TOP'),
                ('GRID',          (0,0),(-1,-1),0.3, colors.HexColor("#D1FAE5")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[WHITE, GREEN_ROW]),
                ('LEFTPADDING',   (0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',    (0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]))
            return t

        def ltable(rows):
            data = [[Paragraph(f"<b>{r[0]}</b>", S['label']),
                     Paragraph(r[1], S['formula'])] for r in rows]
            t = Table(data, colWidths=[50*mm, 110*mm], hAlign='LEFT')
            t.setStyle(TableStyle([
                ('VALIGN',        (0,0),(-1,-1),'TOP'),
                ('GRID',          (0,0),(-1,-1),0.3, colors.HexColor("#A7F3D0")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[WHITE, GREEN_ROW]),
                ('LEFTPADDING',   (0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',    (0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
            ]))
            return t

        SP  = Spacer
        HR  = lambda: HRFlowable(width="100%", thickness=0.6,
                                  color=colors.HexColor("#A7F3D0"),
                                  spaceAfter=6, spaceBefore=4)
        BAN = SectionBanner
        cw  = W - 30*mm

        story = [
            SP(1,18*mm), CoverBlock(cw), SP(1,10*mm),
            Paragraph("This document explains every metric, table, and highlight card in the "
                      "Revenue Metrics Dashboard — a quick reference for the Sales & Operations team.", S['body']),
            SP(1,4*mm),

            # ── Section 1 ──
            BAN("🏆","SECTION 1 — TOP 3 REVENUE HIGHLIGHTS"), SP(1,3*mm),
            Paragraph("Three cards at the top of the dashboard — each shows the single best performer "
                      "in one dimension for the selected date range.", S['body']), SP(1,2*mm),
            btable([
                ("🥇 Top Revenue — Caller",    "Caller with highest Calling Revenue (Enrollment Rev + Balance Rev). Excludes 'direct' and 'bootcamp-direct'."),
                ("🎓 Most Enrollments",         "Caller with most rows where Enrollment = 'New Enrollment'. Covers all agent types."),
                ("🥇 Top Revenue — Collection", "Caller with highest Collection Revenue (Community Collections + Bootcamp Collections). From Collection table only."),
            ]), SP(1,4*mm),

            # ── Section 2 ──
            BAN("💵","SECTION 2 — REVENUE SUMMARY METRICS"), SP(1,3*mm),
            Paragraph("Full-picture revenue breakdown. Filter: Fee Paid > 0 applied globally.", S['body']), SP(1,2*mm),
            ltable([
                ("Total Revenue\n(EXCL. Services)",    "Calling + Bootcamp-Direct + Bootcamp-Collection + Community + Direct/Other + DNA revenue."),
                ("Calling Revenue\n(INCL. Funnel)",    "Fee Paid where Enrollment = 'New Enrollment' OR 'New Enrollment - Balance Payment', caller NOT in {direct, bootcamp-direct}."),
                ("Bootcamp-Direct\nRevenue",            "Fee Paid where Enrollment = 'New Enrollment' AND Caller = 'bootcamp-direct'."),
                ("Bootcamp-Collection\nRevenue",        "Fee Paid where Enrollment = 'Bootcamp Collections - Balance Payments'."),
                ("Community Revenue\n(Direct+Collection)", "Fee Paid from: (a) Community Collections rows, (b) Other Revenue with community Source, (c) New Enrollment by 'direct' with community Source."),
                ("Direct/Other Revenue\n(INCL. Funnel)","Fee Paid where Caller='direct', Source has no 'community', Enrollment in {Other Revenue, New Enrollment, Balance Payment}."),
                ("DNA / Not Updated Yet",               "Fee Paid where Enrollment column is blank. Not yet categorised in BigQuery."),
            ]), SP(1,4*mm),

            # ── Section 3 ──
            BAN("🎓","SECTION 3 — ENROLLMENT SUMMARY METRICS"), SP(1,3*mm),
            Paragraph("Counts ONLY rows where Enrollment = 'New Enrollment'. Balance payments and collections are excluded.", S['body']), SP(1,2*mm),
            ltable([
                ("Total Enrollments",            "All New Enrollment rows across all callers."),
                ("Caller Enrollments",           "New Enrollment rows where Caller is NOT 'direct' or 'bootcamp-direct'."),
                ("Direct Enrollments",           "New Enrollment by 'direct' caller AND Source has no 'community'."),
                ("Bootcamp-Direct Enrollments",  "New Enrollment where Caller = 'bootcamp-direct'."),
                ("Community-Direct Enrollments", "New Enrollment by 'direct' caller AND Source contains 'community'."),
            ]), SP(1,4*mm),

            # ── Section 4 ──
            BAN("📞","SECTION 4 — CALLER REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Calling agents with Calling Revenue > 0, no collection revenue, not Changemakers team. Sorted by Calling Revenue descending.", S['body']), SP(1,2*mm),
            btable([
                ("DESIGNATION",         "Role from team sheet: Academic Counselor, TL, ATL, etc."),
                ("TOTAL TARGET (₹)",    "Sum of monthly targets from team sheet for all months in selected range."),
                ("TILL DAY TARGET (₹)", "Total Target × (Mon-Fri days elapsed ÷ (20 × months in range)). Capped at 1.0."),
                ("ENROLLMENTS",         "Count of New Enrollment rows for this caller."),
                ("ENROLLMENT REV",      "Fee Paid where Enrollment = 'New Enrollment' for this caller."),
                ("BALANCE REV",         "Fee Paid where Enrollment = 'New Enrollment - Balance Payment' for this caller."),
                ("CALLING REVENUE",     "Enrollment Rev + Balance Rev."),
                ("ACHIEVEMENT %",       "Calling Revenue ÷ Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 5 ──
            BAN("🏦","SECTION 5 — COLLECTION CALLER REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Callers with Collection Revenue > 0 and no calling revenue. Changemakers team always here. Sorted by Collection Revenue descending.", S['body']), SP(1,2*mm),
            btable([
                ("COMMUNITY COLLECTION", "Fee Paid where Enrollment = 'Community Collections - Balance Payments'."),
                ("BOOTCAMP COLLECTION",  "Fee Paid where Enrollment = 'Bootcamp Collections - Balance Payments'."),
                ("COLLECTION REVENUE",   "Community Collection + Bootcamp Collection."),
                ("ACHIEVEMENT %",        "Collection Revenue ÷ Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 6 ──
            BAN("📞🏦","SECTION 6 — CALLING + COLLECTION CALLER REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Callers with BOTH Calling Revenue > 0 AND Collection Revenue > 0, not Changemakers. Sorted by Total Revenue descending.", S['body']), SP(1,2*mm),
            btable([
                ("TOTAL REVENUE",  "Calling Revenue + Collection Revenue."),
                ("ACHIEVEMENT %",  "Total Revenue ÷ Total Target × 100."),
                ("Other columns",  "Same definitions as Sections 4 and 5."),
            ]), SP(1,4*mm),

            # ── Section 7 ──
            BAN("📞","SECTION 7 — CALLER REVENUE TEAM PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Section 4 callers grouped by team. Insights & Leaderboard tab only. Sorted by Calling Revenue descending.", S['body']), SP(1,2*mm),
            ltable([
                ("Callers",         "Count of agents from this team in the Caller table."),
                ("Total Target",    "Sum of all agent targets for this team."),
                ("Till Day Target", "Sum of all agent till-day targets for this team."),
                ("Calling Revenue", "Sum of Calling Revenue across all agents in this team."),
                ("Achievement %",   "Team Calling Revenue ÷ Team Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 8 ──
            BAN("🏦","SECTION 8 — COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Section 5 callers grouped by team. Sorted by Collection Revenue descending.", S['body']), SP(1,2*mm),
            ltable([
                ("Community Collection", "Sum of Community Collection revenue for this team."),
                ("Bootcamp Collection",  "Sum of Bootcamp Collection revenue for this team."),
                ("Collection Revenue",   "Community + Bootcamp Collection total for this team."),
                ("Achievement %",        "Team Collection Revenue ÷ Team Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Section 9 ──
            BAN("📞🏦","SECTION 9 — CALLING + COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE"), SP(1,3*mm),
            Paragraph("Section 6 callers grouped by team. Sorted by Total Revenue descending.", S['body']), SP(1,2*mm),
            ltable([
                ("Calling Revenue",      "Sum of Calling Revenue for dual-stream agents in this team."),
                ("Community Collection", "Sum of Community Collection revenue."),
                ("Bootcamp Collection",  "Sum of Bootcamp Collection revenue."),
                ("Total Revenue",        "Calling + Community + Bootcamp Collection for this team."),
                ("Achievement %",        "Team Total Revenue ÷ Team Total Target × 100."),
            ]), SP(1,4*mm),

            # ── Glossary ──
            # ── Section 10 — Pending Revenue ──
            BAN("📊","SECTION 10 — CALLERWISE PENDING REVENUE TAB"), SP(1,3*mm),
            Paragraph("Auto-loads on tab open. Always shows current month + previous month. Not affected by sidebar filters.", S['body']), SP(1,2*mm),
            ltable([
                ("Revenue Pool",             "Sum of Course Price for all booking-fee leads (Full_Installment = 'booking fees') with Fee_paid ≥ ₹999 and positive balance. Excludes drop-sheet leads and students who appear in both months (paid balance)."),
                ("Revenue Collected",        "Sum of Fee_paid (booking fee already received) for the same leads in the pool."),
                ("Balance to be Recovered",  "Course Price − Fee_paid for each lead. Only leads with balance > 0 are shown."),
                ("No. of Leads",             "Count of individual pending-balance leads for this caller/team."),
                ("Leads >48 HRS",            "Count of leads whose enrollment Date is ≤ today − 3 days (IST). Today and the two preceding calendar days are excluded so only genuinely overdue leads are flagged."),
                ("Balance >48 HRS",          "Sum of pending balance for leads that qualify as >48 hrs overdue."),
                ("% Pending >48 HRS",        "Balance >48 HRS ÷ Total Balance × 100 for this row."),
                ("Previous Month columns",   "Same balance and lead count computed from the previous calendar month's pending leads pool."),
                ("Grand Balance / Leads",    "Current month balance + previous month balance (and lead counts). Rows with grand balance = 0 are hidden."),
            ]), SP(1,4*mm),
            BAN("🚫","SECTION 11 — CALLERWISE DROPPED LEADS"), SP(1,3*mm),
            Paragraph("Drop leads are sourced from the Drop Leads Google Sheet. Each drop is attributed to the caller who originally enrolled the student (matched by email then phone against the revenue sheet).", S['body']), SP(1,2*mm),
            ltable([
                ("Current Month Drops",  "Drop form submissions with a timestamp falling in the current calendar month."),
                ("Previous Month Drops", "Drop form submissions with a timestamp falling in the previous calendar month."),
                ("Total Drop Cases",     "Current + Previous month drops for this caller."),
                ("Attribution Logic",    "Email match first; if no email match, phone match. If neither matches any New Enrollment row, attributed to 'Unknown'."),
                ("Team / Vertical",      "Pulled from the team sheet using the attributed caller name as the merge key."),
            ]), SP(1,4*mm),

            # ── Glossary ──
            BAN("📖","KEY TERMS GLOSSARY", color=GREY_DARK), SP(1,3*mm),
            btable([
                ("New Enrollment",       "A fresh admission. Counts toward enrollment metrics and Calling Revenue."),
                ("Balance Payment",      "Remaining fee from a prior enrollment. NOT a new enrollment."),
                ("Bootcamp Collections", "Balance payments from bootcamp-enrolled students."),
                ("Community Collections","Balance payments from community-channel students."),
                ("DNA / Not Updated",    "Rows with blank Enrollment. Tracked separately in summary."),
                ("direct",               "Pseudo-caller for organic closures. Excluded from agent tables."),
                ("bootcamp-direct",      "Pseudo-caller for bootcamp direct admissions. Tracked separately."),
                ("Till Day Target",      "Daily-prorated target: Total Target x (Working Days / (20 x Months)). Working Days = Mon-Fri in selected range."),
                ("Changemakers",         "Special team always routed to Collection table regardless of calling revenue."),
            ], cw=[44*mm, 116*mm]),

            SP(1,8*mm), HR(),
            Paragraph("Designed by Amit Ray  \u00b7  amitray@lawsikho.com  \u00b7  "
                      "For Internal Use of Sales and Operations Team Only. All Rights Reserved.", S['footer']),
        ]

        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=15*mm, rightMargin=15*mm,
            topMargin=14*mm,  bottomMargin=14*mm,
            title="Revenue Metrics — Logic Reference Guide",
            author="Amit Ray",
        )
        doc.build(story)
        return buffer.getvalue()


    # ─────────────────────────────────────────────
    # PENDING REVENUE — CONSTANTS & HELPERS
    # ─────────────────────────────────────────────

    DROP_LEADS_URL_P = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRJ9qaFD5Sc9O1JFH7ijR0JI2SEkAgXM8PJZsWslsASLDnTCA_fwIP0fg_PmtdMm_zs3KwTLI45fPog/pub?gid=1082322056&single=true&output=csv"

    def pending_months():
        today   = date.today()
        c_start = date(today.year, today.month, 1)
        c_end   = today
        p_end   = c_start - timedelta(days=1)
        p_start = date(p_end.year, p_end.month, 1)
        return c_start, c_end, p_start, p_end

    @st.cache_data(ttl=300, show_spinner=False)
    def load_drop_leads():
        try:
            df = pd.read_csv(DROP_LEADS_URL_P)
            df.columns = df.columns.str.strip()
            e_col = next((c for c in df.columns if 'email' in c.lower() and ('drop' in c.lower() or 'student' in c.lower())), None)
            p_col = next((c for c in df.columns if 'phone' in c.lower() or ('number' in c.lower() and 'drop' in c.lower())), None)
            t_col = next((c for c in df.columns if 'timestamp' in c.lower()), None)
            df['drop_email'] = df[e_col].astype(str).str.strip().str.lower() if e_col else ''
            df['drop_phone'] = (df[p_col].astype(str).str.replace(r'\D', '', regex=True).str[-10:]) if p_col else ''
            df['drop_date']  = pd.to_datetime(df[t_col], errors='coerce', dayfirst=True) if t_col else pd.NaT
            return df
        except:
            return pd.DataFrame()

    @st.cache_data(ttl=120, show_spinner=False)
    def fetch_both_months_rev(p_start, c_end):
        q = f"SELECT * FROM `{REV_TABLE_ID}` WHERE Date BETWEEN '{p_start}' AND '{c_end}'"
        df = client.query(q).to_dataframe()
        if df.empty:
            return df
        df['Caller_name']     = df['Caller_name'].astype(str).str.strip()
        df['merge_key']       = df['Caller_name'].str.lower()
        df['Fee_paid']        = pd.to_numeric(df['Fee_paid'],     errors='coerce').fillna(0)
        df['Course_Price']    = pd.to_numeric(df['Course_Price'], errors='coerce').fillna(0)
        df['Email_Id_norm']   = df['Email_Id'].astype(str).str.strip().str.lower()
        df['Contact_No_norm'] = df['Contact_No'].astype(str).str.replace(r'\D', '', regex=True).str[-10:]
        enr_l = df['Enrollment'].astype(str).str.strip().str.lower()
        df['is_new'] = enr_l == 'new enrollment'
        df['is_bal'] = enr_l == 'new enrollment - balance payment'
        df['Date']   = pd.to_datetime(df['Date'], errors='coerce').dt.date
        # Normalize Full_Installment field — "booking fees" = partial payment pending
        df['Full_Installment'] = (
            df['Full_Installment'].astype(str).str.strip().str.lower()
            if 'Full_Installment' in df.columns
            else 'unknown'
        )
        return df

    def pending_leads_for_month(df_m, excl_emails, excl_phones, df_combined=None):
        if df_m.empty:
            return pd.DataFrame()

        # ── Step 1: All rows with Fee_paid >= 999 across ALL enrollment types ──
        df_valid = df_m[df_m['Fee_paid'] >= 999].copy()
        if df_valid.empty:
            return pd.DataFrame()

        # ── Step 2: Count email + phone occurrences across COMBINED two-month pool ──
        # Mirrors original: counts built from ALL rows in both months together
        # so a student who enrolled in month A and paid balance in month B
        # appears twice in the combined pool → count=2 → excluded from both months
        if df_combined is not None and not df_combined.empty:
            df_pool = df_combined[df_combined['Fee_paid'] >= 999]
        else:
            df_pool = df_valid  # fallback: single month only
        email_counts = df_pool['Email_Id_norm'].value_counts()
        phone_counts = df_pool['Contact_No_norm'].value_counts()

        # ── Step 3: Filter to "booking fees" rows only (this month's slice) ──
        if 'Full_Installment' in df_valid.columns:
            pending_rows = df_valid[
                df_valid['Full_Installment'].str.lower().str.strip() == 'booking fees'
            ].copy()
        else:
            # Fallback until pipeline pushes Full_Installment to BigQuery
            pending_rows = df_valid[df_valid['is_new']].copy()

        if pending_rows.empty:
            return pd.DataFrame()

        # ── Step 4: count==1 against COMBINED pool — same as original script ──
        pending_rows = pending_rows[
            pending_rows['Email_Id_norm'].map(email_counts) == 1
        ].copy()
        pending_rows = pending_rows[
            pending_rows['Contact_No_norm'].map(phone_counts) == 1
        ].copy()
        if pending_rows.empty:
            return pd.DataFrame()

        # ── Step 5: Exclude drop sheet (email AND phone) ──
        pending_rows = pending_rows[
            ~pending_rows['Email_Id_norm'].isin(excl_emails) &
            ~pending_rows['Contact_No_norm'].isin(excl_phones)
        ].copy()
        if pending_rows.empty:
            return pd.DataFrame()

        # ── Step 6: Pending balance = Course Price - Fee Paid ──
        pending_rows['balance'] = pending_rows['Course_Price'] - pending_rows['Fee_paid']
        p = pending_rows[pending_rows['balance'] > 0].copy()
        if p.empty:
            return pd.DataFrame()

        india        = pytz.timezone("Asia/Kolkata")
        today_ist    = datetime.now(india).date()
        cut48        = today_ist - timedelta(days=3)
        p['over_48'] = p['Date'] <= cut48
        return p

    def _raw_caller_agg(pending_df):
        """Pure groupby aggregation — NO metadata merge. Team info added later in one pass."""
        if pending_df.empty:
            return pd.DataFrame()
        b48 = (pending_df[pending_df['over_48']]
               .groupby('Caller_name')['balance'].sum()
               .rename('bal_48hr').reset_index())
        agg = (pending_df.groupby('Caller_name')
               .agg(pool     =('Course_Price', 'sum'),
                    collected=('Fee_paid',     'sum'),
                    balance  =('balance',      'sum'),
                    leads    =('balance',      'count'),
                    leads_48 =('over_48',      'sum'))
               .reset_index()
               .merge(b48, on='Caller_name', how='left'))
        agg['bal_48hr'] = agg['bal_48hr'].fillna(0)
        agg['leads_48'] = agg['leads_48'].astype(int)
        return agg

    def build_combined_agg(curr_pending, prev_pending, meta_map, df_rev=None):
        """
        Mirrors attribute_drops_to_callers exactly:
          0. (NEW) Re-attribute each pending lead's Caller_name to the ORIGINAL
             NEW ENROLLMENT CALLER via email/phone lookup from df_rev.
             This is why DROPPED LEADS always shows correct teams — it uses the
             new-enrollment caller, not the raw booking-fees row caller.
          1. Raw-aggregate curr and prev separately (no metadata yet)
          2. Outer-join on Caller_name
          3. ONE meta_map merge at the end → correct team for every caller.
        """
        # ── Step 0: remap Caller_name to canonical new-enrollment caller ──
        if df_rev is not None and not df_rev.empty:
            new_enr = df_rev[df_rev['is_new']].copy()
            email_to_caller = (new_enr.drop_duplicates('Email_Id_norm')
                               .set_index('Email_Id_norm')['Caller_name'].to_dict())
            phone_to_caller = (new_enr.drop_duplicates('Contact_No_norm')
                               .set_index('Contact_No_norm')['Caller_name'].to_dict())

            def _resolve(row):
                e = str(row.get('Email_Id_norm', '')).strip().lower()
                p = str(row.get('Contact_No_norm', '')).strip()
                if e and e in email_to_caller: return email_to_caller[e]
                if p and p in phone_to_caller: return phone_to_caller[p]
                return row['Caller_name']  # fallback: keep original

            def _remap(df):
                if df.empty: return df
                df = df.copy()
                df['Caller_name'] = df.apply(_resolve, axis=1)
                return df

            curr_pending = _remap(curr_pending)
            prev_pending = _remap(prev_pending)

        curr_agg = _raw_caller_agg(curr_pending)
        prev_agg = _raw_caller_agg(prev_pending)

        if curr_agg.empty and prev_agg.empty:
            return pd.DataFrame()

        if not prev_agg.empty:
            prev_slim = prev_agg[['Caller_name', 'balance', 'leads']].rename(
                columns={'balance': 'prev_bal', 'leads': 'prev_leads'})
        else:
            prev_slim = pd.DataFrame(columns=['Caller_name', 'prev_bal', 'prev_leads'])

        if curr_agg.empty:
            combined = prev_slim.copy()
            for c in ['pool', 'collected', 'balance', 'leads', 'leads_48', 'bal_48hr']:
                combined[c] = 0
        elif prev_slim.empty:
            combined = curr_agg.copy()
            combined['prev_bal']   = 0
            combined['prev_leads'] = 0
        else:
            combined = curr_agg.merge(prev_slim, on='Caller_name', how='outer')

        for c in ['pool', 'collected', 'balance', 'leads', 'leads_48', 'bal_48hr', 'prev_bal', 'prev_leads']:
            if c not in combined.columns:
                combined[c] = 0
            combined[c] = combined[c].fillna(0)

        combined['grand_bal']   = combined['balance'] + combined['prev_bal']
        combined['grand_leads'] = combined['leads'].astype(int) + combined['prev_leads'].astype(int)

        # Deduplicate meta_map and merge ONCE for ALL callers (curr + prev + both)
        meta_dedup = (meta_map[['merge_key', 'Caller Name', 'Team Name', 'Vertical']]
                      .drop_duplicates(subset=['merge_key'], keep='first'))
        combined['mk'] = combined['Caller_name'].str.strip().str.lower()
        combined = combined.merge(
            meta_dedup.rename(columns={'merge_key': 'mk'}), on='mk', how='left')
        combined['Caller_name'] = combined['Caller Name'].fillna(combined['Caller_name'])
        combined['Team Name']   = combined['Team Name'].fillna('Others')
        combined['Vertical']    = combined['Vertical'].fillna('Others')
        return combined

    def _pct48(b48, bal):
        return f"{round(b48 / bal * 100)}%" if bal and bal > 0 else "0%"

    

    def attribute_drops_to_callers(drop_df, df_rev, meta_map, c_start, c_end, p_start, p_end, curr_label, prev_label):
        if drop_df.empty or df_rev.empty:
            return pd.DataFrame()
        new_enr = df_rev[df_rev['is_new']].copy()
        email_to_caller = new_enr.drop_duplicates('Email_Id_norm').set_index('Email_Id_norm')['Caller_name'].to_dict()
        phone_to_caller = new_enr.drop_duplicates('Contact_No_norm').set_index('Contact_No_norm')['Caller_name'].to_dict()

        def get_caller(row):
            e = str(row.get('drop_email', '')).strip().lower()
            p = str(row.get('drop_phone', '')).strip()
            if e and e in email_to_caller: return email_to_caller[e]
            if p and p in phone_to_caller: return phone_to_caller[p]
            return 'Unknown'

        df = drop_df.copy()
        df['attributed_caller'] = df.apply(get_caller, axis=1)
        df = df[df['drop_date'].notna()].copy()
        df['drop_d'] = df['drop_date'].dt.date

        curr_drops = (df[(df['drop_d'] >= c_start) & (df['drop_d'] <= c_end)]
                      .groupby('attributed_caller').size().rename('curr_drops'))
        prev_drops = (df[(df['drop_d'] >= p_start) & (df['drop_d'] <= p_end)]
                      .groupby('attributed_caller').size().rename('prev_drops'))

        combined = pd.DataFrame({'curr_drops': curr_drops, 'prev_drops': prev_drops}).fillna(0).reset_index()
        combined.columns = ['Caller_name', 'curr_drops', 'prev_drops']
        combined['curr_drops']   = combined['curr_drops'].astype(int)
        combined['prev_drops']   = combined['prev_drops'].astype(int)
        combined['total_drops']  = combined['curr_drops'] + combined['prev_drops']
        combined['mk']           = combined['Caller_name'].str.strip().str.lower()
        combined = combined.merge(meta_map.rename(columns={'merge_key': 'mk'}), on='mk', how='left')
        combined['Caller_name'] = combined['Caller Name'].fillna(combined['Caller_name'])
        combined['Team Name']   = combined['Team Name'].fillna('Others')
        combined['Vertical']    = combined['Vertical'].fillna('Others')
        return combined.sort_values('total_drops', ascending=False).reset_index(drop=True)

    def render_html_pending_table(combined, mode, curr_label, prev_label, title):
        hdr_style  = "background:#064e3b;color:#fff;font-size:.72rem;font-weight:700;text-transform:uppercase;padding:8px 6px;text-align:center;border:1px solid #065f46;"
        sub_style  = "background:#065f46;color:#fff;font-size:.68rem;font-weight:600;padding:6px 4px;text-align:center;border:1px solid #0d9e6e;"
        data_style = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#ffffff;"
        data_style_alt = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#f0fdf4;"
        team_style = "font-weight:700;background:#1f2937;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #374151;"
        vert_style = "font-weight:700;background:#064e3b;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #065f46;"
        grand_style= "font-weight:700;background:#1e3a5f;color:#fff;font-size:.72rem;padding:8px 5px;text-align:center;border:1px solid #1e3a5f;"
        name_col   = "CALLER NAME" if mode == "caller" else "TEAM NAME"
        extra_th   = '<th rowspan="2" style="' + hdr_style + '">TEAM</th>' if mode == "caller" else ""

        html = (
            "<div style='margin:1.5rem 0 .5rem;text-align:center;'>"
            "<div style='font-size:1rem;font-weight:800;text-transform:uppercase;"
            "letter-spacing:1px;color:#10B981;margin-bottom:.5rem;'>"
            + title +
            "</div></div>"
            "<div style='overflow-x:auto;'>"
            "<table style='width:100%;border-collapse:collapse;'>"
            "<thead>"
            "<tr>"
            "<th rowspan='2' style='" + hdr_style + "'>" + name_col + "</th>"
            + extra_th +
            "<th colspan='7' style='" + hdr_style + "background:#065f46;'>" + curr_label.upper() + "</th>"
            "<th colspan='2' style='" + hdr_style + "background:#92400e;'>" + prev_label.upper() + "</th>"
            "<th colspan='2' style='" + hdr_style + "background:#1e3a5f;'>GRAND TOTAL</th>"
            "</tr>"
            "<tr>"
            "<th style='" + sub_style + "'>REVENUE POOL</th>"
            "<th style='" + sub_style + "'>TOTAL REVENUE COLLECTED</th>"
            "<th style='" + sub_style + "'>BALANCE AMOUNT TO BE RECOVERED</th>"
            "<th style='" + sub_style + "'>NO. OF LEADS (BALANCE TO BE RECOVERED)</th>"
            "<th style='" + sub_style + "background:#0f766e;'>NO. OF LEADS PENDING &gt;48 HRS</th>"
            "<th style='" + sub_style + "background:#0f766e;'>BALANCE RECOVERY PENDING &gt;48 HRS</th>"
            "<th style='" + sub_style + "background:#0f766e;'>PERCENTAGE OF REVENUE PENDING &gt;48 HRS</th>"
            "<th style='" + sub_style + "background:#92400e;'>BALANCE AMOUNT TO BE RECOVERED</th>"
            "<th style='" + sub_style + "background:#92400e;'>NO. OF LEADS (BALANCE AMOUNT TO BE RECOVERED)</th>"
            "<th style='" + sub_style + "background:#1e3a5f;'>AMOUNT TO BE RECOVERED</th>"
            "<th style='" + sub_style + "background:#1e3a5f;'>TOTAL LEAD TO BE RECOVERED</th>"
            "</tr>"
            "</thead><tbody>"
        )

        g = {k: 0 for k in ["pool","collected","balance","leads","leads_48","bal_48hr","prev_bal","prev_leads","grand_bal","grand_leads"]}

        num_keys = ["pool","collected","balance","leads","leads_48","bal_48hr","prev_bal","prev_leads","grand_bal","grand_leads"]

        # Sort verticals by total grand_bal descending
        vert_order = (
            combined.copy()
            .assign(_vert=combined["Vertical"].fillna("Unassigned"))
            .groupby("_vert")["grand_bal"]
            .sum()
            .sort_values(ascending=False)
            .index.tolist()
        )

        for vert in vert_order:
            v_df = combined[combined["Vertical"].fillna("Unassigned") == vert].copy()
            if v_df.empty:
                continue
            # Force numeric on all aggregation columns
            for k in num_keys:
                if k in v_df.columns:
                    v_df[k] = pd.to_numeric(v_df[k], errors='coerce').fillna(0)
            v = {k: 0 for k in g}

            # Sort teams within vertical by total grand_bal descending
            team_order = (
                v_df.assign(_team=v_df["Team Name"].fillna("Unassigned"))
                .groupby("_team")["grand_bal"]
                .sum()
                .sort_values(ascending=False)
                .index.tolist()
            )

            for team in team_order:
                t_df = v_df[v_df["Team Name"].fillna("Unassigned") == team].copy()
                t = {k: 0 for k in g}

                if mode == "caller":
                    for idx, r in t_df.sort_values("balance", ascending=False).iterrows():
                        row_style = data_style_alt if idx % 2 == 0 else data_style
                        html += "<tr>"
                        html += "<td style='" + row_style + "'>" + str(r.get("Caller_name","—")) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(r.get("Team Name","—")) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("pool",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("collected",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("balance",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("leads",0))) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("leads_48",0))) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("bal_48hr",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + _pct48(r.get("bal_48hr",0), r.get("balance",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("prev_bal",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("prev_leads",0))) + "</td>"
                        html += "<td style='" + row_style + "'>" + fmt_inr(r.get("grand_bal",0)) + "</td>"
                        html += "<td style='" + row_style + "'>" + str(int(r.get("grand_leads",0))) + "</td>"
                        html += "</tr>"
                        for k in g: t[k] += r.get(k,0)

                else: # Team mode
                    for k in g: t[k] = t_df[k].sum()

                # Team Total
                html += "<tr>"
                html += "<td style='" + team_style + "'>" + (team if mode=="team" else team+" Total") + "</td>"
                if mode == "caller": html += "<td style='" + team_style + "'>—</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["pool"]) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["collected"]) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["balance"]) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["leads"])) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["leads_48"])) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["bal_48hr"]) + "</td>"
                html += "<td style='" + team_style + "'>" + _pct48(t["bal_48hr"], t["balance"]) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["prev_bal"]) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["prev_leads"])) + "</td>"
                html += "<td style='" + team_style + "'>" + fmt_inr(t["grand_bal"]) + "</td>"
                html += "<td style='" + team_style + "'>" + str(int(t["grand_leads"])) + "</td>"
                html += "</tr>"
                for k in g: v[k] += t[k]

            # Vertical Total
            html += "<tr>"
            html += "<td style='" + vert_style + "'>" + vert + " Total</td>"
            if mode == "caller": html += "<td style='" + vert_style + "'>—</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["pool"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["collected"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["balance"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["leads"])) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["leads_48"])) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["bal_48hr"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + _pct48(v["bal_48hr"], v["balance"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["prev_bal"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["prev_leads"])) + "</td>"
            html += "<td style='" + vert_style + "'>" + fmt_inr(v["grand_bal"]) + "</td>"
            html += "<td style='" + vert_style + "'>" + str(int(v["grand_leads"])) + "</td>"
            html += "</tr>"
            for k in g: g[k] += v[k]

        # Grand Total
        html += "<tr>"
        html += "<td style='" + grand_style + "'>GRAND TOTAL</td>"
        if mode == "caller": html += "<td style='" + grand_style + "'>—</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["pool"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["collected"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["balance"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["leads"])) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["leads_48"])) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["bal_48hr"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + _pct48(g["bal_48hr"], g["balance"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["prev_bal"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["prev_leads"])) + "</td>"
        html += "<td style='" + grand_style + "'>" + fmt_inr(g["grand_bal"]) + "</td>"
        html += "<td style='" + grand_style + "'>" + str(int(g["grand_leads"])) + "</td>"
        html += "</tr>"
        html += "</tbody></table></div>"
        return html


    def render_drop_html(drop_agg, curr_label, prev_label):
        hs = "background:#7c2d12;color:#fff;font-size:.72rem;font-weight:700;text-transform:uppercase;padding:8px 6px;text-align:center;border:1px solid #991b1b;"
        ds = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#ffffff;"
        ds_alt = "font-size:.72rem;padding:6px 5px;text-align:center;border:1px solid #d1fae5;color:#111827;background:#f0fdf4;"
        ts = "font-weight:700;background:#1f2937;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #374151;"
        vs = "font-weight:700;background:#7c2d12;color:#fff;font-size:.72rem;padding:7px 5px;text-align:center;border:1px solid #991b1b;"
        gs = "font-weight:700;background:#1e3a5f;color:#fff;font-size:.72rem;padding:8px 5px;text-align:center;border:1px solid #1e3a5f;"

        html = (
            "<div style='overflow-x:auto;'>"
            "<table style='width:100%;border-collapse:collapse;'>"
            "<thead><tr>"
            "<th style='" + hs + "'>CALLER NAME</th>"
            "<th style='" + hs + "'>TEAM</th>"
            "<th style='" + hs + "'>VERTICAL</th>"
            "<th style='" + hs + "'>" + curr_label.upper() + " DROP CASES</th>"
            "<th style='" + hs + "'>" + prev_label.upper() + " DROP CASES</th>"
            "<th style='" + hs + "'>TOTAL DROP CASES</th>"
            "</tr></thead><tbody>"
        )

        g_c = g_p = g_t = 0
        for vert in sorted(drop_agg["Vertical"].fillna("Others").unique()):
            v_df = drop_agg[drop_agg["Vertical"].fillna("Others") == vert]
            if v_df.empty:
                continue
            vc = vp = vt = 0
            for team in sorted(v_df["Team Name"].fillna("Others").unique()):
                t_df = v_df[v_df["Team Name"].fillna("Others") == team]
                if t_df.empty:
                    continue
                tc = tp = tt = 0
                drop_row_idx = 0
                for _, r in t_df.sort_values('total_drops', ascending=False).iterrows():
                    drow = ds_alt if drop_row_idx % 2 == 1 else ds
                    drop_row_idx += 1
                    html += f"""<tr>
                        <td style='{drow}text-align:center;'>{r['Caller_name']}</td>
                        <td style='{drow}'>{team}</td>
                        <td style='{drow}'>{vert}</td>
                        <td style='{drow}'>{int(r['curr_drops'])}</td>
                        <td style='{drow}'>{int(r['prev_drops'])}</td>
                        <td style='{drow}font-weight:600;color:#DC2626;'>{int(r['total_drops'])}</td>
                    </tr>"""
                    tc += r['curr_drops']; tp += r['prev_drops']; tt += r['total_drops']
                html += (
                    "<tr>"
                    "<td style='" + ts + "'>" + team + " Total</td>"
                    "<td style='" + ts + "'>—</td>"
                    "<td style='" + ts + "'>" + str(vert) + "</td>"
                    "<td style='" + ts + "'>" + str(int(tc)) + "</td>"
                    "<td style='" + ts + "'>" + str(int(tp)) + "</td>"
                    "<td style='" + ts + "'>" + str(int(tt)) + "</td>"
                    "</tr>"
                )
                vc += tc; vp += tp; vt += tt
            html += (
                "<tr>"
                "<td style='" + vs + "'>" + vert + " Total</td>"
                "<td style='" + vs + "'>—</td>"
                "<td style='" + vs + "'>—</td>"
                "<td style='" + vs + "'>" + str(int(vc)) + "</td>"
                "<td style='" + vs + "'>" + str(int(vp)) + "</td>"
                "<td style='" + vs + "'>" + str(int(vt)) + "</td>"
                "</tr>"
            )
            g_c += vc; g_p += vp; g_t += vt
        html += (
            "<tr>"
            "<td style='" + gs + "'>Grand Total</td>"
            "<td style='" + gs + "'>—</td>"
            "<td style='" + gs + "'>—</td>"
            "<td style='" + gs + "'>" + str(int(g_c)) + "</td>"
            "<td style='" + gs + "'>" + str(int(g_p)) + "</td>"
            "<td style='" + gs + "'>" + str(int(g_t)) + "</td>"
            "</tr></tbody></table></div>"
        )
        return html


    # ─────────────────────────────────────────────
    # EXCEL DOWNLOAD BUILDERS — PENDING TAB
    # ─────────────────────────────────────────────

    def build_pending_excel(combined, pend_curr, pend_prev, meta_map_pending, curr_label, prev_label):
        """
        Returns bytes of an xlsx with 2 tabs:
          Tab 1 — Teamwise Pending Revenue
          Tab 2 — Callerwise Pending Revenue
        All monetary values are raw integers (no K/L formatting).
        """
        HDR_FILL   = PatternFill("solid", start_color="064e3b", end_color="064e3b")
        TEAM_FILL  = PatternFill("solid", start_color="1f2937", end_color="1f2937")
        VERT_FILL  = PatternFill("solid", start_color="064e3b", end_color="064e3b")
        GRAND_FILL = PatternFill("solid", start_color="1e3a5f", end_color="1e3a5f")
        ALT_FILL   = PatternFill("solid", start_color="f0fdf4", end_color="f0fdf4")
        WHITE_FILL = PatternFill("solid", start_color="ffffff", end_color="ffffff")
        HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        DATA_FONT  = Font(name="Arial", size=9)
        # BOLD_WHITE and others used in sub-functions or later
        BOLD_WHITE = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        BORDER     = Border(
            left=Side(style='thin', color='D1FAE5'),
            right=Side(style='thin', color='D1FAE5'),
            top=Side(style='thin', color='D1FAE5'),
            bottom=Side(style='thin', color='D1FAE5'),
        )
        CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

        wb = Workbook()

        num_keys = ["pool","collected","balance","leads","leads_48","bal_48hr",
                    "prev_bal","prev_leads","grand_bal","grand_leads"]

        def _safe(v):
            try:
                f = float(v)
                return 0 if pd.isna(f) else f
            except:
                return 0

        def _pct(b48, bal):
            if bal and bal > 0:
                return round(b48 / bal * 100, 1)
            return 0.0

        def write_pending_sheet(ws, mode):
            # ── Header rows ──
            cl = curr_label.upper()
            pl = prev_label.upper()
            name_col_hdr = "CALLER NAME" if mode == "caller" else "TEAM NAME"
            extra = 1 if mode == "caller" else 0   # extra TEAM column for caller mode

            # Row 1 — group headers
            col = 1
            ws.cell(1, col, name_col_hdr).fill = HDR_FILL
            ws.cell(1, col).font = HDR_FONT
            ws.cell(1, col).alignment = CENTER
            ws.cell(1, col).border = BORDER
            col += 1
            if mode == "caller":
                ws.cell(1, col, "TEAM").fill = HDR_FILL
                ws.cell(1, col).font = HDR_FONT
                ws.cell(1, col).alignment = CENTER
                ws.cell(1, col).border = BORDER
                col += 1

            # Current month group (7 cols)
            for i in range(7):
                ws.cell(1, col+i, cl if i == 0 else "").fill = PatternFill("solid", start_color="065f46", end_color="065f46")
                ws.cell(1, col+i).font = HDR_FONT
                ws.cell(1, col+i).alignment = CENTER
                ws.cell(1, col+i).border = BORDER
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+6)
            col += 7

            # Previous month group (2 cols)
            for i in range(2):
                ws.cell(1, col+i, pl if i == 0 else "").fill = PatternFill("solid", start_color="92400e", end_color="92400e")
                ws.cell(1, col+i).font = HDR_FONT
                ws.cell(1, col+i).alignment = CENTER
                ws.cell(1, col+i).border = BORDER
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)
            col += 2

            # Grand total group (2 cols)
            for i in range(2):
                ws.cell(1, col+i, "GRAND TOTAL" if i == 0 else "").fill = PatternFill("solid", start_color="1e3a5f", end_color="1e3a5f")
                ws.cell(1, col+i).font = HDR_FONT
                ws.cell(1, col+i).alignment = CENTER
                ws.cell(1, col+i).border = BORDER
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+1)

            # Row 2 — sub-headers
            sub_headers = [name_col_hdr] + (["TEAM"] if mode == "caller" else []) + [
                "REVENUE POOL (₹)", "COLLECTED (₹)", "BALANCE (₹)", "LEADS",
                "LEADS >48HR", "BALANCE >48HR (₹)", "% PENDING >48HR",
                f"BALANCE ({pl}) (₹)", f"LEADS ({pl})",
                "GRAND BALANCE (₹)", "GRAND LEADS"
            ]
            for c_idx, h in enumerate(sub_headers, 1):
                cell = ws.cell(2, c_idx, h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER

            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 28

            # ── Data rows ──
            row_num = 3
            g = {k: 0.0 for k in num_keys}

            vert_order = (
                combined.assign(_v=combined["Vertical"].fillna("Unassigned"))
                .groupby("_v")["grand_bal"].sum()
                .sort_values(ascending=False).index.tolist()
            )

            for vert in vert_order:
                v_df = combined[combined["Vertical"].fillna("Unassigned") == vert].copy()
                if v_df.empty:
                    continue
                for k in num_keys:
                    if k in v_df.columns:
                        v_df[k] = pd.to_numeric(v_df[k], errors='coerce').fillna(0)
                v = {k: 0.0 for k in num_keys}

                team_order = (
                    v_df.assign(_t=v_df["Team Name"].fillna("Unassigned"))
                    .groupby("_t")["grand_bal"].sum()
                    .sort_values(ascending=False).index.tolist()
                )

                for team in team_order:
                    t_df = v_df[v_df["Team Name"].fillna("Unassigned") == team].copy()
                    if t_df.empty:
                        continue
                    t = {k: 0.0 for k in num_keys}

                    if mode == "caller":
                        alt = False
                        for _, r in t_df.sort_values("balance", ascending=False).iterrows():
                            vals = [
                                str(r.get("Caller_name", "—")),
                                str(team),
                                _safe(r.get("pool", 0)),
                                _safe(r.get("collected", 0)),
                                _safe(r.get("balance", 0)),
                                int(_safe(r.get("leads", 0))),
                                int(_safe(r.get("leads_48", 0))),
                                _safe(r.get("bal_48hr", 0)),
                                _pct(_safe(r.get("bal_48hr", 0)), _safe(r.get("balance", 0))),
                                _safe(r.get("prev_bal", 0)),
                                int(_safe(r.get("prev_leads", 0))),
                                _safe(r.get("grand_bal", 0)),
                                int(_safe(r.get("grand_leads", 0))),
                            ]
                            fill = ALT_FILL if alt else WHITE_FILL
                            alt = not alt
                            for c_idx, v_ in enumerate(vals, 1):
                                cell = ws.cell(row_num, c_idx, v_)
                                cell.fill = fill
                                cell.font = DATA_FONT
                                cell.border = BORDER
                                cell.alignment = LEFT if c_idx <= 2 else CENTER
                            row_num += 1
                            for k in num_keys:
                                t[k] += _safe(r.get(k, 0))
                    else:
                        for k in num_keys:
                            t[k] = pd.to_numeric(t_df[k], errors='coerce').fillna(0).sum() if k in t_df.columns else 0

                    # Team total row
                    t_vals = ([f"{team} Total"] + (["—"] if mode == "caller" else []) +
                              [t["pool"], t["collected"], t["balance"], int(t["leads"]),
                               int(t["leads_48"]), t["bal_48hr"], _pct(t["bal_48hr"], t["balance"]),
                               t["prev_bal"], int(t["prev_leads"]), t["grand_bal"], int(t["grand_leads"])])
                    for c_idx, v_ in enumerate(t_vals, 1):
                        cell = ws.cell(row_num, c_idx, v_)
                        cell.fill = TEAM_FILL
                        cell.font = BOLD_WHITE
                        cell.border = BORDER
                        cell.alignment = LEFT if c_idx == 1 else CENTER
                    row_num += 1

                    for k in num_keys:
                        v[k] += t[k]

                # Vertical total row
                v_vals = ([f"{vert} Total"] + (["—"] if mode == "caller" else []) +
                          [v["pool"], v["collected"], v["balance"], int(v["leads"]),
                           int(v["leads_48"]), v["bal_48hr"], _pct(v["bal_48hr"], v["balance"]),
                           v["prev_bal"], int(v["prev_leads"]), v["grand_bal"], int(v["grand_leads"])])
                for c_idx, v_ in enumerate(v_vals, 1):
                    cell = ws.cell(row_num, c_idx, v_)
                    cell.fill = VERT_FILL
                    cell.font = BOLD_WHITE
                    cell.border = BORDER
                    cell.alignment = LEFT if c_idx == 1 else CENTER
                row_num += 1

                for k in num_keys:
                    g[k] += v[k]

            # Grand total row
            g_vals = (["Grand Total"] + (["—"] if mode == "caller" else []) +
                      [g["pool"], g["collected"], g["balance"], int(g["leads"]),
                       int(g["leads_48"]), g["bal_48hr"], _pct(g["bal_48hr"], g["balance"]),
                       g["prev_bal"], int(g["prev_leads"]), g["grand_bal"], int(g["grand_leads"])])
            for c_idx, v_ in enumerate(g_vals, 1):
                cell = ws.cell(row_num, c_idx, v_)
                cell.fill = GRAND_FILL
                cell.font = BOLD_WHITE
                cell.border = BORDER
                cell.alignment = LEFT if c_idx == 1 else CENTER

            # Column widths
            col_widths = [28] + ([18] if mode == "caller" else []) + [16, 16, 16, 8, 10, 16, 14, 16, 10, 16, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        # Tab 1 — Teamwise
        ws1 = wb.active
        ws1.title = f"Teamwise {curr_label[:3]}{curr_label[-4:]}+{prev_label[:3]}{prev_label[-4:]}"[:31]
        ws1.freeze_panes = "A3"
        write_pending_sheet(ws1, "team")

        # Tab 2 — Callerwise
        ws2 = wb.create_sheet(f"Callerwise {curr_label[:3]}{curr_label[-4:]}+{prev_label[:3]}{prev_label[-4:]}"[:31])
        ws2.freeze_panes = "A3"
        write_pending_sheet(ws2, "caller")

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    def build_pending_leads_excel(pend_curr, pend_prev, meta_map_pending, curr_label, prev_label):
        """
        Returns bytes of an xlsx with 2 tabs — one per month — with full lead details
        including Caller Name, Team Name, Vertical, Course Price, Revenue Collected, Balance.
        """
        HDR_FILL  = PatternFill("solid", start_color="064e3b", end_color="064e3b")
        ALT_FILL  = PatternFill("solid", start_color="f0fdf4", end_color="f0fdf4")
        WHITE_FILL= PatternFill("solid", start_color="ffffff", end_color="ffffff")
        HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        DATA_FONT = Font(name="Arial", size=9)
        CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        BORDER    = Border(
            left=Side(style='thin', color='D1FAE5'),
            right=Side(style='thin', color='D1FAE5'),
            top=Side(style='thin', color='D1FAE5'),
            bottom=Side(style='thin', color='D1FAE5'),
        )

        COLS = ["DATE", "NAME", "CONTACT NO", "EMAIL ID", "COURSE",
                "CALLER NAME", "TEAM NAME", "VERTICAL",
                "COURSE PRICE (₹)", "REVENUE COLLECTED (₹)", "BALANCE (₹)"]
        COL_WIDTHS = [12, 28, 14, 32, 32, 22, 22, 18, 16, 20, 14]

        # Build meta lookup: merge_key → {Team Name, Vertical, Caller Name}
        meta_lkp = {}
        if not meta_map_pending.empty:
            for _, row in meta_map_pending.iterrows():
                k = str(row.get('merge_key', '')).strip().lower()
                meta_lkp[k] = {
                    'Caller Name': row.get('Caller Name', '—'),
                    'Team Name':   row.get('Team Name',   '—'),
                    'Vertical':    row.get('Vertical',    '—'),
                }

        def enrich(df):
            if df.empty:
                return df
            d = df.copy()
            d['_mk'] = d['Caller_name'].astype(str).str.strip().str.lower()
            d['_cn']   = d['_mk'].map(lambda k: meta_lkp.get(k, {}).get('Caller Name', d.loc[d['_mk']==k, 'Caller_name'].iloc[0] if not d[d['_mk']==k].empty else '—'))
            d['_team'] = d['_mk'].map(lambda k: meta_lkp.get(k, {}).get('Team Name', '—'))
            d['_vert'] = d['_mk'].map(lambda k: meta_lkp.get(k, {}).get('Vertical',  '—'))
            return d

        def write_leads_sheet(ws, df, label):
            ws.row_dimensions[1].height = 28
            for c_idx, h in enumerate(COLS, 1):
                cell = ws.cell(1, c_idx, h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER
            for i, w in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            if df.empty:
                ws.cell(2, 1, f"No pending leads found for {label}")
                return

            df = enrich(df)
            alt = False
            for r_idx, (_, r) in enumerate(df.sort_values('Date').iterrows(), 2):
                fill = ALT_FILL if alt else WHITE_FILL
                alt  = not alt
                _raw_phone = str(r.get('Contact_No', '') or '').strip().replace(' ', '')
                try:
                    _phone_val = int(_raw_phone) if _raw_phone.lstrip('+').isdigit() else _raw_phone
                except:
                    _phone_val = _raw_phone
                vals = [
                    str(r.get('Date', '')),
                    str(r.get('Name', '')),
                    _phone_val,
                    str(r.get('Email_Id', '')),
                    str(r.get('Course', '')),
                    str(r.get('_cn', r.get('Caller_name', '—'))),
                    str(r.get('_team', '—')),
                    str(r.get('_vert', '—')),
                    float(r.get('Course_Price', 0) or 0),
                    float(r.get('Fee_paid', 0) or 0),
                    float(r.get('balance', 0) or 0),
                ]
                for c_idx, v_ in enumerate(vals, 1):
                    cell = ws.cell(r_idx, c_idx, v_)
                    cell.fill = fill
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    cell.alignment = LEFT if c_idx <= 6 else CENTER

        wb = Workbook()
        ws1 = wb.active
        ws1.title = f"Pending {curr_label[:3]} {curr_label[-4:]}"[:31]
        write_leads_sheet(ws1, pend_curr, curr_label)

        ws2 = wb.create_sheet(f"Pending {prev_label[:3]} {prev_label[-4:]}"[:31])
        write_leads_sheet(ws2, pend_prev, prev_label)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    def build_drop_leads_excel(drop_df, c_start, c_end, p_start, p_end, curr_label, prev_label, meta_map_pending=None):
        """
        Returns bytes of an xlsx with 2 tabs:
          Tab 1 — Current month dropped leads
          Tab 2 — Previous month dropped leads
        Columns: DATE, EMAIL, PHONE NUMBER, CALLER NAME, TEAM, VERTICAL
        """
        if meta_map_pending is None:
            meta_map_pending = pd.DataFrame()

        HDR_FILL   = PatternFill("solid", start_color="7c2d12", end_color="7c2d12")
        ALT_FILL   = PatternFill("solid", start_color="fff7ed", end_color="fff7ed")
        WHITE_FILL = PatternFill("solid", start_color="ffffff", end_color="ffffff")
        HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        DATA_FONT  = Font(name="Arial", size=9)
        CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT       = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        BORDER     = Border(
            left=Side(style='thin', color='FED7AA'),
            right=Side(style='thin', color='FED7AA'),
            top=Side(style='thin', color='FED7AA'),
            bottom=Side(style='thin', color='FED7AA'),
        )

        COLS       = ["DATE", "EMAIL", "PHONE NUMBER", "CALLER NAME", "TEAM", "VERTICAL"]
        COL_WIDTHS = [14, 36, 16, 28, 22, 18]

        e_col = next((c for c in drop_df.columns if 'email' in c.lower() and ('drop' in c.lower() or 'student' in c.lower())), None)
        p_col = next((c for c in drop_df.columns if 'phone' in c.lower() or ('number' in c.lower() and 'drop' in c.lower())), None)
        t_col = next((c for c in drop_df.columns if 'timestamp' in c.lower()), None)

        def write_drop_sheet(ws, df_slice, label):
            ws.row_dimensions[1].height = 28
            for c_idx, h in enumerate(COLS, 1):
                cell = ws.cell(1, c_idx, h)
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = CENTER
                cell.border = BORDER
            for i, w in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w
            ws.freeze_panes = "A2"

            if df_slice.empty:
                ws.cell(2, 1, f"No dropped leads for {label}")
                return

            alt = False
            for r_idx, (_, r) in enumerate(df_slice.iterrows(), 2):
                fill = ALT_FILL if alt else WHITE_FILL
                alt  = not alt
                drop_date = r.get('drop_date', '')
                if hasattr(drop_date, 'date'):
                    drop_date = drop_date.date()
                _dp = str(r.get('drop_phone', '') or '').strip().replace(' ', '')
                try:
                    _dp_val = int(_dp) if _dp.lstrip('+').isdigit() else _dp
                except:
                    _dp_val = _dp
                vals = [
                    str(drop_date),
                    str(r.get('drop_email', '')),
                    _dp_val,
                    str(r.get('attributed_caller', '—')),
                    str(r.get('_team', '—')),
                    str(r.get('_vert', '—')),
                ]
                for c_idx, v_ in enumerate(vals, 1):
                    cell = ws.cell(r_idx, c_idx, v_)
                    cell.fill = fill
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    cell.alignment = LEFT

        wb  = Workbook()
        ws1 = wb.active
        ws1.title = f"{curr_label[:3]} {curr_label[-4:]} Drops"[:31]

        # Attribute caller to each drop row using revenue data
        df_work = drop_df.copy()
        df_work['drop_d'] = pd.to_datetime(df_work['drop_date'], errors='coerce').dt.date

        # Build email/phone → caller attribution from the revenue table
        try:
            _rev_attr = fetch_both_months_rev(p_start, c_end)
            if not _rev_attr.empty:
                _new_enr_attr = _rev_attr[_rev_attr['is_new']].copy()
                _email_to_caller = _new_enr_attr.drop_duplicates('Email_Id_norm').set_index('Email_Id_norm')['Caller_name'].to_dict()
                _phone_to_caller = _new_enr_attr.drop_duplicates('Contact_No_norm').set_index('Contact_No_norm')['Caller_name'].to_dict()

                def _get_caller(row):
                    e = str(row.get('drop_email', '')).strip().lower()
                    p = str(row.get('drop_phone', '')).strip()
                    if e and e in _email_to_caller: return _email_to_caller[e]
                    if p and p in _phone_to_caller: return _phone_to_caller[p]
                    return 'Unknown'

                df_work['attributed_caller'] = df_work.apply(_get_caller, axis=1)

                # Build caller → team/vertical lookup from meta
                _meta_lkp_drop = {}
                if not meta_map_pending.empty:
                    for _, _mr in meta_map_pending.iterrows():
                        _mk = str(_mr.get('merge_key', '')).strip().lower()
                        _meta_lkp_drop[_mk] = {
                            'Team Name': _mr.get('Team Name', '—'),
                            'Vertical':  _mr.get('Vertical',  '—'),
                        }
                df_work['_team'] = df_work['attributed_caller'].apply(
                    lambda c: _meta_lkp_drop.get(str(c).strip().lower(), {}).get('Team Name', '—'))
                df_work['_vert'] = df_work['attributed_caller'].apply(
                    lambda c: _meta_lkp_drop.get(str(c).strip().lower(), {}).get('Vertical', '—'))
            else:
                df_work['attributed_caller'] = 'Unknown'
                df_work['_team'] = '—'
                df_work['_vert'] = '—'
        except Exception:
            df_work['attributed_caller'] = 'Unknown'
            df_work['_team'] = '—'
            df_work['_vert'] = '—'

        curr_drops = df_work[(df_work['drop_d'] >= c_start) & (df_work['drop_d'] <= c_end)].copy()
        prev_drops = df_work[(df_work['drop_d'] >= p_start) & (df_work['drop_d'] <= p_end)].copy()

        write_drop_sheet(ws1, curr_drops.sort_values('drop_d'), curr_label)

        ws2 = wb.create_sheet(f"{prev_label[:3]} {prev_label[-4:]} Drops"[:31])
        write_drop_sheet(ws2, prev_drops.sort_values('drop_d'), prev_label)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


    # ─────────────────────────────────────────────
    # SIDEBAR
    # ─────────────────────────────────────────────

    min_d, max_d = get_available_dates()

    st.sidebar.markdown("""
    <div style='padding:.5rem 0 .4rem; text-align:center;'>
        <div style='font-size:.72rem; font-weight:700; text-transform:uppercase;
                    letter-spacing:1px; color:var(--text-muted,#6B7280);'>Report Controls</div>
    </div>
    """, unsafe_allow_html=True)

    min_d = pd.Timestamp(min_d).date()
    max_d = pd.Timestamp(max_d).date()

    _s            = date(max_d.year, max_d.month, 1)
    _next_m       = (_s.replace(day=28) + timedelta(days=4)).replace(day=1)
    _month_end    = _next_m - timedelta(days=1)
    default_start = max(_s, min_d)
    default_end   = min(_month_end, max_d)
    if default_start > default_end:
        default_start = default_end

    selected_dates = st.sidebar.date_input(
        "📅 Date Range",
        value=(default_start, default_end),
        min_value=min_d, max_value=max_d, format="DD-MM-YYYY",
        key="rev_date_input"
    )

    if isinstance(selected_dates, tuple) and len(selected_dates) == 2:
        start_date, end_date = selected_dates
    else:
        start_date = end_date = selected_dates if not isinstance(selected_dates, tuple) else selected_dates[0]

    teams_meta, verticals_meta, df_team_mapping = get_metadata()
    # ── Role-aware sidebar filters ──────────────────────────
    _role     = st.session_state.get('rf_role', 'admin')
    _rf_teams = st.session_state.get('rf_teams', [])
    _rf_cname = st.session_state.get('rf_caller_name', '')

    if _role == 'admin':
        selected_vertical = st.sidebar.multiselect("👑 Filter by Vertical", options=verticals_meta, key="rev_vert_multiselect")
        selected_team     = st.sidebar.multiselect("👥 Filter by Team",     options=teams_meta,    key="rev_team_multiselect")
        search_query      = st.sidebar.text_input("👤 Search Caller Name",                         key="rev_search_input")
    elif _role == 'vertical_head':
        selected_team     = _rf_teams
        selected_vertical = []
        search_query      = st.sidebar.text_input("👤 Search Caller Name", key="rev_search_input")
        st.sidebar.caption(f"🔒 Showing: {', '.join(_rf_teams)}")
    elif _role in ('tl', 'trainer'):
        selected_team     = _rf_teams
        selected_vertical = []
        search_query      = st.sidebar.text_input("👤 Search Caller Name", key="rev_search_input")
        st.sidebar.caption(f"🔒 Team: {', '.join(_rf_teams)}")
    else:  # caller
        selected_team     = []
        selected_vertical = []
        search_query      = _rf_cname
        st.sidebar.caption(f"👤 Viewing: {_rf_cname}")

    gen_report = st.sidebar.button("💰 Generate Revenue Report", key="rev_gen_btn")
    gen_pending = st.sidebar.button("📊 Generate Callerwise Pending", key="rev_pending_btn")
    st.sidebar.download_button(
        label="📖 Metrics Guide (PDF)",
        data=generate_helper_pdf_bytes(),
        file_name="Revenue_Metrics_Logic_Guide.pdf",
        mime="application/pdf",
        key="dl_helper_pdf"
    )
  

    # ─────────────────────────────────────────────
    # HEADER BANNER
    # ─────────────────────────────────────────────

    last_update_str = get_last_update()
    display_start   = start_date.strftime('%d-%m-%Y')
    display_end     = end_date.strftime('%d-%m-%Y')

    st.markdown(f"""
    <div class="rv-header">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:.75rem;">
            <div>
                <div class="rv-title">💰 REVENUE METRICS</div>
                <div class="rv-subtitle">REVENUE PERIOD&nbsp;·&nbsp; {display_start} to {display_end}</div>
            </div>
            <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.25rem;">
                <span class="rv-badge"><span class="rv-pulse"></span>LIVE REVENUE DATA</span>
                <span class="rv-badge">🕐 UPDATED AT: {last_update_str}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)


    # ─────────────────────────────────────────────
    # TABS
    # ─────────────────────────────────────────────

    tab1, tab2, tab3 = st.tabs(["💰 Revenue Dashboard", "🧠 Insights & Leaderboard", "📊 Callerwise Pending Revenue"])


    # ══════════════════════════════════════════════
    # TAB 1 — REVENUE DASHBOARD
    # ══════════════════════════════════════════════

    with tab1:
        if gen_report:
            with st.spinner("Fetching revenue data…"):
                df_raw = fetch_revenue_data(start_date, end_date)

                if df_raw.empty:
                    st.warning("No revenue records found for the selected period (Fee Paid > 0).")
                else:
                    # Merge team info
                    df = pd.merge(
                        df_raw,
                        df_team_mapping[['merge_key','Caller Name','Team Name','Vertical']].drop_duplicates('merge_key'),
                        on='merge_key', how='left'
                    )
                    df['Caller_name'] = df['Caller Name'].fillna(df['Caller_name'])

                    # Apply filters
                    if selected_team:
                        df = df[df['Team Name'].isin(selected_team)]
                    if selected_vertical:
                        df = df[df['Vertical'].isin(selected_vertical)]
                    if search_query:
                        df = df[df['Caller_name'].str.contains(search_query, case=False, na=False)]

                    if df.empty:
                        st.error("No records match the selected filters.")
                    else:
                        # ── Classify callers ──
                        calling_df, collection_df, both_df = classify_and_process(
                            df, df_team_mapping, start_date, end_date
                        )

                        # ── Summary metrics ──
                        metrics = compute_summary_metrics(df)

                        # ── TOP 3 HIGHLIGHTS ──
                        section_header("🏆 TOP 3 REVENUE HIGHLIGHTS")
                        top_cols = st.columns(3)

                        # Top Revenue Caller — from calling_df
                        with top_cols[0]:
                            if not calling_df.empty:
                                top_c = calling_df.iloc[0]
                                st.markdown(f"""
                                <div class="metric-card" style="border-top:3px solid var(--accent-primary);">
                                    <div class="metric-label">🥇 TOP REVENUE — CALLER</div>
                                    <div class="metric-value" style="font-size:1.1rem;">{top_c['CALLER NAME']}</div>
                                    <div class="metric-delta">{fmt_inr(top_c['raw_calling_rev'])} Calling Revenue</div>
                                </div>""", unsafe_allow_html=True)
                            else:
                                st.markdown("""<div class="metric-card" style="border-top:3px solid var(--gold);">
                                    <div class="metric-label">🥇 TOP REVENUE — CALLER</div>
                                    <div class="metric-value" style="font-size:1rem;">No Data</div>
                                </div>""", unsafe_allow_html=True)

                        # Most Enrollments — from all agent types combined
                        with top_cols[1]:
                            _agent_list = [d for d in [calling_df, collection_df, both_df] if not d.empty]
                            all_agents = pd.concat(_agent_list, ignore_index=True) if _agent_list else pd.DataFrame()
                            if not all_agents.empty and all_agents['raw_enrollments'].max() > 0:
                                top_enr = all_agents.sort_values('raw_enrollments', ascending=False).iloc[0]
                                st.markdown(f"""
                                <div class="metric-card" style="border-top:3px solid var(--accent-primary);">
                                    <div class="metric-label">🎓 MOST ENROLLMENTS</div>
                                    <div class="metric-value" style="font-size:1.1rem;">{top_enr['CALLER NAME']}</div>
                                    <div class="metric-delta">{top_enr['raw_enrollments']} New Enrollments</div>
                                </div>""", unsafe_allow_html=True)
                            else:
                                st.markdown("""<div class="metric-card" style="border-top:3px solid var(--silver);">
                                    <div class="metric-label">🎓 MOST ENROLLMENTS</div>
                                    <div class="metric-value" style="font-size:1rem;">No Data</div>
                                </div>""", unsafe_allow_html=True)

                        # Top Revenue Collection Caller — from collection_df
                        with top_cols[2]:
                            _coll_list = [d for d in [collection_df] if not d.empty]
                            coll_combined = pd.concat(_coll_list, ignore_index=True) if _coll_list else pd.DataFrame()
                            if not coll_combined.empty:
                                top_coll = coll_combined.sort_values('raw_collection_rev', ascending=False).iloc[0]
                                st.markdown(f"""
                                <div class="metric-card" style="border-top:3px solid var(--accent-primary);">
                                    <div class="metric-label">🥇 TOP REVENUE — COLLECTION</div>
                                    <div class="metric-value" style="font-size:1.1rem;">{top_coll['CALLER NAME']}</div>
                                    <div class="metric-delta">{fmt_inr(top_coll['raw_collection_rev'])} Collection Revenue</div>
                                </div>""", unsafe_allow_html=True)
                            else:
                                st.markdown("""<div class="metric-card" style="border-top:3px solid var(--bronze);">
                                    <div class="metric-label">🥇 TOP REVENUE — COLLECTION</div>
                                    <div class="metric-value" style="font-size:1rem;">No Data</div>
                                </div>""", unsafe_allow_html=True)

                        # ── SUMMARY KPI CARDS ──
                        section_header("💵 REVENUE SUMMARY METRICS")

                        kpis = [
                            ("Total Revenue (EXCL. Services)",                   fmt_inr(metrics['total_rev']),      "💰"),
                            ("Calling Revenue (INCL. Funnel)",                  fmt_inr(metrics['calling_rev']),    "📞"),
                            ("Bootcamp-Direct Revenue",         fmt_inr(metrics['bootcamp_direct_rev']),     "🎓"),
                            ("Bootcamp-Collection Revenue",              fmt_inr(metrics['collection_rev']), "🏦"),
                            ("Community Revenue (Direct + Collection)",               fmt_inr(metrics['community_rev']),  "🌐"),
                            ("Direct/Other Revenue (INCL. Funnel)",                  fmt_inr(metrics['direct_rev']),     "🎯"),
                            ("Details Not Available/ Not Updated Yet",           fmt_inr(metrics['dna_rev']),        "❓"),
                        ]

                        cols = st.columns(len(kpis))
                        for col, (label, val, icon) in zip(cols, kpis):
                            with col:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">{icon} {label}</div>
                                    <div class="metric-value">{val}</div>
                                </div>""", unsafe_allow_html=True)

                        st.divider()

                        # ══════════════════════════════════════
                        # ENROLLMENT SUMMARY METRICS
                        # ══════════════════════════════════════
                        section_header("🎓 ENROLLMENT SUMMARY METRICS")

                        enr_metrics = compute_enrollment_metrics(df)

                        enr_kpis = [
                            ("Total Enrollments",             enr_metrics['total_enr'],     "🎯"),
                            ("Caller Enrollments",            enr_metrics['caller_enr'],    "📞"),
                            ("Direct Enrollments",            enr_metrics['direct_enr'],    "🎯"),
                            ("Bootcamp-Direct Enrollments",   enr_metrics['bootcamp_enr'],  "🎓"),
                            ("Community-Direct Enrollments",  enr_metrics['community_enr'], "🌐"),
                        ]

                        enr_cols = st.columns(len(enr_kpis))
                        for col, (label, val, icon) in zip(enr_cols, enr_kpis):
                            with col:
                                st.markdown(f"""
                                <div class="metric-card">
                                    <div class="metric-label">{icon} {label}</div>
                                    <div class="metric-value">{val}</div>
                                </div>""", unsafe_allow_html=True)

                        st.divider()

                        # ── Download Summary Cards ──
                        _rev_summary = pd.DataFrame([{
                            'Metric'  : label,
                            'Amount'  : val
                        } for label, val, _ in [
                            ("Total Revenue (EXCL. Services)",          metrics['total_rev'],           ""),
                            ("Calling Revenue (INCL. Funnel)",          metrics['calling_rev'],         ""),
                            ("Bootcamp-Direct Revenue",                 metrics['bootcamp_direct_rev'], ""),
                            ("Bootcamp-Collection Revenue",             metrics['collection_rev'],      ""),
                            ("Community Revenue (Direct+Collection)",   metrics['community_rev'],       ""),
                            ("Direct/Other Revenue (INCL. Funnel)",     metrics['direct_rev'],          ""),
                            ("Details Not Available / Not Updated Yet", metrics['dna_rev'],             ""),
                        ]])

                        _enr_summary = pd.DataFrame([{
                            'Metric': label,
                            'Count' : val
                        } for label, val, _ in [
                            ("Total Enrollments",            enr_metrics['total_enr'],     ""),
                            ("Caller Enrollments",           enr_metrics['caller_enr'],    ""),
                            ("Direct Enrollments",           enr_metrics['direct_enr'],    ""),
                            ("Bootcamp-Direct Enrollments",  enr_metrics['bootcamp_enr'],  ""),
                            ("Community-Direct Enrollments", enr_metrics['community_enr'], ""),
                        ]])

                        _combined = pd.concat([
                            _rev_summary.rename(columns={'Amount': 'Value'}),
                            _enr_summary.rename(columns={'Count': 'Value'})
                        ], ignore_index=True)

                        st.download_button(
                            label="📥 Download Revenue & Enrollment Summary",
                            data=_combined.to_csv(index=False).encode('utf-8'),
                            file_name=f"Summary_{display_start}_to_{display_end}.csv",
                            mime='text/csv',
                            key='dl_summary'
                        )

                        st.divider()

                        # ══════════════════════════════════════
                        # TABLE 1 — CALLER REVENUE PERFORMANCE
                        # ══════════════════════════════════════
                        section_header("📞 CALLER REVENUE PERFORMANCE TABLE")

                        calling_display_cols = [
                            'DESIGNATION', 'CALLER NAME', 'TEAM', 'VERTICAL',
                            'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)', 'ENROLLMENTS',
                            'ENROLLMENT REV', 'BALANCE REV',
                            'CALLING REVENUE', 'ACHIEVEMENT %'
                        ]

                        if not calling_df.empty:
                            calling_totals = {
                                'TOTAL TARGET (₹)'    : fmt_inr(calling_df['raw_target'].sum()),
                                'TILL DAY TARGET (₹)' : fmt_inr(calling_df['TILL DAY TARGET (₹)'].sum()),
                                'ENROLLMENTS'         : int(calling_df['raw_enrollments'].sum()),
                                'ENROLLMENT REV'      : fmt_inr(calling_df['ENROLLMENT REV'].sum()),
                                'BALANCE REV'         : fmt_inr(calling_df['BALANCE REV'].sum()),
                                'CALLING REVENUE'     : fmt_inr(calling_df['CALLING REVENUE'].sum()),
                                'ACHIEVEMENT %'       : f"{round(calling_df['CALLING REVENUE'].sum() / calling_df['raw_target'].sum() * 100, 1) if calling_df['raw_target'].sum() > 0 else 0}%",
                            }
                        else:
                            calling_totals = {}

                        render_perf_table(
                            calling_df, calling_display_cols,
                            calling_totals, 'raw_calling_rev', 'calling'
                        )
                        if not calling_df.empty:
                            st.download_button(
                                label="📥 Download Caller Revenue CSV",
                                data=calling_df[calling_display_cols].to_csv(index=False).encode('utf-8'),
                                file_name=f"Caller_Revenue_{display_start}_to_{display_end}.csv",
                                mime='text/csv', key='dl_calling'
                            )

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE 2
                        # ══════════════════════════════════════════════
                        section_header("🏦 COLLECTION CALLER REVENUE PERFORMANCE TABLE")

                        collection_display_cols = [
                            'DESIGNATION', 'CALLER NAME', 'TEAM', 'VERTICAL',
                            'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)', 'ENROLLMENTS',
                            'CALLING REVENUE',
                            'COMMUNITY COLLECTION', 'BOOTCAMP COLLECTION',
                            'COLLECTION REVENUE'
                        ]

                        if not collection_df.empty:
                            collection_totals = {
                                'TOTAL TARGET (₹)'    : fmt_inr(collection_df['raw_target'].sum()),
                                'TILL DAY TARGET (₹)' : fmt_inr(collection_df['TILL DAY TARGET (₹)'].sum()),
                                'ENROLLMENTS'         : int(collection_df['raw_enrollments'].sum()),
                                'CALLING REVENUE'     : fmt_inr(collection_df['CALLING REVENUE'].sum()),
                                'COMMUNITY COLLECTION': fmt_inr(collection_df['COMMUNITY COLLECTION'].sum()),
                                'BOOTCAMP COLLECTION' : fmt_inr(collection_df['BOOTCAMP COLLECTION'].sum()),
                                'COLLECTION REVENUE'  : fmt_inr(collection_df['COLLECTION REVENUE'].sum()),
                            }
                        else:
                            collection_totals = {}

                        render_perf_table(
                            collection_df, collection_display_cols,
                            collection_totals, 'raw_collection_rev', 'collection'
                        )
                        if not collection_df.empty:
                            st.download_button(
                                label="📥 Download Collection Revenue CSV",
                                data=collection_df[collection_display_cols].to_csv(index=False).encode('utf-8'),
                                file_name=f"Collection_Revenue_{display_start}_to_{display_end}.csv",
                                mime='text/csv', key='dl_collection'
                            )

                        st.divider()

                        # ══════════════════════════════════════════════════════════
                        # TABLE 3
                        # ══════════════════════════════════════════════════════════
                        section_header("📞🏦 CALLING + COLLECTION CALLER REVENUE PERFORMANCE TABLE")

                        both_display_cols = [
                            'DESIGNATION', 'CALLER NAME', 'TEAM', 'VERTICAL',
                            'TOTAL TARGET (₹)', 'TILL DAY TARGET (₹)', 'ENROLLMENTS',
                            'CALLING REVENUE', 'COMMUNITY COLLECTION', 'BOOTCAMP COLLECTION',
                            'TOTAL REVENUE', 'ACHIEVEMENT %'
                        ]

                        if not both_df.empty:
                            _both_target     = both_df['raw_target'].sum()
                            _both_till       = both_df['TILL DAY TARGET (₹)'].sum()
                            _both_enr        = int(both_df['raw_enrollments'].sum())
                            _both_comm       = both_df['COMMUNITY COLLECTION'].sum()
                            _both_boot       = both_df['BOOTCAMP COLLECTION'].sum()
                            _both_calling    = both_df['CALLING REVENUE'].sum()
                            _both_collection = both_df['COLLECTION REVENUE'].sum()
                            _both_total      = both_df['raw_revenue'].sum()
                            _both_ach        = round(_both_total / _both_target * 100, 1) if _both_target > 0 else 0
                            both_totals = {
                                'TOTAL TARGET (₹)'    : fmt_inr(_both_target),
                                'TILL DAY TARGET (₹)' : fmt_inr(_both_till),
                                'ENROLLMENTS'         : _both_enr,
                                'COMMUNITY COLLECTION': fmt_inr(_both_comm),
                                'BOOTCAMP COLLECTION' : fmt_inr(_both_boot),
                                'CALLING REVENUE'     : fmt_inr(_both_calling),
                                'COLLECTION REVENUE'  : fmt_inr(_both_collection),
                                'TOTAL REVENUE'       : fmt_inr(_both_total),
                                'ACHIEVEMENT %'       : f"{_both_ach}%",
                            }
                        else:
                            both_totals = {}
                        render_perf_table(
                            both_df, both_display_cols,
                            both_totals, 'raw_revenue', 'both'
                        )
                        if not both_df.empty:
                            st.download_button(
                                label="📥 Download Calling+Collection Revenue CSV",
                                data=both_df[both_display_cols].to_csv(index=False).encode('utf-8'),
                                file_name=f"Both_Revenue_{display_start}_to_{display_end}.csv",
                                mime='text/csv', key='dl_both'
                            )

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>💰</div>
                <div style='font-size:.9rem;font-weight:600;'>Select a date range and click <b>Generate Revenue Report</b></div>
            </div>""", unsafe_allow_html=True)


    # ══════════════════════════════════════════════
    # TAB 2 — INSIGHTS & LEADERBOARD
    # ══════════════════════════════════════════════

    with tab2:
        if gen_report:
            with st.spinner("Analysing revenue patterns…"):
                df_raw = fetch_revenue_data(start_date, end_date)
                if df_raw.empty:
                    st.warning("No revenue data found for the selected period.")
                else:
                    df_ins = pd.merge(
                        df_raw,
                        df_team_mapping[['merge_key','Caller Name','Team Name','Vertical']].drop_duplicates('merge_key'),
                        on='merge_key', how='left'
                    )
                    df_ins['Caller_name'] = df_ins['Caller Name'].fillna(df_ins['Caller_name'])

                    if selected_team:
                        df_ins = df_ins[df_ins['Team Name'].isin(selected_team)]
                    if selected_vertical:
                        df_ins = df_ins[df_ins['Vertical'].isin(selected_vertical)]
                    if search_query:
                        df_ins = df_ins[df_ins['Caller_name'].str.contains(search_query, case=False, na=False)]

                    calling_ins, collection_ins, both_ins = classify_and_process(
                        df_ins, df_team_mapping, start_date, end_date
                    )

                    _ins_list = [d for d in [calling_ins, collection_ins, both_ins] if not d.empty]
                    all_agents_ins = pd.concat(_ins_list, ignore_index=True) if _ins_list else pd.DataFrame()

                    if all_agents_ins.empty:
                        st.warning("Not enough data for insights with current filters.")
                    else:
                        # ── 6 INSIGHTS ──
                        section_header("🧠 REVENUE INSIGHTS")
                        insights = compute_revenue_insights(
                            df_ins, calling_ins, collection_ins, both_ins,
                            start_date, end_date
                        )

                        # ── Pending Revenue Insights (unfiltered, loaded fresh) ──
                        try:
                            _c_start, _c_end, _p_start, _p_end = pending_months()
                            _curr_label_ins = _c_start.strftime("%B %Y")
                            _prev_label_ins = _p_start.strftime("%B %Y")
                            _, _, _df_meta_ins = get_metadata()
                            _meta_map_ins = (_df_meta_ins.sort_values('Month', ascending=False)
                                             .drop_duplicates(subset=['merge_key'], keep='first')
                                             [['merge_key', 'Caller Name', 'Team Name', 'Vertical']].copy())
                            _drop_df_ins  = load_drop_leads()
                            _excl_e = set(_drop_df_ins['drop_email'].dropna().astype(str).unique()) if not _drop_df_ins.empty else set()
                            _excl_p = set(_drop_df_ins['drop_phone'].dropna().astype(str).unique()) if not _drop_df_ins.empty else set()
                            _df_both_ins  = fetch_both_months_rev(_p_start, _c_end)

                            _pending_insight = None
                            _drop_insight    = None

                            if not _df_both_ins.empty:
                                _df_curr_ins = _df_both_ins[_df_both_ins['Date'] >= _c_start].copy()
                                _df_prev_ins = _df_both_ins[(_df_both_ins['Date'] >= _p_start) & (_df_both_ins['Date'] <= _p_end)].copy()
                                _pc = pending_leads_for_month(_df_curr_ins, _excl_e, _excl_p, _df_both_ins)
                                _pp = pending_leads_for_month(_df_prev_ins, _excl_e, _excl_p, _df_both_ins)
                                _cc = build_combined_agg(_pc, _pp, _meta_map_ins, _df_both_ins)

                                if not _cc.empty:
                                    _comb_ins = _cc[_cc['grand_bal'] > 0].copy()
                                    if not _comb_ins.empty:
                                        _top_bal = _comb_ins.sort_values('grand_bal', ascending=False).iloc[0]
                                        _team_bal = (_comb_ins.groupby('Team Name')['grand_bal'].sum()
                                                     .sort_values(ascending=False))
                                        _top_team_bal = _team_bal.index[0] if not _team_bal.empty else '—'
                                        _top_team_bal_amt = _team_bal.iloc[0] if not _team_bal.empty else 0
                                        _pending_body = (
                                            f"{_top_bal['Caller_name']} has the highest combined pending balance of "
                                            f"{fmt_inr(_top_bal['grand_bal'])} across {int(_top_bal['grand_leads'])} lead(s) "
                                            f"({_curr_label_ins} + {_prev_label_ins}). "
                                            f"Top team by pending balance is {_top_team_bal} "
                                            f"with {fmt_inr(_top_team_bal_amt)} outstanding."
                                        )
                                        _pending_insight = {
                                            "type": "warn", "icon": "💰",
                                            "title": (f"Highest Pending Balance — {_top_bal['Caller_name']} "
                                                      f"({_top_bal.get('Team Name','—')})"),
                                            "body": _pending_body
                                        }

                            if not _drop_df_ins.empty and not _df_both_ins.empty:
                                _drop_agg_ins = attribute_drops_to_callers(
                                    _drop_df_ins, _df_both_ins, _meta_map_ins,
                                    _c_start, _c_end, _p_start, _p_end,
                                    _curr_label_ins, _prev_label_ins
                                )
                                if not _drop_agg_ins.empty:
                                    _drop_agg_ins = _drop_agg_ins[_drop_agg_ins['Team Name'] != 'Others'].copy()
                                    if not _drop_agg_ins.empty:
                                        _top_drop = _drop_agg_ins.sort_values('total_drops', ascending=False).iloc[0]
                                        _team_drops = (_drop_agg_ins.groupby('Team Name')['total_drops'].sum()
                                                       .sort_values(ascending=False))
                                        _top_team_drop = _team_drops.index[0] if not _team_drops.empty else '—'
                                        _top_team_drop_cnt = int(_team_drops.iloc[0]) if not _team_drops.empty else 0
                                        _drop_body = (
                                            f"{_top_drop['Caller_name']} has the highest combined drop cases: "
                                            f"{int(_top_drop['curr_drops'])} in {_curr_label_ins} + "
                                            f"{int(_top_drop['prev_drops'])} in {_prev_label_ins} = "
                                            f"{int(_top_drop['total_drops'])} total. "
                                            f"Top team by drops is {_top_team_drop} "
                                            f"with {_top_team_drop_cnt} cases."
                                        )
                                        _drop_insight = {
                                            "type": "bad", "icon": "🚫",
                                            "title": (f"Most Drop Cases — {_top_drop['Caller_name']} "
                                                      f"({_top_drop.get('Team Name','—')})"),
                                            "body": _drop_body
                                        }

                            if _pending_insight:
                                insights.append(_pending_insight)
                            if _drop_insight:
                                insights.append(_drop_insight)
                        except Exception:
                            pass

                        if insights:
                            cols_ins = st.columns(2)
                            for i, ins in enumerate(insights):
                                with cols_ins[i % 2]:
                                    st.markdown(f"""
                                    <div class="insight-card {ins['type']}">
                                        <div style='display:flex;align-items:center;gap:.4rem;'>
                                            <span class="insight-icon">{ins['icon']}</span>
                                            <span class="insight-title">{ins['title']}</span>
                                        </div>
                                        <div class="insight-body">{ins['body']}</div>
                                    </div>""", unsafe_allow_html=True)
                        else:
                            st.info("Not enough variation in data to generate insights.")

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE A — CALLER TEAM LEADERBOARD
                        # ══════════════════════════════════════════════
                        section_header("📞 CALLER REVENUE TEAM PERFORMANCE TABLE")

                        if not calling_ins.empty:
                            lb_calling = (
                                calling_ins.groupby('TEAM')
                                .agg(
                                    Callers   = ('CALLER NAME',       'count'),
                                    Enrollments=('raw_enrollments',   'sum'),
                                    Target    = ('raw_target',        'sum'),
                                    Till_Day  = ('TILL DAY TARGET (₹)','sum'),
                                    Enr_Rev   = ('ENROLLMENT REV',    'sum'),
                                    Bal_Rev   = ('BALANCE REV',       'sum'),
                                    Revenue   = ('raw_calling_rev',   'sum'),
                                )
                                .reset_index()
                                .sort_values('Revenue', ascending=False)
                            )
                            lb_calling['Achievement %'] = lb_calling.apply(
                                lambda r: f"{round(r['Revenue']/r['Target']*100,1)}%" if r['Target'] > 0 else "—", axis=1
                            )
                            lb_calling['Target']   = lb_calling['Target'].apply(fmt_inr)
                            lb_calling['Till_Day'] = lb_calling['Till_Day'].apply(fmt_inr)
                            lb_calling['Enr_Rev']  = lb_calling['Enr_Rev'].apply(fmt_inr)
                            lb_calling['Bal_Rev']  = lb_calling['Bal_Rev'].apply(fmt_inr)
                            lb_calling['Revenue']  = lb_calling['Revenue'].apply(fmt_inr)
                            medals = (["🥇", "🥈", "🥉"] + [""] * len(lb_calling))[:len(lb_calling)]
                            lb_calling.insert(0, "🏅", medals)
                            lb_calling.columns = ['🏅', 'Team', 'Callers', 'Enrollments',
                                                  'Total Target', 'Till Day Target',
                                                  'Enrollment Rev', 'Balance Rev',
                                                  'Calling Revenue', 'Achievement %']
                            st.dataframe(lb_calling.reset_index(drop=True), use_container_width=True, hide_index=True)
                        else:
                            st.info("No calling agent data available.")

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE B — COLLECTION TEAM LEADERBOARD
                        # ══════════════════════════════════════════════
                        section_header("🏦 COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE")

                        if not collection_ins.empty:
                            lb_coll = (
                                collection_ins.groupby('TEAM')
                                .agg(
                                    Callers     = ('CALLER NAME',          'count'),
                                    Enrollments = ('raw_enrollments',      'sum'),
                                    Target      = ('raw_target',           'sum'),
                                    Till_Day    = ('TILL DAY TARGET (₹)',  'sum'),
                                    Calling_Rev = ('CALLING REVENUE',      'sum'),
                                    Comm_Coll   = ('COMMUNITY COLLECTION', 'sum'),
                                    Boot_Coll   = ('BOOTCAMP COLLECTION',  'sum'),
                                    Revenue     = ('raw_collection_rev',   'sum'),
                                )
                                .reset_index()
                                .sort_values('Revenue', ascending=False)
                            )
                            lb_coll['Achievement %'] = lb_coll.apply(
                                lambda r: f"{round(r['Revenue']/r['Target']*100,1)}%" if r['Target'] > 0 else "—", axis=1
                            )
                            lb_coll['Target']      = lb_coll['Target'].apply(fmt_inr)
                            lb_coll['Till_Day']    = lb_coll['Till_Day'].apply(fmt_inr)
                            lb_coll['Calling_Rev'] = lb_coll['Calling_Rev'].apply(fmt_inr)
                            lb_coll['Comm_Coll']   = lb_coll['Comm_Coll'].apply(fmt_inr)
                            lb_coll['Boot_Coll']   = lb_coll['Boot_Coll'].apply(fmt_inr)
                            lb_coll['Revenue']     = lb_coll['Revenue'].apply(fmt_inr)
                            medals = (["🥇", "🥈", "🥉"] + [""] * len(lb_coll))[:len(lb_coll)]
                            lb_coll.insert(0, "🏅", medals)
                            lb_coll.columns = ['🏅', 'Team', 'Callers', 'Enrollments',
                                               'Total Target', 'Till Day Target',
                                               'Calling Revenue', 'Community Collection',
                                               'Bootcamp Collection', 'Collection Revenue', 'Achievement %']
                            st.dataframe(lb_coll.reset_index(drop=True), use_container_width=True, hide_index=True)
                        else:
                            st.info("No collection agent data available.")

                        st.divider()

                        # ══════════════════════════════════════════════
                        # TABLE C — BOTH TEAM LEADERBOARD
                        # ══════════════════════════════════════════════
                        section_header("📞🏦 CALLING + COLLECTION CALLER TEAM REVENUE PERFORMANCE TABLE")

                        if not both_ins.empty:
                            lb_both = (
                                both_ins.groupby('TEAM')
                                .agg(
                                    Callers     = ('CALLER NAME',          'count'),
                                    Enrollments = ('raw_enrollments',      'sum'),
                                    Target      = ('raw_target',           'sum'),
                                    Till_Day    = ('TILL DAY TARGET (₹)',  'sum'),
                                    Calling_Rev = ('CALLING REVENUE',      'sum'),
                                    Comm_Coll   = ('COMMUNITY COLLECTION', 'sum'),
                                    Boot_Coll   = ('BOOTCAMP COLLECTION',  'sum'),
                                    Total_Rev   = ('raw_revenue',          'sum'),
                                )
                                .reset_index()
                                .sort_values('Total_Rev', ascending=False)
                            )
                            lb_both['Achievement %'] = lb_both.apply(
                                lambda r: f"{round(r['Total_Rev']/r['Target']*100,1)}%" if r['Target'] > 0 else "—", axis=1
                            )
                            lb_both['Target']      = lb_both['Target'].apply(fmt_inr)
                            lb_both['Till_Day']    = lb_both['Till_Day'].apply(fmt_inr)
                            lb_both['Calling_Rev'] = lb_both['Calling_Rev'].apply(fmt_inr)
                            lb_both['Comm_Coll']   = lb_both['Comm_Coll'].apply(fmt_inr)
                            lb_both['Boot_Coll']   = lb_both['Boot_Coll'].apply(fmt_inr)
                            lb_both['Total_Rev']   = lb_both['Total_Rev'].apply(fmt_inr)
                            medals = (["🥇", "🥈", "🥉"] + [""] * len(lb_both))[:len(lb_both)]
                            lb_both.insert(0, "🏅", medals)
                            lb_both.columns = ['🏅', 'Team', 'Callers', 'Enrollments',
                                               'Total Target', 'Till Day Target',
                                               'Calling Revenue', 'Community Collection',
                                               'Bootcamp Collection', 'Total Revenue', 'Achievement %']
                            st.dataframe(lb_both.reset_index(drop=True), use_container_width=True, hide_index=True)
                        else:
                            st.info("No calling+collection agent data available.")

        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🧠</div>
                <div style='font-size:.95rem;font-weight:600;'>Click <b>Generate Revenue Report</b> to load insights</div>
            </div>""", unsafe_allow_html=True)


    with tab3:
        c_start, c_end, p_start, p_end = pending_months()
        curr_label = c_start.strftime("%B %Y")
        prev_label = p_start.strftime("%B %Y")

        if gen_pending:
            st.session_state['pending_revenue_loaded'] = True

        if not st.session_state.get('pending_revenue_loaded', False):
            st.markdown("""
            <div style='text-align:center;padding:4rem 1rem;opacity:.6;'>
                <div style='font-size:3rem;margin-bottom:1rem;'>📊</div>
                <div style='font-size:.9rem;font-weight:600;'>Click <b>📊 Generate Callerwise Pending</b> in the sidebar to load this tab</div>
            </div>""", unsafe_allow_html=True)

        else:
            with st.spinner("Loading pending revenue data…"):
                _, _, df_meta_all = get_metadata()
                meta_map_pending  = (df_meta_all.sort_values('Month', ascending=False)
                                     .drop_duplicates(subset=['merge_key'], keep='first')
                                     [['merge_key', 'Caller Name', 'Team Name', 'Vertical']].copy())
    
                drop_df     = load_drop_leads()
                excl_emails = set(drop_df['drop_email'].dropna().astype(str).unique()) if not drop_df.empty else set()
                excl_phones = set(drop_df['drop_phone'].dropna().astype(str).unique()) if not drop_df.empty else set()
                df_both     = fetch_both_months_rev(p_start, c_end)
    
            if df_both.empty:
                st.warning("No revenue data found.")
            else:
                df_curr   = df_both[df_both['Date'] >= c_start].copy()
                df_prev   = df_both[(df_both['Date'] >= p_start) & (df_both['Date'] <= p_end)].copy()
    
                pend_curr = pending_leads_for_month(df_curr, excl_emails, excl_phones, df_both)
                pend_prev = pending_leads_for_month(df_prev, excl_emails, excl_phones, df_both)
    
                combined  = build_combined_agg(pend_curr, pend_prev, meta_map_pending, df_both)
    
                if not combined.empty:
                    combined = combined[~(
                        (combined['Team Name'] == 'Others') &
                        (combined['Vertical'] == 'Others') &
                        (combined['grand_bal'] == 0)
                    )].copy()
                    combined = combined[combined['grand_bal'] > 0].copy()
    
                    if selected_vertical:
                        combined = combined[combined['Vertical'].isin(selected_vertical)].copy()
                    if selected_team:
                        combined = combined[combined['Team Name'].isin(selected_team)].copy()
    
                if combined.empty:
                    st.info("No pending leads found.")
                else:
                    st.markdown(render_html_pending_table(
                        combined, 'team', curr_label, prev_label,
                        f"TEAMWISE PENDING REVENUE {prev_label.upper()} + {curr_label.upper()}"
                    ), unsafe_allow_html=True)
    
                    st.markdown("<div style='margin:2rem 0;'></div>", unsafe_allow_html=True)
    
                    st.markdown(render_html_pending_table(
                        combined, 'caller', curr_label, prev_label,
                        f"CALLERWISE PENDING REVENUE {prev_label.upper()} + {curr_label.upper()}"
                    ), unsafe_allow_html=True)
    
                    st.markdown("<div style='margin:1rem 0;'></div>", unsafe_allow_html=True)
                    dl_col1, dl_col2 = st.columns(2)
    
                    with dl_col1:
                        _pending_xlsx = build_pending_excel(
                            combined, pend_curr, pend_prev,
                            meta_map_pending, curr_label, prev_label
                        )
                        st.download_button(
                            label="📥 Download Teamwise + Callerwise Pending Revenue",
                            data=_pending_xlsx,
                            file_name=f"Pending_Revenue_{prev_label.replace(' ','_')}_{curr_label.replace(' ','_')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_pending_revenue_xlsx'
                        )
    
                    with dl_col2:
                        _leads_xlsx = build_pending_leads_excel(
                            pend_curr, pend_prev,
                            meta_map_pending, curr_label, prev_label
                        )
                        st.download_button(
                            label=f"📥 Download {prev_label} + {curr_label} Pending Leads",
                            data=_leads_xlsx,
                            file_name=f"Pending_Leads_{prev_label.replace(' ','_')}_{curr_label.replace(' ','_')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_pending_leads_xlsx'
                        )
    
                # ══ DROPPED LEADS — inside else: so drop_df and df_both are always defined ══
                st.markdown("<div style='margin:2rem 0;'></div>", unsafe_allow_html=True)
                section_header(f"🚫 CALLERWISE DROPPED LEADS — {prev_label} + {curr_label}")
    
                if not drop_df.empty and not df_both.empty:
                    drop_agg = attribute_drops_to_callers(
                        drop_df, df_both, meta_map_pending,
                        c_start, c_end, p_start, p_end,
                        curr_label, prev_label
                    )
                    if not drop_agg.empty:
                        drop_agg = drop_agg[drop_agg['Team Name'] != 'Others'].copy()
    
                        if selected_vertical:
                            drop_agg = drop_agg[drop_agg['Vertical'].isin(selected_vertical)].copy()
                        if selected_team:
                            drop_agg = drop_agg[drop_agg['Team Name'].isin(selected_team)].copy()
    
                        st.markdown(
                            render_drop_html(drop_agg, curr_label, prev_label),
                            unsafe_allow_html=True
                        )
                    else:
                        st.info("No dropped leads found for current or previous month.")
    
                    if not drop_agg.empty:
                        _drop_xlsx = build_drop_leads_excel(
                            drop_df, c_start, c_end, p_start, p_end,
                            curr_label, prev_label, meta_map_pending
                        )
                        st.download_button(
                            label=f"📥 Download Dropped Leads — {prev_label} + {curr_label}",
                            data=_drop_xlsx,
                            file_name=f"Dropped_Leads_{prev_label.replace(' ','_')}_{curr_label.replace(' ','_')}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            key='dl_drop_leads_xlsx'
                        )
                else:
                    st.info("Drop leads sheet could not be loaded.")

def run_leads_dashboard():
    st.markdown("""
    <style>
    [data-testid="stMainBlockContainer"] {
        max-width: 100% !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }
    .block-container { max-width: 100% !important; }
    </style>
    """, unsafe_allow_html=True)
 
    # ── CONSTANTS ──────────────────────────────────────────────────────────────
    _CSV_URL     = "https://docs.google.com/spreadsheets/d/e/2PACX-1vRT73ztvPNZSvIu5WLxo-3WQ76JMAnt4P9dITd4EAbjSvuDytfgvdfri1WPXotCjm_Etnb80_Q7S-wf/pub?gid=0&single=true&output=csv"
    _LEADS_TABLE = "studious-apex-488820-c3.crm_dashboard.lsq_leads"
 
    _STAGE_MAP = {
        'FRESH'             : ['New Lead', 'Re-enquired Lead', 'Opportunity Created'],
        'DNP'               : ['Call Not Picking Up', 'Call Not Connected'],
        'CBL'               : ['Call Back Later'],
        'FLW-UP'            : ['Follow Up For Closure'],
        'COUNSELLED'        : ['Counselled lead'],
        'DISCOVERY'         : ['Discovery Call Done'],
        'ROADMAP'           : ['Roadmap Done'],
        'MBL'               : ['May buy later'],
        'ACTUALLY-ENROLLED' : ['Actually Enrolled'],
        'INVALID/NTINTRSTD' : ['Irrelevant lead', 'Not Interested', 'Invalid'],
        'BOOKING-RCVD'      : ['Booking fees received'],
        'LOAN-PNDG'         : ['Loan pending'],
        'COLL-DNE'          : ['Collections done'],
        'PRE-SALES'         : ['Pre-Sales Registrations'],
        'COURSE ENROLLED'   : ['Course Enrolled'],
    }
 
    _BREACHED_COL_MAP = {
        'CBL'       : ['Call Back Later'],
        'FLW-UP'    : ['Follow Up For Closure'],
        'COUNSELLED': ['Counselled lead'],
        'DISCOVERY' : ['Discovery Call Done'],
        'ROADMAP'   : ['Roadmap Done'],
    }
 
    _DIALLED_COL_MAP = {
        'CALL NOT PICKING UP': ['Call Not Picking Up'],
        'CALL NOT CONNECTED' : ['Call Not Connected'],
    }
 
    # ── CSS ────────────────────────────────────────────────────────────────────
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
 
:root {
    --ld-accent: #3B82F6; --ld-deep: #1D4ED8; --ld-deepest: #1e3a8a;
    --ld-light: #DBEAFE;
    --radius-sm: 8px; --radius-md: 12px; --radius-lg: 16px;
    --shadow-sm: 0 1px 3px rgba(0,0,0,.08);
    --shadow-md: 0 4px 16px rgba(0,0,0,.10);
    --shadow-lg: 0 8px 32px rgba(0,0,0,.14);
    --transition: all 0.22s cubic-bezier(.4,0,.2,1);
}
[data-testid="stAppViewContainer"]:not([class*="dark"]) {
    --bg-base:#EFF6FF;--bg-surface:#FFFFFF;--bg-muted:#DBEAFE;
    --border:rgba(59,130,246,.12);--text-primary:#111827;
    --text-muted:#6B7280;--metric-bg:#FFFFFF;
}
@media (prefers-color-scheme: dark) {
    :root{--bg-base:#0A0F1E;--bg-surface:#111827;--bg-elevated:#1E293B;
          --bg-muted:#1E3A5F;--border:rgba(59,130,246,.10);
          --text-primary:#EFF6FF;--text-muted:#93C5FD;--metric-bg:#1E293B;}
}
[data-theme="dark"]{
    --bg-base:#0A0F1E!important;--bg-surface:#111827!important;
    --bg-elevated:#1E293B!important;--bg-muted:#1E3A5F!important;
    --border:rgba(59,130,246,.10)!important;--text-primary:#EFF6FF!important;
    --text-muted:#93C5FD!important;--metric-bg:#1E293B!important;
}
html,body,[class*="css"]{font-family:'DM Sans',sans-serif!important;}
footer{visibility:hidden;}
[data-testid="stStatusWidget"]{display:none!important;}
[data-testid="stSidebar"]{border-right:1px solid var(--border,rgba(59,130,246,.12));}
 
.ld-header{
    background:linear-gradient(135deg,#0f172a 0%,#1e3a8a 45%,#1e40af 100%);
    border-radius:var(--radius-lg);padding:1.5rem 2rem 1.2rem;
    margin-bottom:1.2rem;position:relative;overflow:hidden;box-shadow:var(--shadow-lg);
}
.ld-header::before{
    content:"";position:absolute;top:-40px;right:-40px;width:200px;height:200px;
    background:radial-gradient(circle,rgba(59,130,246,.25) 0%,transparent 70%);border-radius:50%;
}
.ld-title{font-size:1.65rem;font-weight:700;color:#FFF;letter-spacing:.5px;margin:0 0 .25rem;}
.ld-subtitle{font-size:.82rem;color:rgba(255,255,255,.6);margin:0;font-family:'DM Mono',monospace;}
.ld-badge{
    display:inline-flex;align-items:center;gap:5px;
    background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.18);
    border-radius:20px;padding:3px 10px;font-size:.73rem;
    color:rgba(255,255,255,.9);font-family:'DM Mono',monospace;
}
.ld-pulse{
    width:6px;height:6px;background:#60A5FA;border-radius:50%;
    display:inline-block;animation:pulse-ld 1.8s ease-in-out infinite;
}
@keyframes pulse-ld{0%,100%{opacity:1;transform:scale(1);}50%{opacity:.5;transform:scale(1.4);}}
 
.ld-metric-card{
    background:var(--metric-bg,#fff);border:1px solid var(--border,rgba(59,130,246,.12));
    border-radius:var(--radius-md);padding:.9rem 1rem;transition:var(--transition);
    box-shadow:var(--shadow-sm);position:relative;overflow:hidden;text-align:center;
}
.ld-metric-card::before{
    content:"";position:absolute;top:0;left:0;width:100%;height:3px;
    background:linear-gradient(90deg,#3B82F6,#60A5FA);opacity:0;transition:opacity .2s;
}
.ld-metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-md);}
.ld-metric-card:hover::before{opacity:1;}
.ld-metric-label{font-size:.68rem;font-weight:600;text-transform:uppercase;letter-spacing:.8px;
    color:var(--text-muted,#6B7280);margin:0 0 .3rem;}
.ld-metric-value{font-size:1.45rem;font-weight:700;color:var(--text-primary,#111827);
    line-height:1;font-family:'DM Mono',monospace;}
 
.ld-section-header{display:flex;align-items:center;gap:.6rem;margin:1.5rem 0 .8rem;}
.ld-section-line{flex:1;height:1px;background:linear-gradient(90deg,#3B82F6,transparent);opacity:.35;}
.ld-section-title{font-size:.78rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;
    color:#3B82F6;white-space:nowrap;text-align:center;}
 
.ld-insight-card{background:var(--metric-bg,#fff);border:1px solid var(--border,rgba(59,130,246,.12));
    border-radius:var(--radius-md);padding:1rem 1.1rem;margin-bottom:.6rem;
    box-shadow:var(--shadow-sm);transition:var(--transition);}
.ld-insight-card:hover{box-shadow:var(--shadow-md);}
.ld-insight-card.good{border-left:3px solid #3B82F6;}
.ld-insight-card.warn{border-left:3px solid #FBBF24;}
.ld-insight-card.bad{border-left:3px solid #F87171;}
.ld-insight-card.info{border-left:3px solid #60A5FA;}
 
div[data-testid="stDataFrame"] thead tr th{
    background:linear-gradient(135deg,#1e3a8a,#1d4ed8)!important;
    color:#fff!important;font-family:'DM Sans',sans-serif!important;
    font-size:.72rem!important;font-weight:700!important;letter-spacing:.6px;
    text-transform:uppercase;white-space:normal!important;word-wrap:break-word!important;
    text-align:center!important;vertical-align:middle!important;
    min-width:80px!important;padding:10px!important;
}
[data-testid="stSidebar"] .stButton>button{
    width:100%;font-family:'DM Sans',sans-serif!important;font-weight:600!important;
    font-size:.82rem!important;border-radius:var(--radius-sm);transition:var(--transition);
    background:linear-gradient(135deg,#1d4ed8,#1e3a8a)!important;
    color:#fff!important;border:none!important;
}
[data-testid="stSidebar"] .stButton>button:hover{opacity:.88!important;transform:translateY(-1px)!important;}
.stDownloadButton>button{
    background:linear-gradient(135deg,#1d4ed8,#1e3a8a)!important;color:#fff!important;
    border:none!important;border-radius:var(--radius-sm)!important;
    font-family:'DM Sans',sans-serif!important;font-weight:600!important;
    font-size:.82rem!important;width:100%!important;transition:var(--transition)!important;
}
hr{border-color:var(--border,rgba(59,130,246,.12))!important;margin:1.2rem 0!important;}
</style>
""", unsafe_allow_html=True)
 
    # ── LOCAL HELPERS ──────────────────────────────────────────────────────────
    def _ld_section_header(label):
        st.markdown(f"""
        <div class="ld-section-header">
            <div class="ld-section-line"></div>
            <span class="ld-section-title">{label}</span>
            <div class="ld-section-line" style="background:linear-gradient(90deg,transparent,#3B82F6)"></div>
        </div>""", unsafe_allow_html=True)
 
    def _style_caller(row):
        if row.get('CALLER') == 'TOTAL':
            return ['font-weight:bold;background-color:#374151;color:#FFFFFF;'] * len(row)
        return [''] * len(row)
 
    def _style_team(row):
        if row.get('TEAM') == 'TOTAL':
            return ['font-weight:bold;background-color:#374151;color:#FFFFFF;'] * len(row)
        return [''] * len(row)
 
    def _merge_team_ld(df, df_meta):
        df_w = df.copy()
        df_w['merge_key'] = df_w['Owner'].astype(str).str.strip().str.lower()
        slim = df_meta[['merge_key','Caller Name','Team Name','Vertical']].drop_duplicates('merge_key')
        merged = pd.merge(df_w, slim, on='merge_key', how='left')
        merged['Owner']     = merged['Caller Name'].fillna(merged['Owner'])
        merged['Team Name'] = merged['Team Name'].fillna('Others')
        merged['Vertical']  = merged['Vertical'].fillna('Others')
        merged = merged[merged['Team Name'] != 'Others']
        return merged
 
    def _append_caller_total_ld(df):
        if df.empty: return df
        nc  = [c for c in df.columns if c not in ('CALLER','TEAM')]
        row = {'CALLER':'TOTAL','TEAM':'—'}
        for c in nc: row[c] = int(df[c].sum())
        return pd.concat([df, pd.DataFrame([row])], ignore_index=True)
 
    def _append_team_total_ld(df):
        if df.empty: return df
        nc  = [c for c in df.columns if c != 'TEAM']
        row = {'TEAM':'TOTAL'}
        for c in nc: row[c] = int(df[c].sum())
        return pd.concat([df, pd.DataFrame([row])], ignore_index=True)
 
    def _show_caller(df, msg="No data."):
        if df.empty: st.info(msg); return
        final = _append_caller_total_ld(df)
        h = min((len(final) + 1) * 35 + 20, 800)
        st.dataframe(final.style.apply(_style_caller, axis=1),
                     use_container_width=True, hide_index=True, height=h)
 
    def _show_team(df, msg="No data."):
        if df.empty: st.info(msg); return
        final = _append_team_total_ld(df)
        h = min((len(final) + 1) * 35 + 20, 600)
        st.dataframe(final.style.apply(_style_team, axis=1),
                     use_container_width=True, hide_index=True, height=h)
        
    def _build_leads_xlsx_bytes_ld(df_rows):
        EXPORT_COLS = [
            'AssignedOn', 'FirstName', 'LastName', 'Email', 'PhoneNumber',
            'Alternate_PhoneNumber', 'Owner', 'ContactStage', 'LastCalledDate',
            'Follow_up_date', 'Enquired_Course', 'Campaign_Name',
            'Phone_call_counter', 'Assigned_On_Call_Counter', 'Team', 'Vertical', 'AssignedBy'
        ]
        df = df_rows.copy()
        if 'Team' not in df.columns and 'Team Name' in df.columns:
            df['Team'] = df['Team Name']
        cols = [c for c in EXPORT_COLS if c in df.columns]
        df   = df[cols].reset_index(drop=True)

        HDR_FILL  = PatternFill("solid", start_color="1e3a8a", end_color="1e3a8a")
        HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        ALT_FILL  = PatternFill("solid", start_color="EFF6FF", end_color="EFF6FF")
        WHT_FILL  = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
        DATA_FONT = Font(name="Calibri", size=10)
        CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        BORDER    = Border(
            left=Side(style='thin', color='BFDBFE'), right=Side(style='thin', color='BFDBFE'),
            top=Side(style='thin',  color='BFDBFE'), bottom=Side(style='thin', color='BFDBFE'),
        )
        COL_WIDTHS = {
            'AssignedOn': 14, 'FirstName': 18, 'LastName': 18, 'Email': 32,
            'PhoneNumber': 14, 'Alternate_PhoneNumber': 18, 'Owner': 24,
            'ContactStage': 24, 'LastCalledDate': 14, 'Follow_up_date': 14,
            'Enquired_Course': 26, 'Campaign_Name': 22, 'Phone_call_counter': 12,
            'Assigned_On_Call_Counter': 14, 'Team': 24, 'Vertical': 18, 'AssignedBy': 20,
        }
        wb = Workbook(); ws = wb.active; ws.title = "Breached Leads"
        for c_idx, col in enumerate(cols, 1):
            cell = ws.cell(1, c_idx, col.replace('_', ' ').upper())
            cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, CENTER, BORDER
        ws.row_dimensions[1].height = 26
        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            fill = ALT_FILL if r_idx % 2 == 0 else WHT_FILL
            for c_idx, col in enumerate(cols, 1):
                val = row[col]
                try:
                    if pd.isna(val): val = ''
                except (TypeError, ValueError): pass
                if hasattr(val, 'strftime'): val = val.strftime('%Y-%m-%d')
                cell = ws.cell(r_idx, c_idx, val)
                cell.fill, cell.font, cell.alignment, cell.border = fill, DATA_FONT, LEFT, BORDER
        for c_idx, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = COL_WIDTHS.get(col, 16)
        ws.freeze_panes = "A2"
        buf = io.BytesIO(); wb.save(buf)
        return buf.getvalue()
 
    def _stage_counts(grp, col_map):
        row, total = {}, 0
        for col, stages in col_map.items():
            cnt = int(grp['ContactStage'].isin(stages).sum())
            row[col] = cnt; total += cnt
        row['TOTAL'] = total
        return row
 
    # ── CACHED DATA FUNCTIONS ──────────────────────────────────────────────────
    @st.cache_data(ttl=120, show_spinner=False)
    def _ld_get_metadata():
        df = pd.read_csv(_CSV_URL)
        df.columns = df.columns.str.strip()
        df['merge_key'] = df['Caller Name'].str.strip().str.lower()
        return sorted(df['Team Name'].dropna().unique()), sorted(df['Vertical'].dropna().unique()), df
 
    @st.cache_data(ttl=300, show_spinner=False)
    def _ld_last_update():
        try:
            r = client.query(
                f"SELECT MAX(updated_at_ampm) AS lu FROM `{_LEADS_TABLE}` WHERE updated_at_ampm IS NOT NULL"
            ).to_dataframe()
            return str(r['lu'].iloc[0]) if not r.empty and r['lu'].iloc[0] else "N/A"
        except Exception: return "N/A"
 
    @st.cache_data(ttl=600, show_spinner=False)
    def _ld_available_dates():
        try:
            r = client.query(
                f"SELECT MIN(AssignedOn) AS mn, MAX(AssignedOn) AS mx FROM `{_LEADS_TABLE}`"
            ).to_dataframe()
            if not r.empty and not pd.isna(r['mn'].iloc[0]):
                return r['mn'].iloc[0], r['mx'].iloc[0]
        except Exception: pass
        return date.today(), date.today()
 
    @st.cache_data(ttl=120, show_spinner=False)
    def _ld_fetch(start_date, end_date):
        q = f"""
            SELECT * FROM `{_LEADS_TABLE}`
            WHERE AssignedOn BETWEEN '{start_date}' AND '{end_date}'
        """
        df = client.query(q).to_dataframe()
        if not df.empty:
            df['Owner']        = df['Owner'].astype(str).str.strip()
            df['ContactStage'] = df['ContactStage'].astype(str).str.strip()
            df['Follow_up_date'] = pd.to_datetime(df['Follow_up_date'], errors='coerce')
            df['LastCalledDate'] = pd.to_datetime(df['LastCalledDate'], errors='coerce')
            df['Assigned_On_Call_Counter'] = pd.to_numeric(
                df['Assigned_On_Call_Counter'], errors='coerce').fillna(0)
        return df
 
    # ── PROCESSING ─────────────────────────────────────────────────────────────
    def _proc_assigned_caller(df_m):
        rows = []
        for owner, g in df_m.groupby('Owner'):
            team = g['Team Name'].mode().iloc[0] if len(g) else 'Others'
            r = {'CALLER': owner, 'TEAM': team}
            r.update(_stage_counts(g, _STAGE_MAP))
            rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _proc_assigned_team(df_m):
        rows = []
        for team, g in df_m.groupby('Team Name'):
            r = {'TEAM': team}; r.update(_stage_counts(g, _STAGE_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _get_breached(df_m):
        now_ts = pd.Timestamp.now().normalize()
        cut    = now_ts - pd.Timedelta(days=3)
        all_s  = [s for v in _BREACHED_COL_MAP.values() for s in v]
        fup    = df_m['Follow_up_date'].isna() | (df_m['Follow_up_date'] < now_ts)
        lcd    = df_m['LastCalledDate'].notna() & (df_m['LastCalledDate'] < cut)
        return df_m[fup & lcd & df_m['ContactStage'].isin(all_s)].copy()
 
    def _proc_breached_caller(df_m):
        df_b = _get_breached(df_m); rows = []
        for owner, g in df_b.groupby('Owner'):
            team = g['Team Name'].mode().iloc[0] if len(g) else 'Others'
            r = {'CALLER': owner, 'TEAM': team}; r.update(_stage_counts(g, _BREACHED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _proc_breached_team(df_m):
        df_b = _get_breached(df_m); rows = []
        for team, g in df_b.groupby('Team Name'):
            r = {'TEAM': team}; r.update(_stage_counts(g, _BREACHED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _get_less_dialled(df_m):
        all_s = [s for v in _DIALLED_COL_MAP.values() for s in v]
        return df_m[(df_m['Assigned_On_Call_Counter'] < 11) & df_m['ContactStage'].isin(all_s)].copy()
 
    def _proc_ld_caller(df_m):
        df_ld = _get_less_dialled(df_m); rows = []
        for owner, g in df_ld.groupby('Owner'):
            team = g['Team Name'].mode().iloc[0] if len(g) else 'Others'
            r = {'CALLER': owner, 'TEAM': team}; r.update(_stage_counts(g, _DIALLED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _proc_ld_team(df_m):
        df_ld = _get_less_dialled(df_m); rows = []
        for team, g in df_ld.groupby('Team Name'):
            r = {'TEAM': team}; r.update(_stage_counts(g, _DIALLED_COL_MAP)); rows.append(r)
        if not rows: return pd.DataFrame()
        return pd.DataFrame(rows).sort_values('TOTAL', ascending=False).reset_index(drop=True)
 
    def _compute_insights(df_ac, df_bc, df_ldc, df_at, df_bt, df_ldt):
        ins = []
        if not df_ac.empty:
            t = df_ac.iloc[0]
            ins.append({"type":"good","icon":"🏆",
                "title":f"Highest Assigned Leads — {t['CALLER']}",
                "body":(f"{t['CALLER']} holds the highest lead count with {t['TOTAL']} leads. "
                        f"Team: {t['TEAM']}. Fresh: {t.get('FRESH',0)} · "
                        f"Counselled: {t.get('COUNSELLED',0)} · Enrolled: {t.get('COURSE ENROLLED',0)}.")})
        if not df_at.empty:
            t = df_at.iloc[0]
            ins.append({"type":"good","icon":"🏅",
                "title":f"Highest Assigned Leads — Team: {t['TEAM']}",
                "body":(f"{t['TEAM']} received the most leads with {t['TOTAL']} assignments. "
                        f"Fresh: {t.get('FRESH',0)} · DNP: {t.get('DNP',0)} · "
                        f"Course Enrolled: {t.get('COURSE ENROLLED',0)}.")})
        if not df_bc.empty:
            t = df_bc.iloc[0]
            ins.append({"type":"bad","icon":"⚠️",
                "title":f"Highest Potential Breached — {t['CALLER']}",
                "body":(f"{t['CALLER']} has {t['TOTAL']} leads with overdue follow-up and no recent call. "
                        f"Team: {t['TEAM']}. CBL: {t.get('CBL',0)} · FLW-UP: {t.get('FLW-UP',0)} · "
                        f"Counselled: {t.get('COUNSELLED',0)}. Immediate action required.")})
        if not df_bt.empty:
            t = df_bt.iloc[0]
            ins.append({"type":"warn","icon":"🚨",
                "title":f"Highest Potential Breached — Team: {t['TEAM']}",
                "body":(f"{t['TEAM']} leads the breached list with {t['TOTAL']} overdue leads. "
                        f"CBL: {t.get('CBL',0)} · FLW-UP: {t.get('FLW-UP',0)} · "
                        f"Counselled: {t.get('COUNSELLED',0)}. Schedule recovery sessions.")})
        if not df_ldc.empty:
            t = df_ldc.iloc[0]
            ins.append({"type":"warn","icon":"📞",
                "title":f"Highest Less Dialled — {t['CALLER']}",
                "body":(f"{t['CALLER']} has {t['TOTAL']} DNP leads with <11 dial attempts. "
                        f"Team: {t['TEAM']}. Not Picking Up: {t.get('CALL NOT PICKING UP',0)} · "
                        f"Not Connected: {t.get('CALL NOT CONNECTED',0)}. Increase dial frequency.")})
        if not df_ldt.empty:
            t = df_ldt.iloc[0]
            ins.append({"type":"info","icon":"📈",
                "title":f"Highest Less Dialled — Team: {t['TEAM']}",
                "body":(f"{t['TEAM']} has {t['TOTAL']} less-dialled leads below 11 attempts. "
                        f"Maximise dial frequency before these leads degrade further.")})
        return ins[:6]
 
    @st.cache_data(show_spinner=False)
    def _ld_pdf_bytes() -> bytes:
        from reportlab.lib.pagesizes import A4 as _A4
        from reportlab.lib import colors as _colors
        from reportlab.lib.units import mm as _mm
        from reportlab.lib.styles import ParagraphStyle as _PS
        from reportlab.platypus import (SimpleDocTemplate as _SDT, Paragraph as _P,
                                         Spacer as _Sp, Table as _T, TableStyle as _TS,
                                         HRFlowable as _HR, Flowable as _F)
        from reportlab.lib.enums import TA_CENTER as _TAC
        import io as _io
 
        buf     = _io.BytesIO()
        BD      = _colors.HexColor("#1e3a8a"); BM=_colors.HexColor("#1d4ed8")
        BP      = _colors.HexColor("#DBEAFE"); BR=_colors.HexColor("#EFF6FF")
        GD      = _colors.HexColor("#374151"); GM=_colors.HexColor("#6B7280")
        W_      = _colors.white;               BK=_colors.HexColor("#111827")
        PW, PH  = _A4
 
        def sty(name,**kw):
            d=dict(fontName='Helvetica',fontSize=9,textColor=BK,spaceAfter=3,leading=14)
            d.update(kw); return _PS(name,**d)
 
        S={'b':sty('b_ld'),'l':sty('l_ld',fontName='Helvetica-Bold',fontSize=8,textColor=BD,spaceAfter=1),
           'f':sty('f_ld',fontName='Helvetica-Oblique',fontSize=8.5,textColor=BM,backColor=BP,leftIndent=8,rightIndent=8),
           'ft':sty('ft_ld',fontSize=7.5,textColor=GM,alignment=_TAC)}
 
        class Cover(_F):
            def __init__(self,w): _F.__init__(self); self.w=w; self.height=90
            def draw(self):
                c=self.canv
                c.setFillColor(BD); c.rect(0,52,self.w,38,fill=1,stroke=0)
                c.setFillColor(BM); c.rect(0,22,self.w,30,fill=1,stroke=0)
                c.setFillColor(_colors.HexColor("#0f172a")); c.rect(0,0,self.w,22,fill=1,stroke=0)
                c.setFillColor(W_); c.setFont("Helvetica-Bold",22)
                c.drawCentredString(self.w/2,66,"LEAD METRICS DASHBOARD")
                c.setFillColor(_colors.HexColor("#BFDBFE")); c.setFont("Helvetica-Bold",11)
                c.drawCentredString(self.w/2,34,"Logic & Metric Reference Guide")
                c.setFillColor(_colors.HexColor("#DBEAFE")); c.setFont("Helvetica",8.5)
                c.drawCentredString(self.w/2,8,
                    "LawSikho & Skill Arbitrage  \u00b7  Internal Use Only")
 
        class Ban(_F):
            def __init__(self,icon,title,color=None,w=None):
                _F.__init__(self); self.icon=icon; self.title=title
                self.color=color or BD; self.w=w or (PW-30*_mm); self.height=22
            def draw(self):
                c=self.canv; c.setFillColor(self.color)
                c.roundRect(0,0,self.w,self.height,4,fill=1,stroke=0)
                c.setFillColor(W_); c.setFont("Helvetica-Bold",11)
                c.drawString(10,6,f"{self.icon}  {self.title}")
 
        def bt(rows,cw=None):
            cw=cw or [44*_mm,116*_mm]
            data=[[_P(f"<b>{r[0]}</b>",S['l']),_P(r[1],S['b'])] for r in rows]
            t=_T(data,colWidths=cw,hAlign='LEFT')
            t.setStyle(_TS([('BACKGROUND',(0,0),(0,-1),BP),('VALIGN',(0,0),(-1,-1),'TOP'),
                ('GRID',(0,0),(-1,-1),0.3,_colors.HexColor("#BFDBFE")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[W_,BR]),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            return t
 
        def lt(rows):
            data=[[_P(f"<b>{r[0]}</b>",S['l']),_P(r[1],S['f'])] for r in rows]
            t=_T(data,colWidths=[52*_mm,108*_mm],hAlign='LEFT')
            t.setStyle(_TS([('VALIGN',(0,0),(-1,-1),'TOP'),
                ('GRID',(0,0),(-1,-1),0.3,_colors.HexColor("#93C5FD")),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[W_,BR]),
                ('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
            return t
 
        SP=_Sp; HR_=lambda:_HR(width="100%",thickness=0.6,color=_colors.HexColor("#93C5FD"),spaceAfter=6,spaceBefore=4)
        cw_=PW-30*_mm
 
        story=[
            SP(1,18*_mm),Cover(cw_),SP(1,10*_mm),
            _P("This document explains every metric, table, and column in the Lead Metrics Dashboard.",S['b']),
            SP(1,6*_mm),
            Ban("📋","SECTION 1 — TABS OVERVIEW"),SP(1,3*_mm),
            bt([("📊 Assigned Leads Report",
                 "Three callerwise tables: (1) Assigned Leads Distribution — all ContactStage buckets per caller; "
                 "(2) Potential Breached Leads — overdue follow-up + no recent call activity; "
                 "(3) Less Dialled Leads — DNP leads with <11 dial attempts."),
                ("🧠 Insights & Teamwise",
                 "Six auto-generated insights + three teamwise versions of all tables.")]),
            SP(1,6*_mm),
            Ban("📊","SECTION 2 — ASSIGNED LEADS DISTRIBUTION"),SP(1,3*_mm),
            _P("One row per caller. Sorted by TOTAL descending. Date filter = AssignedOn field.",S['b']),SP(1,2*_mm),
            lt([
                ("FRESH","New Lead + Re-enquired Lead + Opportunity Created"),
                ("DNP","Call Not Picking Up + Call Not Connected"),
                ("CBL","Call Back Later"),("FLW-UP","Follow Up For Closure"),
                ("COUNSELLED","Counselled lead"),("DISCOVERY","Discovery Call Done"),
                ("ROADMAP","Roadmap Done"),("MBL","May buy later"),
                ("ACTUALLY-ENROLLED","Actually Enrolled"),
                ("INVALID/NTINTRSTD","Irrelevant lead + Not Interested + Invalid"),
                ("BOOKING-RCVD","Booking fees received"),("LOAN-PNDG","Loan pending"),
                ("COLL-DNE","Collections done"),("PRE-SALES","Pre-Sales Registrations"),
                ("COURSE ENROLLED","Course Enrolled"),
                ("TOTAL","Sum of all stage columns = total assigned leads for this caller/team."),
            ]),SP(1,6*_mm),
            Ban("⚠️","SECTION 3 — POTENTIAL BREACHED LEADS"),SP(1,3*_mm),
            lt([
                ("Condition 1","Follow_up_date IS NULL OR Follow_up_date < today (IST)."),
                ("Condition 2","LastCalledDate < today minus 3 days (both conditions required)."),
                ("Stages","CBL · FLW-UP · COUNSELLED · DISCOVERY · ROADMAP only."),
                ("TOTAL","Sum of all five stage columns."),
            ]),SP(1,6*_mm),
            Ban("📞","SECTION 4 — LESS DIALLED LEADS"),SP(1,3*_mm),
            lt([
                ("Filter","Assigned_On_Call_Counter < 11 (fewer than 11 call attempts since assignment)."),
                ("CALL NOT PICKING UP","ContactStage = 'Call Not Picking Up' with <11 attempts."),
                ("CALL NOT CONNECTED","ContactStage = 'Call Not Connected' with <11 attempts."),
                ("TOTAL","Sum of both columns."),
            ]),SP(1,6*_mm),
            Ban("📖","KEY TERMS GLOSSARY",color=GD),SP(1,3*_mm),
            bt([
                ("AssignedOn","Date the lead was assigned. Primary date filter for the dashboard."),
                ("Owner","Caller the lead is assigned to. Mapped via lowercase merge key to Caller Name."),
                ("ContactStage","Current CRM lifecycle stage. Drives all bucketing logic."),
                ("Follow_up_date","Scheduled follow-up date. NULL or past = overdue."),
                ("LastCalledDate","Date last called. Used in breached leads filter."),
                ("Assigned_On_Call_Counter","Call attempts since assignment. Used in less-dialled filter."),
                ("TOTAL row","Bold dark row at bottom summing all numeric columns."),
            ]),SP(1,8*_mm),HR_(),
            _P("Designed by Amit Ray  \u00b7  amitray@lawsikho.in  \u00b7  "
               "For Internal Use of Sales and Operations Team Only. All Rights Reserved.",S['ft']),
        ]
        doc=_SDT(buf,pagesize=_A4,leftMargin=15*_mm,rightMargin=15*_mm,
                 topMargin=14*_mm,bottomMargin=14*_mm,
                 title="Lead Metrics — Logic Reference Guide",author="Amit Ray")
        doc.build(story)
        return buf.getvalue()
 
    # ── SIDEBAR ────────────────────────────────────────────────────────────────
    st.sidebar.markdown("""
    <div style='padding:.5rem 0 .4rem; text-align:center;'>
        <div style='font-size:.72rem; font-weight:700; text-transform:uppercase;
                    letter-spacing:1px; color:var(--text-muted,#6B7280);'>Report Controls</div>
    </div>
    """, unsafe_allow_html=True)
 
    teams_ld, verts_ld, df_meta_ld = _ld_get_metadata()
    min_dr, max_dr = _ld_available_dates()
    min_date_ld = pd.Timestamp(min_dr).date()
    max_date_ld = pd.Timestamp(max_dr).date()
 
    date_range_ld = st.sidebar.date_input(
        "📅 Date Range", value=(max_date_ld, max_date_ld),
        min_value=min_date_ld, max_value=max_date_ld,
        format="DD-MM-YYYY", key="ld_date_range_cd"
    )
    if isinstance(date_range_ld, tuple) and len(date_range_ld) == 2:
        start_date_ld, end_date_ld = date_range_ld
    else:
        start_date_ld = end_date_ld = (
            date_range_ld if not isinstance(date_range_ld, tuple) else date_range_ld[0])
 
    # ── Role-aware sidebar filters ──────────────────────────
    _role     = st.session_state.get('rf_role', 'admin')
    _rf_teams = st.session_state.get('rf_teams', [])
    _rf_cname = st.session_state.get('rf_caller_name', '')

    if _role == 'admin':
        sel_team_ld = st.sidebar.multiselect("👥 Filter by Team",     options=teams_ld, key="ld_team_cd")
        sel_vert_ld = st.sidebar.multiselect("👑 Filter by Vertical", options=verts_ld, key="ld_vert_cd")
        search_ld   = st.sidebar.text_input("👤 Search Caller Name",                    key="ld_search_cd")
    elif _role == 'vertical_head':
        sel_team_ld = _rf_teams
        sel_vert_ld = []
        search_ld   = st.sidebar.text_input("👤 Search Caller Name", key="ld_search_cd")
        st.sidebar.caption(f"🔒 Showing: {', '.join(_rf_teams)}")
    elif _role in ('tl', 'trainer'):
        sel_team_ld = _rf_teams
        sel_vert_ld = []
        search_ld   = st.sidebar.text_input("👤 Search Caller Name", key="ld_search_cd")
        st.sidebar.caption(f"🔒 Team: {', '.join(_rf_teams)}")
    else:  # caller
        sel_team_ld = []
        sel_vert_ld = []
        search_ld   = _rf_cname
        st.sidebar.caption(f"👤 Viewing: {_rf_cname}")
 
    gen_ld = st.sidebar.button("📊 Generate Leads Report", key="ld_gen_cd")
 
    st.sidebar.download_button(
        label="📖 Metrics Guide (PDF)",
        data=_ld_pdf_bytes(),
        file_name="Lead_Metrics_Logic_Guide.pdf",
        mime="application/pdf", key="dl_ld_pdf_cd"
    )
 
    # ── HEADER ─────────────────────────────────────────────────────────────────
    last_upd_ld  = _ld_last_update()
    disp_start   = start_date_ld.strftime('%d-%m-%Y')
    disp_end     = end_date_ld.strftime('%d-%m-%Y')
 
    st.markdown(f"""
    <div class="ld-header">
        <div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:.75rem;">
            <div>
                <div class="ld-title">📊 LEAD METRICS</div>
                <div class="ld-subtitle">ASSIGNMENT PERIOD&nbsp;·&nbsp; {disp_start} to {disp_end}</div>
            </div>
            <div style="display:flex;gap:.5rem;flex-wrap:wrap;align-items:center;margin-top:.25rem;">
                <span class="ld-badge"><span class="ld-pulse"></span>LEADSQUARED DATA</span>
                <span class="ld-badge">🕐 UPDATED AT: {last_upd_ld}</span>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)
 
    # ── TABS ───────────────────────────────────────────────────────────────────
    tab_ld1, tab_ld2 = st.tabs(["📊 Assigned Leads Report", "🧠 Insights & Teamwise"])
 
    # ── TAB 1 ──────────────────────────────────────────────────────────────────
    with tab_ld1:
        if gen_ld:
            with st.spinner("Fetching lead data…"):
                df_raw_ld = _ld_fetch(start_date_ld, end_date_ld)
 
            if df_raw_ld.empty:
                st.warning("No leads found for the selected period.")
            else:
                with st.spinner("Processing…"):
                    df_m_ld = _merge_team_ld(df_raw_ld, df_meta_ld)
                    if sel_team_ld: df_m_ld = df_m_ld[df_m_ld['Team Name'].isin(sel_team_ld)]
                    if sel_vert_ld: df_m_ld = df_m_ld[df_m_ld['Vertical'].isin(sel_vert_ld)]
                    if search_ld:   df_m_ld = df_m_ld[df_m_ld['Owner'].str.contains(search_ld, case=False, na=False)]
 
                if df_m_ld.empty:
                    st.error("No results match the selected filters.")
                else:
                    # Summary KPIs
                    df_valid_ld = df_m_ld[df_m_ld['Team Name'] != "Others"]
                    _ld_section_header("SUMMARY METRICS")
                    fresh_c = int(df_m_ld['ContactStage'].isin(_STAGE_MAP['FRESH']).sum())
                    enrolled_c = int(df_valid_ld['ContactStage'].eq("Actually Enrolled").sum())
                    discovery_c = int(df_valid_ld['ContactStage'].eq("Discovery Call Done").sum())
                    roadmap_c = int(df_valid_ld['ContactStage'].eq("Roadmap Done").sum())
                    followup_c = int(df_valid_ld['ContactStage'].isin(["Follow Up For Closure","Counselled lead"]).sum())
                    roadmap_c = int(df_valid_ld['ContactStage'].eq("Roadmap Done").sum())
                    followup_c = int(df_valid_ld['ContactStage'].isin(["Follow Up For Closure","Counselled lead"]).sum())
                    kc = st.columns(8)
                    for col, (lbl, val, ico) in zip(kc, [
                        ("Total Assigned Leads", f"{len(df_m_ld):,}", "📋"),
                        ("Fresh Leads", f"{fresh_c:,}", "🌱"),
                        ("Lead Conversions", f"{enrolled_c:,}", "🎓"),
                        ("Discovery", f"{discovery_c:,}", "🔍"),
                        ("Roadmap", f"{roadmap_c:,}", "🗺️"),
                        ("Follow Up / Counselled", f"{followup_c:,}", "📞"),
                        ("Active Callers", df_m_ld['Owner'].nunique(), "👤"),
                        ("Active Teams", df_m_ld['Team Name'].nunique(), "👥"),
                    ]):
                        with col:
                            st.markdown(f"""
                            <div class="ld-metric-card">
                                <div class="ld-metric-label">{ico} {lbl}</div>
                                <div class="ld-metric-value">{val}</div>
                            </div>""", unsafe_allow_html=True)
 
                    st.divider()
 
                    # Table 1
                    _ld_section_header("ASSIGNED LEADS DISTRIBUTION")
                    df_ac_ld = _proc_assigned_caller(df_m_ld)
                    _show_caller(df_ac_ld, "No assigned lead data found.")
 
                    st.divider()
 
                    # Table 2
                    _ld_section_header("POTENTIAL BREACHED LEADS AFTER ASSIGNMENT")
                    cut_disp = (pd.Timestamp.now().normalize() - pd.Timedelta(days=3)).strftime('%d-%m-%Y')
                    st.caption(f"Leads Breached and dialled before {cut_disp} · "
                               "of stages Call Back Later, Follow Up, Counselled, Discovery & Roadmap.")
                    df_bc_ld = _proc_breached_caller(df_m_ld)
                    _show_caller(df_bc_ld, "No potential breached leads found.")
 
                    df_bc_raw_ld = _get_breached(df_m_ld)
                    if not df_bc_raw_ld.empty:
                        col_dl_bc, _ = st.columns([1, 3])
                        with col_dl_bc:
                            st.download_button(
                                label="📥 Download Breached Leads (.xlsx)",
                                data=_build_leads_xlsx_bytes_ld(df_bc_raw_ld),
                                file_name=f"Breached_Leads_{disp_start}_to_{disp_end}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_bc_leads_ld_xlsx"
                            )
                    st.divider()
 
                    # Table 3
                    _ld_section_header("LESS DIALLED LEADS AFTER ASSIGNMENT")
                    st.caption("DNP stages (Call Not Picking Up / Call Not Connected) dialled less than 11 times after assignment to counsellor.")
                    df_ldc_ld = _proc_ld_caller(df_m_ld)
                    _show_caller(df_ldc_ld, "No less-dialled leads found.")
 
                    # Persist
                    st.session_state.update({
                        'ld_merged_cd'   : df_m_ld.copy(),
                        'ld_ac_cd'       : df_ac_ld,
                        'ld_bc_cd'       : df_bc_ld,
                        'ld_ldc_cd'      : df_ldc_ld,
                        'ld_generated_cd': True,
                    })
        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>📊</div>
                <div style='font-size:.9rem;font-weight:600;'>
                    Select a date range and click <b>Generate Leads Report</b>
                </div>
            </div>""", unsafe_allow_html=True)
 
    # ── TAB 2 ──────────────────────────────────────────────────────────────────
    with tab_ld2:
        if st.session_state.get('ld_generated_cd') and 'ld_merged_cd' in st.session_state:
            df_m_ld  = st.session_state['ld_merged_cd']
            df_ac_ld = st.session_state.get('ld_ac_cd',  pd.DataFrame())
            df_bc_ld = st.session_state.get('ld_bc_cd',  pd.DataFrame())
            df_ld_ld = st.session_state.get('ld_ldc_cd', pd.DataFrame())
 
            df_at_ld  = _proc_assigned_team(df_m_ld)
            df_bt_ld  = _proc_breached_team(df_m_ld)
            df_ldt_ld = _proc_ld_team(df_m_ld)
 
            _ld_section_header("🧠 GENERATED LEAD INSIGHTS")
            insights_ld = _compute_insights(df_ac_ld, df_bc_ld, df_ld_ld, df_at_ld, df_bt_ld, df_ldt_ld)
            if insights_ld:
                ic = st.columns(2)
                for i, ins in enumerate(insights_ld):
                    with ic[i % 2]:
                        st.markdown(f"""
                        <div class="ld-insight-card {ins['type']}">
                            <div style='display:flex;align-items:center;gap:.4rem;'>
                                <span class="insight-icon">{ins['icon']}</span>
                                <span style='font-size:.82rem;font-weight:700;color:var(--text-primary,#111827);'>
                                    {ins['title']}</span>
                            </div>
                            <div style='font-size:.76rem;color:var(--text-muted,#6B7280);line-height:1.5;'>
                                {ins['body']}</div>
                        </div>""", unsafe_allow_html=True)
            else:
                st.info("Not enough data to generate insights.")
 
            st.divider()
            _ld_section_header("TEAMWISE ASSIGNED LEADS DISTRIBUTION")
            _show_team(df_at_ld, "No teamwise data available.")
 
            st.divider()
            _ld_section_header("TEAMWISE POTENTIAL BREACHED LEADS AFTER ASSIGNMENT")
            st.caption("Potential leads not dialled in last 3 days / having a older follow up date.")
            _show_team(df_bt_ld, "No teamwise breached data found.")
 
            st.divider()
            _ld_section_header("TEAMWISE LESS DIALLED LEADS AFTER ASSIGNMENT")
            st.caption("DNP Leads dialled less than 11 times after assignment to counsellor")
            _show_team(df_ldt_ld, "No teamwise less-dialled data found.")
        else:
            st.markdown("""
            <div style='text-align:center;padding:6rem 1rem;opacity:.6;'>
                <div style='font-size:4rem;margin-bottom:1rem;'>🧠</div>
                <div style='font-size:.9rem;font-weight:600;'>
                    Generate a <b>Leads Report</b> first — insights will appear here automatically.
                </div>
            </div>""", unsafe_allow_html=True)

# --- MAIN APP ROUTER ---
if not st.session_state.get('password_correct', False):
    show_homepage_with_login()
else:
    # ── Shared logged-in CSS ───────────────────────────────────
    st.markdown("""
    <style>
    footer { visibility: hidden; }
    [data-testid="stStatusWidget"] { display: none !important; }
    header[data-testid="stHeader"] { background: transparent !important; }
    </style>
    """, unsafe_allow_html=True)

    # ── Apply role constraints before any dashboard runs ──────
    _apply_role_filters()

    ri   = st.session_state.get('auth_role_info', {'role': 'admin'})
    role = ri.get('role', 'admin')
    disp = ri.get('display_name', st.session_state.get('current_user', ''))

    _ROLE_LABELS = {
        'admin'        : '🛡️ Admin',
        'vertical_head': '🏢 Vertical Head',
        'trainer'      : '🎓 Sales Leader',
        'tl'           : '👑 Team Lead',
        'caller'       : '📞 Caller',
    }

    # ── Previous choice → accent colour ───────────────────────
    _prev = st.session_state.get("dashboard_choice", "Calling Metrics")
    if _prev == "Calling Metrics":
        _lc, _sc, _shc = "#F97316", "rgba(249,115,22,.9)", "rgba(249,115,22,.5)"
    elif _prev == "Revenue Metrics":
        _lc, _sc, _shc = "#10B981", "rgba(16,185,129,.9)", "rgba(16,185,129,.5)"
    else:
        _lc, _sc, _shc = "#3B82F6", "rgba(59,130,246,.9)", "rgba(59,130,246,.5)"

    # ── User info pill ─────────────────────────────────────────
    st.sidebar.markdown(f"""
    <div style='margin:.4rem 0 .6rem;padding:.55rem .7rem;
                background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);
                border-radius:10px;text-align:center;'>
        <div style='font-size:.6rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:1px;color:{_lc};margin-bottom:.2rem;'>
            {_ROLE_LABELS.get(role, role)}
        </div>
        <div style='font-size:.72rem;color:rgba(255,255,255,.75);word-break:break-all;'>
            {disp}
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Sign Out ───────────────────────────────────────────────
    if st.sidebar.button("🚪 Sign Out", key="signout_btn"):
        try:
            supa.auth.sign_out()
        except Exception:
            pass
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

    # ── Logo ───────────────────────────────────────────────────
    st.sidebar.markdown(f"""
    <div style='padding:.7rem 0 .5rem;text-align:center;
                border-bottom:1px solid rgba(128,128,128,.15);'>
        <div style='display:flex;align-items:center;justify-content:center;margin-bottom:.25rem;'>
            <span style='font-size:.95rem;font-weight:700;color:{_lc};letter-spacing:-.4px;'>LawSikho</span>
            <div style='width:1px;height:16px;margin:0 .55rem;
                        background:linear-gradient(180deg,transparent,{_sc},transparent);
                        box-shadow:0 0 5px {_shc};'></div>
            <span style='font-size:.95rem;font-weight:700;color:{_lc};letter-spacing:-.4px;'>Skill Arbitrage</span>
        </div>
        <div style='font-size:.6rem;color:{_lc};letter-spacing:1px;font-family:monospace;font-weight:600;'>
            India Learning 📖 India Earning
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Navigation ─────────────────────────────────────────────
    st.sidebar.markdown("""
    <div style='padding:.5rem 0 .2rem;text-align:center;'>
        <div style='font-size:.72rem;font-weight:700;text-transform:uppercase;
                    letter-spacing:1px;color:var(--text-muted,#6B7280);'>Dashboards Navigation</div>
    </div>
    """, unsafe_allow_html=True)

    choice = st.sidebar.selectbox(
        "Navigation",
        ["Calling Metrics", "Revenue Metrics", "Lead Metrics"],
        key="dashboard_choice",
        label_visibility="collapsed"
    )

    if choice == "Calling Metrics":
        run_calling_dashboard()
    elif choice == "Revenue Metrics":
        run_revenue_dashboard()
    else:
        run_leads_dashboard()

